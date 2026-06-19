"""Read historical IFS operation data for order, machine, and product analysis.

This script intentionally uses past operation history, not the live/ongoing
dispatch list. It reads ShopFloorWorkbenchHandling.Reference_OperationHistory,
then aggregates historical transactions into unique:

    shop order + operation + machine/resource + product

Outputs are written under data/ifs-experiments.
"""

from __future__ import annotations

import argparse
import asyncio
import json
import re
import sys
from collections import defaultdict
from collections.abc import Iterable, Mapping, Sequence
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

import httpx
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from app.core.config import Settings, get_settings  # noqa: E402
from app.integrations.ifs.client import (  # noqa: E402
    _get_all,
    _odata_string,
    _projection_url,
    obtain_access_token,
)

HISTORY_PROJECTION = "ShopFloorWorkbenchHandling"
HISTORY_ENTITY_SET = "Reference_OperationHistory"
PRODUCT_DESCRIPTION_PROJECTION = "InventoryPartInStockHandling"
PRODUCT_DESCRIPTION_ENTITY_SET = "Reference_InventoryPart"
HISTORY_SELECT_FIELDS = (
    "TransactionId",
    "OrderNo",
    "ReleaseNo",
    "SequenceNo",
    "OperationNo",
    "Contract",
    "WorkCenterNo",
    "ResourceId",
    "ResourceDescription",
    "PartNo",
    "DateApplied",
    "TransactionDate",
    "Dated",
    "TimeOfProduction",
    "QtyComplete",
    "QtyScrapped",
    "TransactionCode",
    "OperStatusCode",
    "OrderType",
)
DEFAULT_PROJECTIONS = (
    "ManufacturingPartService",
    "InventoryPartService",
    "InventoryPartInStockHandling",
    "ShopFloorWorkbenchHandling",
    "ShopOrderOperationsHandling",
    "ShopOrdersHandling",
    "WorkCenterOperationsReporting",
    "WorkCenterOperationsReportingHandling",
)
ORDER_FIELD_TOKENS = ("orderno", "shoporder", "shopord")
MACHINE_FIELD_TOKENS = (
    "preferredresourceid",
    "resourceid",
    "resourceno",
    "workcenterno",
    "machine",
    "machineno",
    "mchcode",
)
PRODUCT_FIELD_TOKENS = (
    "partno",
    "product",
    "catalogno",
    "inventorypartno",
)
AGIZ_FIELD = "Ağız"
GRAMAJ_FIELD = "Gramaj"
TRANSACTION_EXCEL_FIELDS = (
    "transaction_id",
    "date_applied",
    "transaction_date",
    "dated",
    "time_of_production",
    "order_no",
    "release_no",
    "sequence_no",
    "operation_no",
    "machine",
    "machine_source",
    "resource_id",
    "work_center_no",
    "resource_description",
    "product_part_no",
    "product_description",
    "product_name",
    AGIZ_FIELD,
    GRAMAJ_FIELD,
    "qty_complete",
    "qty_scrapped",
    "transaction_code",
    "oper_status_code",
    "order_type",
)
COMBINATION_EXCEL_FIELDS = (
    "order_no",
    "release_no",
    "sequence_no",
    "operation_no",
    "machine",
    "machine_source",
    "resource_id",
    "work_center_no",
    "resource_description",
    "product_part_no",
    "product_description",
    "product_name",
    AGIZ_FIELD,
    GRAMAJ_FIELD,
    "first_date_applied",
    "last_date_applied",
    "transaction_count",
    "qty_complete_total",
    "qty_scrapped_total",
    "transaction_codes",
)


def _today() -> date:
    return date.today()


def _default_date_from() -> str:
    return (_today() - timedelta(days=30)).isoformat()


def _default_date_to() -> str:
    return _today().isoformat()


def _stamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _text(value: Any) -> str:
    return str(value or "").strip()


def _parse_date(value: str) -> date | None:
    if not value:
        return None
    try:
        return date.fromisoformat(value[:10])
    except ValueError:
        return None


def _decimal(value: Any) -> Decimal:
    try:
        return Decimal(str(value or "0"))
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _decimal_text(value: Decimal) -> str:
    normalized = value.normalize()
    return format(normalized, "f")


def _edm_namespace(root: ElementTree.Element) -> dict[str, str]:
    namespace = "http://docs.oasis-open.org/odata/ns/edm"
    if root.findall(f".//{{{namespace}}}Schema") or root.tag == f"{{{namespace}}}Schema":
        return {"edm": namespace}
    if root.tag.startswith("{"):
        return {"edm": root.tag[1:].split("}", 1)[0]}
    return {"edm": namespace}


def _fields_matching(properties: Sequence[str], tokens: Sequence[str]) -> list[str]:
    matches: list[str] = []
    for property_name in properties:
        normalized = property_name.lower().replace("_", "")
        if any(token in normalized for token in tokens):
            matches.append(property_name)
    return matches


def classify_fields(properties: Sequence[str]) -> dict[str, list[str]]:
    return {
        "order": _fields_matching(properties, ORDER_FIELD_TOKENS),
        "machine": _fields_matching(properties, MACHINE_FIELD_TOKENS),
        "product": _fields_matching(properties, PRODUCT_FIELD_TOKENS),
    }


def _combination_score(entity_set: Mapping[str, Any]) -> int:
    groups = entity_set.get("field_groups")
    if not isinstance(groups, Mapping):
        return 0
    return sum(1 for name in ("order", "machine", "product") if groups.get(name))


def parse_metadata(xml_text: str) -> dict[str, Any]:
    root = ElementTree.fromstring(xml_text)
    ns = _edm_namespace(root)
    entity_types: dict[str, dict[str, Any]] = {}
    functions: list[dict[str, Any]] = []
    actions: list[dict[str, Any]] = []

    for schema in root.findall(".//edm:Schema", ns):
        schema_namespace = schema.attrib.get("Namespace", "")
        for entity_type in schema.findall("edm:EntityType", ns):
            entity_name = entity_type.attrib.get("Name", "")
            full_name = f"{schema_namespace}.{entity_name}" if schema_namespace else entity_name
            properties = [
                prop.attrib.get("Name", "")
                for prop in entity_type.findall("edm:Property", ns)
                if prop.attrib.get("Name")
            ]
            entity_types[full_name] = {
                "name": entity_name,
                "full_name": full_name,
                "properties": properties,
            }

        for function in schema.findall("edm:Function", ns):
            functions.append(
                {
                    "name": function.attrib.get("Name", ""),
                    "parameters": [
                        parameter.attrib.get("Name", "")
                        for parameter in function.findall("edm:Parameter", ns)
                    ],
                }
            )

        for action in schema.findall("edm:Action", ns):
            actions.append(
                {
                    "name": action.attrib.get("Name", ""),
                    "parameters": [
                        parameter.attrib.get("Name", "")
                        for parameter in action.findall("edm:Parameter", ns)
                    ],
                }
            )

    entity_sets: list[dict[str, Any]] = []
    for entity_set in root.findall(".//edm:EntitySet", ns):
        name = entity_set.attrib.get("Name", "")
        entity_type_name = entity_set.attrib.get("EntityType", "")
        entity_type = entity_types.get(entity_type_name, {})
        properties = list(entity_type.get("properties", ()))
        entity_sets.append(
            {
                "name": name,
                "entity_type": entity_type_name,
                "properties": properties,
                "field_groups": classify_fields(properties),
            }
        )

    return {
        "entity_set_count": len(entity_sets),
        "function_count": len(functions),
        "action_count": len(actions),
        "entity_sets": entity_sets,
        "functions": functions,
        "actions": actions,
    }


def month_prefixes(start: date, end: date) -> list[str]:
    prefixes: list[str] = []
    current = date(start.year, start.month, 1)
    while current <= end:
        prefixes.append(current.strftime("%Y-%m"))
        if current.month == 12:
            current = date(current.year + 1, 1, 1)
        else:
            current = date(current.year, current.month + 1, 1)
    return prefixes


def _function_prefix_filter(field_name: str, prefixes: Sequence[str]) -> str:
    clauses = [
        f"startswith({field_name},{_odata_string(prefix)})"
        for prefix in prefixes
        if prefix
    ]
    if not clauses:
        return ""
    if len(clauses) == 1:
        return clauses[0]
    return "(" + " or ".join(clauses) + ")"


def history_filter(
    settings: Settings,
    month_prefix: str,
    product_prefixes: Sequence[str],
) -> str:
    clauses = [
        # Reference_OperationHistory rejected equality/comparison filters in this
        # IFS environment, but string functions over cast(DateApplied) work.
        f"contains(Contract,{_odata_string(settings.ifs_contract)})",
        f"startswith(cast(DateApplied,Edm.String),{_odata_string(month_prefix)})",
    ]
    product_filter = _function_prefix_filter("PartNo", product_prefixes)
    if product_filter:
        clauses.append(product_filter)
    return " and ".join(clauses)


def normalize_history_row(row: Mapping[str, Any]) -> dict[str, Any]:
    resource_id = _text(row.get("ResourceId"))
    work_center_no = _text(row.get("WorkCenterNo"))
    machine = resource_id or work_center_no
    return {
        "transaction_id": row.get("TransactionId"),
        "date_applied": _text(row.get("DateApplied")),
        "transaction_date": _text(row.get("TransactionDate")),
        "dated": _text(row.get("Dated")),
        "time_of_production": _text(row.get("TimeOfProduction")),
        "order_no": _text(row.get("OrderNo")),
        "release_no": _text(row.get("ReleaseNo")),
        "sequence_no": _text(row.get("SequenceNo")),
        "operation_no": _text(row.get("OperationNo")),
        "machine": machine,
        "machine_source": "ResourceId" if resource_id else "WorkCenterNo",
        "resource_id": resource_id,
        "work_center_no": work_center_no,
        "resource_description": _text(row.get("ResourceDescription")),
        "product_part_no": _text(row.get("PartNo")),
        "product_description": "",
        "product_name": _text(row.get("PartNo")),
        AGIZ_FIELD: "",
        GRAMAJ_FIELD: "",
        "qty_complete": row.get("QtyComplete") or 0,
        "qty_scrapped": row.get("QtyScrapped") or 0,
        "transaction_code": _text(row.get("TransactionCode")),
        "oper_status_code": _text(row.get("OperStatusCode")),
        "order_type": _text(row.get("OrderType")),
    }


def is_complete_combination(row: Mapping[str, Any]) -> bool:
    return bool(row.get("order_no") and row.get("machine") and row.get("product_part_no"))


def date_in_range(row: Mapping[str, Any], start: date, end: date) -> bool:
    applied = _parse_date(_text(row.get("date_applied")))
    return bool(applied and start <= applied <= end)


def has_positive_qty(row: Mapping[str, Any]) -> bool:
    return _decimal(row.get("qty_complete")) > 0 or _decimal(row.get("qty_scrapped")) > 0


def has_machine_source(row: Mapping[str, Any], machine_source: str) -> bool:
    expected = _text(machine_source)
    return not expected or _text(row.get("machine_source")) == expected


def parse_product_name_dimensions(product_name: str) -> tuple[str, str]:
    description = product_name.split(" - ", 1)[1] if " - " in product_name else product_name
    mouth_match = re.search(r"\b(\d+(?:[.,]\d+)?)\s*MM\b", description, re.IGNORECASE)
    weight_match = re.search(r"\b(\d+(?:[.,]\d+)?)\s*GR\b", description, re.IGNORECASE)
    mouth = mouth_match.group(1).replace(",", ".") if mouth_match else ""
    weight = weight_match.group(1).replace(",", ".") if weight_match else ""
    return mouth, weight


def enrich_product_descriptions(
    rows: Sequence[Mapping[str, Any]],
    descriptions: Mapping[str, str],
) -> list[dict[str, Any]]:
    enriched: list[dict[str, Any]] = []
    for row in rows:
        product_part_no = _text(row.get("product_part_no"))
        description = _text(descriptions.get(product_part_no))
        product_name = (
            f"{product_part_no} - {description}"
            if product_part_no and description
            else product_part_no
        )
        mouth, weight = parse_product_name_dimensions(product_name)
        enriched_row = dict(row)
        enriched_row["product_description"] = description
        enriched_row["product_name"] = product_name
        enriched_row[AGIZ_FIELD] = mouth
        enriched_row[GRAMAJ_FIELD] = weight
        enriched.append(enriched_row)
    return enriched


def deduplicate_transactions(rows: Iterable[Mapping[str, Any]]) -> list[dict[str, Any]]:
    deduplicated: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row in rows:
        transaction_id = _text(row.get("transaction_id"))
        key = transaction_id or "|".join(
            _text(row.get(field))
            for field in (
                "date_applied",
                "order_no",
                "release_no",
                "sequence_no",
                "operation_no",
                "machine",
                "product_part_no",
                "transaction_code",
            )
        )
        if key in seen:
            continue
        seen.add(key)
        deduplicated.append(dict(row))
    return deduplicated


def _combo_key(row: Mapping[str, Any]) -> tuple[str, ...]:
    return (
        _text(row.get("order_no")),
        _text(row.get("release_no")),
        _text(row.get("sequence_no")),
        _text(row.get("operation_no")),
        _text(row.get("machine")),
        _text(row.get("product_part_no")),
    )


def _unique_join(values: Iterable[Any]) -> str:
    return "; ".join(dict.fromkeys(value for raw in values if (value := _text(raw))))


def aggregate_combinations(rows: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[tuple[str, ...], list[Mapping[str, Any]]] = defaultdict(list)
    for row in rows:
        groups[_combo_key(row)].append(row)

    combinations: list[dict[str, Any]] = []
    for key, group in groups.items():
        first = group[0]
        dates = sorted(
            date_value
            for row in group
            if (date_value := _text(row.get("date_applied")))
        )
        qty_complete = sum(_decimal(row.get("qty_complete")) for row in group)
        qty_scrapped = sum(_decimal(row.get("qty_scrapped")) for row in group)
        combinations.append(
            {
                "order_no": key[0],
                "release_no": key[1],
                "sequence_no": key[2],
                "operation_no": key[3],
                "machine": key[4],
                "machine_source": _text(first.get("machine_source")),
                "resource_id": _text(first.get("resource_id")),
                "work_center_no": _text(first.get("work_center_no")),
                "resource_description": _text(first.get("resource_description")),
                "product_part_no": key[5],
                "product_description": _text(first.get("product_description")),
                "product_name": _text(first.get("product_name")),
                AGIZ_FIELD: _text(first.get(AGIZ_FIELD)),
                GRAMAJ_FIELD: _text(first.get(GRAMAJ_FIELD)),
                "first_date_applied": dates[0] if dates else "",
                "last_date_applied": dates[-1] if dates else "",
                "transaction_count": len(group),
                "qty_complete_total": _decimal_text(qty_complete),
                "qty_scrapped_total": _decimal_text(qty_scrapped),
                "transaction_codes": _unique_join(
                    row.get("transaction_code") for row in group
                ),
            }
        )
    return sorted(
        combinations,
        key=lambda row: (
            row["first_date_applied"],
            row["order_no"],
            row["operation_no"],
            row["machine"],
        ),
    )


async def fetch_history_month(
    client: httpx.AsyncClient,
    settings: Settings,
    headers: Mapping[str, str],
    *,
    month_prefix: str,
    product_prefixes: Sequence[str],
    row_limit: int,
) -> list[dict[str, Any]]:
    url = _projection_url(
        settings,
        f"{HISTORY_PROJECTION}.svc/{HISTORY_ENTITY_SET}",
    )
    params = {
        "$filter": history_filter(settings, month_prefix, product_prefixes),
        "$select": ",".join(HISTORY_SELECT_FIELDS),
        "$top": str(row_limit),
    }
    return await _get_all(
        client,
        url,
        endpoint_category="past-operation-history",
        headers=headers,
        params=params,
    )


async def fetch_product_descriptions(
    client: httpx.AsyncClient,
    settings: Settings,
    headers: Mapping[str, str],
    *,
    product_prefixes: Sequence[str],
    row_limit: int,
) -> dict[str, str]:
    descriptions: dict[str, str] = {}
    url = _projection_url(
        settings,
        f"{PRODUCT_DESCRIPTION_PROJECTION}.svc/{PRODUCT_DESCRIPTION_ENTITY_SET}",
    )
    for prefix in product_prefixes:
        params = {
            "$filter": (
                f"contains(Contract,{_odata_string(settings.ifs_contract)}) "
                f"and startswith(PartNo,{_odata_string(prefix)})"
            ),
            "$select": "Contract,PartNo,Description",
            "$top": str(row_limit),
        }
        rows = await _get_all(
            client,
            url,
            endpoint_category="product-descriptions",
            headers=headers,
            params=params,
        )
        for row in rows:
            part_no = _text(row.get("PartNo"))
            description = _text(row.get("Description"))
            if part_no and description:
                descriptions[part_no] = description
    return descriptions


async def probe_projection_metadata(
    client: httpx.AsyncClient,
    settings: Settings,
    headers: Mapping[str, str],
    projection: str,
) -> dict[str, Any]:
    projection_name = projection.removesuffix(".svc")
    url = _projection_url(settings, f"{projection_name}.svc/$metadata")
    response = await client.get(url, headers=headers)
    result: dict[str, Any] = {
        "projection": projection_name,
        "metadata_status": response.status_code,
    }
    if response.status_code >= 400:
        result["error"] = response.text[:500]
        return result
    metadata = parse_metadata(response.text)
    result.update(
        {
            "entity_set_count": metadata["entity_set_count"],
            "function_count": metadata["function_count"],
            "action_count": metadata["action_count"],
            "combination_entity_sets": [
                {
                    "name": entity_set["name"],
                    "field_groups": entity_set["field_groups"],
                }
                for entity_set in metadata["entity_sets"]
                if _combination_score(entity_set) == 3
            ][:20],
        }
    )
    return result


def _excel_value(field_name: str, value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    if value is None:
        return None
    if field_name.endswith("_count") or field_name in {
        "transaction_id",
        "operation_no",
        "metadata_status",
        "entity_set_count",
        "function_count",
        "action_count",
        "direct_candidate_count",
    }:
        try:
            return int(value)
        except (TypeError, ValueError):
            return value
    if field_name.startswith("qty_") or field_name.endswith("_total"):
        try:
            return float(_decimal(value))
        except (TypeError, ValueError):
            return value
    if field_name in {AGIZ_FIELD, GRAMAJ_FIELD}:
        try:
            number = float(str(value).replace(",", "."))
        except (TypeError, ValueError):
            return value
        return int(number) if number.is_integer() else number
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _style_table_sheet(
    sheet: Any,
    *,
    fieldnames: Sequence[str],
    row_count: int,
    table_name: str,
) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    if row_count:
        last_column = get_column_letter(len(fieldnames))
        table = Table(
            displayName=table_name,
            ref=f"A1:{last_column}{row_count + 1}",
        )
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        sheet.add_table(table)

    for column_index, field_name in enumerate(fieldnames, start=1):
        letter = get_column_letter(column_index)
        values = [field_name]
        for row in sheet.iter_rows(
            min_row=2,
            max_row=min(row_count + 1, 502),
            min_col=column_index,
            max_col=column_index,
            values_only=True,
        ):
            values.append("" if row[0] is None else str(row[0]))
        width = min(max(max(len(value) for value in values) + 2, 10), 38)
        sheet.column_dimensions[letter].width = width


def _write_table_sheet(
    workbook: Workbook,
    *,
    title: str,
    fieldnames: Sequence[str],
    rows: Sequence[Mapping[str, Any]],
    table_name: str,
) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(list(fieldnames))
    for row in rows:
        sheet.append([
            _excel_value(field_name, row.get(field_name, ""))
            for field_name in fieldnames
        ])
    _style_table_sheet(
        sheet,
        fieldnames=fieldnames,
        row_count=len(rows),
        table_name=table_name,
    )


def _projection_status_rows(
    projection_reports: Sequence[Mapping[str, Any]],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for report in projection_reports:
        candidates = report.get("combination_entity_sets") or []
        candidate_names = [
            str(candidate.get("name"))
            for candidate in candidates
            if isinstance(candidate, Mapping) and candidate.get("name")
        ]
        rows.append(
            {
                "projection": report.get("projection"),
                "metadata_status": report.get("metadata_status"),
                "entity_set_count": report.get("entity_set_count"),
                "function_count": report.get("function_count"),
                "action_count": report.get("action_count"),
                "direct_candidate_count": len(candidate_names),
                "direct_candidate_entity_sets": "; ".join(candidate_names),
                "error": report.get("error", ""),
            }
        )
    return rows


def write_excel_workbook(path: Path, summary: Mapping[str, Any]) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"

    analysis = summary["analysis"]
    summary_rows = [
        ("Generated at", summary["generated_at"]),
        ("Date from", summary["date_from"]),
        ("Date to", summary["date_to"]),
        ("Can get past combination", analysis["can_get_past_combination"]),
        ("Historical rows fetched", summary["raw_history_row_count"]),
        ("Transactions after filters", summary["transaction_count"]),
        ("Unique combinations", summary["combination_count"]),
        ("Complete combinations", analysis["complete_combination_count"]),
        ("Unique orders", analysis["unique_orders"]),
        ("Unique machines", analysis["unique_machines"]),
        ("Unique products", analysis["unique_products"]),
        ("Month row limit", summary["month_row_limit"]),
        ("Part row limit", summary["part_row_limit"]),
        ("Machine source filter", summary["machine_source_filter"] or "none"),
        ("Product descriptions loaded", summary["product_description_count"]),
        (
            "Possible truncated months",
            ", ".join(summary["possible_truncated_months"]) or "none",
        ),
        ("History projection", summary["history_projection"]),
        ("History entity set", summary["history_entity_set"]),
        ("Product description projection", summary["product_description_projection"]),
        ("Product description entity set", summary["product_description_entity_set"]),
        ("Product prefixes", ", ".join(summary["product_prefixes"])),
    ]
    summary_sheet.append(["Metric", "Value"])
    for row in summary_rows:
        summary_sheet.append(list(row))
    summary_sheet.append([])
    summary_sheet.append(["Notes", ""])
    for note in analysis["notes"]:
        summary_sheet.append([note, ""])

    for cell in summary_sheet[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
    summary_sheet.freeze_panes = "A2"
    summary_sheet.column_dimensions["A"].width = 34
    summary_sheet.column_dimensions["B"].width = 80
    for row in summary_sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    _write_table_sheet(
        workbook,
        title="Combinations",
        fieldnames=COMBINATION_EXCEL_FIELDS,
        rows=summary["combinations"],
        table_name="PastCombinations",
    )
    _write_table_sheet(
        workbook,
        title="Transactions",
        fieldnames=TRANSACTION_EXCEL_FIELDS,
        rows=summary["transactions"],
        table_name="PastTransactions",
    )
    _write_table_sheet(
        workbook,
        title="Projection Status",
        fieldnames=(
            "projection",
            "metadata_status",
            "entity_set_count",
            "function_count",
            "action_count",
            "direct_candidate_count",
            "direct_candidate_entity_sets",
            "error",
        ),
        rows=_projection_status_rows(summary["projection_reports"]),
        table_name="ProjectionStatus",
    )

    workbook.save(path)


def write_markdown(path: Path, summary: Mapping[str, Any]) -> None:
    analysis = summary["analysis"]
    lines = [
        "# IFS Past Job Order / Machine / Product Experiment",
        "",
        f"Generated at: {summary['generated_at']}",
        f"Date range: {summary['date_from']} to {summary['date_to']}",
        "",
        "## Result",
        "",
        f"Can get past combination: **{analysis['can_get_past_combination']}**",
        "",
        f"- Historical transactions fetched: {summary['raw_history_row_count']}",
        f"- Transactions after date/quality filters: {summary['transaction_count']}",
        f"- Unique combinations: {summary['combination_count']}",
        f"- Complete combinations: {analysis['complete_combination_count']}",
        f"- Month row limit: {summary['month_row_limit']}",
        f"- Machine source filter: {summary['machine_source_filter'] or 'none'}",
        f"- Product descriptions loaded: {summary['product_description_count']}",
        f"- Possible truncated months: {', '.join(summary['possible_truncated_months']) or 'none'}",
        "",
        "## Notes",
        "",
    ]
    lines.extend(f"- {note}" for note in analysis["notes"])
    lines.extend(
        [
            "",
            "## Output Files",
            "",
            f"- Excel workbook: `{summary['excel_workbook']}`",
            f"- JSON: `{summary['json_report']}`",
        ]
    )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def analyze(
    transactions: Sequence[Mapping[str, Any]],
    combinations: Sequence[Mapping[str, Any]],
    projection_reports: Sequence[Mapping[str, Any]],
) -> dict[str, Any]:
    complete_combinations = [
        row
        for row in combinations
        if row.get("order_no") and row.get("machine") and row.get("product_part_no")
    ]
    available = [
        report.get("projection")
        for report in projection_reports
        if int(report.get("metadata_status") or 0) < 400
    ]
    blocked = [
        {
            "projection": report.get("projection"),
            "status": report.get("metadata_status"),
            "error": report.get("error"),
        }
        for report in projection_reports
        if int(report.get("metadata_status") or 0) >= 400
    ]
    direct_candidates = [
        report.get("projection")
        for report in projection_reports
        if report.get("combination_entity_sets")
    ]
    notes = [
        (
            "Past rows come from "
            "ShopFloorWorkbenchHandling.Reference_OperationHistory, not from "
            "the ongoing dispatch list."
        ),
        (
            "Reference_OperationHistory provides OrderNo, OperationNo, "
            "ResourceId/WorkCenterNo, PartNo, DateApplied, QtyComplete, and "
            "QtyScrapped."
        ),
        (
            "Product names are enriched from "
            "InventoryPartInStockHandling.Reference_InventoryPart as "
            "'PartNo - Description'."
        ),
        (
            "The default run keeps only MM-PET products and rows whose "
            "machine_source is ResourceId."
        ),
        (
            "In this IFS environment, equality/comparison filters on this "
            "history entity were rejected, so the script filters months with "
            "startswith(cast(DateApplied,Edm.String), 'YYYY-MM') and applies "
            "the exact date range client-side."
        ),
        (
            "The history entity did not return @odata.nextLink in paging tests, "
            "so the script requests a high $top per month and reports possible "
            "truncated months when a month reaches that cap."
        ),
    ]
    if blocked:
        blocked_names = ", ".join(
            f"{item['projection']}({item['status']})" for item in blocked
        )
        notes.append(f"Blocked/missing metadata projections: {blocked_names}.")
    if direct_candidates:
        notes.append(
            "Metadata with direct order/machine/product fields was found in: "
            + ", ".join(str(name) for name in direct_candidates)
            + "."
        )

    return {
        "can_get_past_combination": "yes" if complete_combinations else "no",
        "complete_combination_count": len(complete_combinations),
        "available_metadata_projections": available,
        "blocked_or_missing_metadata_projections": blocked,
        "direct_metadata_candidate_projections": direct_candidates,
        "unique_orders": len({row.get("order_no") for row in transactions}),
        "unique_machines": len({row.get("machine") for row in transactions}),
        "unique_products": len({row.get("product_part_no") for row in transactions}),
        "notes": notes,
    }


async def run_experiment(args: argparse.Namespace) -> dict[str, Any]:
    settings = get_settings(validate=False)
    date_from = date.fromisoformat(args.date_from)
    date_to = date.fromisoformat(args.date_to)
    if date_to < date_from:
        raise ValueError("--date-to must be on or after --date-from")

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    async with httpx.AsyncClient(timeout=args.timeout) as client:
        token = await obtain_access_token(settings, client=client)
        headers = {"Authorization": f"Bearer {token}"}
        history_groups = []
        month_counts: dict[str, int] = {}
        for month_prefix in month_prefixes(date_from, date_to):
            month_rows = await fetch_history_month(
                client,
                settings,
                headers,
                month_prefix=month_prefix,
                product_prefixes=args.product_prefix,
                row_limit=args.month_row_limit,
            )
            month_counts[month_prefix] = len(month_rows)
            history_groups.append(month_rows)
        raw_rows = [row for group in history_groups for row in group]

        projection_reports = []
        for projection in args.projection:
            projection_reports.append(
                await probe_projection_metadata(client, settings, headers, projection)
            )
        product_descriptions = await fetch_product_descriptions(
            client,
            settings,
            headers,
            product_prefixes=args.product_prefix,
            row_limit=args.part_row_limit,
        )

    transactions = [normalize_history_row(row) for row in raw_rows]
    transactions = [
        row
        for row in transactions
        if date_in_range(row, date_from, date_to) and is_complete_combination(row)
    ]
    transactions = [
        row
        for row in transactions
        if has_machine_source(row, args.machine_source)
    ]
    if not args.include_zero_qty:
        transactions = [row for row in transactions if has_positive_qty(row)]
    transactions = deduplicate_transactions(transactions)
    transactions = enrich_product_descriptions(transactions, product_descriptions)
    combinations = aggregate_combinations(transactions)

    stamp = _stamp()
    excel_workbook = output_dir / f"ifs_past_combinations_{stamp}.xlsx"
    json_report = output_dir / f"ifs_past_combinations_{stamp}.json"
    markdown_report = output_dir / f"ifs_past_combinations_{stamp}_analysis.md"

    analysis = analyze(transactions, combinations, projection_reports)
    summary: dict[str, Any] = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "date_from": date_from.isoformat(),
        "date_to": date_to.isoformat(),
        "history_projection": HISTORY_PROJECTION,
        "history_entity_set": HISTORY_ENTITY_SET,
        "history_selected_fields": list(HISTORY_SELECT_FIELDS),
        "product_prefixes": list(args.product_prefix),
        "machine_source_filter": args.machine_source,
        "product_description_projection": PRODUCT_DESCRIPTION_PROJECTION,
        "product_description_entity_set": PRODUCT_DESCRIPTION_ENTITY_SET,
        "product_description_count": len(product_descriptions),
        "month_row_limit": args.month_row_limit,
        "part_row_limit": args.part_row_limit,
        "month_row_counts": month_counts,
        "possible_truncated_months": [
            month
            for month, row_count in month_counts.items()
            if row_count >= args.month_row_limit
        ],
        "raw_history_row_count": len(raw_rows),
        "transaction_count": len(transactions),
        "combination_count": len(combinations),
        "transactions": transactions,
        "combinations": combinations,
        "projection_reports": projection_reports,
        "analysis": analysis,
        "excel_workbook": str(excel_workbook),
        "json_report": str(json_report),
        "markdown_report": str(markdown_report),
    }

    write_excel_workbook(excel_workbook, summary)
    json_report.write_text(
        json.dumps(summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    write_markdown(markdown_report, summary)
    return summary


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Fetch historical IFS operation rows and aggregate past "
            "job-order/machine/product combinations."
        )
    )
    parser.add_argument("--date-from", default=_default_date_from())
    parser.add_argument("--date-to", default=_default_date_to())
    parser.add_argument(
        "--product-prefix",
        action="append",
        default=["MM-PET"],
        help="Product PartNo prefix to include. Repeat for multiple prefixes.",
    )
    parser.add_argument(
        "--machine-source",
        default="ResourceId",
        help=(
            "Keep only rows with this machine_source value. Use an empty string "
            "to keep all machine sources."
        ),
    )
    parser.add_argument(
        "--include-zero-qty",
        action="store_true",
        help="Include historical rows where both QtyComplete and QtyScrapped are zero.",
    )
    parser.add_argument(
        "--month-row-limit",
        type=int,
        default=10000,
        help=(
            "Maximum rows requested per YYYY-MM bucket. Increase this if the "
            "report lists a possible truncated month."
        ),
    )
    parser.add_argument(
        "--part-row-limit",
        type=int,
        default=10000,
        help="Maximum product description rows requested per product prefix.",
    )
    parser.add_argument(
        "--projection",
        action="append",
        default=list(DEFAULT_PROJECTIONS),
        help="Projection metadata to probe. Can be repeated.",
    )
    parser.add_argument(
        "--output-dir",
        default=str(ROOT_DIR / "data" / "ifs-experiments"),
    )
    parser.add_argument("--timeout", type=float, default=60.0)
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    try:
        summary = asyncio.run(run_experiment(args))
    except Exception as exc:
        print(f"IFS past-data experiment failed: {exc}", file=sys.stderr)
        return 1

    analysis = summary["analysis"]
    print("IFS past job-order/machine/product experiment complete")
    print(f"Date range: {summary['date_from']} to {summary['date_to']}")
    print(f"Can get past combination: {analysis['can_get_past_combination']}")
    print(f"Transactions: {summary['transaction_count']}")
    print(f"Unique combinations: {summary['combination_count']}")
    print(f"Excel workbook: {summary['excel_workbook']}")
    print(f"JSON: {summary['json_report']}")
    print(f"Markdown: {summary['markdown_report']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
