"""Forecast PET cycle times from process entries enriched with IFS history.

Data responsibilities:
- Process DB entries provide the measured/realized cycle time in seconds.
- IFS operation history provides reliable product text, mouth size, gram, and group.
- The DB cycle table provides the current optimum cycle time benchmark.
"""

from __future__ import annotations

import argparse
import asyncio
import json
import sys
from collections import defaultdict
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import httpx
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from sqlalchemy import Integer, cast, func, select

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from app.config import Settings, get_settings  # noqa: E402
from app.database import create_session  # noqa: E402
from app.features.cycle_reports.service import (  # noqa: E402
    _lookup_optimum_cycle,
    _read_cycle_table,
    normalize_machine_group,
)
from app.integrations.ifs_client import obtain_access_token  # noqa: E402
from app.models import Entry  # noqa: E402
from scripts.experiment_ifs_past_job_order_machine_product import (  # noqa: E402
    AGIZ_FIELD,
    GRAMAJ_FIELD,
    deduplicate_transactions,
    enrich_product_descriptions,
    fetch_history_month,
    fetch_product_descriptions,
    has_machine_source,
    has_positive_qty,
    month_prefixes,
    normalize_history_row,
)

FORECAST_FIELDS = (
    "machine_group",
    "machine",
    "work_center_no",
    "Ağız",
    "Gramaj",
    "product_part_no",
    "product_name",
    "sample_count",
    "entry_count",
    "unique_order_count",
    "avg_actual_cycle_seconds",
    "median_actual_cycle_seconds",
    "p75_actual_cycle_seconds",
    "p90_actual_cycle_seconds",
    "min_actual_cycle_seconds",
    "max_actual_cycle_seconds",
    "current_optimum_cycle_seconds",
    "forecast_cycle_seconds",
    "forecast_basis",
    "median_vs_optimum_seconds",
    "median_vs_optimum_percent",
    "forecast_vs_optimum_seconds",
    "forecast_vs_optimum_percent",
    "first_process_date",
    "last_process_date",
)

MATCHED_ENTRY_FIELDS = (
    "entry_id",
    "process_date",
    "machine",
    "work_order",
    "actual_cycle_seconds",
    "cycle_quality",
    "current_optimum_cycle_seconds",
    "actual_vs_optimum_seconds",
    "actual_vs_optimum_percent",
    "match_status",
    "machine_group",
    "work_center_no",
    "product_part_no",
    "product_name",
    "Ağız",
    "Gramaj",
    "process_product_text",
    "total_cavity_count",
    "active_cavity_count",
)

UNMATCHED_ENTRY_FIELDS = (
    "entry_id",
    "process_date",
    "machine",
    "work_order",
    "actual_cycle_seconds",
    "match_status",
    "process_product_text",
    "total_cavity_count",
    "active_cavity_count",
)


@dataclass(frozen=True)
class ProcessEntryRow:
    entry_id: int
    process_date: str
    machine: str
    work_order: str
    actual_cycle_seconds: Decimal
    process_product_text: str
    total_cavity_count: str
    active_cavity_count: str


def _today() -> date:
    return date.today()


def _stamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _text(value: Any) -> str:
    return str(value or "").strip()


def _decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int):
        return Decimal(value)
    if isinstance(value, float):
        return Decimal(str(value))
    text = _text(value).replace(",", ".")
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _decimal_text(value: Decimal | None, places: str = "0.01") -> str:
    if value is None:
        return ""
    return format(value.quantize(Decimal(places)).normalize(), "f")


def _identifier(value: Any) -> str:
    text = _text(value)
    if text.endswith(".0"):
        return text[:-2]
    return text


def _parse_iso_date(value: str) -> date | None:
    try:
        return date.fromisoformat(value)
    except ValueError:
        return None


def _default_date_range(settings: Settings) -> tuple[str, str]:
    with create_session(settings) as session:
        row = session.execute(
            select(func.min(Entry.process_date), func.max(Entry.process_date)).where(
                Entry.process_date.is_not(None),
                Entry.process_date != "",
            )
        ).one()
    if row[0] and row[1]:
        return str(row[0]), str(row[1])
    today = _today()
    return (today - timedelta(days=30)).isoformat(), today.isoformat()


def _read_process_entries(
    settings: Settings,
    *,
    date_from: date,
    date_to: date,
) -> list[ProcessEntryRow]:
    with create_session(settings) as session:
        entries = session.scalars(
            select(Entry)
            .where(
                Entry.process_date >= date_from.isoformat(),
                Entry.process_date <= date_to.isoformat(),
                Entry.col_l.is_not(None),
                Entry.col_l != "",
                Entry.col_h.is_not(None),
                Entry.col_h != "",
                Entry.col_f.is_not(None),
                Entry.col_f != "",
            )
            .order_by(
                Entry.process_date.asc(),
                cast(Entry.machine_code, Integer).asc(),
                Entry.id.asc(),
            )
        ).all()

    rows: list[ProcessEntryRow] = []
    for entry in entries:
        actual_cycle = _decimal(entry.col_l)
        if actual_cycle is None:
            continue
        rows.append(
            ProcessEntryRow(
                entry_id=entry.id,
                process_date=_text(entry.process_date),
                machine=_identifier(entry.col_f),
                work_order=_identifier(entry.col_h),
                actual_cycle_seconds=actual_cycle,
                process_product_text=_text(entry.col_g),
                total_cavity_count=_text(entry.col_j),
                active_cavity_count=_text(entry.col_k),
            )
        )
    return rows


async def _fetch_ifs_dimension_rows(
    settings: Settings,
    *,
    date_from: date,
    date_to: date,
    product_prefixes: Sequence[str],
    machine_source: str,
    month_row_limit: int,
    part_row_limit: int,
    timeout: float,
) -> list[dict[str, Any]]:
    async with httpx.AsyncClient(timeout=timeout) as client:
        token = await obtain_access_token(settings, client=client)
        headers = {"Authorization": f"Bearer {token}"}
        history_groups = []
        for month_prefix in month_prefixes(date_from, date_to):
            history_groups.append(
                await fetch_history_month(
                    client,
                    settings,
                    headers,
                    month_prefix=month_prefix,
                    product_prefixes=product_prefixes,
                    row_limit=month_row_limit,
                )
            )
        product_descriptions = await fetch_product_descriptions(
            client,
            settings,
            headers,
            product_prefixes=product_prefixes,
            row_limit=part_row_limit,
        )

    raw_rows = [row for group in history_groups for row in group]
    rows = [normalize_history_row(row) for row in raw_rows]
    rows = [
        row
        for row in rows
        if has_machine_source(row, machine_source) and has_positive_qty(row)
    ]
    rows = deduplicate_transactions(rows)
    return enrich_product_descriptions(rows, product_descriptions)


def _unique_dimension_options(
    rows: Sequence[Mapping[str, Any]],
) -> dict[tuple[str, str], list[dict[str, Any]]]:
    grouped: dict[tuple[str, str], dict[tuple[str, str, str, str], dict[str, Any]]] = (
        defaultdict(dict)
    )
    for row in rows:
        order_no = _identifier(row.get("order_no"))
        machine = _identifier(row.get("machine"))
        if not order_no or not machine:
            continue
        option_key = (
            _text(row.get("product_part_no")),
            _text(row.get("product_name")),
            _text(row.get(AGIZ_FIELD)),
            _text(row.get(GRAMAJ_FIELD)),
        )
        if not all(option_key):
            continue
        grouped[(order_no, machine)][option_key] = {
            "work_order": order_no,
            "machine": machine,
            "work_center_no": _text(row.get("work_center_no")),
            "machine_group": normalize_machine_group(row.get("work_center_no")),
            "product_part_no": option_key[0],
            "product_name": option_key[1],
            "Ağız": option_key[2],
            "Gramaj": option_key[3],
            "resource_description": _text(row.get("resource_description")),
        }
    return {
        key: list(options.values())
        for key, options in grouped.items()
    }


def _match_dimension(
    process_row: ProcessEntryRow,
    dimension_index: Mapping[tuple[str, str], Sequence[Mapping[str, Any]]],
) -> tuple[Mapping[str, Any] | None, str]:
    exact_options = list(
        dimension_index.get((process_row.work_order, process_row.machine), ())
    )
    if len(exact_options) == 1:
        return exact_options[0], "exact_order_machine"
    if len(exact_options) > 1:
        return None, "ambiguous_order_machine"
    return None, "no_ifs_match"


def _percentile(values: Sequence[Decimal], percentile: Decimal) -> Decimal | None:
    if not values:
        return None
    ordered = sorted(values)
    if len(ordered) == 1:
        return ordered[0]
    rank = (Decimal(len(ordered) - 1) * percentile) / Decimal("100")
    lower_index = int(rank.to_integral_value(rounding="ROUND_FLOOR"))
    upper_index = int(rank.to_integral_value(rounding="ROUND_CEILING"))
    if lower_index == upper_index:
        return ordered[lower_index]
    fraction = rank - Decimal(lower_index)
    return ordered[lower_index] + (ordered[upper_index] - ordered[lower_index]) * fraction


def _average(values: Sequence[Decimal]) -> Decimal | None:
    if not values:
        return None
    return sum(values) / Decimal(len(values))


def _ratio_delta(value: Decimal | None, benchmark: Decimal | None) -> Decimal | None:
    if value is None or benchmark is None or benchmark == 0:
        return None
    return (value / benchmark) - Decimal("1")


def _lookup_optimum(
    cycle_entries: Any,
    *,
    work_center_no: str,
    machine: str,
    agiz: str,
    gramaj: str,
) -> tuple[Decimal | None, str | None]:
    neck = _decimal(agiz)
    gram = _decimal(gramaj)
    if neck is None or gram is None:
        return None, "Ağız/Gramaj bulunamadı"
    return _lookup_optimum_cycle(
        cycle_entries,
        machine_group=work_center_no,
        machine_no=machine,
        neck_diameter=neck,
        gram=gram,
    )


def _cycle_quality(actual_cycle: Decimal, optimum: Decimal | None) -> str:
    if actual_cycle <= 0:
        return "invalid_non_positive"
    if actual_cycle > Decimal("120"):
        return "outlier_over_120s"
    if optimum is not None and optimum > 0 and actual_cycle > optimum * Decimal("3"):
        return "outlier_over_3x_optimum"
    return "usable"


def _build_entry_rows(
    process_rows: Sequence[ProcessEntryRow],
    dimension_index: Mapping[tuple[str, str], Sequence[Mapping[str, Any]]],
    cycle_entries: Any,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    matched: list[dict[str, Any]] = []
    unmatched: list[dict[str, Any]] = []

    for process_row in process_rows:
        dimension, match_status = _match_dimension(process_row, dimension_index)
        if dimension is None:
            unmatched.append(
                {
                    "entry_id": process_row.entry_id,
                    "process_date": process_row.process_date,
                    "machine": process_row.machine,
                    "work_order": process_row.work_order,
                    "actual_cycle_seconds": process_row.actual_cycle_seconds,
                    "match_status": match_status,
                    "process_product_text": process_row.process_product_text,
                    "total_cavity_count": process_row.total_cavity_count,
                    "active_cavity_count": process_row.active_cavity_count,
                }
            )
            continue

        optimum, optimum_status = _lookup_optimum(
            cycle_entries,
            work_center_no=_text(dimension.get("work_center_no")),
            machine=process_row.machine,
            agiz=_text(dimension.get("Ağız")),
            gramaj=_text(dimension.get("Gramaj")),
        )
        delta = (
            process_row.actual_cycle_seconds - optimum
            if optimum is not None
            else None
        )
        cycle_quality = _cycle_quality(process_row.actual_cycle_seconds, optimum)
        row = {
            "entry_id": process_row.entry_id,
            "process_date": process_row.process_date,
            "machine": process_row.machine,
            "work_order": process_row.work_order,
            "actual_cycle_seconds": process_row.actual_cycle_seconds,
            "cycle_quality": cycle_quality,
            "current_optimum_cycle_seconds": optimum,
            "actual_vs_optimum_seconds": delta,
            "actual_vs_optimum_percent": _ratio_delta(
                process_row.actual_cycle_seconds,
                optimum,
            ),
            "match_status": optimum_status or match_status,
            "machine_group": dimension.get("machine_group"),
            "work_center_no": dimension.get("work_center_no"),
            "product_part_no": dimension.get("product_part_no"),
            "product_name": dimension.get("product_name"),
            "Ağız": dimension.get("Ağız"),
            "Gramaj": dimension.get("Gramaj"),
            "process_product_text": process_row.process_product_text,
            "total_cavity_count": process_row.total_cavity_count,
            "active_cavity_count": process_row.active_cavity_count,
        }
        matched.append(row)
    return matched, unmatched


def _build_forecast_rows(matched_rows: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[tuple[str, ...], list[Mapping[str, Any]]] = defaultdict(list)
    for row in matched_rows:
        if _text(row.get("cycle_quality")) != "usable":
            continue
        key = (
            _text(row.get("machine_group")),
            _text(row.get("machine")),
            _text(row.get("work_center_no")),
            _text(row.get("Ağız")),
            _text(row.get("Gramaj")),
            _text(row.get("product_part_no")),
            _text(row.get("product_name")),
        )
        groups[key].append(row)

    forecast_rows: list[dict[str, Any]] = []
    for key, rows in groups.items():
        actuals = [
            value
            for row in rows
            if (value := _decimal(row.get("actual_cycle_seconds"))) is not None
        ]
        optimums = [
            value
            for row in rows
            if (value := _decimal(row.get("current_optimum_cycle_seconds"))) is not None
        ]
        dates = sorted(_text(row.get("process_date")) for row in rows if row.get("process_date"))
        avg_actual = _average(actuals)
        median_actual = _percentile(actuals, Decimal("50"))
        p75_actual = _percentile(actuals, Decimal("75"))
        p90_actual = _percentile(actuals, Decimal("90"))
        min_actual = min(actuals) if actuals else None
        max_actual = max(actuals) if actuals else None
        optimum = optimums[0] if optimums and len(set(optimums)) == 1 else None
        if len(actuals) >= 5:
            forecast = p75_actual
            basis = "p75_realized_cycle"
        else:
            forecast = median_actual
            basis = "median_realized_cycle_low_sample"
        median_delta = (
            median_actual - optimum
            if median_actual is not None and optimum is not None
            else None
        )
        forecast_delta = (
            forecast - optimum
            if forecast is not None and optimum is not None
            else None
        )
        forecast_rows.append(
            {
                "machine_group": key[0],
                "machine": key[1],
                "work_center_no": key[2],
                "Ağız": key[3],
                "Gramaj": key[4],
                "product_part_no": key[5],
                "product_name": key[6],
                "sample_count": len(actuals),
                "entry_count": len(rows),
                "unique_order_count": len({_text(row.get("work_order")) for row in rows}),
                "avg_actual_cycle_seconds": avg_actual,
                "median_actual_cycle_seconds": median_actual,
                "p75_actual_cycle_seconds": p75_actual,
                "p90_actual_cycle_seconds": p90_actual,
                "min_actual_cycle_seconds": min_actual,
                "max_actual_cycle_seconds": max_actual,
                "current_optimum_cycle_seconds": optimum,
                "forecast_cycle_seconds": forecast,
                "forecast_basis": basis,
                "median_vs_optimum_seconds": median_delta,
                "median_vs_optimum_percent": _ratio_delta(median_actual, optimum),
                "forecast_vs_optimum_seconds": forecast_delta,
                "forecast_vs_optimum_percent": _ratio_delta(forecast, optimum),
                "first_process_date": dates[0] if dates else "",
                "last_process_date": dates[-1] if dates else "",
            }
        )

    return sorted(
        forecast_rows,
        key=lambda row: (
            _text(row.get("machine_group")),
            _text(row.get("machine")),
            _text(row.get("product_part_no")),
        ),
    )


def _excel_value(field_name: str, value: Any) -> Any:
    if isinstance(value, Decimal):
        if field_name.endswith("_percent"):
            return float(value)
        return float(value)
    if value is None:
        return None
    if field_name.endswith("_count") or field_name in {"entry_id"}:
        try:
            return int(value)
        except (TypeError, ValueError):
            return value
    if field_name in {"Ağız", "Gramaj"} or field_name.endswith("_seconds"):
        number = _decimal(value)
        return float(number) if number is not None else value
    return value


def _style_table_sheet(
    sheet: Any,
    *,
    fieldnames: Sequence[str],
    row_count: int,
    table_name: str,
) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    if row_count:
        last_column = get_column_letter(len(fieldnames))
        table = Table(displayName=table_name, ref=f"A1:{last_column}{row_count + 1}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)

    for column_index, field_name in enumerate(fieldnames, start=1):
        letter = get_column_letter(column_index)
        values = [field_name]
        for row in sheet.iter_rows(
            min_row=2,
            max_row=min(row_count + 1, 302),
            min_col=column_index,
            max_col=column_index,
            values_only=True,
        ):
            values.append("" if row[0] is None else str(row[0]))
        sheet.column_dimensions[letter].width = min(
            max(max(len(value) for value in values) + 2, 10),
            46,
        )

    for field_name in fieldnames:
        col = fieldnames.index(field_name) + 1
        letter = get_column_letter(col)
        if field_name.endswith("_percent"):
            for cell in sheet[f"{letter}2:{letter}{row_count + 1}"]:
                cell[0].number_format = "0.0%"
        elif field_name.endswith("_seconds") or field_name in {"Ağız", "Gramaj"}:
            for cell in sheet[f"{letter}2:{letter}{row_count + 1}"]:
                cell[0].number_format = "0.00"


def _write_table_sheet(
    workbook: Workbook,
    *,
    title: str,
    fieldnames: Sequence[str],
    rows: Sequence[Mapping[str, Any]],
    table_name: str,
) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(list(fieldnames))
    for row in rows:
        sheet.append([
            _excel_value(field_name, row.get(field_name, ""))
            for field_name in fieldnames
        ])
    _style_table_sheet(
        sheet,
        fieldnames=fieldnames,
        row_count=len(rows),
        table_name=table_name,
    )


def _write_workbook(path: Path, summary: Mapping[str, Any]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Summary"
    rows = [
        ("Generated at", summary["generated_at"]),
        ("Date from", summary["date_from"]),
        ("Date to", summary["date_to"]),
        ("Process entries", summary["process_entry_count"]),
        ("Matched entries", summary["matched_entry_count"]),
        ("Unmatched entries", summary["unmatched_entry_count"]),
        ("Forecast combinations", summary["forecast_count"]),
        ("IFS dimension rows", summary["ifs_dimension_row_count"]),
        ("Product prefixes", ", ".join(summary["product_prefixes"])),
        ("Machine source", summary["machine_source"]),
        ("Forecast rule", "p75 realized cycle if sample >= 5, otherwise median"),
    ]
    sheet.append(["Metric", "Value"])
    for row in rows:
        sheet.append(list(row))
    sheet.append([])
    sheet.append(["Notes", ""])
    for note in summary["notes"]:
        sheet.append([note, ""])
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
    sheet.freeze_panes = "A2"
    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 90
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    _write_table_sheet(
        workbook,
        title="Forecast",
        fieldnames=FORECAST_FIELDS,
        rows=summary["forecast_rows"],
        table_name="CycleForecast",
    )
    _write_table_sheet(
        workbook,
        title="Matched Entries",
        fieldnames=MATCHED_ENTRY_FIELDS,
        rows=summary["matched_rows"],
        table_name="MatchedEntries",
    )
    _write_table_sheet(
        workbook,
        title="Unmatched Entries",
        fieldnames=UNMATCHED_ENTRY_FIELDS,
        rows=summary["unmatched_rows"],
        table_name="UnmatchedEntries",
    )
    workbook.save(path)


async def run_forecast(args: argparse.Namespace) -> dict[str, Any]:
    settings = get_settings(validate=False)
    default_from, default_to = _default_date_range(settings)
    date_from = _parse_iso_date(args.date_from or default_from)
    date_to = _parse_iso_date(args.date_to or default_to)
    if date_from is None or date_to is None:
        raise ValueError("--date-from and --date-to must be YYYY-MM-DD")
    if date_to < date_from:
        raise ValueError("--date-to must be on or after --date-from")

    process_rows = _read_process_entries(
        settings,
        date_from=date_from,
        date_to=date_to,
    )
    ifs_rows = await _fetch_ifs_dimension_rows(
        settings,
        date_from=date_from,
        date_to=date_to,
        product_prefixes=args.product_prefix,
        machine_source=args.machine_source,
        month_row_limit=args.month_row_limit,
        part_row_limit=args.part_row_limit,
        timeout=args.timeout,
    )
    dimension_index = _unique_dimension_options(ifs_rows)
    cycle_entries = _read_cycle_table(settings)
    matched_rows, unmatched_rows = _build_entry_rows(
        process_rows,
        dimension_index,
        cycle_entries,
    )
    forecast_rows = _build_forecast_rows(matched_rows)

    stamp = _stamp()
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    workbook_path = output_dir / f"cycle_time_forecast_{stamp}.xlsx"
    json_path = output_dir / f"cycle_time_forecast_{stamp}.json"
    summary: dict[str, Any] = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "date_from": date_from.isoformat(),
        "date_to": date_to.isoformat(),
        "product_prefixes": list(args.product_prefix),
        "machine_source": args.machine_source,
        "process_entry_count": len(process_rows),
        "matched_entry_count": len(matched_rows),
        "unmatched_entry_count": len(unmatched_rows),
        "forecast_count": len(forecast_rows),
        "ifs_dimension_row_count": len(ifs_rows),
        "forecast_rows": forecast_rows,
        "matched_rows": matched_rows,
        "unmatched_rows": unmatched_rows,
        "notes": [
            "Actual realized cycle time is read from process entries col_l, in seconds.",
            "Product text, Ağız, Gramaj, and work center are joined from IFS past operation history by work_order + machine/resource.",
            "Current optimum cycle time is read from the DB machine cycle table with the existing lookup rules.",
            "Forecast cycle is p75 realized cycle when sample count is at least 5; otherwise median realized cycle.",
        ],
        "excel_workbook": str(workbook_path),
        "json_report": str(json_path),
    }
    _write_workbook(workbook_path, summary)
    json_path.write_text(
        json.dumps(summary, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )
    return summary


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Forecast cycle times from process DB entries enriched by IFS product dimensions."
    )
    parser.add_argument("--date-from", default="")
    parser.add_argument("--date-to", default="")
    parser.add_argument("--product-prefix", action="append", default=["MM-PET"])
    parser.add_argument("--machine-source", default="ResourceId")
    parser.add_argument("--month-row-limit", type=int, default=50000)
    parser.add_argument("--part-row-limit", type=int, default=10000)
    parser.add_argument("--timeout", type=float, default=60.0)
    parser.add_argument(
        "--output-dir",
        default=str(ROOT_DIR / "data" / "ifs-experiments"),
    )
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    try:
        summary = asyncio.run(run_forecast(args))
    except Exception as exc:
        print(f"Cycle forecast failed: {exc}", file=sys.stderr)
        return 1

    print("Cycle forecast complete")
    print(f"Date range: {summary['date_from']} to {summary['date_to']}")
    print(f"Process entries: {summary['process_entry_count']}")
    print(f"Matched entries: {summary['matched_entry_count']}")
    print(f"Forecast combinations: {summary['forecast_count']}")
    print(f"Excel workbook: {summary['excel_workbook']}")
    print(f"JSON: {summary['json_report']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
