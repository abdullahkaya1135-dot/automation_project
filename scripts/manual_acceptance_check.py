from __future__ import annotations

import argparse
import json
import sys
from datetime import UTC, datetime
from http.cookiejar import CookieJar
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.request import HTTPCookieProcessor, Request, build_opener
from uuid import uuid4

from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import selectinload

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from app.config import Settings, SettingsError, get_settings  # noqa: E402
from app.database import create_session, sqlite_health  # noqa: E402
from app.models import (  # noqa: E402
    SYNC_STATUS_FAILED_EXCEL,
    SYNC_STATUS_PENDING_EXCEL,
    SYNC_STATUS_SYNCED,
    AuxiliarySystemsSubmission,
    Entry,
)
from app.modules.auth.service import ROLE_ADMIN, role_pins  # noqa: E402
from app.modules.auxiliary_systems.row_builder import (  # noqa: E402
    normalize_auxiliary_value,
)
from app.modules.auxiliary_systems.workbook_service import (  # noqa: E402
    check_auxiliary_systems_reachable,
)
from app.modules.process_entry.workbook_io import check_excel_reachable  # noqa: E402
from app.shared.workbook_utils import INVALID_BACKUP_FILENAME_CHARS  # noqa: E402


def _ok(message: str) -> None:
    print(f"[OK] {message}")


def _warn(message: str) -> None:
    print(f"[WARN] {message}")


def _fail(message: str) -> None:
    print(f"[FAIL] {message}")


def _json_request(
    opener,
    url: str,
    *,
    method: str = "GET",
    body: dict | None = None,
    timeout: int = 10,
) -> dict:
    headers = {"Accept": "application/json"}
    data = None
    if body is not None:
        headers["Content-Type"] = "application/json"
        data = json.dumps(body).encode("utf-8")
    request = Request(url, data=data, headers=headers, method=method)
    with opener.open(request, timeout=timeout) as response:
        response_body = response.read().decode("utf-8")
    return json.loads(response_body) if response_body else {}


def _text_request(opener, url: str, *, timeout: int = 10) -> str:
    request = Request(url, headers={"Accept": "text/html"}, method="GET")
    with opener.open(request, timeout=timeout) as response:
        return response.read().decode("utf-8", errors="replace")


def _request_status(opener, url: str) -> int:
    request = Request(url, headers={"Accept": "text/html"}, method="GET")
    try:
        with opener.open(request, timeout=5) as response:
            response.read()
            return int(response.status)
    except HTTPError as exc:
        exc.read()
        return int(exc.code)


def _pin_for_api_check(settings: Settings) -> str | None:
    pins = role_pins(settings)
    return (
        settings.app_pin
        or pins.get(ROLE_ADMIN)
        or pins.get("operator")
        or next(iter(pins.values()), None)
    )


def check_http(base_url: str, settings: Settings) -> bool:
    success = True
    base_url = base_url.rstrip("/")
    opener = build_opener(HTTPCookieProcessor(CookieJar()))

    try:
        health = _json_request(opener, f"{base_url}/health")
    except (HTTPError, URLError, TimeoutError, json.JSONDecodeError) as exc:
        _fail(f"App health endpoint is not reachable at {base_url}: {exc}")
        return False

    sqlite_ok = bool(health.get("sqlite", {}).get("ok"))
    _ok(f"App health is reachable; status={health.get('status')}")
    if sqlite_ok:
        _ok("Health reports SQLite as available")
    else:
        _fail(f"Health reports SQLite issue: {health.get('sqlite')}")
        success = False

    login_pin = _pin_for_api_check(settings)
    if not login_pin:
        _fail("No APP_PIN or APP_ROLE_PINS value is available for API login")
        return False

    try:
        _json_request(
            opener,
            f"{base_url}/api/login",
            method="POST",
            body={"pin": login_pin},
        )
        bootstrap = _json_request(opener, f"{base_url}/api/bootstrap")
    except (HTTPError, URLError, TimeoutError, json.JSONDecodeError) as exc:
        _fail(f"Authenticated API check failed: {exc}")
        return False

    excel = bootstrap.get("excel") or {}
    _ok("Authenticated API works")
    if excel.get("available"):
        _ok("Bootstrap reports Excel as available")
    else:
        _warn(f"Bootstrap reports Excel unavailable: {excel.get('last_error')}")
    return success


ROLE_PAGE_CHECKS = {
    "operator": {
        "path": "/operator",
        "required": [
            'data-page="operator"',
            'id="entry-form"',
            'id="sync-phone-outbox"',
        ],
        "forbidden": [
            'id="auxiliary-form"',
            'id="retry-sync"',
            'id="run-ifs-return-candidates"',
        ],
        "denied_path": "/supervisor",
    },
    "utility": {
        "path": "/utility",
        "required": [
            'data-page="utility"',
            'id="auxiliary-form"',
        ],
        "forbidden": [
            'id="entry-form"',
            'id="retry-sync"',
            'id="run-ifs-return-candidates"',
        ],
    },
    "supervisor": {
        "path": "/supervisor",
        "required": [
            'data-page="supervisor"',
            'id="retry-sync"',
            'id="retry-auxiliary-sync"',
        ],
        "forbidden": [
            'id="entry-form"',
            'id="auxiliary-form"',
            'id="run-ifs-return-candidates"',
        ],
    },
    "planning": {
        "path": "/planning",
        "required": [
            'data-page="planning"',
            'id="create-cycle-report"',
            'id="run-ifs-return-candidates"',
        ],
        "forbidden": [
            'id="entry-form"',
            'id="auxiliary-form"',
            'id="retry-sync"',
        ],
    },
}


def _login_for_role(base_url: str, role: str, pin: str):
    opener = build_opener(HTTPCookieProcessor(CookieJar()))
    payload = _json_request(
        opener,
        f"{base_url}/api/login",
        method="POST",
        body={"pin": pin},
    )
    if payload.get("role") != role:
        raise RuntimeError(f"PIN for {role} logged in as {payload.get('role')}")
    return opener, payload


def check_role_pages(base_url: str, settings: Settings) -> bool:
    base_url = base_url.rstrip("/")
    pins = role_pins(settings)
    success = True

    for role, check in ROLE_PAGE_CHECKS.items():
        pin = pins.get(role)
        if not pin:
            _fail(f"APP_ROLE_PINS is missing a {role} PIN")
            success = False
            continue

        try:
            opener, payload = _login_for_role(base_url, role, pin)
            page = _text_request(opener, f"{base_url}{check['path']}")
        except (HTTPError, URLError, TimeoutError, json.JSONDecodeError, RuntimeError) as exc:
            _fail(f"{role} workspace check failed: {exc}")
            success = False
            continue

        if payload.get("default_path") == check["path"]:
            _ok(f"{role} login redirects to {check['path']}")
        else:
            _fail(
                f"{role} default_path is {payload.get('default_path')}; "
                f"expected {check['path']}"
            )
            success = False

        missing = [text for text in check["required"] if text not in page]
        unexpected = [text for text in check["forbidden"] if text in page]
        if missing or unexpected:
            _fail(
                f"{role} page content mismatch; "
                f"missing={missing}, unexpected={unexpected}"
            )
            success = False
        else:
            _ok(f"{role} workspace contains only expected controls")

        denied_path = check.get("denied_path")
        if denied_path:
            status_code = _request_status(opener, f"{base_url}{denied_path}")
            if status_code == 403:
                _ok(f"{role} is denied access to {denied_path}")
            else:
                _fail(
                    f"{role} access to {denied_path} returned {status_code}; "
                    "expected 403"
                )
                success = False

    admin_pin = pins.get(ROLE_ADMIN) or settings.app_pin
    if admin_pin:
        try:
            opener, payload = _login_for_role(base_url, ROLE_ADMIN, admin_pin)
            statuses = {
                path: _request_status(opener, f"{base_url}{path}")
                for path in ("/operator", "/utility", "/supervisor", "/planning")
            }
        except (HTTPError, URLError, TimeoutError, json.JSONDecodeError, RuntimeError) as exc:
            _fail(f"admin workspace check failed: {exc}")
            success = False
        else:
            if payload.get("default_path") == "/operator" and set(statuses.values()) == {200}:
                _ok("admin can open all role workspaces")
            else:
                _fail(
                    "admin workspace access mismatch; "
                    f"default_path={payload.get('default_path')}, statuses={statuses}"
                )
                success = False
    else:
        _warn("Skipping admin workspace check; no admin role PIN or APP_PIN configured")

    return success


def latest_entry(settings: Settings) -> Entry | None:
    with create_session(settings) as session:
        return session.scalars(
            select(Entry)
            .options(selectinload(Entry.tour_context))
            .order_by(Entry.created_at.desc(), Entry.id.desc())
        ).first()


def check_sqlite(settings: Settings) -> tuple[bool, Entry | None]:
    health = sqlite_health(settings)
    if not health.get("ok"):
        _fail(f"SQLite health check failed: {health.get('error')}")
        return False, None

    _ok(f"SQLite database is available at {settings.sqlite_path}")
    entry = latest_entry(settings)
    if entry is None:
        _fail("No SQLite entry found. Submit one machine row before final acceptance.")
        return False, None

    _ok(
        "Latest SQLite entry "
        f"id={entry.id}, machine={entry.col_f}, "
        f"work_order={entry.col_h}, sync_status={entry.sync_status}"
    )
    if entry.last_error:
        _warn(f"Latest entry sync error: {entry.last_error}")
    return True, entry


def latest_entries(settings: Settings, limit: int) -> list[Entry]:
    with create_session(settings) as session:
        return list(
            session.scalars(
                select(Entry)
                .options(selectinload(Entry.tour_context))
                .order_by(Entry.created_at.desc(), Entry.id.desc())
                .limit(limit)
            ).all()
        )


def entry_count(settings: Settings) -> int:
    with create_session(settings) as session:
        return int(session.scalar(select(func.count()).select_from(Entry)) or 0)


def latest_auxiliary_submissions(
    settings: Settings,
    limit: int,
) -> list[AuxiliarySystemsSubmission]:
    with create_session(settings) as session:
        return list(
            session.scalars(
                select(AuxiliarySystemsSubmission)
                .order_by(
                    AuxiliarySystemsSubmission.created_at.desc(),
                    AuxiliarySystemsSubmission.id.desc(),
                )
                .limit(limit)
            ).all()
        )


def auxiliary_submission_count(settings: Settings) -> int:
    with create_session(settings) as session:
        return int(
            session.scalar(
                select(func.count()).select_from(AuxiliarySystemsSubmission)
            )
            or 0
        )


def auxiliary_unsynced_count(settings: Settings) -> int:
    with create_session(settings) as session:
        return int(
            session.scalar(
                select(func.count())
                .select_from(AuxiliarySystemsSubmission)
                .where(
                    AuxiliarySystemsSubmission.sync_status.in_(
                        (SYNC_STATUS_PENDING_EXCEL, SYNC_STATUS_FAILED_EXCEL)
                    )
                )
            )
            or 0
        )


def check_excel(settings: Settings, entry: Entry | None, expected_row: int | None) -> bool:
    status = check_excel_reachable(settings)
    if not status.available:
        if expected_row is not None:
            _fail(f"Excel is unavailable, so row {expected_row} cannot be verified: {status.error}")
            return False
        _warn(f"Excel is unavailable: {status.error}")
        return True

    _ok("Excel workbook is reachable")
    if entry is None:
        return False
    if entry.excel_row_number is None:
        if expected_row is not None:
            _fail(f"Latest entry is not synced to expected Excel row {expected_row}")
            return False
        _warn("Latest entry is not synced to Excel yet")
        return True

    success = True
    if expected_row is not None and entry.excel_row_number != expected_row:
        _fail(
            f"Latest entry Excel row is {entry.excel_row_number}; "
            f"expected {expected_row}"
        )
        success = False
    else:
        _ok(f"Latest entry recorded Excel row {entry.excel_row_number}")

    workbook = load_workbook(settings.excel_path, read_only=True, data_only=True)
    try:
        if settings.sheet_name not in workbook.sheetnames:
            _fail(f"Sheet was not found: {settings.sheet_name}")
            return False
        worksheet = workbook[settings.sheet_name]
        row = next(
            worksheet.iter_rows(
                min_row=entry.excel_row_number,
                max_row=entry.excel_row_number,
                max_col=8,
                values_only=True,
            )
        )
    finally:
        workbook.close()

    if row[5] == entry.col_f and row[7] == entry.col_h:
        _ok("Excel row matches the latest SQLite machine and work order")
    else:
        _fail(
            "Excel row does not match SQLite. "
            f"Excel F:H={row[5:8]}, SQLite F/H={(entry.col_f, entry.col_h)}"
        )
        success = False
    return success


def _backup_files_for_source(
    settings: Settings,
    source_path: Path,
    fallback_stem: str,
) -> list[Path]:
    backup_dir = Path(settings.backup_dir)
    if not backup_dir.exists():
        return []

    safe_stem = INVALID_BACKUP_FILENAME_CHARS.sub("_", source_path.stem).strip()
    if not safe_stem:
        safe_stem = fallback_stem
    pattern = f"{safe_stem}_*{source_path.suffix or '.xlsx'}"
    return sorted(backup_dir.glob(pattern), key=lambda path: path.stat().st_mtime)


def _process_backup_files(settings: Settings) -> list[Path]:
    return _backup_files_for_source(settings, Path(settings.excel_path), "workbook")


def _auxiliary_backup_files(settings: Settings) -> list[Path]:
    return _backup_files_for_source(
        settings,
        Path(settings.auxiliary_systems_target_path),
        "yardimci_sistemler",
    )


def _bulk_sync_acceptance_records(batch_size: int) -> list[dict]:
    batch_id = uuid4().hex[:12]
    recorded_at = datetime.now(UTC).isoformat()
    tour_client_id = f"acceptance-tour-{batch_id}"
    records = [
        {
            "type": "tour_context",
            "client_request_id": tour_client_id,
            "client_recorded_at": recorded_at,
            "body": {
                "client_request_id": tour_client_id,
                "client_recorded_at": recorded_at,
                "ambient_temp": "24,5",
                "production_engineer": "Acceptance",
                "shift_chief": "Selman",
            },
        }
    ]
    for index in range(batch_size):
        entry_client_id = f"acceptance-entry-{batch_id}-{index + 1:03d}"
        records.append(
            {
                "type": "entry",
                "client_request_id": entry_client_id,
                "depends_on_client_request_id": tour_client_id,
                "client_recorded_at": recorded_at,
                "body": {
                    "client_request_id": entry_client_id,
                    "client_recorded_at": recorded_at,
                    "payload_schema_version": 2,
                    "payload": {
                        "col_f": f"101-{index + 1:03d}",
                        "col_g": f"Acceptance Product {index + 1:03d}",
                        "col_h": f"ACCEPT-WO-{index + 1:03d}",
                        "col_i": "HM-02-ACCEPT",
                        "col_j": "16",
                        "col_k": "12",
                        "col_l": "12,5",
                    },
                    "status": "acceptance",
                    "notes": "manual acceptance bulk exercise",
                },
            }
        )
    return records


def exercise_bulk_sync(base_url: str, settings: Settings, batch_size: int) -> bool:
    if batch_size < 1:
        return True

    login_pin = _pin_for_api_check(settings)
    if not login_pin:
        _fail("No APP_PIN or APP_ROLE_PINS value is available for bulk sync login")
        return False

    base_url = base_url.rstrip("/")
    opener = build_opener(HTTPCookieProcessor(CookieJar()))
    try:
        _json_request(
            opener,
            f"{base_url}/api/login",
            method="POST",
            body={"pin": login_pin},
        )
    except (HTTPError, URLError, TimeoutError, json.JSONDecodeError) as exc:
        _fail(f"Bulk sync login failed: {exc}")
        return False

    before_count = entry_count(settings)
    before_backup_names = {path.name for path in _process_backup_files(settings)}
    records = _bulk_sync_acceptance_records(batch_size)
    request_body = {"records": records, "sync_excel": True}
    timeout = max(30, batch_size)

    try:
        first_payload = _json_request(
            opener,
            f"{base_url}/api/offline/bulk-sync",
            method="POST",
            body=request_body,
            timeout=timeout,
        )
    except (HTTPError, URLError, TimeoutError, json.JSONDecodeError) as exc:
        _fail(f"Bulk sync request failed: {exc}")
        return False

    success = True
    if (
        first_payload.get("saved_count") == batch_size + 1
        and first_payload.get("synced_count") == batch_size
        and first_payload.get("failed_count") == 0
        and first_payload.get("excel_pending") is False
    ):
        _ok(f"Bulk sync saved one tour context and {batch_size} process entries")
    else:
        _fail(f"Unexpected bulk sync response: {first_payload}")
        success = False

    after_first_count = entry_count(settings)
    if after_first_count == before_count + batch_size:
        _ok(f"SQLite entry count increased by {batch_size}")
    else:
        _fail(
            f"SQLite entry count is {after_first_count}; "
            f"expected {before_count + batch_size}"
        )
        success = False

    after_first_backups = _process_backup_files(settings)
    new_backups = [
        path for path in after_first_backups if path.name not in before_backup_names
    ]
    if len(new_backups) == 1:
        _ok(f"Bulk sync created one process workbook backup: {new_backups[0].name}")
    else:
        _fail(
            f"Bulk sync created {len(new_backups)} new process backups; expected 1"
        )
        success = False

    try:
        replay_payload = _json_request(
            opener,
            f"{base_url}/api/offline/bulk-sync",
            method="POST",
            body=request_body,
            timeout=timeout,
        )
    except (HTTPError, URLError, TimeoutError, json.JSONDecodeError) as exc:
        _fail(f"Bulk sync replay failed: {exc}")
        return False

    replay_records = replay_payload.get("records") or []
    replay_entry_count = entry_count(settings)
    replay_backup_names = {path.name for path in _process_backup_files(settings)}
    first_backup_names = {path.name for path in after_first_backups}
    if (
        replay_entry_count == after_first_count
        and replay_backup_names == first_backup_names
        and len(replay_records) == batch_size + 1
        and all(record.get("idempotent_replay") for record in replay_records)
    ):
        _ok("Replaying the same bulk sync did not duplicate rows or backups")
    else:
        _fail(
            "Bulk sync replay was not idempotent; "
            f"entry_count={replay_entry_count}, "
            f"new_backups={sorted(replay_backup_names - first_backup_names)}, "
            f"payload={replay_payload}"
        )
        success = False

    return check_latest_batch(settings, batch_size) and success


def _auxiliary_retry_acceptance_records(batch_size: int) -> list[dict]:
    batch_id = uuid4().hex[:12]
    recorded_at = datetime.now(UTC).isoformat()
    records = []
    for index in range(batch_size):
        submission_client_id = f"acceptance-auxiliary-{batch_id}-{index + 1:03d}"
        records.append(
            {
                "type": "auxiliary_submission",
                "client_request_id": submission_client_id,
                "client_recorded_at": recorded_at,
                "body": {
                    "client_request_id": submission_client_id,
                    "client_recorded_at": recorded_at,
                    "recorded_date": f"2026-06-{(index % 28) + 1:02d}",
                    "payload": {
                        "tower_frequency": str(50 + index),
                        "tower_set_pressure": "3,6",
                        "compressor_low_716_pressure": str(11 + index),
                    },
                },
            }
        )
    return records


def exercise_auxiliary_retry(
    base_url: str,
    settings: Settings,
    batch_size: int,
) -> bool:
    if batch_size < 1:
        return True
    if batch_size > 100:
        _fail("--exercise-auxiliary-retry supports at most 100 rows per run")
        return False

    login_pin = _pin_for_api_check(settings)
    if not login_pin:
        _fail("No APP_PIN or APP_ROLE_PINS value is available for auxiliary retry login")
        return False

    existing_unsynced = auxiliary_unsynced_count(settings)
    if existing_unsynced:
        _fail(
            "Auxiliary retry exercise requires a clean pending queue; "
            f"found {existing_unsynced} unsynced auxiliary submission(s)"
        )
        return False

    base_url = base_url.rstrip("/")
    opener = build_opener(HTTPCookieProcessor(CookieJar()))
    try:
        _json_request(
            opener,
            f"{base_url}/api/login",
            method="POST",
            body={"pin": login_pin},
        )
    except (HTTPError, URLError, TimeoutError, json.JSONDecodeError) as exc:
        _fail(f"Auxiliary retry login failed: {exc}")
        return False

    before_count = auxiliary_submission_count(settings)
    before_backup_names = {path.name for path in _auxiliary_backup_files(settings)}
    request_body = {
        "records": _auxiliary_retry_acceptance_records(batch_size),
        "sync_excel": False,
    }
    timeout = max(30, batch_size)

    try:
        queue_payload = _json_request(
            opener,
            f"{base_url}/api/offline/bulk-sync",
            method="POST",
            body=request_body,
            timeout=timeout,
        )
    except (HTTPError, URLError, TimeoutError, json.JSONDecodeError) as exc:
        _fail(f"Auxiliary offline queue request failed: {exc}")
        return False

    success = True
    queued_records = queue_payload.get("records") or []
    if (
        queue_payload.get("saved_count") == batch_size
        and queue_payload.get("synced_count") == 0
        and queue_payload.get("failed_count") == batch_size
        and queue_payload.get("excel_pending") is True
        and len(queued_records) == batch_size
        and all(record.get("synced_to_excel") is False for record in queued_records)
    ):
        _ok(f"Queued {batch_size} auxiliary submissions without Excel sync")
    else:
        _fail(f"Unexpected auxiliary offline queue response: {queue_payload}")
        success = False

    after_queue_count = auxiliary_submission_count(settings)
    if after_queue_count == before_count + batch_size:
        _ok(f"Auxiliary SQLite count increased by {batch_size}")
    else:
        _fail(
            f"Auxiliary SQLite count is {after_queue_count}; "
            f"expected {before_count + batch_size}"
        )
        success = False

    after_queue_backup_names = {path.name for path in _auxiliary_backup_files(settings)}
    if after_queue_backup_names == before_backup_names:
        _ok("Queueing auxiliary submissions did not touch the workbook")
    else:
        _fail(
            "Queueing auxiliary submissions unexpectedly created workbook backups: "
            f"{sorted(after_queue_backup_names - before_backup_names)}"
        )
        success = False

    try:
        retry_payload = _json_request(
            opener,
            f"{base_url}/api/auxiliary-systems/sync/retry",
            method="POST",
            timeout=max(30, batch_size * 2),
        )
    except (HTTPError, URLError, TimeoutError, json.JSONDecodeError) as exc:
        _fail(f"Auxiliary retry request failed: {exc}")
        return False

    retry_results = retry_payload.get("results") or []
    if (
        retry_payload.get("attempted") == batch_size
        and retry_payload.get("synced") == batch_size
        and retry_payload.get("failed") == 0
        and retry_payload.get("remaining") == 0
        and retry_payload.get("stopped_on_error") is False
        and len(retry_results) == batch_size
        and all(result.get("success") is True for result in retry_results)
    ):
        _ok(f"Auxiliary retry synced {batch_size} queued submissions")
    else:
        _fail(f"Unexpected auxiliary retry response: {retry_payload}")
        success = False

    after_retry_backups = _auxiliary_backup_files(settings)
    new_backups = [
        path for path in after_retry_backups if path.name not in after_queue_backup_names
    ]
    if len(new_backups) == 1:
        _ok(f"Auxiliary retry created one workbook backup: {new_backups[0].name}")
    else:
        _fail(
            f"Auxiliary retry created {len(new_backups)} new backups; expected 1"
        )
        success = False

    return check_latest_auxiliary_batch(settings, batch_size) and success


def check_latest_batch(settings: Settings, batch_size: int) -> bool:
    if batch_size < 1:
        return True

    entries = latest_entries(settings, batch_size)
    if len(entries) != batch_size:
        _fail(
            f"Only {len(entries)} SQLite entries were found; "
            f"expected latest batch size {batch_size}"
        )
        return False

    success = True
    row_numbers = [entry.excel_row_number for entry in entries]
    missing_rows = [entry.id for entry in entries if entry.excel_row_number is None]
    if missing_rows:
        _fail(f"Latest batch contains entries without Excel rows: {missing_rows}")
        success = False

    unsynced = [
        (entry.id, entry.sync_status, entry.last_error)
        for entry in entries
        if entry.sync_status != "synced"
    ]
    if unsynced:
        _fail(f"Latest batch contains unsynced entries: {unsynced}")
        success = False

    synced_rows = [row for row in row_numbers if row is not None]
    if len(synced_rows) != len(set(synced_rows)):
        _fail(f"Latest batch has duplicate Excel row numbers: {synced_rows}")
        success = False
    if synced_rows:
        sorted_rows = sorted(synced_rows)
        expected_rows = list(range(sorted_rows[0], sorted_rows[0] + len(sorted_rows)))
        if sorted_rows == expected_rows:
            _ok(
                "Latest batch Excel rows are sequential: "
                f"{sorted_rows[0]}-{sorted_rows[-1]}"
            )
        else:
            _fail(
                "Latest batch Excel rows are not sequential: "
                f"actual={sorted_rows}, expected={expected_rows}"
            )
            success = False

    status = check_excel_reachable(settings)
    if not status.available:
        _fail(f"Excel is unavailable, so latest batch rows cannot be verified: {status.error}")
        return False

    if not synced_rows:
        return False

    workbook = load_workbook(settings.excel_path, read_only=True, data_only=True)
    try:
        if settings.sheet_name not in workbook.sheetnames:
            _fail(f"Sheet was not found: {settings.sheet_name}")
            return False
        worksheet = workbook[settings.sheet_name]
        for entry in entries:
            if entry.excel_row_number is None:
                continue
            row = next(
                worksheet.iter_rows(
                    min_row=entry.excel_row_number,
                    max_row=entry.excel_row_number,
                    max_col=8,
                    values_only=True,
                )
            )
            if row[5] != entry.col_f or row[7] != entry.col_h:
                _fail(
                    f"Batch row mismatch for entry {entry.id}; "
                    f"Excel F/H={(row[5], row[7])}, "
                    f"SQLite F/H={(entry.col_f, entry.col_h)}"
                )
                success = False
    finally:
        workbook.close()

    if success:
        _ok(f"Latest {batch_size} SQLite entries match their Excel rows")
    return success


def _auxiliary_payload_from_submission(
    submission: AuxiliarySystemsSubmission,
) -> dict:
    try:
        payload = json.loads(submission.payload_json)
    except json.JSONDecodeError:
        return {}
    return payload if isinstance(payload, dict) else {}


def check_latest_auxiliary_batch(settings: Settings, batch_size: int) -> bool:
    if batch_size < 1:
        return True

    submissions = latest_auxiliary_submissions(settings, batch_size)
    if len(submissions) != batch_size:
        _fail(
            f"Only {len(submissions)} auxiliary submissions were found; "
            f"expected latest batch size {batch_size}"
        )
        return False

    success = True
    missing_blocks = [
        submission.id
        for submission in submissions
        if submission.excel_start_row is None or submission.excel_end_row is None
    ]
    if missing_blocks:
        _fail(
            "Latest auxiliary batch contains submissions without Excel blocks: "
            f"{missing_blocks}"
        )
        success = False

    unsynced = [
        (submission.id, submission.sync_status, submission.last_error)
        for submission in submissions
        if submission.sync_status != SYNC_STATUS_SYNCED
    ]
    if unsynced:
        _fail(f"Latest auxiliary batch contains unsynced submissions: {unsynced}")
        success = False

    blocks: list[tuple[int, int]] = []
    for submission in submissions:
        if submission.excel_start_row is None or submission.excel_end_row is None:
            continue
        blocks.append((submission.excel_start_row, submission.excel_end_row))

    if len(blocks) != len(set(blocks)):
        _fail(f"Latest auxiliary batch has duplicate Excel blocks: {blocks}")
        success = False

    invalid_lengths = [
        block for block in blocks if block[1] - block[0] + 1 != 15
    ]
    if invalid_lengths:
        _fail(f"Auxiliary block length mismatch: {invalid_lengths}")
        success = False

    if blocks:
        sorted_blocks = sorted(blocks)
        first_start = sorted_blocks[0][0]
        expected_blocks = [
            (first_start + (15 * index), first_start + (15 * index) + 14)
            for index in range(len(sorted_blocks))
        ]
        if sorted_blocks == expected_blocks:
            _ok(
                "Latest auxiliary Excel blocks are sequential: "
                f"{sorted_blocks[0][0]}-{sorted_blocks[-1][1]}"
            )
        else:
            _fail(
                "Latest auxiliary Excel blocks are not sequential: "
                f"actual={sorted_blocks}, expected={expected_blocks}"
            )
            success = False

    status = check_auxiliary_systems_reachable(settings)
    if not status.target_available:
        _fail(
            "Auxiliary workbook is unavailable, so latest blocks cannot be verified: "
            f"{status.target_error}"
        )
        return False

    if not blocks:
        return False

    workbook = load_workbook(
        settings.auxiliary_systems_target_path,
        read_only=True,
        data_only=True,
    )
    try:
        if settings.auxiliary_systems_sheet_name not in workbook.sheetnames:
            _fail(
                f"Auxiliary sheet was not found: "
                f"{settings.auxiliary_systems_sheet_name}"
            )
            return False
        worksheet = workbook[settings.auxiliary_systems_sheet_name]
        for submission in submissions:
            if submission.excel_start_row is None or submission.excel_end_row is None:
                continue
            payload = _auxiliary_payload_from_submission(submission)
            first_row = next(
                worksheet.iter_rows(
                    min_row=submission.excel_start_row,
                    max_row=submission.excel_start_row,
                    max_col=9,
                    values_only=True,
                )
            )
            last_row = next(
                worksheet.iter_rows(
                    min_row=submission.excel_end_row,
                    max_row=submission.excel_end_row,
                    max_col=9,
                    values_only=True,
                )
            )
            expected_frequency = normalize_auxiliary_value(
                payload.get("tower_frequency")
            )
            expected_low_pressure = normalize_auxiliary_value(
                payload.get("compressor_low_716_pressure")
            )
            if (
                first_row[0] != submission.recorded_date
                or first_row[2] != expected_frequency
                or last_row[8] != expected_low_pressure
            ):
                _fail(
                    f"Auxiliary block mismatch for submission {submission.id}; "
                    f"first_row_date/frequency={(first_row[0], first_row[2])}, "
                    f"last_row_pressure={last_row[8]}, "
                    f"expected={(submission.recorded_date, expected_frequency, expected_low_pressure)}"
                )
                success = False
    finally:
        workbook.close()

    if success:
        _ok(
            f"Latest {batch_size} auxiliary submissions match their Excel blocks"
        )
    return success


def check_backups(settings: Settings) -> bool:
    backup_dir = Path(settings.backup_dir)
    if not backup_dir.exists():
        _fail(f"Backup directory does not exist: {backup_dir}")
        return False

    backups = _process_backup_files(settings)
    if not backups:
        _fail(f"No timestamped backup found in {backup_dir}")
        return False

    newest = backups[-1]
    _ok(f"Newest backup: {newest.name}")
    if len(backups) <= settings.backup_keep_count:
        _ok(
            f"Backup retention is within limit: "
            f"{len(backups)}/{settings.backup_keep_count}"
        )
        return True

    _fail(
        f"Backup retention exceeded: {len(backups)} files for "
        f"keep_count={settings.backup_keep_count}"
    )
    return False


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Verify the manual acceptance state after an operator submits one "
            "machine row through the running FastAPI app."
        )
    )
    parser.add_argument(
        "--url",
        default="http://127.0.0.1:8080",
        help="Base URL for the running app.",
    )
    parser.add_argument(
        "--expected-row",
        type=int,
        default=None,
        help="Expected Excel row for the latest synced entry, for example 1726.",
    )
    parser.add_argument(
        "--skip-http",
        action="store_true",
        help="Skip health/login/bootstrap checks against the running app.",
    )
    parser.add_argument(
        "--skip-backups",
        action="store_true",
        help="Skip backup directory and retention checks.",
    )
    parser.add_argument(
        "--check-role-pages",
        action="store_true",
        help="Verify role PIN login, workspace visibility, and page permissions.",
    )
    parser.add_argument(
        "--latest-batch-size",
        type=int,
        default=0,
        help=(
            "Verify the latest N process entries are synced to unique sequential "
            "Excel rows and match SQLite machine/work-order values."
        ),
    )
    parser.add_argument(
        "--exercise-bulk-sync",
        type=int,
        default=0,
        metavar="N",
        help=(
            "Create one phone-style offline bulk-sync batch with N process "
            "entries against the running app, verify one backup, then replay "
            "the same batch to prove idempotency. Use only with safe workbook "
            "copies."
        ),
    )
    parser.add_argument(
        "--exercise-auxiliary-retry",
        type=int,
        default=0,
        metavar="N",
        help=(
            "Queue N auxiliary submissions through offline bulk sync without "
            "Excel, then call the supervisor retry API and verify one workbook "
            "backup plus sequential 15-row Excel blocks. Use only with safe "
            "workbook copies."
        ),
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    checks: list[bool] = []

    try:
        settings = get_settings()
    except SettingsError as exc:
        _fail(str(exc))
        return 1

    _ok("Runtime settings loaded")
    if not args.skip_http:
        checks.append(check_http(args.url, settings))
        if args.check_role_pages:
            checks.append(check_role_pages(args.url, settings))
        if args.exercise_bulk_sync:
            checks.append(exercise_bulk_sync(args.url, settings, args.exercise_bulk_sync))
        if args.exercise_auxiliary_retry:
            checks.append(
                exercise_auxiliary_retry(
                    args.url,
                    settings,
                    args.exercise_auxiliary_retry,
                )
            )
    else:
        if args.check_role_pages:
            _fail("--check-role-pages requires HTTP checks; remove --skip-http")
            checks.append(False)
        if args.exercise_bulk_sync:
            _fail("--exercise-bulk-sync requires HTTP checks; remove --skip-http")
            checks.append(False)
        if args.exercise_auxiliary_retry:
            _fail("--exercise-auxiliary-retry requires HTTP checks; remove --skip-http")
            checks.append(False)

    sqlite_ok, entry = check_sqlite(settings)
    checks.append(sqlite_ok)
    checks.append(check_excel(settings, entry, args.expected_row))
    if args.latest_batch_size:
        checks.append(check_latest_batch(settings, args.latest_batch_size))
    if not args.skip_backups:
        checks.append(check_backups(settings))

    if all(checks):
        _ok("Manual acceptance checks passed")
        return 0

    _fail("Manual acceptance checks need attention")
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
