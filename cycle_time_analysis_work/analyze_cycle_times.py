import json
import math
import re
import sqlite3
import sys
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path
from statistics import mean, median, stdev

import openpyxl


WORK_DIR = Path(__file__).resolve().parent
PROJECT_DIR = WORK_DIR.parent
if str(PROJECT_DIR) not in sys.path:
    sys.path.insert(0, str(PROJECT_DIR))

from app.features.cycle_reports.service import (  # noqa: E402
    CycleTableEntry,
    _lookup_optimum_cycle,
    normalize_machine_group,
)

SOURCE = WORK_DIR / "PROSES 2026_FK_source_copy.xlsx"
DB_PATH = PROJECT_DIR / "data" / "process_entries.sqlite3"
OUTPUT_JSON = WORK_DIR / "cycle_time_analysis_data.json"

REQUESTED_START_ROW = 1498
REQUESTED_END_ROW = 1867

MOUTH_RE = re.compile(r"(\d+(?:[,.]\d+)?)\s*MM\b", re.I)
WEIGHT_RE = re.compile(r"(\d+(?:[,.]\d+)?)\s*GR\b", re.I)


def to_float(value):
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if math.isnan(float(value)):
            return None
        return float(value)
    text = str(value).strip().replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def to_decimal(value):
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int):
        return Decimal(value)
    if isinstance(value, float):
        return Decimal(str(value))
    text = str(value).strip().replace(",", ".")
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def clean_number(value):
    if value is None:
        return None
    if abs(value - round(value)) < 1e-9:
        return int(round(value))
    return round(value, 3)


def identifier(value):
    text = "" if value is None else str(value).strip()
    if text.endswith(".0"):
        return text[:-2]
    return text


def parse_product(text):
    mouth = MOUTH_RE.search(text or "")
    weight = WEIGHT_RE.search(text or "")
    return (
        clean_number(to_float(mouth.group(1)) if mouth else None),
        clean_number(to_float(weight.group(1)) if weight else None),
    )


def quantile(values, q):
    """Excel-compatible inclusive quantile approximation."""
    vals = sorted(values)
    if not vals:
        return None
    if len(vals) == 1:
        return vals[0]
    pos = (len(vals) - 1) * q
    lo = math.floor(pos)
    hi = math.ceil(pos)
    if lo == hi:
        return vals[lo]
    frac = pos - lo
    return vals[lo] + (vals[hi] - vals[lo]) * frac


def stats(values):
    vals = sorted(values)
    if not vals:
        return {
            "n": 0,
            "min": None,
            "q1": None,
            "median": None,
            "mean": None,
            "q3": None,
            "max": None,
            "stdev": None,
            "cv": None,
        }
    avg = mean(vals)
    sd = stdev(vals) if len(vals) > 1 else 0
    return {
        "n": len(vals),
        "min": round(min(vals), 3),
        "q1": round(quantile(vals, 0.25), 3),
        "median": round(median(vals), 3),
        "mean": round(avg, 3),
        "q3": round(quantile(vals, 0.75), 3),
        "max": round(max(vals), 3),
        "stdev": round(sd, 3),
        "cv": round(sd / avg, 4) if avg else None,
    }


def select_source_sheet(workbook):
    candidates = []
    for ws in workbook.worksheets:
        match_count = 0
        numeric_cycles = 0
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=12, values_only=True):
            product = row[1]
            cycle = to_float(row[6])
            if isinstance(product, str) and MOUTH_RE.search(product) and WEIGHT_RE.search(product):
                match_count += 1
                if cycle is not None:
                    numeric_cycles += 1
        if match_count:
            candidates.append((numeric_cycles, match_count, ws))
    if not candidates:
        raise RuntimeError("No worksheet has column G values matching the MM/GR product pattern.")
    return max(candidates, key=lambda item: (item[0], item[1]))[2]


def confidence(clean_values):
    if not clean_values:
        return "No usable sample"
    s = stats(clean_values)
    n = s["n"]
    cv = s["cv"] if s["cv"] is not None else 0
    if n >= 4 and cv <= 0.05:
        return "High"
    if n >= 3 and cv <= 0.10:
        return "Medium"
    return "Low"


def recommendation_note(clean_values, removed):
    if not clean_values:
        return "No usable cycle"
    clean_stats = stats(clean_values)
    if len(clean_values) < 3:
        if len(clean_values) == 2:
            spread = max(clean_values) - min(clean_values)
            center = median(clean_values)
            if spread > max(2, 0.15 * center):
                return "Verify: conflicting limited sample"
        return "Verify: limited sample"
    if clean_stats["cv"] is not None and clean_stats["cv"] > 0.10:
        return "Verify: variable cycles"
    if removed:
        return "Outlier removed"
    return "Use"


def append_observation(
    observations,
    skipped,
    *,
    source_type,
    source_sheet,
    source_row,
    machine,
    product,
    total_cavities,
    active_cavities,
    cycle,
    process_date="",
    work_order="",
):
    if not product:
        skipped.append(
            {
                "source_type": source_type,
                "source_sheet": source_sheet,
                "source_row": source_row,
                "process_date": process_date,
                "work_order": work_order,
                "reason": "Blank product",
                "machine_code": identifier(machine),
                "product": product,
            }
        )
        return

    mouth, weight = parse_product(str(product))
    cycle_value = to_float(cycle)
    if mouth is None or weight is None:
        skipped.append(
            {
                "source_type": source_type,
                "source_sheet": source_sheet,
                "source_row": source_row,
                "process_date": process_date,
                "work_order": work_order,
                "reason": "Could not parse Ağız/Gramaj from product text",
                "machine_code": identifier(machine),
                "product": product,
            }
        )
        return
    if cycle_value is None or cycle_value <= 0:
        skipped.append(
            {
                "source_type": source_type,
                "source_sheet": source_sheet,
                "source_row": source_row,
                "process_date": process_date,
                "work_order": work_order,
                "reason": "Missing or non-positive cycle time",
                "machine_code": identifier(machine),
                "product": product,
                "ağız": mouth,
                "gramaj": weight,
                "cycle_time": cycle,
            }
        )
        return

    source_key = f"{source_type}:{source_sheet}:{source_row}"
    observations.append(
        {
            "observation_id": source_key,
            "source_type": source_type,
            "source_sheet": source_sheet,
            "source_row": source_row,
            "process_date": process_date,
            "work_order": work_order,
            "machine_code": identifier(machine),
            "product": str(product).strip(),
            "ağız": mouth,
            "gramaj": weight,
            "total_cavities": clean_number(to_float(total_cavities)),
            "active_cavities": clean_number(to_float(active_cavities)),
            "cycle_time_sec": round(cycle_value, 3),
        }
    )


def read_workbook_observations(source_sheet):
    observations = []
    skipped = []
    for source_row, row in enumerate(
        source_sheet.iter_rows(min_row=2, max_row=source_sheet.max_row, min_col=6, max_col=12, values_only=True),
        2,
    ):
        machine, product, work_order, _, total_cavities, active_cavities, cycle = row
        append_observation(
            observations,
            skipped,
            source_type="Old Workbook",
            source_sheet=source_sheet.title,
            source_row=source_row,
            machine=machine,
            product=product,
            total_cavities=total_cavities,
            active_cavities=active_cavities,
            cycle=cycle,
            work_order=identifier(work_order),
        )
    return observations, skipped


def read_db_observations():
    observations = []
    skipped = []
    stats_row = {"total_rows": 0}
    if not DB_PATH.exists():
        skipped.append(
            {
                "source_type": "Process DB",
                "source_sheet": "entries",
                "source_row": "",
                "reason": f"Database not found: {DB_PATH}",
            }
        )
        return observations, skipped, stats_row

    connection = sqlite3.connect(DB_PATH)
    connection.row_factory = sqlite3.Row
    query = """
        select id, process_date, machine_code, col_f, col_g, col_h, col_j, col_k, col_l
        from entries
        order by process_date asc, id asc
    """
    rows = list(connection.execute(query))
    stats_row["total_rows"] = len(rows)
    for row in rows:
        append_observation(
            observations,
            skipped,
            source_type="Process DB",
            source_sheet="entries",
            source_row=row["id"],
            machine=row["col_f"] or row["machine_code"],
            product=row["col_g"],
            total_cavities=row["col_j"],
            active_cavities=row["col_k"],
            cycle=row["col_l"],
            process_date=row["process_date"] or "",
            work_order=identifier(row["col_h"]),
        )
    return observations, skipped, stats_row


def read_machine_groups(connection):
    query = """
        select m.machine_code, g.group_name
        from machine_group_machines gm
        join machines m on m.id = gm.machine_id
        join machine_groups g on g.id = gm.machine_group_id
    """
    return {
        identifier(row["machine_code"]): normalize_machine_group(row["group_name"])
        for row in connection.execute(query)
    }


def read_cycle_entries(connection):
    rows = list(
        connection.execute(
            """
            select source_row_number, col_a, col_b, col_c, col_d, col_g
            from machine_cycle_table_rows
            order by source_row_number asc
            """
        )
    )

    current_machine = ""
    current_group = ""
    current_neck = None
    exact_entries = defaultdict(list)
    lookup_entries = defaultdict(list)

    for row in rows:
        machine_no = identifier(row["col_a"])
        if machine_no:
            current_machine = machine_no

        group_key = normalize_machine_group(row["col_b"])
        if group_key:
            current_group = group_key

        neck = to_decimal(row["col_c"])
        if neck is not None:
            current_neck = neck
        else:
            neck = current_neck

        gram = to_decimal(row["col_d"])
        cycle = to_decimal(row["col_g"])
        if not current_group or neck is None or gram is None or cycle is None:
            continue

        key = (current_group, neck.normalize(), gram.normalize())
        entry = {
            "source_row_number": row["source_row_number"],
            "machine_no": current_machine,
            "machine_group": current_group,
            "neck": neck,
            "gram": gram,
            "cycle": cycle,
        }
        exact_entries[key].append(entry)
        lookup_entries[key].append(
            CycleTableEntry(
                machine_no=current_machine,
                machine_group_key=current_group,
                neck_diameter=neck,
                gram=gram,
                estimated_cycle_time=cycle,
            )
        )

    return dict(exact_entries), dict(lookup_entries), len(rows)


def read_db_comparison_context():
    if not DB_PATH.exists():
        return {
            "machine_groups": {},
            "exact_entries": {},
            "lookup_entries": {},
            "cycle_table_row_count": 0,
        }

    connection = sqlite3.connect(DB_PATH)
    connection.row_factory = sqlite3.Row
    try:
        machine_groups = read_machine_groups(connection)
        exact_entries, lookup_entries, cycle_table_row_count = read_cycle_entries(connection)
    finally:
        connection.close()

    return {
        "machine_groups": machine_groups,
        "exact_entries": exact_entries,
        "lookup_entries": lookup_entries,
        "cycle_table_row_count": cycle_table_row_count,
    }


def decimal_to_float(value):
    return None if value is None else float(value)


def ratio_delta(value, benchmark):
    if value is None or benchmark is None or benchmark == 0:
        return None
    return (value / benchmark) - 1


def resolve_exact_optimum(entries, machine_code):
    if not entries:
        return None, "no_exact_match", ""

    cycles = {entry["cycle"] for entry in entries}
    source_refs = ", ".join(
        f"row {entry['source_row_number']} (M{entry['machine_no']}: {entry['cycle']})"
        for entry in entries
    )
    if len(cycles) == 1:
        return next(iter(cycles)), "exact_match", source_refs

    machine_cycles = {
        entry["cycle"]
        for entry in entries
        if identifier(entry["machine_no"]) == identifier(machine_code)
    }
    if len(machine_cycles) == 1:
        return next(iter(machine_cycles)), "exact_machine_resolved", source_refs

    return None, "exact_ambiguous", source_refs


def compare_with_db_optimum(summary, comparison_context):
    machine_code = identifier(summary["machine_code"])
    machine_group = comparison_context["machine_groups"].get(machine_code)
    recommendation = to_decimal(summary["recommended_cycle_sec"])
    neck = to_decimal(summary["ağız"])
    gram = to_decimal(summary["gramaj"])

    base = {
        "machine_group": machine_group or "",
        "db_optimum_cycle_sec": None,
        "db_optimum_match_status": "machine_group_missing" if not machine_group else "",
        "db_optimum_delta_sec": None,
        "db_optimum_delta_percent": None,
        "db_optimum_source_rows": "",
        "diagnostic_optimum_cycle_sec": None,
        "diagnostic_match_status": "",
    }
    if not machine_group or neck is None or gram is None:
        return base

    key = (machine_group, neck.normalize(), gram.normalize())
    exact_entries = comparison_context["exact_entries"].get(key, [])
    optimum, status, source_refs = resolve_exact_optimum(exact_entries, machine_code)

    base["db_optimum_match_status"] = status
    base["db_optimum_source_rows"] = source_refs
    if optimum is not None:
        delta = recommendation - optimum if recommendation is not None else None
        base["db_optimum_cycle_sec"] = decimal_to_float(optimum)
        base["db_optimum_delta_sec"] = decimal_to_float(delta)
        base["db_optimum_delta_percent"] = decimal_to_float(
            ratio_delta(recommendation, optimum)
        )
        return base

    diagnostic_optimum, diagnostic_status = _lookup_optimum_cycle(
        comparison_context["lookup_entries"],
        machine_group=machine_group,
        machine_no=machine_code,
        neck_diameter=neck,
        gram=gram,
    )
    base["diagnostic_optimum_cycle_sec"] = decimal_to_float(diagnostic_optimum)
    base["diagnostic_match_status"] = diagnostic_status or "app_lookup_match"
    return base


def main():
    workbook = openpyxl.load_workbook(SOURCE, data_only=True, read_only=True)
    source_sheet = select_source_sheet(workbook)
    max_workbook_row = max(ws.max_row or 0 for ws in workbook.worksheets)

    workbook_observations, workbook_skipped = read_workbook_observations(source_sheet)
    db_observations, db_skipped, db_stats = read_db_observations()
    comparison_context = read_db_comparison_context()
    observations = workbook_observations + db_observations
    skipped = workbook_skipped + db_skipped

    all_cycles = [obs["cycle_time_sec"] for obs in observations]
    global_stats = stats(all_cycles)
    global_iqr = (global_stats["q3"] or 0) - (global_stats["q1"] or 0)
    global_lower = max(0, (global_stats["q1"] or 0) - 3 * global_iqr)
    global_upper = (global_stats["q3"] or 0) + 3 * global_iqr

    grouped = defaultdict(list)
    for obs in observations:
        key = (obs["machine_code"], obs["ağız"], obs["gramaj"])
        grouped[key].append(obs)

    issue_rows = []
    summaries = []
    detailed = []

    for key in sorted(grouped.keys(), key=lambda item: (str(item[0]), float(item[1]), float(item[2]))):
        group_observations = grouped[key]
        raw_values = [obs["cycle_time_sec"] for obs in group_observations]
        raw_stats = stats(raw_values)

        flags_by_id = defaultdict(list)
        for obs in group_observations:
            value = obs["cycle_time_sec"]
            if value < global_lower or value > global_upper:
                flags_by_id[obs["observation_id"]].append(
                    f"Global robust outlier ({round(global_lower, 2)}-{round(global_upper, 2)} sec)"
                )

        values_after_global = [
            obs["cycle_time_sec"] for obs in group_observations if not flags_by_id[obs["observation_id"]]
        ]
        if len(values_after_global) >= 4:
            group_q1 = quantile(values_after_global, 0.25)
            group_q3 = quantile(values_after_global, 0.75)
            group_iqr = group_q3 - group_q1
            lower = max(0, group_q1 - 1.5 * group_iqr)
            upper = group_q3 + 1.5 * group_iqr
            center = median(values_after_global)
            meaningful_delta = max(2, 0.15 * center)
            for obs in group_observations:
                if flags_by_id[obs["observation_id"]]:
                    continue
                value = obs["cycle_time_sec"]
                if (value < lower or value > upper) and abs(value - center) >= meaningful_delta:
                    flags_by_id[obs["observation_id"]].append(
                        f"Group IQR outlier ({round(lower, 2)}-{round(upper, 2)} sec)"
                    )

        clean_values = [
            obs["cycle_time_sec"] for obs in group_observations if not flags_by_id[obs["observation_id"]]
        ]
        clean_stats = stats(clean_values)
        recommendation = clean_stats["median"]
        products = sorted({obs["product"] for obs in group_observations})
        removed = len(raw_values) - len(clean_values)
        workbook_n = sum(1 for obs in group_observations if obs["source_type"] == "Old Workbook")
        db_n = sum(1 for obs in group_observations if obs["source_type"] == "Process DB")
        source_refs = []
        for obs in group_observations:
            if obs["source_type"] == "Process DB":
                source_refs.append(f"DB#{obs['source_row']}")
            else:
                source_refs.append(f"WB row {obs['source_row']}")

        summary = {
                "machine_code": key[0],
                "ağız": key[1],
                "gramaj": key[2],
                "recommended_cycle_sec": recommendation,
                "raw_sample_size": len(raw_values),
                "usable_sample_size": len(clean_values),
                "removed_sample_size": removed,
                "workbook_sample_size": workbook_n,
                "database_sample_size": db_n,
                "confidence": confidence(clean_values),
                "raw_min": raw_stats["min"],
                "raw_median": raw_stats["median"],
                "raw_mean": raw_stats["mean"],
                "raw_max": raw_stats["max"],
                "usable_min": clean_stats["min"],
                "usable_mean": clean_stats["mean"],
                "usable_max": clean_stats["max"],
                "usable_stdev": clean_stats["stdev"],
                "usable_cv": clean_stats["cv"],
                "source_rows": ", ".join(source_refs),
                "products_observed": " | ".join(products[:4]) + (" | ..." if len(products) > 4 else ""),
                "note": recommendation_note(clean_values, removed),
            }
        summary.update(compare_with_db_optimum(summary, comparison_context))
        summaries.append(summary)

        for obs in group_observations:
            flag = "; ".join(flags_by_id[obs["observation_id"]])
            detailed.append(
                {
                    **obs,
                    "status": "Excluded" if flag else "Usable",
                    "flag_reason": flag,
                    "group_key": f"{key[0]} / {key[1]}MM / {key[2]}GR",
                }
            )
            if flag:
                issue_rows.append(
                    {
                        "issue_type": "Excluded outlier",
                        "source_type": obs["source_type"],
                        "source_sheet": obs["source_sheet"],
                        "source_row": obs["source_row"],
                        "process_date": obs["process_date"],
                        "work_order": obs["work_order"],
                        "machine_code": obs["machine_code"],
                        "ağız": obs["ağız"],
                        "gramaj": obs["gramaj"],
                        "cycle_time_sec": obs["cycle_time_sec"],
                        "reason": flag,
                        "product": obs["product"],
                    }
                )

    for row in skipped:
        issue_rows.append({"issue_type": "Skipped source row", **row})

    if max_workbook_row < REQUESTED_START_ROW:
        issue_rows.insert(
            0,
            {
                "issue_type": "Requested row range unavailable",
                "source_type": "Old Workbook",
                "source_sheet": "Workbook",
                "source_row": None,
                "reason": (
                    f"Requested rows {REQUESTED_START_ROW}-{REQUESTED_END_ROW}, "
                    f"but the workbook's largest sheet ends at row {max_workbook_row}. "
                    f"Analysis uses worksheet '{source_sheet.title}', where columns F/G/J/K/L match the requested fields."
                ),
            },
        )

    db_match_status_counts = dict(
        sorted(
            {
                status: sum(
                    1
                    for row in summaries
                    if row["db_optimum_match_status"] == status
                )
                for status in {row["db_optimum_match_status"] for row in summaries}
            }.items()
        )
    )

    payload = {
        "metadata": {
            "source_path": str(SOURCE),
            "original_source_path": r"\\fileserver\PRODUCTION\ÜRETİM RAPOR FK\PROSES-YARDIMCI SİSTEMLER\PROSES 2026_FK.xlsx",
            "database_path": str(DB_PATH),
            "database_total_rows": db_stats["total_rows"],
            "source_sheet": source_sheet.title,
            "source_sheet_max_row": source_sheet.max_row,
            "max_workbook_row": max_workbook_row,
            "requested_rows": f"{REQUESTED_START_ROW}-{REQUESTED_END_ROW}",
            "observation_count": len(observations),
            "workbook_observation_count": len(workbook_observations),
            "database_observation_count": len(db_observations),
            "skipped_source_rows": len(skipped),
            "group_count": len(summaries),
            "cycle_table_row_count": comparison_context["cycle_table_row_count"],
            "db_optimum_exact_match_count": sum(
                1
                for row in summaries
                if row["db_optimum_match_status"]
                in {"exact_match", "exact_machine_resolved"}
            ),
            "db_optimum_exact_ambiguous_count": db_match_status_counts.get(
                "exact_ambiguous",
                0,
            ),
            "db_optimum_no_exact_match_count": db_match_status_counts.get(
                "no_exact_match",
                0,
            ),
            "db_optimum_machine_group_missing_count": db_match_status_counts.get(
                "machine_group_missing",
                0,
            ),
            "db_optimum_match_status_counts": db_match_status_counts,
            "global_cycle_q1": global_stats["q1"],
            "global_cycle_q3": global_stats["q3"],
            "global_outlier_lower": round(global_lower, 3),
            "global_outlier_upper": round(global_upper, 3),
            "method": (
                "Parsed Ağız from the first number before MM and Gramaj from the first number before GR in product text. "
                "Grouped by machine code, Ağız and Gramaj. Cycle time is column/field L only; J and K are total/active cavity counts. "
                "Recommended cycle is the median of usable cycle times after removing non-positive values, global robust outliers, "
                "and practically meaningful group IQR outliers where the group sample supports it. Current DB optimum comparison "
                "uses exact normalized machine group + Ağız + Gramaj matches; app-style nearest-match lookup is reported only "
                "as a diagnostic when exact matching does not return an authoritative optimum."
            ),
        },
        "summaries": summaries,
        "observations": detailed,
        "issues": issue_rows,
    }
    OUTPUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(payload["metadata"], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
