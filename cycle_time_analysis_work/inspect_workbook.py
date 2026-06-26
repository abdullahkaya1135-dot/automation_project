import json
import re
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

import openpyxl


SOURCE = Path(__file__).with_name("PROSES 2026_FK_source_copy.xlsx")


def main() -> None:
    wb = openpyxl.load_workbook(SOURCE, data_only=True, read_only=True)
    print("openpyxl sheets")
    for ws in wb.worksheets:
        print(json.dumps({"sheet": ws.title, "max_row": ws.max_row, "max_col": ws.max_column}, ensure_ascii=False))

    print("xlsx xml dimensions")
    with zipfile.ZipFile(SOURCE) as zf:
        workbook_xml = ET.fromstring(zf.read("xl/workbook.xml"))
        rels_xml = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        ns = {
            "a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
            "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        }
        relmap = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels_xml}
        for sheet in workbook_xml.find("a:sheets", ns):
            name = sheet.attrib["name"]
            rid = sheet.attrib["{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"]
            target = relmap[rid].lstrip("/")
            xml_path = target if target.startswith("xl/") else f"xl/{target}"
            data = zf.read(xml_path).decode("utf-8", errors="replace")
            dim = re.search(r'<dimension ref="([^"]+)"', data)
            max_row = 0
            for match in re.finditer(r'<row[^>]* r="(\d+)"', data):
                max_row = max(max_row, int(match.group(1)))
            print(
                json.dumps(
                    {
                        "sheet": name,
                        "xml": xml_path,
                        "dimension": dim.group(1) if dim else None,
                        "max_row_in_xml": max_row,
                    },
                    ensure_ascii=False,
                )
            )


if __name__ == "__main__":
    main()
