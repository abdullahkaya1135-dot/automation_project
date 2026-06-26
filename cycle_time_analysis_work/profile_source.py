import json
import re
from pathlib import Path

import openpyxl


SOURCE = Path(__file__).with_name("PROSES 2026_FK_source_copy.xlsx")
PRODUCT_RE = re.compile(r"\b\d+(?:[,.]\d+)?\s*MM\b.*\b\d+(?:[,.]\d+)?\s*GR\b", re.I)


def main() -> None:
    wb = openpyxl.load_workbook(SOURCE, data_only=True, read_only=True)
    ws = wb.worksheets[6]
    print(json.dumps({"using_sheet": ws.title, "max_row": ws.max_row, "max_col": ws.max_column}, ensure_ascii=False))

    sample_starts = {1, 2, 3, 4, 5, 10, 100, 1113, 1120, 1200, 1300, 1400, 1475}
    samples = {}
    matching = []
    values_by_row = {}

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=26, values_only=True), 1):
        if any(start <= row_idx <= min(start + 4, ws.max_row) for start in sample_starts):
            samples[row_idx] = row[5:12]
        g = row[6]
        if isinstance(g, str) and PRODUCT_RE.search(g):
            matching.append(row_idx)
            values_by_row[row_idx] = [row[5], row[6], row[9], row[10], row[11]]

    for start in sorted(sample_starts):
        print(f"--- rows {start} to {min(start + 4, ws.max_row)} F:L")
        for row_idx in range(start, min(start + 4, ws.max_row) + 1):
            if row_idx in samples:
                print(row_idx, json.dumps(samples[row_idx], ensure_ascii=False, default=str))

    print("--- rows with G contains MM and GR")
    print(json.dumps({"count": len(matching), "first10": matching[:10], "last10": matching[-10:]}, ensure_ascii=False))
    for row_idx in matching[:5] + matching[-5:]:
        print(row_idx, json.dumps(values_by_row[row_idx], ensure_ascii=False, default=str))


if __name__ == "__main__":
    main()
