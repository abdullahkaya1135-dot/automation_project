from datetime import date as Date
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from app.core.config import Settings
from app.core.database import create_session, init_db
from app.features.production_loss.service import (
    ProductionLossReportError,
    create_production_loss_report,
    get_production_loss_report,
    normalize_ifs_actual_rows,
    normalize_inventory_label_rows,
    normalize_operation_history_rows,
)
from app.models import (
    AmountControlShift,
    Entry,
    Machine,
    MachineBreakdown,
    ProductionLossIfsActual,
    ProductionLossReport,
    TourContext,
)


def _settings(tmp_path: Path) -> Settings:
    return Settings(
        excel_path=str(tmp_path / "process.xlsx"),
        app_pin="1234",
        sqlite_path=str(tmp_path / "process_entries.sqlite3"),
        backup_dir=str(tmp_path / "backups"),
        cycle_table_path=str(tmp_path / "cycle.xlsx"),
        report_output_dir=str(tmp_path / "reports"),
    )


def _create_cycle_table(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sayfa1"
    worksheet.append(
        [
            "Machine No",
            "Machine Name",
            "Neck",
            "GR",
            "Mold Code",
            "Production Date",
            "Estimated Cycle Time",
        ]
    )
    worksheet.append([174, "DPH70", 28, 35, "MOLD-A", None, 14])
    workbook.save(path)
    workbook.close()


def _seed_sources(settings: Settings) -> None:
    with create_session(settings) as session:
        machine = session.query(Machine).filter_by(machine_code="174").one()
        session.add_all(
            [
                AmountControlShift(
                    record_date="2026-06-08",
                    machine=machine,
                    job_order="WO-1",
                    shift="08.00-16.00",
                    worker_names="Operator One",
                    produced_quantity=999,
                ),
                Entry(
                    col_a="08.06.2026",
                    col_f="174",
                    col_g="Fallback Product",
                    col_h="WO-1",
                    col_k="2",
                    col_l="18",
                    process_date="2026-06-08",
                    machine_code="174",
                    sync_status="synced",
                    submitted_at=datetime(2026, 6, 8, 10, 0),
                ),
            ]
        )
        session.flush()
        amount_shift = (
            session.query(AmountControlShift)
            .filter_by(machine_id=machine.id, job_order="WO-1", shift="08.00-16.00")
            .one()
        )
        session.add(
            MachineBreakdown(
                machine=machine,
                amount_control_shift=amount_shift,
                produced_product="Fallback Product",
                stop_reason="Planned stop",
                duration_minutes=30,
            )
        )
        session.commit()


def _seed_process_entry(
    settings: Settings,
    *,
    machine_code: str = "174",
    job_order: str = "WO-1",
    active_cavities: str = "2",
    cycle_time: str | None = "12,5",
    product: str = "Fallback Product",
    submitted_at: datetime | None = None,
) -> None:
    with create_session(settings) as session:
        session.add(
            Entry(
                col_a="08.06.2026",
                col_f=machine_code,
                col_g=product,
                col_h=job_order,
                col_k=active_cavities,
                col_l=cycle_time,
                process_date="2026-06-08",
                machine_code=machine_code,
                sync_status="synced",
                submitted_at=submitted_at or datetime(2026, 6, 8, 10, 0),
            )
        )
        session.commit()


def _operation_history_row(
    *,
    transaction_id: str,
    job_order: str,
    quantity: int = 100,
    machine_code: str = "174",
    part_no: str = "MM-PET-35",
    time_of_production: str = "2026-06-08T05:00:00Z",
) -> dict[str, str]:
    return {
        "TransactionId": transaction_id,
        "TimeOfProduction": time_of_production,
        "QtyComplete": str(quantity),
        "ResourceId": machine_code,
        "OrderNo": job_order,
        "PartNo": part_no,
    }


def _label_stock_row(
    *,
    handling_unit_id: str,
    location_no: str,
    lot_batch_no: str = "2758-*-*-1",
    part_no: str = "MM-PET-35",
    part_description: str = "PET 28MM 35GR",
    qty_onhand: str = "0",
    pack_quantity: str = "0",
    last_activity: str = "2026-06-08T03:30:00+03:00",
    receipt_date: str = "2026-06-08T02:45:00+03:00",
) -> dict[str, str]:
    return {
        "Contract": "S01",
        "PartNo": part_no,
        "PartNoDesc": part_description,
        "LocationNo": location_no,
        "LotBatchNo": lot_batch_no,
        "QtyOnhand": qty_onhand,
        "Cf_Palet_Ici_Miktar": pack_quantity,
        "LastActivityDate": last_activity,
        "ReceiptDate": receipt_date,
        "HandlingUnitId": handling_unit_id,
    }


def _ifs_actual_row(
    *,
    job_order: str,
    machine_code: str = "174",
    work_center_no: str = "70DPH",
    part_no: str = "PET-35",
    part_description: str = "PET 28MM 35GR",
    actual_start: str = "2026-06-08T08:00:00+03:00",
    actual_finish: str = "2026-06-08T09:00:00+03:00",
    net_machine_minutes: int = 60,
) -> dict[str, object]:
    return {
        "OrderNo": job_order,
        "ReleaseNo": "*",
        "SequenceNo": "*",
        "OperationNo": 10,
        "PreferredResourceId": machine_code,
        "WorkCenterNo": work_center_no,
        "PartNo": part_no,
        "PartNoDesc": part_description,
        "ActualStartDate": actual_start,
        "ActualFinishDate": actual_finish,
        "NetMachineMinutes": net_machine_minutes,
    }


def _stub_production_loss_fetches(
    monkeypatch: pytest.MonkeyPatch,
    *,
    actual_rows: list[dict[str, object]],
    operation_history_rows: list[dict[str, str]],
    expected_order_numbers: list[str],
) -> None:
    async def fake_fetch_actuals(_settings, order_numbers):
        assert order_numbers == expected_order_numbers
        return actual_rows

    monkeypatch.setattr(
        "app.features.production_loss.service.fetch_shop_order_operation_actual_rows",
        fake_fetch_actuals,
    )

    async def fake_fetch_operation_history(
        _settings,
        *,
        date_from_utc,
        date_to_utc,
    ):
        assert date_from_utc.isoformat() == "2026-06-07T21:00:00+00:00"
        assert date_to_utc.isoformat() == "2026-06-08T21:00:00+00:00"
        return operation_history_rows

    monkeypatch.setattr(
        "app.features.production_loss.service.fetch_shop_order_operation_history_rows",
        fake_fetch_operation_history,
    )

    async def fake_fetch_part_descriptions(_settings, part_numbers):
        assert part_numbers == ["MM-PET-35"]
        return {"MM-PET-35": "PET 28MM 35GR"}

    monkeypatch.setattr(
        "app.features.production_loss.service.fetch_inventory_part_descriptions",
        fake_fetch_part_descriptions,
    )


def test_normalize_ifs_actual_rows_uses_realized_operation_statistics_fields():
    actuals = normalize_ifs_actual_rows(
        [
            {
                "OrderNo": "2718",
                "ReleaseNo": "*",
                "SequenceNo": "*",
                "OperationNo": "10",
                "WorkCenterNo": "PF-4",
                "PartNo": "MM-PET0156-01-045Y-085",
                "PlanStart": "2026-06-24T02:29:52Z",
                "PlanFinished": "2026-06-29T21:00:00Z",
                "RealStart": "2026-06-10T12:16:05Z",
                "RealFinished": "2026-06-16T21:33:30Z",
                "PartNoRef": {
                    "Description": "PET0156-01 4000ML SEFFAF SEFFAF 45MM YIVLI 85GR"
                },
                "WorkCenterNoRef": {"Description": "PF-4_141"},
            }
        ]
    )

    assert len(actuals) == 1
    actual = actuals[0]
    assert actual.job_order == "2718"
    assert actual.machine_code == "141"
    assert actual.work_center_no == "PF-4"
    assert actual.part_description == "PET0156-01 4000ML SEFFAF SEFFAF 45MM YIVLI 85GR"
    assert actual.actual_start_at == datetime(2026, 6, 10, 12, 16, 5)
    assert actual.actual_finish_at == datetime(2026, 6, 16, 21, 33, 30)


def test_normalize_operation_history_rows_uses_local_time_for_date_and_shift(tmp_path):
    settings = _settings(tmp_path)
    rows = [
        {
            "TransactionId": "TX-LOCAL-1",
            "TimeOfProduction": "2026-06-07T21:30:00Z",
            "QtyComplete": "12",
            "ResourceId": "174",
            "OrderNo": "WO-1",
            "PartNo": "MM-PET-35",
        },
        {
            "TransactionId": "TX-LOCAL-2",
            "TimeOfProduction": "2026-06-08T05:00:00Z",
            "QtyComplete": "18",
            "ResourceId": "174",
            "OrderNo": "WO-1",
            "PartNo": "MM-PET-35",
        },
        {
            "TransactionId": "TX-OUTSIDE",
            "TimeOfProduction": "2026-06-08T21:30:00Z",
            "QtyComplete": "99",
            "ResourceId": "174",
            "OrderNo": "WO-1",
            "PartNo": "MM-PET-35",
        },
        {
            "TransactionId": "TX-ZERO",
            "TimeOfProduction": "2026-06-08T06:00:00Z",
            "QtyComplete": "0",
            "ResourceId": "174",
            "OrderNo": "WO-1",
            "PartNo": "MM-PET-35",
        },
    ]

    events, parse_error_count, out_of_range_count = normalize_operation_history_rows(
        settings,
        rows,
        Date(2026, 6, 8),
        Date(2026, 6, 8),
        product_descriptions={"MM-PET-35": "PET 28MM 35GR"},
    )

    assert parse_error_count == 1
    assert out_of_range_count == 1
    assert [(event.record_date, event.shift, event.quantity) for event in events] == [
        ("2026-06-08", "24.00-08.00", 12),
        ("2026-06-08", "08.00-16.00", 18),
    ]
    assert events[0].time_of_production.isoformat(timespec="minutes") == (
        "2026-06-08T00:30"
    )
    assert events[0].product_description == "PET 28MM 35GR"


def test_normalize_inventory_label_rows_dedupes_stock_journey_by_physical_label(
    tmp_path,
):
    settings = _settings(tmp_path)
    rows = [
        _label_stock_row(
            handling_unit_id="HU-2758",
            location_no="U01",
            qty_onhand="12",
            last_activity="2026-06-08T03:30:00+03:00",
            receipt_date="2026-06-08T02:45:00+03:00",
        ),
        _label_stock_row(
            handling_unit_id="HU-2758",
            location_no="U03",
            qty_onhand="0",
            last_activity="2026-06-08T11:30:00+03:00",
        ),
        _label_stock_row(
            handling_unit_id="HU-2759",
            location_no="U01",
            lot_batch_no="2759-*-*-1",
            qty_onhand="0",
            pack_quantity="36",
            last_activity="2026-06-08T06:45:00+03:00",
            receipt_date="2026-06-08T06:45:00+03:00",
        ),
        _label_stock_row(
            handling_unit_id="HU-2760",
            location_no="U03",
            lot_batch_no="2760-*-*-1",
            qty_onhand="24",
            last_activity="2026-06-08T06:45:00+03:00",
        ),
    ]

    events, parse_error_count, out_of_range_count, detail_counts = (
        normalize_inventory_label_rows(
            settings,
            rows,
            Date(2026, 6, 8),
            Date(2026, 6, 8),
        )
    )

    assert parse_error_count == 0
    assert out_of_range_count == 0
    assert detail_counts == {
        "physical_unit_count": 3,
        "missing_u01_count": 1,
        "zero_quantity_count": 0,
    }
    assert [
        (
            event.result_key,
            event.record_date,
            event.shift,
            event.machine_code,
            event.job_order,
            event.quantity,
            event.label_time.isoformat(timespec="minutes"),
        )
        for event in events
    ] == [
        (
            "stock:HU-2758:MM-PET-35:2758-*-*-1",
            "2026-06-08",
            "24.00-08.00",
            "UNKNOWN",
            "2758",
            12,
            "2026-06-08T02:45",
        ),
        (
            "stock:HU-2759:MM-PET-35:2759-*-*-1",
            "2026-06-08",
            "24.00-08.00",
            "UNKNOWN",
            "2759",
            36,
            "2026-06-08T06:45",
        ),
    ]


def test_normalize_inventory_label_rows_uses_receipt_date_for_record_date_and_shift(
    tmp_path,
):
    settings = _settings(tmp_path)
    rows = [
        _label_stock_row(
            handling_unit_id="HU-2800",
            location_no="U01",
            lot_batch_no="2800-*-*-1",
            qty_onhand="0",
            pack_quantity="48",
            receipt_date="2026-06-08T07:59:00+03:00",
            last_activity="2026-06-09T08:01:00+03:00",
        ),
        _label_stock_row(
            handling_unit_id="HU-2800",
            location_no="100B14",
            lot_batch_no="2800-*-*-1",
            qty_onhand="48",
            receipt_date="2026-06-08T07:59:00+03:00",
            last_activity="2026-06-09T08:01:00+03:00",
        ),
        _label_stock_row(
            handling_unit_id="HU-2801",
            location_no="U01",
            lot_batch_no="2801-*-*-1",
            qty_onhand="24",
            receipt_date="2026-06-09T02:00:00+03:00",
            last_activity="2026-06-08T12:00:00+03:00",
        ),
        _label_stock_row(
            handling_unit_id="HU-2802",
            location_no="U01",
            lot_batch_no="2802-*-*-1",
            qty_onhand="12",
            receipt_date="",
            last_activity="2026-06-08T12:00:00+03:00",
        ),
    ]

    events, parse_error_count, out_of_range_count, detail_counts = (
        normalize_inventory_label_rows(
            settings,
            rows,
            Date(2026, 6, 8),
            Date(2026, 6, 8),
        )
    )

    assert parse_error_count == 1
    assert out_of_range_count == 1
    assert detail_counts == {
        "physical_unit_count": 3,
        "missing_u01_count": 0,
        "zero_quantity_count": 0,
    }
    assert len(events) == 1
    event = events[0]
    assert event.record_date == "2026-06-08"
    assert event.shift == "24.00-08.00"
    assert event.label_time.isoformat(timespec="minutes") == "2026-06-08T07:59"
    assert event.quantity == 48


def test_create_production_loss_report_prefers_label_stock_rows(
    tmp_path,
    monkeypatch,
):
    settings = _settings(tmp_path)
    init_db(settings)
    _seed_process_entry(
        settings,
        machine_code="174",
        job_order="2758",
        cycle_time="18",
        product="PET 28MM 35GR",
    )

    async def fake_fetch_label_stock(_settings, *, date_from, date_to):
        assert date_from == Date(2026, 6, 8)
        assert date_to == Date(2026, 6, 8)
        return [
            _label_stock_row(
                handling_unit_id="284844",
                location_no="U01",
                qty_onhand="0",
                pack_quantity="384",
                last_activity="2026-06-08T01:13:30+03:00",
                receipt_date="2026-06-08T00:59:04+03:00",
            ),
            _label_stock_row(
                handling_unit_id="284844",
                location_no="100B14",
                qty_onhand="384",
                pack_quantity="384",
                last_activity="2026-06-08T08:08:54+03:00",
                receipt_date="2026-06-08T00:59:04+03:00",
            ),
            _label_stock_row(
                handling_unit_id="284844",
                location_no="CEPDEPO",
                qty_onhand="0",
                pack_quantity="384",
                last_activity="2026-06-08T08:08:54+03:00",
                receipt_date="2026-06-08T00:59:04+03:00",
            ),
        ]

    monkeypatch.setattr(
        "app.features.production_loss.service.fetch_production_label_stock_rows",
        fake_fetch_label_stock,
    )

    async def fake_fetch_operation_history(*_args, **_kwargs):
        raise AssertionError("operation history should not be used when labels exist")

    monkeypatch.setattr(
        "app.features.production_loss.service.fetch_shop_order_operation_history_rows",
        fake_fetch_operation_history,
    )

    async def fake_fetch_actuals(_settings, order_numbers):
        assert order_numbers == ["2758"]
        return [
            _ifs_actual_row(
                job_order="2758",
                machine_code="174",
                actual_start="2026-06-08T00:00:00+03:00",
                actual_finish="2026-06-08T02:00:00+03:00",
                net_machine_minutes=120,
            )
        ]

    monkeypatch.setattr(
        "app.features.production_loss.service.fetch_shop_order_operation_actual_rows",
        fake_fetch_actuals,
    )

    result = create_production_loss_report(
        settings,
        date_from="2026-06-08",
        date_to="2026-06-08",
    )

    assert result.source_summary["quantity_source"] == "ifs-inventory-label-stock"
    assert result.source_summary["label_stock_row_count"] == 3
    assert result.source_summary["label_physical_unit_count"] == 1
    assert result.source_summary["label_event_count"] == 1
    assert result.source_summary["label_quantity_total"] == 384
    assert result.source_summary["process_match_count"] == 1
    assert result.row_count == 1

    row = result.rows[0]
    assert row["machine_code"] == "174"
    assert row["job_order"] == "2758"
    assert row["lot_batch_no"] == "2758-*-*-1"
    assert row["shift_2400_0800_actual_quantity"] == 384
    assert row["shift_0800_1600_actual_quantity"] == 0
    assert row["daily_total_quantity"] == 384


def test_create_production_loss_report_matches_label_cycle_by_job_order(
    tmp_path,
    monkeypatch,
):
    settings = _settings(tmp_path)
    init_db(settings)
    _seed_process_entry(
        settings,
        machine_code="174",
        job_order="2758",
        cycle_time="18",
        product="PET 28MM 35GR",
    )

    async def fake_fetch_label_stock(_settings, *, date_from, date_to):
        assert date_from == Date(2026, 6, 8)
        assert date_to == Date(2026, 6, 8)
        return [
            _label_stock_row(
                handling_unit_id="284844",
                location_no="U01",
                qty_onhand="0",
                pack_quantity="384",
            ),
            _label_stock_row(
                handling_unit_id="284844",
                location_no="100B14",
                qty_onhand="384",
                pack_quantity="384",
            ),
        ]

    monkeypatch.setattr(
        "app.features.production_loss.service.fetch_production_label_stock_rows",
        fake_fetch_label_stock,
    )

    async def fake_fetch_operation_history(*_args, **_kwargs):
        raise AssertionError("operation history should not be used when labels exist")

    monkeypatch.setattr(
        "app.features.production_loss.service.fetch_shop_order_operation_history_rows",
        fake_fetch_operation_history,
    )

    async def fake_fetch_actuals(_settings, order_numbers):
        assert order_numbers == ["2758"]
        return [
            _ifs_actual_row(
                job_order="2758",
                machine_code="",
                work_center_no="PKT",
                actual_start="2026-06-08T00:00:00+03:00",
                actual_finish="2026-06-08T00:00:00+03:00",
                net_machine_minutes=0,
            ),
            _ifs_actual_row(
                job_order="2758",
                machine_code="999",
                work_center_no="70DPH",
                actual_start="2026-06-08T00:00:00+03:00",
                actual_finish="2026-06-08T02:00:00+03:00",
                net_machine_minutes=120,
            ),
        ]

    monkeypatch.setattr(
        "app.features.production_loss.service.fetch_shop_order_operation_actual_rows",
        fake_fetch_actuals,
    )

    result = create_production_loss_report(
        settings,
        date_from="2026-06-08",
        date_to="2026-06-08",
    )

    assert result.row_count == 1
    assert result.warning_count == 0
    assert result.source_summary["process_match_count"] == 1
    assert result.source_summary["realized_cycle_valid_count"] == 1

    row = result.rows[0]
    assert row["machine_code"] == "999"
    assert row["job_order"] == "2758"
    assert row["active_cavities"] == "2"
    assert row["cycle_time_seconds"] == "18"
    assert row["shift_2400_0800_actual_quantity"] == 384
    assert row["shift_2400_0800_machine_minutes"] == "120"
    assert row["shift_2400_0800_optimum_quantity"] == "800"
    assert row["shift_2400_0800_loss_quantity"] == "416"


def test_create_production_loss_report_uses_realized_entry_cycle_for_loss_calculation(
    tmp_path,
    monkeypatch,
):
    settings = _settings(tmp_path)
    _create_cycle_table(tmp_path / "cycle.xlsx")
    init_db(settings)
    _seed_sources(settings)

    async def fake_fetch_actuals(_settings, order_numbers):
        assert order_numbers == ["WO-1"]
        return [
            {
                "OrderNo": "WO-1",
                "ReleaseNo": "*",
                "SequenceNo": "*",
                "OperationNo": 10,
                "PreferredResourceId": "174",
                "WorkCenterNo": "70DPH",
                "PartNo": "PET-35",
                "ActualStartDate": "2026-06-08T08:00:00+03:00",
                "ActualFinishDate": "2026-06-08T10:00:00+03:00",
                "NetMachineMinutes": 100,
            }
        ]

    monkeypatch.setattr(
        "app.features.production_loss.service.fetch_shop_order_operation_actual_rows",
        fake_fetch_actuals,
    )

    async def fake_fetch_operation_history(
        _settings,
        *,
        date_from_utc,
        date_to_utc,
    ):
        assert date_from_utc.isoformat() == "2026-06-07T21:00:00+00:00"
        assert date_to_utc.isoformat() == "2026-06-08T21:00:00+00:00"
        return [
            {
                "TransactionId": "TX-1",
                "TimeOfProduction": "2026-06-08T05:00:00Z",
                "QtyComplete": "40",
                "ResourceId": "174",
                "OrderNo": "WO-1",
                "PartNo": "MM-PET-35",
            },
            {
                "TransactionId": "TX-2",
                "TimeOfProduction": "2026-06-08T05:30:00Z",
                "QtyComplete": "60",
                "ResourceId": "174",
                "OrderNo": "WO-1",
                "PartNo": "MM-PET-35",
            },
            {
                "TransactionId": "TX-3",
                "TimeOfProduction": "2026-06-08T13:00:00Z",
                "QtyComplete": "80",
                "ResourceId": "174",
                "OrderNo": "WO-1",
                "PartNo": "MM-PET-35",
            },
            {
                "TransactionId": "TX-4",
                "TimeOfProduction": "2026-06-08T21:00:00Z",
                "QtyComplete": "999",
                "ResourceId": "174",
                "OrderNo": "WO-1",
                "PartNo": "MM-PET-35",
            },
        ]

    monkeypatch.setattr(
        "app.features.production_loss.service.fetch_shop_order_operation_history_rows",
        fake_fetch_operation_history,
    )

    async def fake_fetch_part_descriptions(_settings, part_numbers):
        assert part_numbers == ["MM-PET-35"]
        return {"MM-PET-35": "PET 28MM 35GR"}

    monkeypatch.setattr(
        "app.features.production_loss.service.fetch_inventory_part_descriptions",
        fake_fetch_part_descriptions,
    )

    result = create_production_loss_report(
        settings,
        date_from="2026-06-08",
        date_to="2026-06-08",
    )

    assert result.row_count == 1
    assert result.warning_count == 0
    assert result.source_summary["quantity_source"] == "ifs-operation-history"
    assert result.source_summary["operation_history_event_count"] == 3
    assert result.source_summary["operation_history_shift_group_count"] == 2
    assert result.source_summary["operation_history_quantity_total"] == 180
    assert result.source_summary["operation_history_out_of_range_count"] == 1
    assert result.source_summary["operation_history_product_description_count"] == 1
    assert result.source_summary["operation_history_product_description_error"] is None
    assert result.output_path.exists()
    row = result.rows[0]
    assert row["machine_code"] == "174"
    assert row["machine_name"] == "70DPH"
    assert row["product_description"] == "PET 28MM 35GR"
    assert row["gram"] == "35"
    assert row["daily_total_quantity"] == 180
    assert row["active_cavities"] == "2"
    assert row["cycle_time_seconds"] == "18"
    assert row["net_machine_minutes"] == "90"
    assert row["gross_elapsed_minutes"] == "120"
    assert row["shift_0800_1600_actual_quantity"] == 100
    assert row["shift_0800_1600_machine_minutes"] == "90"
    assert row["shift_0800_1600_optimum_quantity"] == "600"
    assert row["shift_0800_1600_loss_quantity"] == "500"
    assert row["shift_1600_2400_actual_quantity"] == 80
    assert row["shift_1600_2400_machine_minutes"] == "0"
    assert row["shift_1600_2400_optimum_quantity"] == "0"
    assert row["shift_1600_2400_loss_quantity"] == "-80"
    assert row["optimum_net_quantity"] == "600"
    assert row["production_loss_net"] == "420"
    assert row["optimum_gross_quantity"] == "800"
    assert row["production_loss_gross"] == "620"
    assert row["break_loss_quantity"] == "200"
    assert row["local_breakdown_minutes"] == 30

    workbook = load_workbook(result.output_path)
    worksheet = workbook["Production Loss"]
    assert worksheet["A2"].value == "2026-06-08"
    assert worksheet["T2"].value == 180
    assert worksheet["V2"].value == 18
    assert worksheet["Y2"].value == 600
    assert worksheet["Z2"].value == 420
    assert worksheet["AA2"].value == 800
    assert worksheet["AB2"].value == 620
    assert worksheet["AC2"].value == 200
    workbook.close()

    with create_session(settings) as session:
        amount_shift = (
            session.query(AmountControlShift)
            .filter_by(job_order="WO-1", shift="08.00-16.00")
            .one()
        )
        amount_shift.produced_quantity = 1
        session.commit()

    snapshot = get_production_loss_report(settings, result.report_id)
    assert snapshot is not None
    assert snapshot["rows"][0]["daily_total_quantity"] == 180


def test_create_production_loss_report_splits_standalone_timed_breakdown_by_shift(
    tmp_path,
    monkeypatch,
):
    settings = _settings(tmp_path)
    init_db(settings)
    with create_session(settings) as session:
        machine = session.query(Machine).filter_by(machine_code="174").one()
        context = TourContext(
            date="08.06.2026",
            ambient_temp="24",
            production_engineer="Operator",
            shift_chief="Selman",
            shift="08.00-16.00",
        )
        entry = Entry(
            tour_context=context,
            col_a="08.06.2026",
            col_f="174",
            col_g="PET 28MM 35GR",
            col_h="WO-1",
            col_k="1",
            col_l="60",
            process_date="2026-06-08",
            machine_code="174",
            sync_status="synced",
            submitted_at=datetime(2026, 6, 8, 10, 0),
        )
        session.add(
            MachineBreakdown(
                machine=machine,
                entry=entry,
                produced_product="PET 28MM 35GR",
                stop_reason="Standalone stop",
                duration_minutes=60,
                stopped_at=datetime(2026, 6, 8, 4, 30),
                resumed_at=datetime(2026, 6, 8, 5, 30),
            )
        )
        session.commit()

    _stub_production_loss_fetches(
        monkeypatch,
        actual_rows=[
            _ifs_actual_row(
                job_order="WO-1",
                actual_start="2026-06-08T07:00:00+03:00",
                actual_finish="2026-06-08T09:00:00+03:00",
                net_machine_minutes=120,
            )
        ],
        operation_history_rows=[
            _operation_history_row(
                transaction_id="TX-1",
                job_order="WO-1",
                quantity=10,
                time_of_production="2026-06-08T04:30:00Z",
            ),
            _operation_history_row(
                transaction_id="TX-2",
                job_order="WO-1",
                quantity=20,
                time_of_production="2026-06-08T05:30:00Z",
            ),
        ],
        expected_order_numbers=["WO-1"],
    )

    result = create_production_loss_report(
        settings,
        date_from="2026-06-08",
        date_to="2026-06-08",
    )

    row = result.rows[0]
    assert row["shift_2400_0800_actual_quantity"] == 10
    assert row["shift_2400_0800_machine_minutes"] == "30"
    assert row["shift_2400_0800_optimum_quantity"] == "30"
    assert row["shift_2400_0800_loss_quantity"] == "20"
    assert row["shift_0800_1600_actual_quantity"] == 20
    assert row["shift_0800_1600_machine_minutes"] == "30"
    assert row["shift_0800_1600_optimum_quantity"] == "30"
    assert row["shift_0800_1600_loss_quantity"] == "10"
    assert row["net_machine_minutes"] == "60"
    assert row["gross_elapsed_minutes"] == "120"
    assert row["production_loss_net"] == "30"
    assert row["production_loss_gross"] == "90"
    assert row["break_loss_quantity"] == "60"
    assert row["local_breakdown_minutes"] == 60


def test_create_production_loss_report_allocates_untimed_standalone_breakdown_by_unique_shift(
    tmp_path,
    monkeypatch,
):
    settings = _settings(tmp_path)
    init_db(settings)
    _seed_process_entry(settings, job_order="WO-1", cycle_time="60", active_cavities="1")
    _seed_process_entry(settings, job_order="WO-2", cycle_time="60", active_cavities="1")
    _seed_process_entry(settings, job_order="WO-3", cycle_time="60", active_cavities="1")
    with create_session(settings) as session:
        machine = session.query(Machine).filter_by(machine_code="174").one()
        unique_context = TourContext(
            date="08.06.2026",
            ambient_temp="24",
            production_engineer="Operator",
            shift_chief="Selman",
            shift="08.00-16.00",
        )
        ambiguous_context = TourContext(
            date="08.06.2026",
            ambient_temp="24",
            production_engineer="Operator",
            shift_chief="Selman",
            shift="16.00-24.00",
        )
        unique_entry = Entry(
            tour_context=unique_context,
            col_a="08.06.2026",
            col_f="174",
            col_g="PET 28MM 35GR",
            process_date="2026-06-08",
            machine_code="174",
            sync_status="synced",
        )
        ambiguous_entry = Entry(
            tour_context=ambiguous_context,
            col_a="08.06.2026",
            col_f="174",
            col_g="PET 28MM 35GR",
            process_date="2026-06-08",
            machine_code="174",
            sync_status="synced",
        )
        session.add_all(
            [
                MachineBreakdown(
                    machine=machine,
                    entry=unique_entry,
                    produced_product="PET 28MM 35GR",
                    stop_reason="No order unique shift",
                    duration_minutes=20,
                ),
                MachineBreakdown(
                    machine=machine,
                    entry=ambiguous_entry,
                    produced_product="PET 28MM 35GR",
                    stop_reason="No order ambiguous shift",
                    duration_minutes=40,
                ),
            ]
        )
        session.commit()

    _stub_production_loss_fetches(
        monkeypatch,
        actual_rows=[
            _ifs_actual_row(
                job_order="WO-1",
                actual_start="2026-06-08T08:00:00+03:00",
                actual_finish="2026-06-08T09:00:00+03:00",
            ),
            _ifs_actual_row(
                job_order="WO-2",
                actual_start="2026-06-08T16:00:00+03:00",
                actual_finish="2026-06-08T17:00:00+03:00",
            ),
            _ifs_actual_row(
                job_order="WO-3",
                actual_start="2026-06-08T17:00:00+03:00",
                actual_finish="2026-06-08T18:00:00+03:00",
            ),
        ],
        operation_history_rows=[
            _operation_history_row(
                transaction_id="TX-1",
                job_order="WO-1",
                time_of_production="2026-06-08T05:30:00Z",
            ),
            _operation_history_row(
                transaction_id="TX-2",
                job_order="WO-2",
                time_of_production="2026-06-08T13:30:00Z",
            ),
            _operation_history_row(
                transaction_id="TX-3",
                job_order="WO-3",
                time_of_production="2026-06-08T14:30:00Z",
            ),
        ],
        expected_order_numbers=["WO-1", "WO-2", "WO-3"],
    )

    result = create_production_loss_report(
        settings,
        date_from="2026-06-08",
        date_to="2026-06-08",
    )

    rows = {row["job_order"]: row for row in result.rows}
    assert rows["WO-1"]["local_breakdown_minutes"] == 20
    assert rows["WO-1"]["shift_0800_1600_machine_minutes"] == "40"
    assert rows["WO-2"]["local_breakdown_minutes"] == 0
    assert rows["WO-2"]["shift_1600_2400_machine_minutes"] == "60"
    assert rows["WO-3"]["local_breakdown_minutes"] == 0
    assert rows["WO-3"]["shift_1600_2400_machine_minutes"] == "60"


def test_create_production_loss_report_does_not_double_count_linked_breakdown(
    tmp_path,
    monkeypatch,
):
    settings = _settings(tmp_path)
    init_db(settings)
    with create_session(settings) as session:
        machine = session.query(Machine).filter_by(machine_code="174").one()
        context = TourContext(
            date="08.06.2026",
            ambient_temp="24",
            production_engineer="Operator",
            shift_chief="Selman",
            shift="08.00-16.00",
        )
        entry = Entry(
            tour_context=context,
            col_a="08.06.2026",
            col_f="174",
            col_g="PET 28MM 35GR",
            col_h="WO-1",
            col_k="1",
            col_l="60",
            process_date="2026-06-08",
            machine_code="174",
            sync_status="synced",
            submitted_at=datetime(2026, 6, 8, 10, 0),
        )
        amount_shift = AmountControlShift(
            record_date="2026-06-08",
            machine=machine,
            job_order="WO-1",
            shift="08.00-16.00",
            worker_names="Operator One",
            produced_quantity=100,
        )
        session.add_all(
            [
                amount_shift,
                MachineBreakdown(
                    machine=machine,
                    entry=entry,
                    amount_control_shift=amount_shift,
                    produced_product="PET 28MM 35GR",
                    stop_reason="Linked stop",
                    duration_minutes=15,
                ),
            ]
        )
        session.commit()

    _stub_production_loss_fetches(
        monkeypatch,
        actual_rows=[
            _ifs_actual_row(
                job_order="WO-1",
                actual_start="2026-06-08T08:00:00+03:00",
                actual_finish="2026-06-08T09:00:00+03:00",
            )
        ],
        operation_history_rows=[
            _operation_history_row(
                transaction_id="TX-1",
                job_order="WO-1",
                time_of_production="2026-06-08T05:30:00Z",
            )
        ],
        expected_order_numbers=["WO-1"],
    )

    result = create_production_loss_report(
        settings,
        date_from="2026-06-08",
        date_to="2026-06-08",
    )

    row = result.rows[0]
    assert row["local_breakdown_minutes"] == 15
    assert row["shift_0800_1600_machine_minutes"] == "45"
    assert row["production_loss_net"] == "-55"
    assert row["production_loss_gross"] == "-40"
    assert row["break_loss_quantity"] == "15"


def test_create_production_loss_report_uses_realized_operation_statistics_window(
    tmp_path,
    monkeypatch,
):
    settings = _settings(tmp_path)
    init_db(settings)
    _seed_process_entry(settings, job_order="WO-1", cycle_time="20")
    _stub_production_loss_fetches(
        monkeypatch,
        actual_rows=[
            {
                "OrderNo": "WO-1",
                "ReleaseNo": "*",
                "SequenceNo": "*",
                "OperationNo": "10",
                "WorkCenterNo": "PF-4",
                "PartNo": "MM-PET-35",
                "RealStart": "2026-06-08T13:00:00Z",
                "RealFinished": "2026-06-08T15:00:00Z",
                "PlanStart": "2026-06-24T02:29:52Z",
                "PlanFinished": "2026-06-29T21:00:00Z",
                "PartNoRef": {"Description": "PET 28MM 35GR"},
                "WorkCenterNoRef": {"Description": "PF-4_174"},
            }
        ],
        operation_history_rows=[
            _operation_history_row(
                transaction_id="TX-1",
                job_order="WO-1",
                quantity=80,
                time_of_production="2026-06-08T13:30:00Z",
            )
        ],
        expected_order_numbers=["WO-1"],
    )

    result = create_production_loss_report(
        settings,
        date_from="2026-06-08",
        date_to="2026-06-08",
    )

    row = result.rows[0]
    assert row["machine_code"] == "174"
    assert row["machine_name"] == "PF-4"
    assert row["shift_1600_2400_actual_quantity"] == 80
    assert row["shift_1600_2400_machine_minutes"] == "120"
    assert row["shift_1600_2400_optimum_quantity"] == "720"
    assert row["shift_1600_2400_loss_quantity"] == "640"
    assert row["net_machine_minutes"] == "120"


def test_create_production_loss_report_calculates_completed_shifts_for_open_actual(
    tmp_path,
    monkeypatch,
):
    settings = _settings(tmp_path)
    init_db(settings)
    _seed_process_entry(
        settings,
        job_order="WO-1",
        cycle_time="60",
        active_cavities="1",
    )
    monkeypatch.setattr(
        "app.features.production_loss.service._current_local_datetime",
        lambda _settings: datetime(2026, 6, 8, 17, 0),
    )
    _stub_production_loss_fetches(
        monkeypatch,
        actual_rows=[
            {
                "OrderNo": "WO-1",
                "PartDescription": "PET 28MM 35GR",
                "RealStart": "2026-06-08T05:00:00Z",
                "RealFinished": None,
                "RealMachRunTime": 480,
                "InterruptionTime": 0,
            }
        ],
        operation_history_rows=[
            _operation_history_row(
                transaction_id="TX-1",
                job_order="WO-1",
                quantity=100,
                time_of_production="2026-06-08T07:00:00Z",
            )
        ],
        expected_order_numbers=["WO-1"],
    )

    result = create_production_loss_report(
        settings,
        date_from="2026-06-08",
        date_to="2026-06-08",
    )

    row = result.rows[0]
    assert row["product_description"] == "PET 28MM 35GR"
    assert row["shift_0800_1600_actual_quantity"] == 100
    assert row["shift_0800_1600_machine_minutes"] == "480"
    assert row["shift_0800_1600_optimum_quantity"] == "480"
    assert row["shift_0800_1600_loss_quantity"] == "380"
    assert row["shift_1600_2400_actual_quantity"] == 0
    assert row["shift_1600_2400_machine_minutes"] is None
    assert row["net_machine_minutes"] == "480"
    assert "continuing production" in row["warnings"]


def test_create_production_loss_report_ignores_planned_window_for_empty_shift_loss(
    tmp_path,
    monkeypatch,
):
    settings = _settings(tmp_path)
    init_db(settings)
    _seed_process_entry(settings, job_order="WO-1", cycle_time="20")
    actual_row = _ifs_actual_row(
        job_order="WO-1",
        actual_start="2026-06-08T08:00:00+03:00",
        actual_finish="2026-06-08T10:00:00+03:00",
        net_machine_minutes=120,
    )
    actual_row.update(
        {
            "StartDate": "2026-06-08T16:00:00+03:00",
            "FinishDate": "2026-06-08T18:00:00+03:00",
        }
    )
    _stub_production_loss_fetches(
        monkeypatch,
        actual_rows=[actual_row],
        operation_history_rows=[
            _operation_history_row(
                transaction_id="TX-1",
                job_order="WO-1",
                quantity=80,
                time_of_production="2026-06-08T05:30:00Z",
            )
        ],
        expected_order_numbers=["WO-1"],
    )

    result = create_production_loss_report(
        settings,
        date_from="2026-06-08",
        date_to="2026-06-08",
    )

    row = result.rows[0]
    assert row["shift_0800_1600_actual_quantity"] == 80
    assert row["shift_0800_1600_machine_minutes"] == "120"
    assert row["shift_0800_1600_optimum_quantity"] == "720"
    assert row["shift_0800_1600_loss_quantity"] == "640"
    assert row["shift_1600_2400_actual_quantity"] == 0
    assert row["shift_1600_2400_machine_minutes"] == "0"
    assert row["shift_1600_2400_optimum_quantity"] == "0"
    assert row["shift_1600_2400_loss_quantity"] == "0"


@pytest.mark.parametrize(
    ("cycle_time", "missing_count", "invalid_count"),
    [
        (None, 1, 0),
        ("", 1, 0),
        ("not-a-number", 0, 1),
        ("0", 0, 1),
        ("-1", 0, 1),
    ],
    ids=["missing", "blank", "invalid", "zero", "negative"],
)
def test_create_production_loss_report_fails_without_any_valid_realized_cycle(
    tmp_path,
    monkeypatch,
    cycle_time,
    missing_count,
    invalid_count,
):
    settings = _settings(tmp_path)
    _create_cycle_table(tmp_path / "cycle.xlsx")
    init_db(settings)
    _seed_process_entry(settings, job_order="WO-1", cycle_time=cycle_time)
    _stub_production_loss_fetches(
        monkeypatch,
        actual_rows=[_ifs_actual_row(job_order="WO-1")],
        operation_history_rows=[
            _operation_history_row(transaction_id="TX-1", job_order="WO-1")
        ],
        expected_order_numbers=["WO-1"],
    )

    with pytest.raises(ProductionLossReportError) as exc_info:
        create_production_loss_report(
            settings,
            date_from="2026-06-08",
            date_to="2026-06-08",
        )

    message = str(exc_info.value)
    assert "Hesaplanabilir Production Loss satiri yok" in message
    assert "1 satir gerceklesen cevrim eksik/gecersiz oldugu icin atlandi" in message
    assert f"({missing_count} eksik, {invalid_count} gecersiz)" in message
    with create_session(settings) as session:
        assert session.query(ProductionLossReport).count() == 0
        assert session.query(ProductionLossIfsActual).count() == 0
    assert not Path(settings.report_output_dir).exists()


def test_create_production_loss_report_skips_rows_without_realized_cycle(
    tmp_path,
    monkeypatch,
):
    settings = _settings(tmp_path)
    _create_cycle_table(tmp_path / "cycle.xlsx")
    init_db(settings)
    _seed_process_entry(settings, job_order="WO-1", cycle_time="20")
    _stub_production_loss_fetches(
        monkeypatch,
        actual_rows=[
            _ifs_actual_row(job_order="WO-1"),
            _ifs_actual_row(job_order="WO-2"),
        ],
        operation_history_rows=[
            _operation_history_row(transaction_id="TX-1", job_order="WO-1"),
            _operation_history_row(
                transaction_id="TX-2",
                job_order="WO-2",
                quantity=40,
            ),
        ],
        expected_order_numbers=["WO-1", "WO-2"],
    )

    result = create_production_loss_report(
        settings,
        date_from="2026-06-08",
        date_to="2026-06-08",
    )

    assert result.row_count == 1
    assert result.warning_count == 1
    row = result.rows[0]
    assert row["job_order"] == "WO-1"
    assert row["cycle_time_seconds"] == "20"
    assert row["optimum_net_quantity"] == "360"
    assert row["production_loss_net"] == "260"
    assert {row["job_order"] for row in result.rows} == {"WO-1"}
    assert result.source_summary["realized_cycle_source"] == "process-entry-col-l"
    assert result.source_summary["realized_cycle_valid_count"] == 1
    assert result.source_summary["realized_cycle_missing_count"] == 1
    assert result.source_summary["realized_cycle_invalid_count"] == 0
    assert result.source_summary["realized_cycle_skipped_count"] == 1
    assert result.source_summary["realized_cycle_included_quantity_total"] == 100
    assert result.source_summary["realized_cycle_skipped_quantity_total"] == 40
    assert (
        "gerceklesen cevrim eksik/gecersiz"
        in result.source_summary["realized_cycle_skip_message"]
    )
    skip_detail = result.source_summary["realized_cycle_skip_details"][0]
    assert skip_detail == {
        "record_date": "2026-06-08",
        "machine_code": "174",
        "job_order": "WO-2",
        "entry_id": None,
        "raw_cycle": None,
        "reason": "missing-process-entry",
        "message": "Gerceklesen cevrim eksik",
    }
    assert result.source_summary["realized_cycle_skip_details"] == [
        {
            "record_date": "2026-06-08",
            "machine_code": "174",
            "job_order": "WO-2",
            "entry_id": None,
            "raw_cycle": None,
            "reason": "missing-process-entry",
            "message": "Gerceklesen cevrim eksik",
        }
    ]
    snapshot = get_production_loss_report(settings, result.report_id)
    assert snapshot is not None
    assert snapshot["source_summary"] == result.source_summary


def test_create_production_loss_report_uses_latest_realized_cycle_entry(
    tmp_path,
    monkeypatch,
):
    settings = _settings(tmp_path)
    init_db(settings)
    _seed_process_entry(
        settings,
        job_order="WO-1",
        cycle_time="30",
        submitted_at=datetime(2026, 6, 8, 9, 0),
    )
    _seed_process_entry(
        settings,
        job_order="WO-1",
        cycle_time="10",
        submitted_at=datetime(2026, 6, 8, 11, 0),
    )
    _seed_process_entry(
        settings,
        job_order="WO-2",
        cycle_time="5",
        submitted_at=datetime(2026, 6, 8, 12, 0),
    )
    _stub_production_loss_fetches(
        monkeypatch,
        actual_rows=[_ifs_actual_row(job_order="WO-1")],
        operation_history_rows=[
            _operation_history_row(transaction_id="TX-1", job_order="WO-1")
        ],
        expected_order_numbers=["WO-1"],
    )

    result = create_production_loss_report(
        settings,
        date_from="2026-06-08",
        date_to="2026-06-08",
    )

    assert result.row_count == 1
    assert result.warning_count == 0
    row = result.rows[0]
    assert row["cycle_time_seconds"] == "10"
    assert row["optimum_net_quantity"] == "720"
    assert row["production_loss_net"] == "620"


def test_create_production_loss_report_ignores_missing_cycle_workbook_with_realized_cycles(
    tmp_path,
    monkeypatch,
):
    settings = _settings(tmp_path)
    init_db(settings)
    assert not (tmp_path / "cycle.xlsx").exists()
    _seed_process_entry(settings, job_order="WO-1", cycle_time="12,5")
    _stub_production_loss_fetches(
        monkeypatch,
        actual_rows=[_ifs_actual_row(job_order="WO-1")],
        operation_history_rows=[
            _operation_history_row(transaction_id="TX-1", job_order="WO-1")
        ],
        expected_order_numbers=["WO-1"],
    )

    result = create_production_loss_report(
        settings,
        date_from="2026-06-08",
        date_to="2026-06-08",
    )

    assert result.row_count == 1
    assert result.warning_count == 0
    assert result.rows[0]["cycle_time_seconds"] == "12.5"
    assert result.rows[0]["production_loss_net"] == "476"
    assert result.output_path.exists()


def test_create_production_loss_report_ignores_cycle_workbook_without_cycle_table(
    tmp_path,
    monkeypatch,
):
    settings = _settings(tmp_path)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Unrelated"
    worksheet.append(["not", "the", "cycle", "table"])
    workbook.save(tmp_path / "cycle.xlsx")
    workbook.close()
    init_db(settings)
    _seed_process_entry(settings, job_order="WO-1", cycle_time="12,5")
    _stub_production_loss_fetches(
        monkeypatch,
        actual_rows=[_ifs_actual_row(job_order="WO-1")],
        operation_history_rows=[
            _operation_history_row(transaction_id="TX-1", job_order="WO-1")
        ],
        expected_order_numbers=["WO-1"],
    )

    result = create_production_loss_report(
        settings,
        date_from="2026-06-08",
        date_to="2026-06-08",
    )

    assert result.row_count == 1
    assert result.warning_count == 0
    assert result.rows[0]["cycle_time_seconds"] == "12.5"
    assert result.output_path.exists()


def test_create_production_loss_report_requires_operation_history_rows(tmp_path):
    settings = _settings(tmp_path)
    _create_cycle_table(tmp_path / "cycle.xlsx")
    init_db(settings)

    with pytest.raises(
        ProductionLossReportError,
        match=(r"^Secilen tarih araliginda IFS etiket stok hareketi bulunamadi\.$"),
    ):
        create_production_loss_report(
            settings,
            date_from="2026-06-08",
            date_to="2026-06-08",
            refresh_ifs=False,
            refresh_labels=False,
        )


def test_create_production_loss_report_preserves_operation_history_fetch_error(
    tmp_path,
    monkeypatch,
):
    settings = _settings(tmp_path)
    init_db(settings)

    async def fake_fetch_label_stock(
        _settings,
        *,
        date_from,
        date_to,
    ):
        return []

    monkeypatch.setattr(
        "app.features.production_loss.service.fetch_production_label_stock_rows",
        fake_fetch_label_stock,
    )

    async def fake_fetch_operation_history(
        _settings,
        *,
        date_from_utc,
        date_to_utc,
    ):
        raise RuntimeError("IFS operation history unavailable")

    monkeypatch.setattr(
        "app.features.production_loss.service.fetch_shop_order_operation_history_rows",
        fake_fetch_operation_history,
    )

    with pytest.raises(ProductionLossReportError) as exc_info:
        create_production_loss_report(
            settings,
            date_from="2026-06-08",
            date_to="2026-06-08",
            refresh_ifs=False,
        )

    message = str(exc_info.value)
    assert "Yedek IFS operasyon gecmisi de alinamadi" in message
    assert "IFS operation history unavailable" in message
