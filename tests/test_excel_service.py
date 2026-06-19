from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.workbook.workbook import Workbook as OpenpyxlWorkbook

from app.core.config import Settings
from app.core.database import commit_session, create_session, init_db
from app.features.process_entries.sync import (
    attempt_excel_append,
    retry_unsynced_entries,
)
from app.models import (
    SYNC_STATUS_FAILED_EXCEL,
    SYNC_STATUS_PENDING_EXCEL,
    Entry,
)
from app.services.excel_service import (
    WorkbookLockedError,
    WorkbookMissingError,
    WorkbookPermissionError,
    WorkbookStructureError,
    append_entries_to_workbook,
    append_entry_to_workbook,
    check_excel_reachable,
    detect_last_value_row,
    normalize_excel_value,
    validate_headers,
)

HEADERS = [
    "Date",
    "Ambient Temp",
    "Production Engineer",
    "Shift Chief",
    "Shift",
    "Machine",
    "Product",
    "Work Order",
    "Raw Material",
    "Total Cavity Count",
    "Active Cavity Count",
    "Cycle Time",
    "Cooling Time",
    "Injection Time",
    "Blow Time",
    "Conditioner Temp",
    "Dryer Temp",
    "Injection Pressure",
    "Injection Speed",
    "Holding Time",
    "Holding Speed",
    "Holding Pressure",
    "Clamp Force",
    "Oven Temperatures",
    "Mold Temperatures",
]


def _create_workbook(path: Path, *, last_data_row: int = 1725) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "PROSES 2026"
    worksheet.append(HEADERS)
    worksheet.cell(row=last_data_row, column=1, value="existing")
    worksheet.cell(row=1_048_506, column=25).fill = PatternFill(
        fill_type="solid",
        fgColor="FFFF00",
    )
    workbook.save(path)
    workbook.close()


def _settings(workbook_path: Path, tmp_path: Path, *, keep_count: int = 20) -> Settings:
    return Settings(
        excel_path=str(workbook_path),
        app_pin="1234",
        sqlite_path=str(tmp_path / "process_entries.sqlite3"),
        backup_dir=str(tmp_path / "backups"),
        backup_keep_count=keep_count,
    )


def _payload() -> dict[str, str]:
    return {
        f"col_{letter}": f"value-{letter}"
        for letter in "abcdefghijklmnopqrstuvwxy"
    } | {
        "col_a": "2026-06-08",
        "col_b": "24,5",
        "col_f": "101",
        "col_g": "Product",
        "col_h": "001234",
        "col_i": "HM-02",
        "col_j": "16",
        "col_k": "12",
        "col_l": "12.5",
        "col_m": "7",
        "col_n": "4",
        "col_o": "6",
        "col_p": "45",
        "col_q": "180",
        "col_r": "70-69",
        "col_s": "50-49",
        "col_t": "1.5",
        "col_u": "25",
        "col_v": "30",
        "col_w": "125",
        "col_x": "180-190",
        "col_y": "",
    }


def test_detect_last_value_row_ignores_formatting_only_rows(tmp_path):
    workbook_path = tmp_path / "process.xlsx"
    _create_workbook(workbook_path)

    workbook = load_workbook(workbook_path)
    worksheet = workbook["PROSES 2026"]

    assert worksheet.max_row == 1_048_506
    assert detect_last_value_row(worksheet) == 1725

    workbook.close()


def test_validate_headers_rejects_changed_workbook_shape(tmp_path):
    workbook_path = tmp_path / "process.xlsx"
    _create_workbook(workbook_path)

    workbook = load_workbook(workbook_path)
    worksheet = workbook["PROSES 2026"]
    worksheet["F1"] = "Unexpected"

    with pytest.raises(WorkbookStructureError):
        validate_headers(worksheet)

    workbook.close()


def test_check_excel_reachable_reports_header_mismatch(tmp_path):
    workbook_path = tmp_path / "process.xlsx"
    _create_workbook(workbook_path)

    workbook = load_workbook(workbook_path)
    worksheet = workbook["PROSES 2026"]
    worksheet["G1"] = "Unexpected"
    workbook.save(workbook_path)
    workbook.close()

    status = check_excel_reachable(_settings(workbook_path, tmp_path))

    assert status.available is False
    assert "A:Y" in status.error


def test_normalize_excel_value_handles_decimal_and_range_text():
    assert normalize_excel_value("24,5", "B") == 24.5
    assert normalize_excel_value("12.5", "L") == 12.5
    assert normalize_excel_value("7", "J") == 7
    assert normalize_excel_value("180-190", "X") == "180-190"
    assert normalize_excel_value("001234", "H") == "001234"
    assert normalize_excel_value("  ", "Y") is None


def test_append_entry_to_workbook_writes_next_real_row_and_creates_backup(tmp_path):
    workbook_path = tmp_path / "process.xlsx"
    _create_workbook(workbook_path)
    settings = _settings(workbook_path, tmp_path)

    row_number = append_entry_to_workbook(settings, _payload())

    assert row_number == 1726
    backup_files = list((tmp_path / "backups").glob("process_*.xlsx"))
    assert len(backup_files) == 1

    workbook = load_workbook(workbook_path)
    worksheet = workbook["PROSES 2026"]
    header_values = [
        worksheet.cell(row=1, column=index).value for index in range(1, 26)
    ]
    row_values = [
        worksheet.cell(row=1726, column=index).value for index in range(1, 26)
    ]
    expected_row_values = [
        "2026-06-08",
        24.5,
        "value-c",
        "value-d",
        "value-e",
        "101",
        "Product",
        "001234",
        "HM-02",
        16,
        12,
        12.5,
        7,
        4,
        6,
        45,
        180,
        None,
        None,
        None,
        None,
        None,
        None,
        "180-190",
        None,
    ]
    assert header_values == HEADERS
    assert row_values == expected_row_values
    workbook.close()

    backup = load_workbook(backup_files[0])
    backup_worksheet = backup["PROSES 2026"]
    assert backup_worksheet.cell(row=1726, column=1).value is None
    backup.close()


def test_append_entries_to_workbook_writes_bulk_rows_with_one_backup(tmp_path):
    workbook_path = tmp_path / "process.xlsx"
    _create_workbook(workbook_path)
    settings = _settings(workbook_path, tmp_path)
    first_payload = _payload()
    second_payload = _payload() | {"col_f": "102", "col_h": "WO-2"}

    row_numbers = append_entries_to_workbook(
        settings,
        [first_payload, second_payload],
    )

    assert row_numbers == [1726, 1727]
    assert len(list((tmp_path / "backups").glob("process_*.xlsx"))) == 1

    workbook = load_workbook(workbook_path)
    worksheet = workbook["PROSES 2026"]
    assert worksheet.cell(row=1726, column=6).value == "101"
    assert worksheet.cell(row=1727, column=6).value == "102"
    assert worksheet.cell(row=1727, column=8).value == "WO-2"
    workbook.close()


def test_append_entry_to_workbook_prunes_old_backups(tmp_path):
    workbook_path = tmp_path / "process.xlsx"
    _create_workbook(workbook_path)
    settings = _settings(workbook_path, tmp_path, keep_count=2)

    append_entry_to_workbook(settings, _payload())
    append_entry_to_workbook(settings, _payload())
    append_entry_to_workbook(settings, _payload())

    assert len(list((tmp_path / "backups").glob("process_*.xlsx"))) == 2


def test_concurrent_appends_are_serialized_with_unique_rows(tmp_path):
    workbook_path = tmp_path / "process.xlsx"
    _create_workbook(workbook_path)
    settings = _settings(workbook_path, tmp_path)

    payloads = []
    for index in range(4):
        payload = _payload()
        payload["col_f"] = f"M-{index + 1:02d}"
        payload["col_h"] = f"WO-{index + 1:02d}"
        payloads.append(payload)

    with ThreadPoolExecutor(max_workers=4) as executor:
        row_numbers = list(
            executor.map(
                lambda payload: append_entry_to_workbook(settings, payload),
                payloads,
            )
        )

    assert sorted(row_numbers) == [1726, 1727, 1728, 1729]

    workbook = load_workbook(workbook_path)
    worksheet = workbook["PROSES 2026"]
    written_machines = {
        worksheet.cell(row=row_number, column=6).value
        for row_number in row_numbers
    }
    written_orders = {
        worksheet.cell(row=row_number, column=8).value
        for row_number in row_numbers
    }
    workbook.close()

    assert written_machines == {"M-01", "M-02", "M-03", "M-04"}
    assert written_orders == {"WO-01", "WO-02", "WO-03", "WO-04"}


def test_retry_reuses_existing_matching_excel_row_after_partial_success(tmp_path):
    workbook_path = tmp_path / "process.xlsx"
    _create_workbook(workbook_path)
    settings = _settings(workbook_path, tmp_path)
    init_db(settings)

    with create_session(settings) as session:
        entry = Entry(
            sync_status=SYNC_STATUS_PENDING_EXCEL,
            **_payload(),
        )
        session.add(entry)
        commit_session(session)

        row_number = append_entry_to_workbook(settings, entry)
        assert row_number == 1726

        result = attempt_excel_append(
            settings,
            session,
            entry,
            failure_status=SYNC_STATUS_FAILED_EXCEL,
        )

        assert result["success"] is True
        assert result["excel_row_number"] == 1726
        assert entry.sync_status == "synced"
        assert entry.excel_row_number == 1726
        assert entry.last_error is None

    workbook = load_workbook(workbook_path)
    worksheet = workbook["PROSES 2026"]
    assert worksheet.cell(row=1726, column=6).value == "101"
    assert worksheet.cell(row=1727, column=6).value is None
    workbook.close()

    assert len(list((tmp_path / "backups").glob("process_*.xlsx"))) == 1


def test_retry_unsynced_entries_bulk_appends_selected_process_day(tmp_path):
    workbook_path = tmp_path / "process.xlsx"
    _create_workbook(workbook_path)
    settings = _settings(workbook_path, tmp_path)
    init_db(settings)

    with create_session(settings) as session:
        same_day_first = Entry(
            sync_status=SYNC_STATUS_PENDING_EXCEL,
            process_date="2026-06-08",
            machine_code="101",
            **_payload(),
        )
        same_day_second = Entry(
            sync_status=SYNC_STATUS_PENDING_EXCEL,
            process_date="2026-06-08",
            machine_code="102",
            **(_payload() | {"col_f": "102", "col_h": "WO-2"}),
        )
        other_day = Entry(
            sync_status=SYNC_STATUS_PENDING_EXCEL,
            process_date="2026-06-09",
            machine_code="103",
            **(_payload() | {"col_a": "2026-06-09", "col_f": "103"}),
        )
        session.add_all([same_day_first, same_day_second, other_day])
        commit_session(session)

    result = retry_unsynced_entries(settings, process_date="2026-06-08")

    assert result["attempted"] == 2
    assert result["synced"] == 2
    assert result["failed"] == 0
    assert result["remaining"] == 0
    assert result["process_date"] == "2026-06-08"
    assert [item["excel_row_number"] for item in result["results"]] == [1726, 1727]
    assert len(list((tmp_path / "backups").glob("process_*.xlsx"))) == 1

    workbook = load_workbook(workbook_path)
    worksheet = workbook["PROSES 2026"]
    assert worksheet.cell(row=1726, column=6).value == "101"
    assert worksheet.cell(row=1727, column=6).value == "102"
    assert worksheet.cell(row=1728, column=6).value is None
    workbook.close()

    with create_session(settings) as session:
        entries = session.query(Entry).order_by(Entry.machine_code).all()
        assert [entry.sync_status for entry in entries] == [
            "synced",
            "synced",
            SYNC_STATUS_PENDING_EXCEL,
        ]


def test_append_entry_to_workbook_blanks_first_section_injection_settings(tmp_path):
    workbook_path = tmp_path / "process.xlsx"
    _create_workbook(workbook_path)
    settings = _settings(workbook_path, tmp_path)
    payload = _payload() | {"col_f": "271"}

    append_entry_to_workbook(settings, payload)

    workbook = load_workbook(workbook_path)
    worksheet = workbook["PROSES 2026"]
    assert worksheet.cell(row=1726, column=6).value == "271"
    assert worksheet.cell(row=1726, column=15).value == 6
    assert worksheet.cell(row=1726, column=16).value == 45
    assert worksheet.cell(row=1726, column=17).value == 180
    assert [worksheet.cell(row=1726, column=column).value for column in range(18, 24)] == [
        None,
        None,
        None,
        None,
        None,
        None,
    ]
    workbook.close()


def test_append_entry_to_workbook_blanks_second_section_conditioning_fields(tmp_path):
    workbook_path = tmp_path / "process.xlsx"
    _create_workbook(workbook_path)
    settings = _settings(workbook_path, tmp_path)
    payload = _payload() | {"col_f": "203"}

    append_entry_to_workbook(settings, payload)

    workbook = load_workbook(workbook_path)
    worksheet = workbook["PROSES 2026"]
    assert worksheet.cell(row=1726, column=6).value == "203"
    assert [worksheet.cell(row=1726, column=column).value for column in range(15, 18)] == [
        None,
        None,
        None,
    ]
    assert worksheet.cell(row=1726, column=18).value == "70-69"
    assert worksheet.cell(row=1726, column=19).value == "50-49"
    assert worksheet.cell(row=1726, column=20).value == 1.5
    assert worksheet.cell(row=1726, column=23).value == 125
    workbook.close()


def test_append_entry_to_workbook_raises_typed_error_for_missing_file(tmp_path):
    settings = _settings(tmp_path / "missing.xlsx", tmp_path)

    with pytest.raises(WorkbookMissingError):
        append_entry_to_workbook(settings, _payload())


def test_append_entry_to_workbook_raises_typed_error_for_open_permission_denied(
    monkeypatch,
    tmp_path,
):
    settings = _settings(tmp_path / "process.xlsx", tmp_path)

    def raise_permission_error(*_args, **_kwargs):
        raise PermissionError("denied")

    monkeypatch.setattr("app.services.excel_service.load_workbook", raise_permission_error)

    with pytest.raises(WorkbookPermissionError):
        append_entry_to_workbook(settings, _payload())


def test_append_entry_to_workbook_raises_locked_error_when_save_is_denied(
    monkeypatch,
    tmp_path,
):
    workbook_path = tmp_path / "process.xlsx"
    _create_workbook(workbook_path)
    settings = _settings(workbook_path, tmp_path)

    def raise_permission_error(_workbook, _filename):
        raise PermissionError("locked")

    monkeypatch.setattr(OpenpyxlWorkbook, "save", raise_permission_error)

    with pytest.raises(WorkbookLockedError):
        append_entry_to_workbook(settings, _payload())

    assert len(list((tmp_path / "backups").glob("process_*.xlsx"))) == 1
