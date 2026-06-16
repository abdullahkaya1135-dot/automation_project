from openpyxl import Workbook
from openpyxl.styles import PatternFill

from app.modules.workbook_rows import detect_last_value_row_in_columns


def test_detect_last_value_row_in_columns_ignores_formatting_and_later_columns():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.cell(row=1, column=1, value="header")
    worksheet.cell(row=5, column=2, value="actual")
    worksheet.cell(row=8, column=4, value="outside-scan")
    worksheet.cell(row=100, column=2).fill = PatternFill(
        fill_type="solid",
        fgColor="FFFF00",
    )

    assert worksheet.max_row == 100
    assert detect_last_value_row_in_columns(worksheet, 3) == 5
    assert detect_last_value_row_in_columns(worksheet, 4) == 8

    workbook.close()
