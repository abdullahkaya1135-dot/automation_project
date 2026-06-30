from collections.abc import Generator
from datetime import datetime, timedelta, timezone
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.main import create_app

HEADERS = [
    "Date",
    "Ambient Temp",
    "Production Engineer",
    "Shift Chief",
    "Shift",
    "Machine",
    "Product",
    "Work Order",
    "Raw Material",
    "Total Cavity Count",
    "Active Cavity Count",
    "Cycle Time",
    "Cooling Time",
    "Injection Time",
    "Blow Time",
    "Conditioner Temp",
    "Dryer Temp",
    "Injection Pressure",
    "Injection Speed",
    "Holding Time",
    "Holding Speed",
    "Holding Pressure",
    "Clamp Force",
    "Oven Temperatures",
    "Mold Temperatures",
]


AUXILIARY_HEADERS = [
    "TARİH",
    "MAKİNE ",
    "ÇIKIŞ FREKANSI (HZ)",
    "SET DEĞERİ (BAR)",
    "GERİ BESLEME (BAR)",
    "ISI SET DEĞERİ ©",
    "GİRİŞ ISISI ©",
    "ÇIKIŞ ISISI ©",
    "KOMP BASINÇ DEĞERİ (BAR)",
]


ISTANBUL_TIMEZONE = timezone(timedelta(hours=3), "Europe/Istanbul")


DEFAULT_REQUEST_TIME = datetime(2026, 6, 8, 9, 15, tzinfo=ISTANBUL_TIMEZONE)


def _create_workbook(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "PROSES 2026"
    worksheet.append(HEADERS)
    workbook.save(path)
    workbook.close()


def _ifs_operations() -> list[dict[str, object]]:
    return [
        {
            "OrderNo": "WO-1",
            "PreferredResourceId": "M-01",
            "WorkCenterNo": "SP25",
            "PartNoDesc": "PET0001-01 400ML SEFFAF 28MM YIVLI 6GR",
        },
        {
            "OrderNo": "WO-2",
            "PreferredResourceId": "M-02",
            "WorkCenterNo": "SP25",
            "PartNoDesc": "PET0002-01 400ML SEFFAF 28MM YIVLI 8GR",
        },
        {
            "OrderNo": "WO-1",
            "PreferredResourceId": "M-01",
            "WorkCenterNo": "SP25",
            "PartNoDesc": "PET0001-01 400ML SEFFAF 28MM YIVLI 6GR",
        },
    ]


def _create_cycle_table(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sayfa1"
    worksheet.append(
        [
            "YENİ MAKİNA NO - New Machine Code",
            "MAKİNA ADI Machine Name",
            "AĞIZ ÇAP  Neck Diameter",
            "GR",
            "PREFORM MOLD CODE",
            "PREFORM KALIBI ÜRETİM YILI PRODUCTION DATE",
            "ESTIMATED CYCLE TIME",
            "BPH",
        ]
    )
    worksheet.append(["101", "SP25", 28, 6, None, None, 12, None])
    workbook.save(path)
    workbook.close()


def _create_auxiliary_workbook(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "YARDIMCI TESİSLER TAKİP"
    worksheet.append(AUXILIARY_HEADERS)
    workbook.save(path)
    workbook.close()


def _create_auxiliary_form(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "FORM"
    worksheet["A1"] = "YARDIMCI TESİSLER KONTROL FORMU"
    workbook.save(path)
    workbook.close()


def _freeze_request_time(monkeypatch, value: datetime = DEFAULT_REQUEST_TIME) -> None:
    monkeypatch.setattr("app.domain.shifts.request_datetime", lambda _settings: value)


@pytest.fixture
def client(monkeypatch, tmp_path) -> Generator[TestClient]:
    workbook_path = tmp_path / "process.xlsx"
    cycle_table_path = tmp_path / "cycle.xlsx"
    auxiliary_target_path = tmp_path / "auxiliary.xlsx"
    auxiliary_form_path = tmp_path / "auxiliary_form.xlsx"
    _create_workbook(workbook_path)
    _create_cycle_table(cycle_table_path)
    _create_auxiliary_workbook(auxiliary_target_path)
    _create_auxiliary_form(auxiliary_form_path)
    _freeze_request_time(monkeypatch)

    async def fake_fetch_pet_ongoing_operations(_settings):
        return _ifs_operations()

    monkeypatch.setattr(
        "app.features.bootstrap.api.fetch_pet_ongoing_operations",
        fake_fetch_pet_ongoing_operations,
    )
    monkeypatch.setattr(
        "app.features.cycle_reports.service.fetch_pet_ongoing_operations",
        fake_fetch_pet_ongoing_operations,
    )
    monkeypatch.setattr(
        "app.features.ifs.api.fetch_pet_ongoing_operations",
        fake_fetch_pet_ongoing_operations,
    )
    monkeypatch.setenv("EXCEL_PATH", str(workbook_path))
    monkeypatch.setenv("APP_PIN", "1234")
    monkeypatch.setenv("SESSION_SECRET", "test-session-secret")
    monkeypatch.setenv("SQLITE_PATH", str(tmp_path / "process_entries.sqlite3"))
    monkeypatch.setenv("BACKUP_DIR", str(tmp_path / "backups"))
    monkeypatch.setenv("CYCLE_TABLE_PATH", str(cycle_table_path))
    monkeypatch.setenv("REPORT_OUTPUT_DIR", str(tmp_path / "reports"))
    monkeypatch.setenv("AUXILIARY_SYSTEMS_FORM_PATH", str(auxiliary_form_path))
    monkeypatch.setenv("AUXILIARY_SYSTEMS_TARGET_PATH", str(auxiliary_target_path))
    monkeypatch.setenv("AUXILIARY_SYSTEMS_SHEET_NAME", "YARDIMCI TESİSLER TAKİP")

    with TestClient(create_app()) as test_client:
        login_response = test_client.post("/api/login", json={"pin": "1234"})
        assert login_response.status_code == 200
        yield test_client


def _tour_context(client: TestClient) -> int:
    response = client.post(
        "/api/tour-context",
        json={
            "ambient_temp": "24,5",
            "production_engineer": "Barış Çetik",
            "shift_chief": "Selman",
        },
    )

    assert response.status_code == 201
    return response.json()["id"]


def _amount_control_body(**overrides):
    body = {
        "record_date": "2026-06-08",
        "machine_code": "101",
        "job_order": "WO-1",
        "shift": "08.00-16.00",
        "worker_names": "Operator One, Operator Two",
        "produced_quantity": 1200,
        "breakdowns": [
            {
                "produced_product": "Product 101",
                "stop_reason": "Hydraulic pressure fault",
                "duration_minutes": 45,
                "stopped_at": "2026-06-08T10:00:00+03:00",
                "resumed_at": "2026-06-08T10:45:00+03:00",
            },
            {
                "produced_product": "Product 101",
                "stop_reason": "Mold change",
                "duration_minutes": 30,
            },
        ],
    }
    body.update(overrides)
    return body


def _breakdown_body(**overrides):
    body = {
        "record_date": "2026-06-08",
        "machine_code": "101",
        "shift": "00.00-08.00",
        "reason": "Hydraulic pressure fault",
        "duration_minutes": 45,
        "job_order": "WO-1",
        "produced_product": "Product 101",
        "stopped_at": "2026-06-08T00:30:00+03:00",
        "resumed_at": "2026-06-08T01:15:00+03:00",
    }
    body.update(overrides)
    return body
