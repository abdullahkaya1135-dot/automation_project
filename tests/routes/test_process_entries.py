import logging
from datetime import datetime

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.core.config import Settings
from app.core.database import create_session
from app.main import create_app
from app.models import Entry, ProductionEngineer
from tests.routes.helpers import _freeze_request_time, _tour_context


def test_entry_submit_saves_locally_and_queues_excel_sync(client, tmp_path):
    context_id = _tour_context(client)

    response = client.post(
        "/api/entries",
        json={
            "tour_context_id": context_id,
            "payload_schema_version": 2,
            "payload": {
                "col_f": "101",
                "col_g": "Product 101",
                "col_h": "WO-1",
                "col_j": "16",
                "col_k": "12",
                "col_l": "12,5",
                "col_r": "99x2",
                "col_x": "270x3,276x2,275",
                "col_y": "30x2,40",
            },
            "status": "ok",
            "notes": "note",
            "mold_info": "mold",
        },
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["saved_locally"] is True
    assert payload["saved_to_database"] is True
    assert payload["synced_to_excel"] is False
    assert payload["entry"]["sync_status"] == "pending_excel"
    assert payload["entry"]["excel_row_number"] is None
    assert payload["entry"]["process_date"] == "2026-06-08"
    assert payload["entry"]["machine_code"] == "101"
    assert payload["entry"]["production_engineer_id"] is not None
    assert payload["entry"]["payload"]["col_a"] == "08.06.2026"
    assert payload["entry"]["payload"]["col_c"] == "Barış Çetik"
    assert payload["entry"]["payload"]["col_e"] == "08.00-16.00"
    assert payload["entry"]["payload"]["col_f"] == "101"
    assert payload["entry"]["payload"]["col_h"] == "WO-1"
    assert payload["entry"]["payload"]["col_i"] is None
    assert payload["entry"]["payload"]["col_l"] == "12,5"
    assert payload["entry"]["payload"]["col_r"] is None
    assert payload["entry"]["payload"]["col_x"] == "270-270-270-276-276-275"
    assert payload["entry"]["payload"]["col_y"] == "30-30-40"

    workbook = load_workbook(tmp_path / "process.xlsx")
    worksheet = workbook["PROSES 2026"]
    assert worksheet.max_row == 1
    workbook.close()

    entries_response = client.get("/api/entries")
    assert entries_response.status_code == 200
    entries_payload = entries_response.json()["entries"]
    assert len(entries_payload) == 1
    assert entries_payload[0]["payload"]["col_l"] == "12,5"

    settings = Settings(
        excel_path=str(tmp_path / "process.xlsx"),
        app_pin="1234",
        sqlite_path=str(tmp_path / "process_entries.sqlite3"),
        backup_dir=str(tmp_path / "backups"),
    )
    with create_session(settings) as session:
        entry = session.query(Entry).one()
        assert entry.sync_status == "pending_excel"
        assert entry.excel_row_number is None
        assert entry.last_error is None
        assert entry.synced_at is None
        assert entry.process_date == "2026-06-08"
        assert entry.machine_code == "101"
        assert entry.col_l == "12,5"
        assert entry.production_engineer_id is not None


def test_process_excel_sync_button_bulk_appends_selected_day(client, tmp_path):
    context_id = _tour_context(client)
    for machine_code in ("101", "102"):
        response = client.post(
            "/api/entries",
            json={
                "tour_context_id": context_id,
                "payload_schema_version": 2,
                "payload": {
                    "col_f": machine_code,
                    "col_g": f"Product {machine_code}",
                    "col_h": f"WO-{machine_code}",
                    "col_j": "16",
                    "col_k": "12",
                    "col_l": "12,5",
                },
            },
        )
        assert response.status_code == 201

    response = client.post(
        "/api/sync/retry",
        params={"process_date": "2026-06-08"},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["database_backed"] is True
    assert payload["process_date"] == "2026-06-08"
    assert payload["attempted"] == 2
    assert payload["synced"] == 2
    assert payload["failed"] == 0
    assert payload["remaining"] == 0
    assert [result["excel_row_number"] for result in payload["results"]] == [2, 3]

    workbook = load_workbook(tmp_path / "process.xlsx")
    worksheet = workbook["PROSES 2026"]
    assert worksheet.cell(row=2, column=6).value == "101"
    assert worksheet.cell(row=3, column=6).value == "102"
    assert worksheet.cell(row=2, column=12).value == 12.5
    assert worksheet.cell(row=3, column=12).value == 12.5
    workbook.close()

    settings = Settings(
        excel_path=str(tmp_path / "process.xlsx"),
        app_pin="1234",
        sqlite_path=str(tmp_path / "process_entries.sqlite3"),
        backup_dir=str(tmp_path / "backups"),
    )
    with create_session(settings) as session:
        entries = session.query(Entry).order_by(Entry.machine_code).all()
        assert [entry.sync_status for entry in entries] == ["synced", "synced"]
        assert [entry.excel_row_number for entry in entries] == [2, 3]
        assert [entry.col_l for entry in entries] == ["12,5", "12,5"]


def test_entry_submit_is_idempotent_for_client_request_id(client, tmp_path):
    context_id = _tour_context(client)
    request_body = {
        "client_request_id": "entry-client-request-1",
        "client_recorded_at": "2026-06-08T09:20:00+03:00",
        "tour_context_id": context_id,
        "payload_schema_version": 2,
        "payload": {
            "col_f": "M-01",
            "col_g": "WO-1",
            "col_h": "16",
            "col_l": "12,5",
        },
    }

    first_response = client.post("/api/entries", json=request_body)
    replay_response = client.post("/api/entries", json=request_body)

    assert first_response.status_code == 201
    assert replay_response.status_code == 200
    first_payload = first_response.json()
    replay_payload = replay_response.json()
    assert replay_payload["idempotent_replay"] is True
    assert replay_payload["entry"]["id"] == first_payload["entry"]["id"]
    assert replay_payload["entry"]["client_request_id"] == "entry-client-request-1"
    assert first_payload["entry"]["payload"]["col_l"] == "12,5"
    assert replay_payload["entry"]["payload"]["col_l"] == "12,5"

    workbook = load_workbook(tmp_path / "process.xlsx")
    worksheet = workbook["PROSES 2026"]
    assert worksheet.max_row == 1
    workbook.close()

    settings = Settings(
        excel_path=str(tmp_path / "process.xlsx"),
        app_pin="1234",
        sqlite_path=str(tmp_path / "process_entries.sqlite3"),
        backup_dir=str(tmp_path / "backups"),
    )
    with create_session(settings) as session:
        entry = session.query(Entry).one()
        assert entry.col_l == "12,5"


def test_entries_list_orders_by_process_date_machine_and_engineer(client, tmp_path):
    settings = Settings(
        excel_path=str(tmp_path / "process.xlsx"),
        app_pin="1234",
        sqlite_path=str(tmp_path / "process_entries.sqlite3"),
        backup_dir=str(tmp_path / "backups"),
    )
    with create_session(settings) as session:
        engineers = {
            engineer.full_name: engineer.id
            for engineer in session.query(ProductionEngineer).all()
        }
        session.add_all(
            [
                Entry(
                    col_a="17.06.2026",
                    col_c="Fevzi Kılınç",
                    col_f="102",
                    process_date="2026-06-17",
                    machine_code="102",
                    production_engineer_id=engineers["Fevzi Kılınç"],
                    sync_status="synced",
                ),
                Entry(
                    col_a="16.06.2026",
                    col_c="Abdullah Kaya",
                    col_f="101",
                    process_date="2026-06-16",
                    machine_code="101",
                    production_engineer_id=engineers["Abdullah Kaya"],
                    sync_status="synced",
                ),
                Entry(
                    col_a="16.06.2026",
                    col_c="Barış Çetik",
                    col_f="102",
                    process_date="2026-06-16",
                    machine_code="102",
                    production_engineer_id=engineers["Barış Çetik"],
                    sync_status="synced",
                ),
            ]
        )
        session.commit()

    response = client.get("/api/entries")

    assert response.status_code == 200
    entries = response.json()["entries"]
    assert [
        (entry["process_date"], entry["machine_code"], entry["payload"]["col_c"])
        for entry in entries
    ] == [
        ("2026-06-16", "101", "Abdullah Kaya"),
        ("2026-06-16", "102", "Barış Çetik"),
        ("2026-06-17", "102", "Fevzi Kılınç"),
    ]


def test_entries_list_can_order_by_recent_submission(client, tmp_path):
    settings = Settings(
        excel_path=str(tmp_path / "process.xlsx"),
        app_pin="1234",
        sqlite_path=str(tmp_path / "process_entries.sqlite3"),
        backup_dir=str(tmp_path / "backups"),
    )
    with create_session(settings) as session:
        session.add_all(
            [
                Entry(
                    col_a="18.06.2026",
                    col_c="Engineer A",
                    col_f="101",
                    process_date="2026-06-18",
                    machine_code="101",
                    sync_status="synced",
                    submitted_at=datetime(2026, 6, 18, 6, 13),
                ),
                Entry(
                    col_a="18.06.2026",
                    col_c="Engineer B",
                    col_f="171",
                    process_date="2026-06-18",
                    machine_code="171",
                    sync_status="synced",
                    submitted_at=datetime(2026, 6, 18, 6, 17),
                ),
                Entry(
                    col_a="18.06.2026",
                    col_c="Engineer C",
                    col_f="103",
                    process_date="2026-06-18",
                    machine_code="103",
                    sync_status="synced",
                    submitted_at=datetime(2026, 6, 18, 6, 15),
                ),
            ]
        )
        session.commit()

    response = client.get("/api/entries", params={"sort": "recent", "limit": 2})

    assert response.status_code == 200
    entries = response.json()["entries"]
    assert [entry["machine_code"] for entry in entries] == ["171", "103"]


def test_second_section_entry_blanks_first_section_only_columns(client):
    context_id = _tour_context(client)

    response = client.post(
        "/api/entries",
        json={
            "tour_context_id": context_id,
            "payload_schema_version": 2,
            "payload": {
                "col_f": "203",
                "col_g": "Product 203",
                "col_h": "WO-203",
                "col_j": "16",
                "col_o": "6",
                "col_p": "45",
                "col_q": "180",
                "col_r": "70x2,69",
                "col_s": "49x2",
                "col_t": "1,5",
                "col_u": "25",
                "col_v": "30",
                "col_w": "125",
                "col_x": "230-195",
            },
        },
    )

    assert response.status_code == 201
    entry_payload = response.json()["entry"]["payload"]
    assert entry_payload["col_i"] is None
    assert entry_payload["col_o"] is None
    assert entry_payload["col_p"] is None
    assert entry_payload["col_q"] is None
    assert entry_payload["col_r"] == "70-70-69"
    assert entry_payload["col_s"] == "49-49"
    assert response.json()["entry"]["machine_code"] == "203"


def test_legacy_entry_payload_is_canonicalized_before_database_save(client, caplog):
    context_id = _tour_context(client)
    caplog.set_level(logging.INFO, logger="app.features.process_entries.payloads")

    response = client.post(
        "/api/entries",
        json={
            "tour_context_id": context_id,
            "payload": {
                "col_f": "101",
                "col_g": "WO-LEGACY",
                "col_h": "16",
                "col_i": "12",
                "col_j": "12,5",
                "col_p": "270x2",
                "col_q": "30x2",
            },
        },
    )

    assert response.status_code == 201
    entry_payload = response.json()["entry"]["payload"]
    assert entry_payload["col_g"] is None
    assert entry_payload["col_h"] == "WO-LEGACY"
    assert entry_payload["col_j"] == "16"
    assert entry_payload["col_k"] == "12"
    assert entry_payload["col_l"] == "12,5"
    assert entry_payload["col_x"] == "270-270"
    assert entry_payload["col_y"] == "30-30"
    assert response.json()["entry"]["process_date"] == "2026-06-08"
    assert response.json()["entry"]["machine_code"] == "101"
    records = [
        record
        for record in caplog.records
        if record.name == "app.features.process_entries.payloads"
    ]
    assert len(records) == 1
    assert records[0].legacy_payload_field_count == 17
    assert records[0].canonical_payload_field_count == 25
    assert records[0].entry_payload_schema_version == 2
    assert "WO-LEGACY" not in records[0].getMessage()
    assert "270x2" not in records[0].getMessage()
    assert "30x2" not in records[0].getMessage()
    assert "WO-LEGACY" not in caplog.text
    assert "270x2" not in caplog.text
    assert "30x2" not in caplog.text


def test_entry_submit_uses_database_when_process_excel_is_missing(
    monkeypatch, tmp_path
):
    workbook_path = tmp_path / "process.xlsx"
    _freeze_request_time(monkeypatch)
    monkeypatch.setenv("EXCEL_PATH", str(workbook_path))
    monkeypatch.setenv("APP_PIN", "1234")
    monkeypatch.setenv("SESSION_SECRET", "test-session-secret")
    monkeypatch.setenv("SQLITE_PATH", str(tmp_path / "process_entries.sqlite3"))
    monkeypatch.setenv("BACKUP_DIR", str(tmp_path / "backups"))

    with TestClient(create_app()) as client:
        login_response = client.post("/api/login", json={"pin": "1234"})
        assert login_response.status_code == 200
        context_id = _tour_context(client)

        pending_response = client.post(
            "/api/entries",
            json={
                "tour_context_id": context_id,
                "payload_schema_version": 2,
                "payload": {
                    "col_f": "M-01",
                    "col_g": "Product M-01",
                    "col_h": "WO-1",
                    "col_l": "12,5",
                },
            },
        )

        assert pending_response.status_code == 201
        pending_payload = pending_response.json()
        assert pending_payload["synced_to_excel"] is False
        assert pending_payload["saved_to_database"] is True
        assert pending_payload["entry"]["sync_status"] == "pending_excel"
        assert pending_payload["entry"]["last_error"] is None

        pending_entries_response = client.get(
            "/api/entries",
            params={"sync_status": "pending_excel"},
        )
        assert pending_entries_response.status_code == 200
        pending_entries = pending_entries_response.json()["entries"]
        assert len(pending_entries) == 1
        assert pending_entries[0]["payload"]["col_l"] == "12,5"

        retry_response = client.post(
            "/api/sync/retry",
            params={"process_date": "2026-06-08"},
        )

    assert retry_response.status_code == 200
    retry_payload = retry_response.json()
    assert retry_payload["database_backed"] is True
    assert retry_payload["process_date"] == "2026-06-08"
    assert retry_payload["attempted"] == 1
    assert retry_payload["synced"] == 0
    assert retry_payload["failed"] == 1
    assert retry_payload["remaining"] == 1

    settings = Settings(
        excel_path=str(workbook_path),
        app_pin="1234",
        sqlite_path=str(tmp_path / "process_entries.sqlite3"),
        backup_dir=str(tmp_path / "backups"),
    )
    with create_session(settings) as session:
        entry = session.query(Entry).one()
        assert entry.sync_status == "failed_excel"
        assert entry.excel_row_number is None
        assert entry.last_error
        assert entry.synced_at is None
        assert entry.col_l == "12,5"
