from openpyxl import load_workbook

from app.core.config import Settings
from app.core.database import create_session
from app.models import AuxiliarySystemsSubmission


def test_auxiliary_systems_submission_saves_locally_and_appends_block(
    client,
    tmp_path,
):
    response = client.post(
        "/api/auxiliary-systems/submissions",
        json={
            "recorded_date": "2026-06-08",
            "payload": {
                "tower_frequency": "50",
                "tower_set_pressure": "3,6",
                "tower_feedback_pressure": "3.5",
                "termokar_chiller_1_temp_set": "12",
                "termokar_chiller_1_inlet_temp": "13,9",
                "termokar_chiller_1_outlet_temp": "12.3",
                "compressor_high_708_pressure": "31,4",
                "compressor_low_716_pressure": "11.4",
                "oil_cooling_water_tank_checked": True,
            },
        },
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["saved_locally"] is True
    assert payload["synced_to_excel"] is True
    assert payload["submission"]["sync_status"] == "synced"
    assert payload["submission"]["excel_start_row"] == 2
    assert payload["submission"]["excel_end_row"] == 16
    assert payload["submission"]["payload"]["oil_cooling_water_tank_checked"] is True

    workbook = load_workbook(tmp_path / "auxiliary.xlsx")
    worksheet = workbook["YARDIMCI TESİSLER TAKİP"]
    assert worksheet.cell(row=2, column=1).value == "08.06.2026"
    assert worksheet.cell(row=2, column=2).value == "1- ELEKTROMOTOR (KULE)"
    assert worksheet.cell(row=2, column=3).value == 50
    assert worksheet.cell(row=2, column=4).value == 3.6
    assert worksheet.cell(row=2, column=5).value == 3.5
    assert worksheet.cell(row=4, column=6).value == 12
    assert worksheet.cell(row=4, column=7).value == 13.9
    assert worksheet.cell(row=4, column=8).value == 12.3
    assert worksheet.cell(row=8, column=9).value == 31.4
    assert worksheet.cell(row=16, column=9).value == 11.4
    workbook.close()

    submissions_response = client.get("/api/auxiliary-systems/submissions")
    assert submissions_response.status_code == 200
    assert len(submissions_response.json()["submissions"]) == 1


def test_auxiliary_submission_is_idempotent_for_client_request_id(client, tmp_path):
    request_body = {
        "client_request_id": "auxiliary-client-request-1",
        "client_recorded_at": "2026-06-08T09:25:00+03:00",
        "recorded_date": "2026-06-08",
        "payload": {
            "tower_frequency": "50",
            "tower_set_pressure": "3,6",
            "tower_feedback_pressure": "3.5",
        },
    }

    first_response = client.post(
        "/api/auxiliary-systems/submissions",
        json=request_body,
    )
    replay_response = client.post(
        "/api/auxiliary-systems/submissions",
        json=request_body,
    )

    assert first_response.status_code == 201
    assert replay_response.status_code == 200
    first_payload = first_response.json()
    replay_payload = replay_response.json()
    assert replay_payload["idempotent_replay"] is True
    assert replay_payload["submission"]["id"] == first_payload["submission"]["id"]
    assert (
        replay_payload["submission"]["client_request_id"]
        == "auxiliary-client-request-1"
    )

    workbook = load_workbook(tmp_path / "auxiliary.xlsx")
    worksheet = workbook["YARDIMCI TESİSLER TAKİP"]
    assert worksheet.cell(row=2, column=2).value == "1- ELEKTROMOTOR (KULE)"
    assert worksheet.cell(row=17, column=2).value is None
    workbook.close()

    settings = Settings(
        excel_path=str(tmp_path / "process.xlsx"),
        app_pin="1234",
        sqlite_path=str(tmp_path / "process_entries.sqlite3"),
        backup_dir=str(tmp_path / "backups"),
    )
    with create_session(settings) as session:
        assert session.query(AuxiliarySystemsSubmission).count() == 1
