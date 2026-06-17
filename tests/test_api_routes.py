from collections.abc import Generator
from datetime import datetime, timedelta, timezone
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app.config import Settings
from app.database import create_session
from app.main import create_app
from app.models import (
    AmountControlShift,
    AuxiliarySystemsSubmission,
    Entry,
    MachineBreakdown,
    ProductionEngineer,
)

HEADERS = [
    "Date",
    "Ambient Temp",
    "Production Engineer",
    "Shift Chief",
    "Shift",
    "Machine",
    "Product",
    "Work Order",
    "Raw Material",
    "Total Cavity Count",
    "Active Cavity Count",
    "Cycle Time",
    "Cooling Time",
    "Injection Time",
    "Blow Time",
    "Conditioner Temp",
    "Dryer Temp",
    "Injection Pressure",
    "Injection Speed",
    "Holding Time",
    "Holding Speed",
    "Holding Pressure",
    "Clamp Force",
    "Oven Temperatures",
    "Mold Temperatures",
]
AUXILIARY_HEADERS = [
    "TARİH",
    "MAKİNE ",
    "ÇIKIŞ FREKANSI (HZ)",
    "SET DEĞERİ (BAR)",
    "GERİ BESLEME (BAR)",
    "ISI SET DEĞERİ ©",
    "GİRİŞ ISISI ©",
    "ÇIKIŞ ISISI ©",
    "KOMP BASINÇ DEĞERİ (BAR)",
]
ISTANBUL_TIMEZONE = timezone(timedelta(hours=3), "Europe/Istanbul")
DEFAULT_REQUEST_TIME = datetime(2026, 6, 8, 9, 15, tzinfo=ISTANBUL_TIMEZONE)
PROTECTED_PAGE_ROUTES = ("/", "/process", "/auxiliary", "/amount-control", "/reports")
SHARED_PAGE_ASSET_MARKERS = (
    'href="/static/css/app.css?v=',
    'type="module" src="/static/js/app.js?v=',
)
PAGE_SECTION_EXPECTATIONS = {
    "/": (
        'data-page="dashboard"',
        'id="dashboard-heading"',
        'class="dashboard-grid"',
        'class="dashboard-card',
        'href="/process"',
        'href="/auxiliary"',
        'href="/amount-control"',
        'href="/reports"',
    ),
    "/process": (
        'data-page="process"',
        'id="tour-context-form"',
        'id="entry-form"',
    ),
    "/auxiliary": (
        'data-page="auxiliary"',
        'id="auxiliary-form"',
    ),
    "/amount-control": (
        'data-page="amount-control"',
        'id="amount-control-form"',
    ),
    "/reports": (
        'data-page="reports"',
        'id="phone-sync-heading"',
        'id="pending-entries"',
        'id="create-cycle-report"',
        'id="run-ifs-return-candidates"',
        'id="print-ifs-return-candidates"',
        'id="ifs-return-print-area"',
    ),
}
PAGE_SECTION_EXCLUSIONS = {
    "/": (
        'id="tour-context-form"',
        'id="entry-form"',
        'id="auxiliary-form"',
        'id="amount-control-form"',
        'id="pending-entries"',
        'id="create-cycle-report"',
        'id="run-ifs-return-candidates"',
    ),
    "/process": (
        'id="dashboard-heading"',
        'id="auxiliary-form"',
        'id="amount-control-form"',
        'id="pending-entries"',
        'id="create-cycle-report"',
        'id="run-ifs-return-candidates"',
    ),
    "/auxiliary": (
        'id="dashboard-heading"',
        'id="tour-context-form"',
        'id="entry-form"',
        'id="amount-control-form"',
        'id="pending-entries"',
        'id="create-cycle-report"',
        'id="run-ifs-return-candidates"',
    ),
    "/amount-control": (
        'id="dashboard-heading"',
        'id="tour-context-form"',
        'id="entry-form"',
        'id="auxiliary-form"',
        'id="pending-entries"',
        'id="create-cycle-report"',
        'id="run-ifs-return-candidates"',
    ),
    "/reports": (
        'id="dashboard-heading"',
        'id="tour-context-form"',
        'id="entry-form"',
        'id="auxiliary-form"',
        'id="amount-control-form"',
    ),
}
SERVICE_WORKER_ASSET_MARKERS = (
    "/manifest.webmanifest",
    "/static/css/app.css?v=",
    "/static/js/api.js?v=",
    "/static/js/app.js?v=",
    "/static/js/modules/amount-control.js?v=",
    "/static/js/modules/main-page.js?v=",
    "/static/js/modules/offline.js?v=",
)


def _create_workbook(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "PROSES 2026"
    worksheet.append(HEADERS)
    workbook.save(path)
    workbook.close()


def _ifs_operations() -> list[dict[str, object]]:
    return [
        {
            "OrderNo": "WO-1",
            "PreferredResourceId": "M-01",
            "WorkCenterNo": "SP25",
            "PartNoDesc": "PET0001-01 400ML SEFFAF 28MM YIVLI 6GR",
        },
        {
            "OrderNo": "WO-2",
            "PreferredResourceId": "M-02",
            "WorkCenterNo": "SP25",
            "PartNoDesc": "PET0002-01 400ML SEFFAF 28MM YIVLI 8GR",
        },
        {
            "OrderNo": "WO-1",
            "PreferredResourceId": "M-01",
            "WorkCenterNo": "SP25",
            "PartNoDesc": "PET0001-01 400ML SEFFAF 28MM YIVLI 6GR",
        },
    ]


def _create_cycle_table(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sayfa1"
    worksheet.append(
        [
            "YENİ MAKİNA NO - New Machine Code",
            "MAKİNA ADI Machine Name",
            "AĞIZ ÇAP  Neck Diameter",
            "GR",
            "PREFORM MOLD CODE",
            "PREFORM KALIBI ÜRETİM YILI PRODUCTION DATE",
            "ESTIMATED CYCLE TIME",
            "BPH",
        ]
    )
    worksheet.append(["M-01", "SP25", 28, 6, None, None, 12, None])
    workbook.save(path)
    workbook.close()


def _create_auxiliary_workbook(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "YARDIMCI TESİSLER TAKİP"
    worksheet.append(AUXILIARY_HEADERS)
    workbook.save(path)
    workbook.close()


def _create_auxiliary_form(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "FORM"
    worksheet["A1"] = "YARDIMCI TESİSLER KONTROL FORMU"
    workbook.save(path)
    workbook.close()


def _freeze_request_time(monkeypatch, value: datetime = DEFAULT_REQUEST_TIME) -> None:
    monkeypatch.setattr("app.domain.shifts.request_datetime", lambda _settings: value)


@pytest.fixture
def client(monkeypatch, tmp_path) -> Generator[TestClient]:
    workbook_path = tmp_path / "process.xlsx"
    cycle_table_path = tmp_path / "cycle.xlsx"
    auxiliary_target_path = tmp_path / "auxiliary.xlsx"
    auxiliary_form_path = tmp_path / "auxiliary_form.xlsx"
    _create_workbook(workbook_path)
    _create_cycle_table(cycle_table_path)
    _create_auxiliary_workbook(auxiliary_target_path)
    _create_auxiliary_form(auxiliary_form_path)
    _freeze_request_time(monkeypatch)

    async def fake_fetch_pet_ongoing_operations(_settings):
        return _ifs_operations()

    monkeypatch.setattr(
        "app.routers.bootstrap.fetch_pet_ongoing_operations",
        fake_fetch_pet_ongoing_operations,
    )
    monkeypatch.setattr(
        "app.services.cycle_report_service.fetch_pet_ongoing_operations",
        fake_fetch_pet_ongoing_operations,
    )
    monkeypatch.setenv("EXCEL_PATH", str(workbook_path))
    monkeypatch.setenv("APP_PIN", "1234")
    monkeypatch.setenv("SESSION_SECRET", "test-session-secret")
    monkeypatch.setenv("SQLITE_PATH", str(tmp_path / "process_entries.sqlite3"))
    monkeypatch.setenv("BACKUP_DIR", str(tmp_path / "backups"))
    monkeypatch.setenv("CYCLE_TABLE_PATH", str(cycle_table_path))
    monkeypatch.setenv("REPORT_OUTPUT_DIR", str(tmp_path / "reports"))
    monkeypatch.setenv("AUXILIARY_SYSTEMS_FORM_PATH", str(auxiliary_form_path))
    monkeypatch.setenv("AUXILIARY_SYSTEMS_TARGET_PATH", str(auxiliary_target_path))
    monkeypatch.setenv("AUXILIARY_SYSTEMS_SHEET_NAME", "YARDIMCI TESİSLER TAKİP")

    with TestClient(create_app()) as test_client:
        login_response = test_client.post("/api/login", json={"pin": "1234"})
        assert login_response.status_code == 200
        yield test_client


def _tour_context(client: TestClient) -> int:
    response = client.post(
        "/api/tour-context",
        json={
            "ambient_temp": "24,5",
            "production_engineer": "Barış Çetik",
            "shift_chief": "Selman",
        },
    )

    assert response.status_code == 201
    return response.json()["id"]


def _amount_control_body(**overrides):
    body = {
        "record_date": "2026-06-08",
        "machine_code": "101",
        "job_order": "WO-1",
        "shift": "08.00-16.00",
        "worker_names": "Operator One, Operator Two",
        "produced_quantity": 1200,
        "breakdowns": [
            {
                "produced_product": "Product 101",
                "stop_reason": "Hydraulic pressure fault",
                "duration_minutes": 45,
                "stopped_at": "2026-06-08T10:00:00+03:00",
                "resumed_at": "2026-06-08T10:45:00+03:00",
            },
            {
                "produced_product": "Product 101",
                "stop_reason": "Mold change",
                "duration_minutes": 30,
            },
        ],
    }
    body.update(overrides)
    return body


def test_entry_submit_saves_locally_and_appends_to_excel(client, tmp_path):
    context_id = _tour_context(client)

    response = client.post(
        "/api/entries",
        json={
            "tour_context_id": context_id,
            "payload_schema_version": 2,
            "payload": {
                "col_f": "101",
                "col_g": "Product 101",
                "col_h": "WO-1",
                "col_i": "HM-02-01",
                "col_j": "16",
                "col_k": "12",
                "col_l": "12,5",
                "col_r": "99x2",
                "col_x": "270x3,276x2,275",
                "col_y": "30x2,40",
            },
            "status": "ok",
            "notes": "note",
            "mold_info": "mold",
        },
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["saved_locally"] is True
    assert payload["saved_to_database"] is True
    assert payload["synced_to_excel"] is False
    assert payload["entry"]["sync_status"] == "synced"
    assert payload["entry"]["excel_row_number"] is None
    assert payload["entry"]["process_date"] == "2026-06-08"
    assert payload["entry"]["machine_code"] == "101"
    assert payload["entry"]["production_engineer_id"] is not None
    assert payload["entry"]["payload"]["col_a"] == "08.06.2026"
    assert payload["entry"]["payload"]["col_c"] == "Barış Çetik"
    assert payload["entry"]["payload"]["col_e"] == "08.00-16.00"
    assert payload["entry"]["payload"]["col_f"] == "101"
    assert payload["entry"]["payload"]["col_h"] == "WO-1"
    assert payload["entry"]["payload"]["col_r"] is None
    assert payload["entry"]["payload"]["col_x"] == "270-270-270-276-276-275"
    assert payload["entry"]["payload"]["col_y"] == "30-30-40"

    workbook = load_workbook(tmp_path / "process.xlsx")
    worksheet = workbook["PROSES 2026"]
    assert worksheet.max_row == 1
    workbook.close()

    entries_response = client.get("/api/entries")
    assert entries_response.status_code == 200
    assert len(entries_response.json()["entries"]) == 1

    settings = Settings(
        excel_path=str(tmp_path / "process.xlsx"),
        app_pin="1234",
        sqlite_path=str(tmp_path / "process_entries.sqlite3"),
        backup_dir=str(tmp_path / "backups"),
    )
    with create_session(settings) as session:
        entry = session.query(Entry).one()
        assert entry.sync_status == "synced"
        assert entry.excel_row_number is None
        assert entry.last_error is None
        assert entry.synced_at is not None
        assert entry.process_date == "2026-06-08"
        assert entry.machine_code == "101"
        assert entry.production_engineer_id is not None


def test_entry_submit_is_idempotent_for_client_request_id(client, tmp_path):
    context_id = _tour_context(client)
    request_body = {
        "client_request_id": "entry-client-request-1",
        "client_recorded_at": "2026-06-08T09:20:00+03:00",
        "tour_context_id": context_id,
        "payload": {
            "col_f": "M-01",
            "col_g": "WO-1",
            "col_h": "16",
        },
    }

    first_response = client.post("/api/entries", json=request_body)
    replay_response = client.post("/api/entries", json=request_body)

    assert first_response.status_code == 201
    assert replay_response.status_code == 200
    first_payload = first_response.json()
    replay_payload = replay_response.json()
    assert replay_payload["idempotent_replay"] is True
    assert replay_payload["entry"]["id"] == first_payload["entry"]["id"]
    assert replay_payload["entry"]["client_request_id"] == "entry-client-request-1"

    workbook = load_workbook(tmp_path / "process.xlsx")
    worksheet = workbook["PROSES 2026"]
    assert worksheet.max_row == 1
    workbook.close()

    settings = Settings(
        excel_path=str(tmp_path / "process.xlsx"),
        app_pin="1234",
        sqlite_path=str(tmp_path / "process_entries.sqlite3"),
        backup_dir=str(tmp_path / "backups"),
    )
    with create_session(settings) as session:
        assert session.query(Entry).count() == 1


def test_entries_list_orders_by_process_date_machine_and_engineer(client, tmp_path):
    settings = Settings(
        excel_path=str(tmp_path / "process.xlsx"),
        app_pin="1234",
        sqlite_path=str(tmp_path / "process_entries.sqlite3"),
        backup_dir=str(tmp_path / "backups"),
    )
    with create_session(settings) as session:
        engineers = {
            engineer.full_name: engineer.id
            for engineer in session.query(ProductionEngineer).all()
        }
        session.add_all(
            [
                Entry(
                    col_a="17.06.2026",
                    col_c="Fevzi Kılınç",
                    col_f="102",
                    process_date="2026-06-17",
                    machine_code="102",
                    production_engineer_id=engineers["Fevzi Kılınç"],
                    sync_status="synced",
                ),
                Entry(
                    col_a="16.06.2026",
                    col_c="Abdullah Kaya",
                    col_f="101",
                    process_date="2026-06-16",
                    machine_code="101",
                    production_engineer_id=engineers["Abdullah Kaya"],
                    sync_status="synced",
                ),
                Entry(
                    col_a="16.06.2026",
                    col_c="Barış Çetik",
                    col_f="102",
                    process_date="2026-06-16",
                    machine_code="102",
                    production_engineer_id=engineers["Barış Çetik"],
                    sync_status="synced",
                ),
            ]
        )
        session.commit()

    response = client.get("/api/entries")

    assert response.status_code == 200
    entries = response.json()["entries"]
    assert [
        (entry["process_date"], entry["machine_code"], entry["payload"]["col_c"])
        for entry in entries
    ] == [
        ("2026-06-16", "101", "Abdullah Kaya"),
        ("2026-06-16", "102", "Barış Çetik"),
        ("2026-06-17", "102", "Fevzi Kılınç"),
    ]


def test_second_section_entry_blanks_first_section_only_columns(client):
    context_id = _tour_context(client)

    response = client.post(
        "/api/entries",
        json={
            "tour_context_id": context_id,
            "payload_schema_version": 2,
            "payload": {
                "col_f": "203",
                "col_g": "Product 203",
                "col_h": "WO-203",
                "col_i": "HM-02-203",
                "col_j": "16",
                "col_o": "6",
                "col_p": "45",
                "col_q": "180",
                "col_r": "70x2,69",
                "col_s": "49x2",
                "col_t": "1,5",
                "col_u": "25",
                "col_v": "30",
                "col_w": "125",
                "col_x": "230-195",
            },
        },
    )

    assert response.status_code == 201
    entry_payload = response.json()["entry"]["payload"]
    assert entry_payload["col_o"] is None
    assert entry_payload["col_p"] is None
    assert entry_payload["col_q"] is None
    assert entry_payload["col_r"] == "70-70-69"
    assert entry_payload["col_s"] == "49-49"
    assert response.json()["entry"]["machine_code"] == "203"


def test_legacy_entry_payload_is_canonicalized_before_database_save(client):
    context_id = _tour_context(client)

    response = client.post(
        "/api/entries",
        json={
            "tour_context_id": context_id,
            "payload": {
                "col_f": "101",
                "col_g": "WO-LEGACY",
                "col_h": "16",
                "col_i": "12",
                "col_j": "12,5",
                "col_p": "270x2",
                "col_q": "30x2",
            },
        },
    )

    assert response.status_code == 201
    entry_payload = response.json()["entry"]["payload"]
    assert entry_payload["col_g"] is None
    assert entry_payload["col_h"] == "WO-LEGACY"
    assert entry_payload["col_j"] == "16"
    assert entry_payload["col_k"] == "12"
    assert entry_payload["col_l"] == "12,5"
    assert entry_payload["col_x"] == "270-270"
    assert entry_payload["col_y"] == "30-30"
    assert response.json()["entry"]["process_date"] == "2026-06-08"
    assert response.json()["entry"]["machine_code"] == "101"


def test_auxiliary_systems_submission_saves_locally_and_appends_block(
    client,
    tmp_path,
):
    response = client.post(
        "/api/auxiliary-systems/submissions",
        json={
            "recorded_date": "2026-06-08",
            "payload": {
                "tower_frequency": "50",
                "tower_set_pressure": "3,6",
                "tower_feedback_pressure": "3.5",
                "termokar_chiller_1_temp_set": "12",
                "termokar_chiller_1_inlet_temp": "13,9",
                "termokar_chiller_1_outlet_temp": "12.3",
                "compressor_high_708_pressure": "31,4",
                "compressor_low_716_pressure": "11.4",
                "oil_cooling_water_tank_checked": True,
            },
        },
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["saved_locally"] is True
    assert payload["synced_to_excel"] is True
    assert payload["submission"]["sync_status"] == "synced"
    assert payload["submission"]["excel_start_row"] == 2
    assert payload["submission"]["excel_end_row"] == 16
    assert payload["submission"]["payload"]["oil_cooling_water_tank_checked"] is True

    workbook = load_workbook(tmp_path / "auxiliary.xlsx")
    worksheet = workbook["YARDIMCI TESİSLER TAKİP"]
    assert worksheet.cell(row=2, column=1).value == "08.06.2026"
    assert worksheet.cell(row=2, column=2).value == "1- ELEKTROMOTOR (KULE)"
    assert worksheet.cell(row=2, column=3).value == 50
    assert worksheet.cell(row=2, column=4).value == 3.6
    assert worksheet.cell(row=2, column=5).value == 3.5
    assert worksheet.cell(row=4, column=6).value == 12
    assert worksheet.cell(row=4, column=7).value == 13.9
    assert worksheet.cell(row=4, column=8).value == 12.3
    assert worksheet.cell(row=8, column=9).value == 31.4
    assert worksheet.cell(row=16, column=9).value == 11.4
    workbook.close()

    submissions_response = client.get("/api/auxiliary-systems/submissions")
    assert submissions_response.status_code == 200
    assert len(submissions_response.json()["submissions"]) == 1


def test_auxiliary_submission_is_idempotent_for_client_request_id(client, tmp_path):
    request_body = {
        "client_request_id": "auxiliary-client-request-1",
        "client_recorded_at": "2026-06-08T09:25:00+03:00",
        "recorded_date": "2026-06-08",
        "payload": {
            "tower_frequency": "50",
            "tower_set_pressure": "3,6",
            "tower_feedback_pressure": "3.5",
        },
    }

    first_response = client.post(
        "/api/auxiliary-systems/submissions",
        json=request_body,
    )
    replay_response = client.post(
        "/api/auxiliary-systems/submissions",
        json=request_body,
    )

    assert first_response.status_code == 201
    assert replay_response.status_code == 200
    first_payload = first_response.json()
    replay_payload = replay_response.json()
    assert replay_payload["idempotent_replay"] is True
    assert replay_payload["submission"]["id"] == first_payload["submission"]["id"]
    assert (
        replay_payload["submission"]["client_request_id"]
        == "auxiliary-client-request-1"
    )

    workbook = load_workbook(tmp_path / "auxiliary.xlsx")
    worksheet = workbook["YARDIMCI TESİSLER TAKİP"]
    assert worksheet.cell(row=2, column=2).value == "1- ELEKTROMOTOR (KULE)"
    assert worksheet.cell(row=17, column=2).value is None
    workbook.close()

    settings = Settings(
        excel_path=str(tmp_path / "process.xlsx"),
        app_pin="1234",
        sqlite_path=str(tmp_path / "process_entries.sqlite3"),
        backup_dir=str(tmp_path / "backups"),
    )
    with create_session(settings) as session:
        assert session.query(AuxiliarySystemsSubmission).count() == 1


def test_amount_control_shift_create_list_detail_and_breakdown_links(client, tmp_path):
    response = client.post(
        "/api/amount-control/shifts",
        json=_amount_control_body(client_request_id="amount-client-request-1"),
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["saved_locally"] is True
    assert payload["idempotent_replay"] is False
    amount_shift = payload["shift"]
    assert amount_shift["client_request_id"] == "amount-client-request-1"
    assert amount_shift["record_date"] == "2026-06-08"
    assert amount_shift["machine_code"] == "101"
    assert amount_shift["job_order"] == "WO-1"
    assert amount_shift["shift"] == "08.00-16.00"
    assert amount_shift["worker_names"] == "Operator One, Operator Two"
    assert amount_shift["produced_quantity"] == 1200
    assert len(amount_shift["breakdowns"]) == 2
    assert amount_shift["breakdowns"][0]["entry_id"] is None
    assert (
        amount_shift["breakdowns"][0]["amount_control_shift_id"]
        == amount_shift["id"]
    )
    assert amount_shift["breakdowns"][0]["machine_code"] == "101"

    list_response = client.get(
        "/api/amount-control/shifts",
        params={
            "record_date": "2026-06-08",
            "machine_code": "101",
            "job_order": "WO-1",
        },
    )
    assert list_response.status_code == 200
    listed_shifts = list_response.json()["shifts"]
    assert len(listed_shifts) == 1
    assert listed_shifts[0]["id"] == amount_shift["id"]

    detail_response = client.get(f"/api/amount-control/shifts/{amount_shift['id']}")
    assert detail_response.status_code == 200
    assert detail_response.json()["breakdowns"][1]["stop_reason"] == "Mold change"

    settings = Settings(
        excel_path=str(tmp_path / "process.xlsx"),
        app_pin="1234",
        sqlite_path=str(tmp_path / "process_entries.sqlite3"),
        backup_dir=str(tmp_path / "backups"),
    )
    with create_session(settings) as session:
        assert session.query(AmountControlShift).count() == 1
        assert session.query(MachineBreakdown).count() == 2
        breakdown = session.query(MachineBreakdown).first()
        assert breakdown is not None
        assert breakdown.entry_id is None
        assert breakdown.amount_control_shift_id == amount_shift["id"]


def test_amount_control_shift_is_idempotent_for_client_request_id(
    client,
    tmp_path,
):
    request_body = _amount_control_body(
        client_request_id="amount-client-request-2",
        breakdowns=[],
    )

    first_response = client.post("/api/amount-control/shifts", json=request_body)
    replay_response = client.post("/api/amount-control/shifts", json=request_body)

    assert first_response.status_code == 201
    assert replay_response.status_code == 200
    first_payload = first_response.json()
    replay_payload = replay_response.json()
    assert replay_payload["idempotent_replay"] is True
    assert replay_payload["shift"]["id"] == first_payload["shift"]["id"]
    assert (
        replay_payload["shift"]["client_request_id"]
        == "amount-client-request-2"
    )

    settings = Settings(
        excel_path=str(tmp_path / "process.xlsx"),
        app_pin="1234",
        sqlite_path=str(tmp_path / "process_entries.sqlite3"),
        backup_dir=str(tmp_path / "backups"),
    )
    with create_session(settings) as session:
        assert session.query(AmountControlShift).count() == 1


def test_amount_control_shift_rejects_duplicate_business_key(client):
    first_response = client.post(
        "/api/amount-control/shifts",
        json=_amount_control_body(client_request_id="amount-client-request-3"),
    )
    duplicate_response = client.post(
        "/api/amount-control/shifts",
        json=_amount_control_body(
            client_request_id="amount-client-request-4",
            worker_names="Operator Three",
            produced_quantity=900,
            breakdowns=[],
        ),
    )

    assert first_response.status_code == 201
    assert duplicate_response.status_code == 409
    assert "zaten var" in duplicate_response.json()["detail"]


def test_amount_control_shift_rejects_unknown_machine(client):
    response = client.post(
        "/api/amount-control/shifts",
        json=_amount_control_body(machine_code="999", breakdowns=[]),
    )

    assert response.status_code == 422
    assert "makine" in response.json()["detail"]


@pytest.mark.parametrize(
    "overrides",
    [
        {"record_date": "08.06.2026"},
        {"shift": "00.00-08.00"},
        {"job_order": " "},
        {"worker_names": " "},
        {"produced_quantity": -1},
        {
            "breakdowns": [
                {
                    "produced_product": "Product 101",
                    "stop_reason": "Fault",
                    "duration_minutes": 0,
                }
            ]
        },
        {
            "breakdowns": [
                {
                    "produced_product": "Product 101",
                    "stop_reason": "Fault",
                    "duration_minutes": 1,
                    "stopped_at": "2026-06-08T11:00:00+03:00",
                    "resumed_at": "2026-06-08T10:00:00+03:00",
                }
            ]
        },
    ],
)
def test_amount_control_shift_validates_payload(client, overrides):
    response = client.post(
        "/api/amount-control/shifts",
        json=_amount_control_body(**overrides),
    )

    assert response.status_code == 422


def test_amount_control_shift_detail_returns_404_for_missing_id(client):
    response = client.get("/api/amount-control/shifts/999999")

    assert response.status_code == 404


def test_cycle_report_endpoint_creates_today_report(client):
    context_id = _tour_context(client)
    entry_response = client.post(
        "/api/entries",
        json={
            "tour_context_id": context_id,
            "payload": {
                "col_f": "M-01",
                "col_g": "WO-1",
                "col_h": "16",
                "col_i": "12",
                "col_j": "12,5",
            },
        },
    )
    assert entry_response.status_code == 201

    response = client.post("/api/cycle-report/today")

    assert response.status_code == 200
    payload = response.json()
    assert payload["row_count"] == 1
    assert payload["matched_count"] == 1
    assert payload["warning_count"] == 0
    assert payload["date"] == "2026-06-08"

    output_path = Path(payload["output_path"])
    assert output_path.exists()
    workbook = load_workbook(output_path)
    worksheet = workbook["Çevrim Kontrol"]
    assert [cell.value for cell in worksheet[2]] == [
        "SP25",
        "M-01",
        28,
        6,
        16,
        12,
        12.5,
        12,
        "Uygun",
    ]
    workbook.close()


def test_ifs_stock_endpoint_returns_u1_hm02_stock(client, monkeypatch):
    async def fake_fetch(_settings):
        return [
            {
                "Contract": "S01",
                "PartNo": "HM-02-A",
                "PartNoRef": {"Description": "Raw Material"},
                "LocationNo": "U1",
                "LotBatchNo": "L1",
                "AvailableQty": 12.5,
                "QtyOnhand": 15,
                "UoM": "kg",
                "ObjId": "obj-1",
            }
        ]

    monkeypatch.setattr("app.routers.ifs.fetch_u1_hm02_stock", fake_fetch)

    response = client.get("/api/ifs/u1-hm02-stock")

    assert response.status_code == 200
    assert response.json() == {
        "stock_count": 1,
        "stock": [
            {
                "contract": "S01",
                "part_no": "HM-02-A",
                "material_name": "Raw Material",
                "location_no": "U1",
                "lot_batch_no": "L1",
                "available_qty": 12.5,
                "qty_onhand": 15,
                "uom": "kg",
                "obj_id": "obj-1",
                "configuration_id": None,
                "serial_no": None,
                "eng_chg_level": None,
                "waiv_dev_rej_no": None,
                "activity_seq": None,
                "handling_unit_id": None,
            }
        ],
    }


def test_ifs_operations_endpoint_returns_pet_ongoing_operations(client, monkeypatch):
    async def fake_fetch(_settings):
        return [
            {
                "OrderNo": "2615",
                "ReleaseNo": "*",
                "SequenceNo": "*",
                "OperationNo": 10,
                "Contract": "S01",
                "WorkCenterNo": "SP25",
                "PartNo": "MM-PET0048-12-024Y-025",
                "PartNoDesc": "Bottle",
                "PreferredResourceId": "135",
                "OperationNoDesc": "Run",
                "RemainingQty": 42,
            }
        ]

    monkeypatch.setattr("app.routers.ifs.fetch_pet_ongoing_operations", fake_fetch)

    response = client.get("/api/ifs/pet-ongoing-operations")

    assert response.status_code == 200
    assert response.json() == {
        "operation_count": 1,
        "operations": [
            {
                "order_no": "2615",
                "release_no": "*",
                "sequence_no": "*",
                "operation_no": 10,
                "contract": "S01",
                "work_center_no": "SP25",
                "part_no": "MM-PET0048-12-024Y-025",
                "part_description": "Bottle",
                "preferred_resource_id": "135",
                "operation_description": "Run",
                "remaining_qty": 42,
            }
        ],
    }


def test_ifs_operation_materials_endpoint_returns_hm02_materials(client, monkeypatch):
    called_with: dict[str, object] = {}

    async def fake_fetch(_settings, operation):
        called_with["operation"] = operation
        return [
            {
                "OrderNo": "2615",
                "ReleaseNo": "*",
                "SequenceNo": "*",
                "LineItemNo": 2,
                "OperationNo": 10,
                "PartNo": "HM-02-01-01-525",
                "IssueToLoc": "U1",
                "QtyRequired": 25.578,
                "QtyAssigned": 0,
                "QtyIssued": 0,
                "QtyRemainingToReserve": 25.578,
                "QtyAvailable": 100.72,
                "PrintUnit": "kg",
                "SoPartNo": "MM-PET0048-12-024Y-025",
                "Cf_Tercihedilenkaynak": "135",
            }
        ]

    monkeypatch.setattr("app.routers.ifs.fetch_operation_hm02_materials", fake_fetch)

    response = client.get(
        "/api/ifs/operation-hm02-materials",
        params={"order_no": "2615", "operation_no": 10},
    )

    assert response.status_code == 200
    assert called_with["operation"] == {
        "OrderNo": "2615",
        "ReleaseNo": "*",
        "SequenceNo": "*",
        "OperationNo": 10,
    }
    assert response.json() == {
        "material_count": 1,
        "materials": [
            {
                "order_no": "2615",
                "release_no": "*",
                "sequence_no": "*",
                "line_item_no": 2,
                "operation_no": 10,
                "part_no": "HM-02-01-01-525",
                "issue_to_location": "U1",
                "qty_required": 25.578,
                "qty_assigned": 0,
                "qty_issued": 0,
                "qty_remaining_to_reserve": 25.578,
                "qty_available": 100.72,
                "print_unit": "kg",
                "produced_part_no": "MM-PET0048-12-024Y-025",
                "machine": "135",
            }
        ],
    }


def test_ifs_used_materials_endpoint_returns_used_hm02_set(client, monkeypatch):
    async def fake_fetch(_settings):
        return {
            "operation_count": 2,
            "used_material_count": 2,
            "used_part_count": 1,
            "used_parts": ["HM-02-01-01-525"],
            "used_hm02_part_count": 1,
            "used_hm02_parts": ["HM-02-01-01-525"],
            "used_materials": [
                {
                    "OrderNo": "2615",
                    "ReleaseNo": "*",
                    "SequenceNo": "*",
                    "LineItemNo": 2,
                    "OperationNo": 10,
                    "PartNo": "HM-02-01-01-525",
                    "IssueToLoc": "U1",
                    "QtyRequired": 25.578,
                    "QtyAssigned": 0,
                    "QtyIssued": 0,
                    "QtyRemainingToReserve": 25.578,
                    "QtyAvailable": 100.72,
                    "PrintUnit": "kg",
                    "SoPartNo": "MM-PET0048-12-024Y-025",
                    "Cf_Tercihedilenkaynak": "135",
                },
                {
                    "OrderNo": "2616",
                    "ReleaseNo": "*",
                    "SequenceNo": "*",
                    "LineItemNo": 4,
                    "OperationNo": 20,
                    "PartNo": "HM-02-01-01-525",
                    "IssueToLoc": "U1",
                    "QtyRequired": 12,
                    "QtyAssigned": 0,
                    "QtyIssued": 0,
                    "QtyRemainingToReserve": 12,
                    "QtyAvailable": 50,
                    "PrintUnit": "kg",
                    "SoPartNo": "MM-PET0001",
                    "Cf_Tercihedilenkaynak": "136",
                },
            ],
        }

    monkeypatch.setattr("app.routers.ifs.fetch_used_hm02_materials", fake_fetch)

    response = client.get("/api/ifs/used-hm02-materials")

    assert response.status_code == 200
    payload = response.json()
    assert payload["operation_count"] == 2
    assert payload["used_material_count"] == 2
    assert payload["used_part_count"] == 1
    assert payload["used_parts"] == ["HM-02-01-01-525"]
    assert payload["used_hm02_part_count"] == 1
    assert payload["used_hm02_parts"] == ["HM-02-01-01-525"]
    assert [row["part_no"] for row in payload["used_materials"]] == [
        "HM-02-01-01-525",
        "HM-02-01-01-525",
    ]
    assert payload["used_materials"][0]["machine"] == "135"
    assert payload["used_materials"][1]["produced_part_no"] == "MM-PET0001"


def test_ifs_return_candidates_endpoint_returns_unused_u1_stock(client, monkeypatch):
    async def fake_find(_settings):
        return {
            "generated_at": "2026-06-10T14:30:00+03:00",
            "stock_count": 3,
            "operation_count": 1,
            "active_used_material_count": 1,
            "active_used_part_count": 1,
            "active_used_hm02_part_count": 1,
            "planning_order_count": 1,
            "planning_operation_count": 1,
            "planning_used_material_count": 1,
            "planning_used_part_count": 1,
            "planning_used_hm02_part_count": 1,
            "used_material_count": 2,
            "used_part_count": 2,
            "used_hm02_part_count": 2,
            "return_candidate_count": 1,
            "return_candidates": [
                {
                    "Contract": "S01",
                    "PartNo": "HM-04-C",
                    "PartNoRef": {"Description": "Candidate Material"},
                    "LocationNo": "U1",
                    "LotBatchNo": "L1",
                    "AvailableQty": 12.5,
                    "QtyOnhand": 15,
                    "UoM": "kg",
                    "ObjId": "obj-1",
                },
            ],
            "used_parts": ["HM-02-A", "HM-03-B"],
            "used_hm02_parts": ["HM-02-A", "HM-03-B"],
            "used_materials": [
                {
                    "OrderNo": "2615",
                    "ReleaseNo": "*",
                    "SequenceNo": "*",
                    "LineItemNo": 2,
                    "OperationNo": 10,
                    "PartNo": "HM-02-A",
                    "IssueToLoc": "U1",
                    "QtyRequired": 25.578,
                    "QtyAssigned": 0,
                    "QtyIssued": 0,
                    "QtyRemainingToReserve": 25.578,
                    "QtyAvailable": 100.72,
                    "PrintUnit": "kg",
                    "SoPartNo": "MM-PET0048-12-024Y-025",
                    "Cf_Tercihedilenkaynak": "135",
                },
                {
                    "OrderNo": "2579",
                    "ReleaseNo": "*",
                    "SequenceNo": "*",
                    "LineItemNo": 3,
                    "OperationNo": 20,
                    "PartNo": "HM-03-B",
                    "IssueToLoc": "U1",
                    "QtyRequired": 10,
                    "QtyAssigned": 0,
                    "QtyIssued": 0,
                    "QtyRemainingToReserve": 10,
                    "QtyAvailable": 40,
                    "PrintUnit": "kg",
                    "SoPartNo": "MM-PET0002",
                    "Cf_Tercihedilenkaynak": "172",
                }
            ],
        }

    monkeypatch.setattr("app.routers.ifs.find_u1_return_candidates", fake_find)

    response = client.get("/api/ifs/u1-return-candidates")

    assert response.status_code == 200
    payload = response.json()
    assert payload["generated_at"] == "2026-06-10T14:30:00+03:00"
    assert payload["stock_count"] == 3
    assert payload["operation_count"] == 1
    assert payload["active_used_material_count"] == 1
    assert payload["active_used_part_count"] == 1
    assert payload["active_used_hm02_part_count"] == 1
    assert payload["planning_order_count"] == 1
    assert payload["planning_operation_count"] == 1
    assert payload["planning_used_material_count"] == 1
    assert payload["planning_used_part_count"] == 1
    assert payload["planning_used_hm02_part_count"] == 1
    assert payload["used_material_count"] == 2
    assert payload["used_part_count"] == 2
    assert payload["used_hm02_part_count"] == 2
    assert payload["return_candidate_count"] == 1
    assert payload["used_parts"] == ["HM-02-A", "HM-03-B"]
    assert payload["used_hm02_parts"] == ["HM-02-A", "HM-03-B"]
    assert [row["lot_batch_no"] for row in payload["return_candidates"]] == [
        "L1",
    ]
    assert [row["part_no"] for row in payload["return_candidates"]] == [
        "HM-04-C",
    ]
    assert [row["material_name"] for row in payload["return_candidates"]] == [
        "Candidate Material",
    ]
    assert all("obj_id" not in row for row in payload["return_candidates"])
    assert payload["used_materials"][0]["part_no"] == "HM-02-A"
    assert payload["used_materials"][0]["machine"] == "135"
    assert payload["used_materials"][1]["part_no"] == "HM-03-B"


def test_bootstrap_and_health_endpoints_report_current_state(client):
    bootstrap_response = client.get("/api/bootstrap")
    assert bootstrap_response.status_code == 200
    bootstrap_payload = bootstrap_response.json()
    assert bootstrap_payload["excel"]["available"] is True
    assert bootstrap_payload["excel"]["database_backed"] is True
    assert [engineer["full_name"] for engineer in bootstrap_payload["production_engineers"]] == [
        "Barış Çetik",
        "Abdullah Kaya",
        "Fevzi Kılınç",
    ]
    assert len(bootstrap_payload["machines"]) == 47
    assert bootstrap_payload["machines"][0]["machine_code"] == "101"
    assert bootstrap_payload["auxiliary_systems"]["form_available"] is True
    assert bootstrap_payload["auxiliary_systems"]["target_available"] is True
    assert bootstrap_payload["shop_order_source"]["available"] is True
    assert bootstrap_payload["shop_order_source"]["operation_count"] == 2
    assert bootstrap_payload["shop_order_source"]["order_count"] == 2
    assert bootstrap_payload["shop_order_source"]["resource_count"] == 2
    assert bootstrap_payload["shop_order_source"]["options"] == [
        {
            "order_no": "WO-1",
            "resource_id": "M-01",
            "release_no": "*",
            "sequence_no": "*",
            "operation_no": None,
            "part_no": "",
            "part_description": "PET0001-01 400ML SEFFAF 28MM YIVLI 6GR",
        },
        {
            "order_no": "WO-2",
            "resource_id": "M-02",
            "release_no": "*",
            "sequence_no": "*",
            "operation_no": None,
            "part_no": "",
            "part_description": "PET0002-01 400ML SEFFAF 28MM YIVLI 8GR",
        },
    ]
    assert bootstrap_payload["current_tour_timing"]["date"] == "08.06.2026"
    assert bootstrap_payload["current_tour_timing"]["shift"] == "08.00-16.00"
    assert bootstrap_payload["current_auxiliary_date"] == "08.06.2026"
    assert bootstrap_payload["latest_tour_context"] is None
    assert bootstrap_payload["last_sync_error"] is None
    assert bootstrap_payload["last_auxiliary_sync_error"] is None

    context_id = _tour_context(client)
    refreshed_bootstrap_response = client.get("/api/bootstrap")
    assert refreshed_bootstrap_response.status_code == 200
    assert refreshed_bootstrap_response.json()["latest_tour_context"]["id"] == context_id

    health_response = client.get("/health")
    assert health_response.status_code == 200
    health_payload = health_response.json()
    assert health_payload["status"] == "ok"
    assert health_payload["sqlite"]["ok"] is True
    assert health_payload["process_data"]["ok"] is True
    assert health_payload["process_data"]["database_backed"] is True
    assert health_payload["auxiliary_systems"]["form_ok"] is True
    assert health_payload["auxiliary_systems"]["target_ok"] is True
    assert health_payload["excel_write_lock"]["locked"] is False
    assert health_payload["excel_write_lock"]["waiting"] == 0


def test_page_routes_render_expected_split_sections_and_shared_assets(client):
    for path, expected_markers in PAGE_SECTION_EXPECTATIONS.items():
        response = client.get(path)

        assert response.status_code == 200
        assert response.headers["Cache-Control"] == "no-store, max-age=0"
        for marker in (*SHARED_PAGE_ASSET_MARKERS, *expected_markers):
            assert marker in response.text
        for marker in PAGE_SECTION_EXCLUSIONS[path]:
            assert marker not in response.text


def test_service_worker_uses_route_specific_navigation_cache_and_shared_assets(
    client,
):
    response = client.get("/service-worker.js")

    assert response.status_code == 200
    assert response.headers["Cache-Control"] == "no-cache"
    assert response.headers["Service-Worker-Allowed"] == "/"
    source = response.text
    assert "process-offline-shell-" in source
    assert 'request.mode === "navigate"' in source
    assert "url.pathname" in source
    assert 'networkFirst(request, "/")' not in source
    assert "cache.put(cacheKey, response.clone())" in source
    assert "cache.match(cacheKey)" in source
    for route in PROTECTED_PAGE_ROUTES:
        assert f'"{route}"' in source
    for marker in SERVICE_WORKER_ASSET_MARKERS:
        assert marker in source


def test_manifest_start_url_and_scope_stay_at_root(client):
    response = client.get("/manifest.webmanifest")

    assert response.status_code == 200
    manifest = response.json()
    assert manifest["start_url"] == "/"
    assert manifest["scope"] == "/"


def test_bootstrap_can_load_shop_orders_from_ifs(client, monkeypatch):
    async def fake_fetch(_settings):
        return [
            {
                "OrderNo": "2615",
                "PreferredResourceId": "135",
                "WorkCenterNo": "SP25",
                "PartNoDesc": "Bottle",
            },
            {
                "OrderNo": "2616",
                "PreferredResourceId": "136",
                "WorkCenterNo": "SP26",
                "PartNoDesc": "Bottle 2",
            },
        ]

    monkeypatch.setattr("app.routers.bootstrap.fetch_pet_ongoing_operations", fake_fetch)

    response = client.get("/api/bootstrap")

    assert response.status_code == 200
    source = response.json()["shop_order_source"]
    assert source["available"] is True
    assert source["source"] == "ifs-token"
    assert source["operation_count"] == 2
    assert source["order_count"] == 2
    assert source["resource_count"] == 2
    assert source["options"] == [
        {
            "order_no": "2615",
            "resource_id": "135",
            "release_no": "*",
            "sequence_no": "*",
            "operation_no": None,
            "part_no": "",
            "part_description": "Bottle",
        },
        {
            "order_no": "2616",
            "resource_id": "136",
            "release_no": "*",
            "sequence_no": "*",
            "operation_no": None,
            "part_no": "",
            "part_description": "Bottle 2",
        },
    ]


def test_tour_context_uses_request_time_for_date_and_shift(client, monkeypatch):
    _freeze_request_time(
        monkeypatch,
        datetime(2026, 6, 9, 2, 30, tzinfo=ISTANBUL_TIMEZONE),
    )

    response = client.post(
        "/api/tour-context",
        json={
            "date": "1999-01-01",
            "ambient_temp": "24,5",
            "production_engineer": "Barış Çetik",
            "shift_chief": "Selman",
            "shift": "16.00-24.00",
        },
    )

    assert response.status_code == 201
    context = response.json()["tour_context"]
    assert context["date"] == "09.06.2026"
    assert context["shift"] == "24.00-08.00"


def test_tour_context_uses_client_recorded_at_for_offline_shift(
    client,
    monkeypatch,
):
    _freeze_request_time(
        monkeypatch,
        datetime(2026, 6, 8, 16, 5, tzinfo=ISTANBUL_TIMEZONE),
    )

    response = client.post(
        "/api/tour-context",
        json={
            "client_request_id": "tour-client-request-1",
            "client_recorded_at": "2026-06-08T15:55:00+03:00",
            "ambient_temp": "24,5",
            "production_engineer": "Barış Çetik",
            "shift_chief": "Selman",
        },
    )

    assert response.status_code == 201
    context = response.json()["tour_context"]
    assert context["client_request_id"] == "tour-client-request-1"
    assert context["date"] == "08.06.2026"
    assert context["shift"] == "08.00-16.00"


def test_tour_context_is_idempotent_for_client_request_id(client):
    request_body = {
        "client_request_id": "tour-client-request-2",
        "client_recorded_at": "2026-06-08T09:30:00+03:00",
        "ambient_temp": "24,5",
        "production_engineer": "Barış Çetik",
        "shift_chief": "Selman",
    }

    first_response = client.post("/api/tour-context", json=request_body)
    replay_response = client.post("/api/tour-context", json=request_body)

    assert first_response.status_code == 201
    assert replay_response.status_code == 200
    first_payload = first_response.json()
    replay_payload = replay_response.json()
    assert replay_payload["idempotent_replay"] is True
    assert replay_payload["id"] == first_payload["id"]
    assert (
        replay_payload["tour_context"]["client_request_id"]
        == "tour-client-request-2"
    )


def test_tour_context_rejects_unknown_shift_chief(client):
    response = client.post(
        "/api/tour-context",
        json={
            "ambient_temp": "24,5",
            "production_engineer": "Barış Çetik",
            "shift_chief": "Chief",
        },
    )

    assert response.status_code == 422
    assert "Selman, Serkan, Hakan" in response.json()["detail"]


@pytest.mark.parametrize(
    ("request_time", "expected_shift"),
    [
        (datetime(2026, 6, 8, 7, 59, tzinfo=ISTANBUL_TIMEZONE), "24.00-08.00"),
        (datetime(2026, 6, 8, 8, 0, tzinfo=ISTANBUL_TIMEZONE), "08.00-16.00"),
        (datetime(2026, 6, 8, 15, 59, tzinfo=ISTANBUL_TIMEZONE), "08.00-16.00"),
        (datetime(2026, 6, 8, 16, 0, tzinfo=ISTANBUL_TIMEZONE), "16.00-24.00"),
    ],
)
def test_shift_boundaries(request_time, expected_shift):
    from app.domain.shifts import shift_for_request_time

    assert shift_for_request_time(request_time) == expected_shift


def test_entry_submit_uses_database_when_process_excel_is_missing(monkeypatch, tmp_path):
    workbook_path = tmp_path / "process.xlsx"
    _freeze_request_time(monkeypatch)
    monkeypatch.setenv("EXCEL_PATH", str(workbook_path))
    monkeypatch.setenv("APP_PIN", "1234")
    monkeypatch.setenv("SESSION_SECRET", "test-session-secret")
    monkeypatch.setenv("SQLITE_PATH", str(tmp_path / "process_entries.sqlite3"))
    monkeypatch.setenv("BACKUP_DIR", str(tmp_path / "backups"))

    with TestClient(create_app()) as client:
        login_response = client.post("/api/login", json={"pin": "1234"})
        assert login_response.status_code == 200
        context_id = _tour_context(client)

        pending_response = client.post(
            "/api/entries",
            json={
                "tour_context_id": context_id,
                "payload": {
                    "col_f": "M-01",
                    "col_g": "WO-1",
                },
            },
        )

        assert pending_response.status_code == 201
        pending_payload = pending_response.json()
        assert pending_payload["synced_to_excel"] is False
        assert pending_payload["saved_to_database"] is True
        assert pending_payload["entry"]["sync_status"] == "synced"
        assert pending_payload["entry"]["last_error"] is None

        pending_entries_response = client.get(
            "/api/entries",
            params={"sync_status": "pending_excel"},
        )
        assert pending_entries_response.status_code == 200
        pending_entries = pending_entries_response.json()["entries"]
        assert pending_entries == []

        retry_response = client.post("/api/sync/retry")

    assert retry_response.status_code == 200
    retry_payload = retry_response.json()
    assert retry_payload["database_backed"] is True
    assert retry_payload["attempted"] == 0
    assert retry_payload["synced"] == 0
    assert retry_payload["remaining"] == 0

    settings = Settings(
        excel_path=str(workbook_path),
        app_pin="1234",
        sqlite_path=str(tmp_path / "process_entries.sqlite3"),
        backup_dir=str(tmp_path / "backups"),
    )
    with create_session(settings) as session:
        entry = session.query(Entry).one()
        assert entry.sync_status == "synced"
        assert entry.excel_row_number is None
        assert entry.last_error is None
        assert entry.synced_at is not None
