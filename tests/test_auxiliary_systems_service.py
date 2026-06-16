import json
import shutil
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.workbook.workbook import Workbook as OpenpyxlWorkbook

from app.config import Settings
from app.database import commit_session, create_session, init_db
from app.models import (
    SYNC_STATUS_FAILED_EXCEL,
    SYNC_STATUS_PENDING_EXCEL,
    AuxiliarySystemsSubmission,
)
from app.modules.auxiliary_systems.sync_service import attempt_auxiliary_excel_append
from app.modules.auxiliary_systems.workbook_service import (
    AuxiliaryWorkbookSubmission,
    append_auxiliary_submissions_to_workbook,
    append_auxiliary_systems_to_workbook,
)

AUXILIARY_HEADERS = [
    "TARİH",
    "MAKİNE ",
    "ÇIKIŞ FREKANSI (HZ)",
    "SET DEĞERİ (BAR)",
    "GERİ BESLEME (BAR)",
    "ISI SET DEĞERİ ©",
    "GİRİŞ ISISI ©",
    "ÇIKIŞ ISISI ©",
    "KOMP BASINÇ DEĞERİ (BAR)",
]


def _create_auxiliary_workbook(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "YARDIMCI TESİSLER TAKİP"
    worksheet.append(AUXILIARY_HEADERS)
    workbook.save(path)
    workbook.close()


def _settings(workbook_path: Path, tmp_path: Path) -> Settings:
    return Settings(
        excel_path=str(tmp_path / "process.xlsx"),
        app_pin="1234",
        sqlite_path=str(tmp_path / "process_entries.sqlite3"),
        backup_dir=str(tmp_path / "backups"),
        auxiliary_systems_form_path=str(tmp_path / "auxiliary_form.xlsx"),
        auxiliary_systems_target_path=str(workbook_path),
        auxiliary_systems_sheet_name="YARDIMCI TESİSLER TAKİP",
    )


def test_append_auxiliary_systems_to_workbook_writes_daily_block(tmp_path):
    workbook_path = tmp_path / "auxiliary.xlsx"
    _create_auxiliary_workbook(workbook_path)
    settings = _settings(workbook_path, tmp_path)

    start_row, end_row = append_auxiliary_systems_to_workbook(
        settings,
        {
            "tower_frequency": "45,5",
            "tower_set_pressure": "3.6",
            "tower_feedback_pressure": "3,5",
            "chiller_motor_frequency": "50",
            "chiller_motor_set_pressure": "4",
            "chiller_motor_feedback_pressure": "3.8",
            "termokar_chiller_1_temp_set": "12",
            "termokar_chiller_1_inlet_temp": "13,9",
            "termokar_chiller_1_outlet_temp": "12.3",
            "itech_temp_set": "13",
            "itech_current_temp": "14,2",
            "compressor_high_708_pressure": "31,4",
            "compressor_low_716_pressure": "11.4",
        },
        "2026-06-10",
    )

    assert (start_row, end_row) == (2, 16)
    assert len(list((tmp_path / "backups").glob("auxiliary_*.xlsx"))) == 1

    workbook = load_workbook(workbook_path)
    worksheet = workbook["YARDIMCI TESİSLER TAKİP"]
    assert worksheet.cell(row=2, column=1).value == "10.06.2026"
    assert worksheet.cell(row=2, column=2).value == "1- ELEKTROMOTOR (KULE)"
    assert worksheet.cell(row=2, column=3).value == 45.5
    assert worksheet.cell(row=2, column=4).value == 3.6
    assert worksheet.cell(row=2, column=5).value == 3.5
    assert worksheet.cell(row=3, column=2).value == "2- ELEKTROMOTOR (CHİLLER)"
    assert worksheet.cell(row=3, column=3).value == 50
    assert worksheet.cell(row=4, column=6).value == 12
    assert worksheet.cell(row=4, column=7).value == 13.9
    assert worksheet.cell(row=4, column=8).value == 12.3
    assert worksheet.cell(row=7, column=2).value == "ITECH SOĞUTUCU"
    assert worksheet.cell(row=7, column=6).value == 13
    assert worksheet.cell(row=7, column=7).value == 14.2
    assert worksheet.cell(row=8, column=9).value == 31.4
    assert worksheet.cell(row=16, column=9).value == 11.4
    workbook.close()


def test_append_auxiliary_submissions_to_workbook_uses_one_open_save_and_backup(
    monkeypatch,
    tmp_path,
):
    workbook_path = tmp_path / "auxiliary.xlsx"
    _create_auxiliary_workbook(workbook_path)
    settings = _settings(workbook_path, tmp_path)
    submissions = [
        AuxiliaryWorkbookSubmission(
            payload={
                "tower_frequency": f"{45 + index}",
                "tower_set_pressure": "3.6",
                "compressor_low_716_pressure": f"{11 + index}",
            },
            recorded_date=f"2026-06-{10 + index:02d}",
            submission_id=index + 1,
        )
        for index in range(3)
    ]

    load_count = 0
    save_count = 0
    backup_count = 0
    real_load_workbook = load_workbook
    real_save = OpenpyxlWorkbook.save
    real_copyfile = shutil.copyfile

    def counted_load_workbook(*args, **kwargs):
        nonlocal load_count
        load_count += 1
        return real_load_workbook(*args, **kwargs)

    def counted_save(self, filename):
        nonlocal save_count
        save_count += 1
        return real_save(self, filename)

    def counted_copyfile(source, destination):
        nonlocal backup_count
        backup_count += 1
        return real_copyfile(source, destination)

    monkeypatch.setattr(
        "app.modules.auxiliary_systems.workbook_io.load_workbook",
        counted_load_workbook,
    )
    monkeypatch.setattr(OpenpyxlWorkbook, "save", counted_save)
    monkeypatch.setattr(
        "app.modules.auxiliary_systems.workbook_io.shutil.copyfile",
        counted_copyfile,
    )

    results = append_auxiliary_submissions_to_workbook(settings, submissions)

    assert [(result.start_row, result.end_row) for result in results] == [
        (2, 16),
        (17, 31),
        (32, 46),
    ]
    assert load_count == 1
    assert save_count == 1
    assert backup_count == 1
    assert len(list((tmp_path / "backups").glob("auxiliary_*.xlsx"))) == 1

    workbook = load_workbook(workbook_path)
    worksheet = workbook[settings.auxiliary_systems_sheet_name]
    assert worksheet.cell(row=2, column=1).value == "10.06.2026"
    assert worksheet.cell(row=17, column=1).value == "11.06.2026"
    assert worksheet.cell(row=32, column=1).value == "12.06.2026"
    assert worksheet.cell(row=46, column=9).value == 13
    workbook.close()


def test_retry_reuses_existing_auxiliary_block_after_partial_success(tmp_path):
    workbook_path = tmp_path / "auxiliary.xlsx"
    _create_auxiliary_workbook(workbook_path)
    settings = _settings(workbook_path, tmp_path)
    init_db(settings)
    payload = {
        "tower_frequency": "45,5",
        "tower_set_pressure": "3.6",
        "tower_feedback_pressure": "3,5",
        "compressor_low_716_pressure": "11.4",
    }

    with create_session(settings) as session:
        submission = AuxiliarySystemsSubmission(
            recorded_date="2026-06-10",
            payload_json=json.dumps(payload),
            sync_status=SYNC_STATUS_PENDING_EXCEL,
        )
        session.add(submission)
        commit_session(session)

        start_row, end_row = append_auxiliary_systems_to_workbook(
            settings,
            payload,
            submission.recorded_date,
        )
        assert (start_row, end_row) == (2, 16)

        result = attempt_auxiliary_excel_append(
            settings,
            session,
            submission,
            failure_status=SYNC_STATUS_FAILED_EXCEL,
        )

        assert result["success"] is True
        assert result["excel_start_row"] == 2
        assert result["excel_end_row"] == 16
        assert submission.sync_status == "synced"
        assert submission.excel_start_row == 2
        assert submission.excel_end_row == 16
        assert submission.last_error is None

    workbook = load_workbook(workbook_path)
    worksheet = workbook[settings.auxiliary_systems_sheet_name]
    assert worksheet.cell(row=2, column=2).value == "1- ELEKTROMOTOR (KULE)"
    assert worksheet.cell(row=17, column=2).value is None
    workbook.close()

    assert len(list((tmp_path / "backups").glob("auxiliary_*.xlsx"))) == 1
