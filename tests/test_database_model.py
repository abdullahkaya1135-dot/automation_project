import sqlite3

import pytest
from openpyxl import Workbook
from sqlalchemy import inspect, text

from app.core.config import Settings
from app.core.database import (
    DatabaseCommitError,
    create_session,
    init_db,
    session_scope,
)
from app.models import (
    SYNC_STATUS_PENDING_EXCEL,
    AmountControlShift,
    Entry,
    Machine,
    MachineBreakdown,
    MachineCycleTableRow,
    MachineGroup,
    MachineGroupMachine,
    ProductionEngineer,
    ProductionLossLabelEvent,
    TourContext,
)


def _create_cycle_seed_workbook(path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sayfa1"
    worksheet.append(
        [
            "Machine No",
            "Machine Name",
            "Neck",
            "GR",
            "Mold Code",
            "Production Date",
            "Estimated Cycle Time",
        ]
    )
    worksheet.append([174, "DPH70", 28, 35, "MOLD-A", None, 14])
    worksheet.append([None, None, None, 40, "MOLD-B", None, 16])
    worksheet.merge_cells("A2:A3")
    worksheet.merge_cells("B2:B3")
    worksheet.merge_cells("C2:C3")
    workbook.save(path)
    workbook.close()


def test_sqlite_model_creates_required_tables_and_columns(tmp_path):
    settings = Settings(
        excel_path="dummy.xlsx",
        app_pin="1234",
        sqlite_path=str(tmp_path / "process_entries.sqlite3"),
    )

    init_db(settings)

    with create_session(settings) as session:
        inspector = inspect(session.get_bind())
        table_names = set(inspector.get_table_names())
        entry_columns = {column["name"] for column in inspector.get_columns("entries")}
        machine_columns = {
            column["name"] for column in inspector.get_columns("machines")
        }
        machine_group_columns = {
            column["name"] for column in inspector.get_columns("machine_groups")
        }
        machine_group_machine_columns = {
            column["name"] for column in inspector.get_columns("machine_group_machines")
        }
        production_engineer_columns = {
            column["name"] for column in inspector.get_columns("production_engineers")
        }
        cycle_table_columns = {
            column["name"]
            for column in inspector.get_columns("machine_cycle_table_rows")
        }
        breakdown_columns = {
            column["name"] for column in inspector.get_columns("machine_breakdowns")
        }
        amount_control_columns = {
            column["name"] for column in inspector.get_columns("amount_control_shifts")
        }
        production_loss_report_columns = {
            column["name"]
            for column in inspector.get_columns("production_loss_reports")
        }
        production_loss_row_columns = {
            column["name"]
            for column in inspector.get_columns("production_loss_report_rows")
        }
        production_loss_ifs_columns = {
            column["name"]
            for column in inspector.get_columns("production_loss_ifs_actuals")
        }
        production_loss_label_event_columns = {
            column["name"]
            for column in inspector.get_columns("production_loss_label_events")
        }
        amount_control_foreign_keys = inspector.get_foreign_keys(
            "amount_control_shifts"
        )
        amount_control_check_constraints = {
            constraint["name"]
            for constraint in inspector.get_check_constraints("amount_control_shifts")
        }
        amount_control_indexes = {
            index["name"] for index in inspector.get_indexes("amount_control_shifts")
        }
        breakdown_foreign_keys = inspector.get_foreign_keys("machine_breakdowns")
        production_loss_row_foreign_keys = inspector.get_foreign_keys(
            "production_loss_report_rows"
        )
        production_loss_label_event_check_constraints = {
            constraint["name"]
            for constraint in inspector.get_check_constraints(
                "production_loss_label_events"
            )
        }
        production_loss_label_event_indexes = {
            index["name"]
            for index in inspector.get_indexes("production_loss_label_events")
        }
        breakdown_check_constraints = {
            constraint["name"]
            for constraint in inspector.get_check_constraints("machine_breakdowns")
        }

    assert table_names >= {
        "tour_contexts",
        "entries",
        "machines",
        "machine_groups",
        "machine_group_machines",
        "machine_cycle_table_rows",
        "production_engineers",
        "machine_breakdowns",
        "amount_control_shifts",
        "production_loss_reports",
        "production_loss_report_rows",
        "production_loss_ifs_actuals",
        "production_loss_label_events",
    }
    assert "machines_cache" not in table_names
    assert {f"col_{letter}" for letter in "abcdefghijklmnopqrstuvwxy"} <= entry_columns
    assert {
        "client_request_id",
        "client_recorded_at",
        "tour_context_id",
        "status",
        "notes",
        "mold_info",
        "sync_status",
        "excel_row_number",
        "last_error",
        "submitted_at",
        "created_at",
        "updated_at",
        "synced_at",
    } <= entry_columns
    assert {
        "id",
        "machine_code",
        "hall_number",
        "hall_name",
        "display_order",
        "created_at",
        "updated_at",
    } <= machine_columns
    assert {
        "id",
        "group_name",
        "display_order",
        "created_at",
        "updated_at",
    } <= machine_group_columns
    assert {
        "machine_group_id",
        "machine_id",
        "created_at",
        "updated_at",
    } <= machine_group_machine_columns
    assert {
        "id",
        "full_name",
        "display_order",
        "active",
        "created_at",
        "updated_at",
    } <= production_engineer_columns
    assert {
        "id",
        "source_row_number",
        "col_a",
        "col_b",
        "col_c",
        "col_d",
        "col_e",
        "col_g",
        "created_at",
        "updated_at",
    } <= cycle_table_columns
    assert {
        "process_date",
        "machine_code",
        "production_engineer_id",
    } <= entry_columns
    assert {
        "id",
        "machine_id",
        "entry_id",
        "amount_control_shift_id",
        "produced_product",
        "stop_reason",
        "duration_minutes",
        "stopped_at",
        "resumed_at",
        "created_at",
        "updated_at",
    } <= breakdown_columns
    assert {
        "id",
        "client_request_id",
        "client_recorded_at",
        "record_date",
        "machine_id",
        "job_order",
        "shift",
        "worker_names",
        "produced_quantity",
        "created_at",
        "updated_at",
    } <= amount_control_columns
    assert {
        "id",
        "date_from",
        "date_to",
        "generated_at",
        "output_path",
        "row_count",
        "warning_count",
        "source_summary_json",
        "created_at",
        "updated_at",
    } <= production_loss_report_columns
    assert {
        "id",
        "report_id",
        "record_date",
        "machine_id",
        "machine_code",
        "machine_name",
        "job_order",
        "lot_batch_no",
        "product_no",
        "product_description",
        "gram",
        "active_cavities",
        "cycle_time_seconds",
        "shift_2400_0800_quantity",
        "shift_2400_0800_actual_quantity",
        "shift_2400_0800_machine_minutes",
        "shift_2400_0800_optimum_quantity",
        "shift_2400_0800_loss_quantity",
        "shift_0800_1600_quantity",
        "shift_0800_1600_actual_quantity",
        "shift_0800_1600_machine_minutes",
        "shift_0800_1600_optimum_quantity",
        "shift_0800_1600_loss_quantity",
        "shift_1600_2400_quantity",
        "shift_1600_2400_actual_quantity",
        "shift_1600_2400_machine_minutes",
        "shift_1600_2400_optimum_quantity",
        "shift_1600_2400_loss_quantity",
        "daily_total_quantity",
        "net_machine_minutes",
        "gross_elapsed_minutes",
        "gross_start_at",
        "gross_finish_at",
        "optimum_net_quantity",
        "production_loss_net",
        "optimum_gross_quantity",
        "production_loss_gross",
        "break_loss_quantity",
        "local_breakdown_minutes",
        "warnings",
    } <= production_loss_row_columns
    assert {
        "id",
        "job_order",
        "machine_code",
        "work_center_no",
        "part_no",
        "part_description",
        "actual_start_at",
        "actual_finish_at",
        "net_machine_minutes",
        "raw_json",
        "source_updated_at",
    } <= production_loss_ifs_columns
    assert {
        "id",
        "result_key",
        "report_id",
        "print_job_id",
        "archive_exec_time",
        "label_time",
        "record_date",
        "shift",
        "machine_code",
        "job_order",
        "part_no",
        "product_description",
        "quantity",
        "package_id",
        "lot_batch_no",
        "pallet_no",
        "sequence_no",
        "raw_xml",
        "source_updated_at",
        "created_at",
        "updated_at",
    } <= production_loss_label_event_columns
    assert {
        (
            tuple(foreign_key["constrained_columns"]),
            foreign_key["referred_table"],
            tuple(foreign_key["referred_columns"]),
            foreign_key["options"].get("ondelete"),
        )
        for foreign_key in amount_control_foreign_keys
    } >= {
        (("machine_id",), "machines", ("id",), "RESTRICT"),
    }
    assert {
        "ck_amount_control_shifts_record_date",
        "ck_amount_control_shifts_job_order",
        "ck_amount_control_shifts_shift",
        "ck_amount_control_shifts_worker_names",
        "ck_amount_control_shifts_produced_quantity",
    } <= amount_control_check_constraints
    assert {
        "ix_amount_control_shifts_record_date",
        "ix_amount_control_shifts_machine_id",
        "ix_amount_control_shifts_created_at",
        "ux_amount_control_shifts_client_request_id",
        "ux_amount_control_shifts_business_key",
    } <= amount_control_indexes
    assert {
        (
            tuple(foreign_key["constrained_columns"]),
            foreign_key["referred_table"],
            tuple(foreign_key["referred_columns"]),
            foreign_key["options"].get("ondelete"),
        )
        for foreign_key in breakdown_foreign_keys
    } >= {
        (("machine_id",), "machines", ("id",), "RESTRICT"),
        (("entry_id",), "entries", ("id",), "SET NULL"),
        (
            ("amount_control_shift_id",),
            "amount_control_shifts",
            ("id",),
            "SET NULL",
        ),
    }
    assert {
        (
            tuple(foreign_key["constrained_columns"]),
            foreign_key["referred_table"],
            tuple(foreign_key["referred_columns"]),
            foreign_key["options"].get("ondelete"),
        )
        for foreign_key in production_loss_row_foreign_keys
    } >= {
        (
            ("report_id",),
            "production_loss_reports",
            ("id",),
            "CASCADE",
        ),
        (("machine_id",), "machines", ("id",), "SET NULL"),
    }
    assert {
        "ck_production_loss_label_events_result_key",
        "ck_production_loss_label_events_record_date",
        "ck_production_loss_label_events_shift",
        "ck_production_loss_label_events_machine_code",
        "ck_production_loss_label_events_job_order",
        "ck_production_loss_label_events_quantity",
    } <= production_loss_label_event_check_constraints
    assert {
        "ux_production_loss_label_events_result_key",
        "ix_production_loss_label_events_lookup",
        "ix_production_loss_label_events_archive_time",
    } <= production_loss_label_event_indexes
    assert {
        "ck_machine_breakdowns_produced_product",
        "ck_machine_breakdowns_stop_reason",
        "ck_machine_breakdowns_duration_minutes",
        "ck_machine_breakdowns_stop_time_order",
    } <= breakdown_check_constraints


def test_init_db_seeds_production_engineers(tmp_path):
    settings = Settings(
        excel_path="dummy.xlsx",
        app_pin="1234",
        sqlite_path=str(tmp_path / "process_entries.sqlite3"),
    )

    init_db(settings)
    init_db(settings)

    with create_session(settings) as session:
        engineers = (
            session.query(ProductionEngineer)
            .order_by(ProductionEngineer.display_order)
            .all()
        )

    assert [engineer.full_name for engineer in engineers] == [
        "Barış Çetik",
        "Abdullah Kaya",
        "Fevzi Kılınç",
    ]
    assert [engineer.active for engineer in engineers] == [True, True, True]


def test_init_db_seeds_factory_machines_by_hall(tmp_path):
    settings = Settings(
        excel_path="dummy.xlsx",
        app_pin="1234",
        sqlite_path=str(tmp_path / "process_entries.sqlite3"),
    )

    init_db(settings)
    init_db(settings)

    with create_session(settings) as session:
        machines = (
            session.query(Machine)
            .order_by(Machine.hall_number, Machine.display_order)
            .all()
        )

    assert len(machines) == 47
    assert {
        hall_number: sum(
            1 for machine in machines if machine.hall_number == hall_number
        )
        for hall_number in range(1, 5)
    } == {1: 15, 2: 12, 3: 11, 4: 9}
    assert [
        machine.machine_code for machine in machines if machine.hall_number == 1
    ] == [
        "101",
        "102",
        "103",
        "104",
        "171",
        "172",
        "173",
        "174",
        "175",
        "131",
        "132",
        "133",
        "134",
        "135",
        "138",
    ]
    assert [
        machine.machine_code for machine in machines if machine.hall_number == 2
    ] == [
        "176",
        "177",
        "178",
        "179",
        "271",
        "161",
        "162",
        "163",
        "164",
        "165",
        "181",
        "182",
    ]
    assert [
        machine.machine_code for machine in machines if machine.hall_number == 3
    ] == [
        "302",
        "303",
        "401",
        "402",
        "404",
        "405",
        "501",
        "502",
        "503",
        "505",
        "506",
    ]
    assert [
        machine.machine_code for machine in machines if machine.hall_number == 4
    ] == [
        "202",
        "205",
        "206",
        "207",
        "209",
        "210",
        "212",
        "213",
        "301",
    ]


def test_init_db_seeds_machine_groups_from_joined_machines(tmp_path):
    settings = Settings(
        excel_path="dummy.xlsx",
        app_pin="1234",
        sqlite_path=str(tmp_path / "process_entries.sqlite3"),
    )

    init_db(settings)
    init_db(settings)

    with create_session(settings) as session:
        group_names = [
            name
            for (name,) in session.query(MachineGroup.group_name)
            .order_by(MachineGroup.display_order)
            .all()
        ]
        link_count = session.query(MachineGroupMachine).count()

    assert group_names == [
        "SP25",
        "SP80",
        "PF6/2",
        "PF8/4",
        "50MB",
        "12M",
        "DPH70",
        "ENJK.",
        "PREFORM",
        "PET \u015e\u0130\u015e.",
        "UNILOY",
        "5L\u0130K",
        "\u0130LA\u00c7",
    ]
    assert link_count == 47

    with create_session(settings) as session:
        session.add(
            Machine(
                machine_code="141",
                hall_number=2,
                hall_name="Hole 2",
                display_order=13,
            )
        )
        session.commit()

    init_db(settings)

    with create_session(settings) as session:
        group_names = [
            name
            for (name,) in session.query(MachineGroup.group_name)
            .order_by(MachineGroup.display_order)
            .all()
        ]
        group_links = session.execute(
            text(
                "SELECT mg.group_name, m.machine_code "
                "FROM machine_groups mg "
                "INNER JOIN machine_group_machines mgm "
                "ON mgm.machine_group_id = mg.id "
                "INNER JOIN machines m ON m.id = mgm.machine_id "
                "ORDER BY mg.display_order, m.machine_code"
            )
        ).all()

    assert group_names == [
        "SP25",
        "SP80",
        "PF4/1",
        "PF6/2",
        "PF8/4",
        "50MB",
        "12M",
        "DPH70",
        "ENJK.",
        "PREFORM",
        "PET \u015e\u0130\u015e.",
        "UNILOY",
        "5L\u0130K",
        "\u0130LA\u00c7",
    ]
    assert len(group_links) == 48
    assert ("PF4/1", "141") in group_links


def test_init_db_seeds_machine_cycle_table_rows_from_workbook(tmp_path):
    cycle_table_path = tmp_path / "cycle.xlsx"
    _create_cycle_seed_workbook(cycle_table_path)
    settings = Settings(
        excel_path="dummy.xlsx",
        app_pin="1234",
        sqlite_path=str(tmp_path / "process_entries.sqlite3"),
        cycle_table_path=str(cycle_table_path),
    )

    init_db(settings)
    init_db(settings)

    with create_session(settings) as session:
        rows = (
            session.query(MachineCycleTableRow)
            .order_by(MachineCycleTableRow.source_row_number)
            .all()
        )

    assert len(rows) == 2
    assert [
        (
            row.source_row_number,
            row.col_a,
            row.col_b,
            row.col_c,
            row.col_d,
            row.col_e,
            row.col_g,
        )
        for row in rows
    ] == [
        (2, "174", "DPH70", "28", "35", "MOLD-A", "14"),
        (3, "174", "DPH70", "28", "40", "MOLD-B", "16"),
    ]


def test_init_db_migrates_existing_tables_for_offline_idempotency(tmp_path):
    sqlite_path = tmp_path / "process_entries.sqlite3"
    connection = sqlite3.connect(sqlite_path)
    try:
        connection.execute("CREATE TABLE tour_contexts (id INTEGER PRIMARY KEY)")
        connection.execute(
            "CREATE TABLE entries ("
            "id INTEGER PRIMARY KEY, "
            "sync_status VARCHAR(32), "
            "tour_context_id INTEGER, "
            "created_at DATETIME)"
        )
        connection.execute(
            "CREATE TABLE auxiliary_systems_submissions ("
            "id INTEGER PRIMARY KEY, "
            "sync_status VARCHAR(32), "
            "recorded_date VARCHAR(32), "
            "created_at DATETIME)"
        )
        connection.execute(
            "CREATE TABLE amount_control_shifts (id INTEGER PRIMARY KEY)"
        )
        connection.commit()
    finally:
        connection.close()

    settings = Settings(
        excel_path="dummy.xlsx",
        app_pin="1234",
        sqlite_path=str(sqlite_path),
    )

    init_db(settings)

    with create_session(settings) as session:
        inspector = inspect(session.get_bind())
        for table_name in (
            "tour_contexts",
            "entries",
            "auxiliary_systems_submissions",
            "amount_control_shifts",
        ):
            columns = {column["name"] for column in inspector.get_columns(table_name)}
            assert {"client_request_id", "client_recorded_at"} <= columns


def test_init_db_migrates_production_loss_shift_snapshot_columns(tmp_path):
    sqlite_path = tmp_path / "process_entries.sqlite3"
    connection = sqlite3.connect(sqlite_path)
    try:
        connection.execute(
            "CREATE TABLE production_loss_reports ("
            "id INTEGER PRIMARY KEY, "
            "date_from VARCHAR(10) NOT NULL, "
            "date_to VARCHAR(10) NOT NULL, "
            "generated_at DATETIME NOT NULL, "
            "row_count INTEGER NOT NULL, "
            "warning_count INTEGER NOT NULL, "
            "source_summary_json TEXT NOT NULL, "
            "created_at DATETIME NOT NULL, "
            "updated_at DATETIME NOT NULL)"
        )
        connection.execute(
            "CREATE TABLE production_loss_report_rows ("
            "id INTEGER PRIMARY KEY, "
            "report_id INTEGER NOT NULL, "
            "record_date VARCHAR(10) NOT NULL, "
            "machine_code VARCHAR(16) NOT NULL, "
            "job_order TEXT NOT NULL, "
            "shift_2400_0800_quantity INTEGER NOT NULL, "
            "shift_0800_1600_quantity INTEGER NOT NULL, "
            "shift_1600_2400_quantity INTEGER NOT NULL, "
            "daily_total_quantity INTEGER NOT NULL, "
            "local_breakdown_minutes INTEGER NOT NULL, "
            "created_at DATETIME NOT NULL, "
            "updated_at DATETIME NOT NULL)"
        )
        connection.execute(
            "INSERT INTO production_loss_reports ("
            "id, date_from, date_to, generated_at, row_count, warning_count, "
            "source_summary_json, created_at, updated_at"
            ") VALUES ("
            "1, '2026-06-17', '2026-06-17', CURRENT_TIMESTAMP, 1, 0, '{}', "
            "CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)"
        )
        connection.execute(
            "INSERT INTO production_loss_report_rows ("
            "id, report_id, record_date, machine_code, job_order, "
            "shift_2400_0800_quantity, shift_0800_1600_quantity, "
            "shift_1600_2400_quantity, daily_total_quantity, "
            "local_breakdown_minutes, created_at, updated_at"
            ") VALUES ("
            "1, 1, '2026-06-17', '101', 'WO-1', 10, 20, 30, 60, 0, "
            "CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)"
        )
        connection.commit()
    finally:
        connection.close()

    settings = Settings(
        excel_path="dummy.xlsx",
        app_pin="1234",
        sqlite_path=str(sqlite_path),
    )

    init_db(settings)
    init_db(settings)

    connection = sqlite3.connect(sqlite_path)
    try:
        columns = {
            row[1]
            for row in connection.execute(
                "PRAGMA table_info(production_loss_report_rows)"
            )
        }
        row = connection.execute(
            "SELECT "
            "shift_2400_0800_quantity, shift_2400_0800_actual_quantity, "
            "shift_2400_0800_machine_minutes, "
            "shift_2400_0800_optimum_quantity, shift_2400_0800_loss_quantity, "
            "shift_0800_1600_quantity, shift_0800_1600_actual_quantity, "
            "shift_0800_1600_machine_minutes, "
            "shift_0800_1600_optimum_quantity, shift_0800_1600_loss_quantity, "
            "shift_1600_2400_quantity, shift_1600_2400_actual_quantity, "
            "shift_1600_2400_machine_minutes, "
            "shift_1600_2400_optimum_quantity, shift_1600_2400_loss_quantity "
            "FROM production_loss_report_rows WHERE id = 1"
        ).fetchone()
    finally:
        connection.close()

    assert {
        "lot_batch_no",
        "shift_2400_0800_actual_quantity",
        "shift_2400_0800_machine_minutes",
        "shift_2400_0800_optimum_quantity",
        "shift_2400_0800_loss_quantity",
        "shift_0800_1600_actual_quantity",
        "shift_0800_1600_machine_minutes",
        "shift_0800_1600_optimum_quantity",
        "shift_0800_1600_loss_quantity",
        "shift_1600_2400_actual_quantity",
        "shift_1600_2400_machine_minutes",
        "shift_1600_2400_optimum_quantity",
        "shift_1600_2400_loss_quantity",
    } <= columns
    assert row == (
        10,
        10,
        None,
        None,
        None,
        20,
        20,
        None,
        None,
        None,
        30,
        30,
        None,
        None,
        None,
    )


def test_init_db_canonicalizes_legacy_entry_columns_once(tmp_path):
    sqlite_path = tmp_path / "process_entries.sqlite3"
    connection = sqlite3.connect(sqlite_path)
    try:
        connection.execute(
            "CREATE TABLE entries ("
            "id INTEGER PRIMARY KEY, "
            "col_f TEXT, "
            "col_g TEXT, "
            "col_h TEXT, "
            "col_i TEXT, "
            "col_j TEXT, "
            "col_k TEXT, "
            "col_l TEXT, "
            "col_m TEXT, "
            "col_n TEXT, "
            "col_o TEXT, "
            "col_p TEXT, "
            "col_q TEXT, "
            "col_x TEXT, "
            "col_y TEXT, "
            "sync_status VARCHAR(32), "
            "created_at DATETIME)"
        )
        connection.execute(
            "INSERT INTO entries ("
            "id, col_f, col_g, col_h, col_i, col_j, col_p, col_q, sync_status"
            ") VALUES (1, '101', 'WO-1', '16', '12', '12.5', '270x2', '30x2', 'synced')"
        )
        connection.commit()
    finally:
        connection.close()

    settings = Settings(
        excel_path="dummy.xlsx",
        app_pin="1234",
        sqlite_path=str(sqlite_path),
    )

    init_db(settings)
    init_db(settings)

    connection = sqlite3.connect(sqlite_path)
    try:
        row = connection.execute(
            "SELECT col_g, col_h, col_j, col_k, col_l, col_x, col_y "
            "FROM entries WHERE id = 1"
        ).fetchone()
    finally:
        connection.close()

    assert row == (None, "WO-1", "16", "12", "12.5", "270x2", "30x2")


def test_init_db_backfills_entry_process_metadata(tmp_path):
    sqlite_path = tmp_path / "process_entries.sqlite3"
    connection = sqlite3.connect(sqlite_path)
    try:
        connection.execute(
            "CREATE TABLE entries ("
            "id INTEGER PRIMARY KEY, "
            "col_a TEXT, "
            "col_c TEXT, "
            "col_f TEXT, "
            "sync_status VARCHAR(32), "
            "created_at DATETIME)"
        )
        connection.execute(
            "INSERT INTO entries (id, col_a, col_c, col_f, sync_status) "
            "VALUES (1, '17.06.2026', 'BARIŞ ÇETİK', 'M-101', 'synced')"
        )
        connection.commit()
    finally:
        connection.close()

    settings = Settings(
        excel_path="dummy.xlsx",
        app_pin="1234",
        sqlite_path=str(sqlite_path),
    )

    init_db(settings)

    connection = sqlite3.connect(sqlite_path)
    try:
        row = connection.execute(
            "SELECT process_date, machine_code, col_c, production_engineer_id "
            "FROM entries WHERE id = 1"
        ).fetchone()
    finally:
        connection.close()

    assert row == ("2026-06-17", "101", "Barış Çetik", 1)


def test_production_loss_label_event_persists(tmp_path):
    settings = Settings(
        excel_path="dummy.xlsx",
        app_pin="1234",
        sqlite_path=str(tmp_path / "process_entries.sqlite3"),
    )
    init_db(settings)

    with session_scope(settings) as session:
        machine = session.query(Machine).filter_by(machine_code="101").one()
        session.add(
            ProductionLossLabelEvent(
                result_key="label-row-1",
                report_id="REPORT-1",
                print_job_id="PRINT-1",
                record_date="2026-06-17",
                shift="08.00-16.00",
                machine_code=machine.machine_code,
                job_order="WO-1",
                part_no="PET-35",
                product_description="PET 28MM 35GR",
                quantity=120,
                package_id="PKG-1",
                lot_batch_no="LOT-1",
                raw_xml="<LabelId>label-row-1</LabelId>",
            )
        )

    with create_session(settings) as session:
        label_event = session.query(ProductionLossLabelEvent).one()
        assert label_event.result_key == "label-row-1"
        assert label_event.report_id == "REPORT-1"
        assert label_event.machine_code == "101"
        assert label_event.job_order == "WO-1"
        assert label_event.shift == "08.00-16.00"
        assert label_event.part_no == "PET-35"
        assert label_event.quantity == 120
        assert label_event.lot_batch_no == "LOT-1"


def test_machine_breakdown_persists_with_machine_and_entry_links(tmp_path):
    settings = Settings(
        excel_path="dummy.xlsx",
        app_pin="1234",
        sqlite_path=str(tmp_path / "process_entries.sqlite3"),
    )
    init_db(settings)

    with session_scope(settings) as session:
        machine = session.query(Machine).filter_by(machine_code="101").one()
        context = TourContext(
            date="2026-06-17",
            ambient_temp="24",
            production_engineer="Barış Çetik",
            shift_chief="Selman",
            shift="08.00-16.00",
        )
        entry = Entry(
            tour_context=context,
            col_a="17.06.2026",
            col_f="101",
            col_g="Product 101",
            sync_status=SYNC_STATUS_PENDING_EXCEL,
        )
        session.add(
            MachineBreakdown(
                machine=machine,
                entry=entry,
                produced_product="Product 101",
                stop_reason="Hydraulic pressure fault",
                duration_minutes=45,
            )
        )

    with create_session(settings) as session:
        breakdown = session.query(MachineBreakdown).one()
        assert breakdown.machine.machine_code == "101"
        assert breakdown.entry is not None
        assert breakdown.produced_product == "Product 101"
        assert breakdown.stop_reason == "Hydraulic pressure fault"
        assert breakdown.duration_minutes == 45


def test_machine_breakdown_rejects_invalid_machine_and_duration(tmp_path):
    settings = Settings(
        excel_path="dummy.xlsx",
        app_pin="1234",
        sqlite_path=str(tmp_path / "process_entries.sqlite3"),
    )
    init_db(settings)

    with create_session(settings) as session:
        machine_id = session.query(Machine.id).filter_by(machine_code="101").scalar()

    with pytest.raises(DatabaseCommitError), session_scope(settings) as session:
        session.add(
            MachineBreakdown(
                machine_id=machine_id,
                produced_product="Product 101",
                stop_reason="",
                duration_minutes=15,
            )
        )

    with pytest.raises(DatabaseCommitError), session_scope(settings) as session:
        session.add(
            MachineBreakdown(
                machine_id=machine_id,
                produced_product="Product 101",
                stop_reason="Hydraulic pressure fault",
                duration_minutes=0,
            )
        )

    with pytest.raises(DatabaseCommitError), session_scope(settings) as session:
        session.add(
            MachineBreakdown(
                machine_id=999_999,
                produced_product="Product 101",
                stop_reason="Hydraulic pressure fault",
                duration_minutes=15,
            )
        )


def test_amount_control_shift_enforces_business_rules(tmp_path):
    settings = Settings(
        excel_path="dummy.xlsx",
        app_pin="1234",
        sqlite_path=str(tmp_path / "process_entries.sqlite3"),
    )
    init_db(settings)

    with create_session(settings) as session:
        machine_id = session.query(Machine.id).filter_by(machine_code="101").scalar()

    with session_scope(settings) as session:
        session.add(
            AmountControlShift(
                record_date="2026-06-17",
                machine_id=machine_id,
                job_order="WO-1",
                shift="08.00-16.00",
                worker_names="Operator One",
                produced_quantity=1200,
            )
        )

    invalid_rows = [
        {
            "job_order": "",
            "shift": "08.00-16.00",
            "worker_names": "Operator One",
            "produced_quantity": 1,
        },
        {
            "job_order": "WO-2",
            "shift": "00.00-08.00",
            "worker_names": "Operator One",
            "produced_quantity": 1,
        },
        {
            "job_order": "WO-2",
            "shift": "08.00-16.00",
            "worker_names": " ",
            "produced_quantity": 1,
        },
        {
            "job_order": "WO-2",
            "shift": "08.00-16.00",
            "worker_names": "Operator One",
            "produced_quantity": -1,
        },
    ]
    for row in invalid_rows:
        with pytest.raises(DatabaseCommitError), session_scope(settings) as session:
            session.add(
                AmountControlShift(
                    record_date="2026-06-17",
                    machine_id=machine_id,
                    **row,
                )
            )

    with pytest.raises(DatabaseCommitError), session_scope(settings) as session:
        session.add(
            AmountControlShift(
                record_date="2026-06-17",
                machine_id=999_999,
                job_order="WO-2",
                shift="08.00-16.00",
                worker_names="Operator One",
                produced_quantity=1,
            )
        )

    with pytest.raises(DatabaseCommitError), session_scope(settings) as session:
        session.add(
            AmountControlShift(
                record_date="2026-06-17",
                machine_id=machine_id,
                job_order="WO-1",
                shift="08.00-16.00",
                worker_names="Operator Two",
                produced_quantity=900,
            )
        )


def test_machine_breakdowns_link_to_amount_control_shift(tmp_path):
    settings = Settings(
        excel_path="dummy.xlsx",
        app_pin="1234",
        sqlite_path=str(tmp_path / "process_entries.sqlite3"),
    )
    init_db(settings)

    with session_scope(settings) as session:
        machine = session.query(Machine).filter_by(machine_code="101").one()
        amount_shift = AmountControlShift(
            record_date="2026-06-17",
            machine=machine,
            job_order="WO-1",
            shift="08.00-16.00",
            worker_names="Operator One, Operator Two",
            produced_quantity=1200,
        )
        session.add_all(
            [
                amount_shift,
                MachineBreakdown(
                    machine=machine,
                    amount_control_shift=amount_shift,
                    produced_product="Product 101",
                    stop_reason="Hydraulic pressure fault",
                    duration_minutes=45,
                ),
                MachineBreakdown(
                    machine=machine,
                    amount_control_shift=amount_shift,
                    produced_product="Product 101",
                    stop_reason="Mold change",
                    duration_minutes=30,
                ),
            ]
        )

    with create_session(settings) as session:
        amount_shift = session.query(AmountControlShift).one()
        assert amount_shift.machine.machine_code == "101"
        assert len(amount_shift.machine_breakdowns) == 2
        assert {
            breakdown.stop_reason for breakdown in amount_shift.machine_breakdowns
        } == {"Hydraulic pressure fault", "Mold change"}


def test_deleting_amount_control_shift_nulls_breakdown_link(tmp_path):
    settings = Settings(
        excel_path="dummy.xlsx",
        app_pin="1234",
        sqlite_path=str(tmp_path / "process_entries.sqlite3"),
    )
    init_db(settings)

    with session_scope(settings) as session:
        machine = session.query(Machine).filter_by(machine_code="101").one()
        amount_shift = AmountControlShift(
            record_date="2026-06-17",
            machine=machine,
            job_order="WO-1",
            shift="08.00-16.00",
            worker_names="Operator One",
            produced_quantity=1200,
        )
        session.add(amount_shift)
        session.flush()
        session.add(
            MachineBreakdown(
                machine=machine,
                amount_control_shift_id=amount_shift.id,
                produced_product="Product 101",
                stop_reason="Hydraulic pressure fault",
                duration_minutes=45,
            )
        )

    with session_scope(settings) as session:
        amount_shift = session.query(AmountControlShift).one()
        session.delete(amount_shift)

    with create_session(settings) as session:
        breakdown = session.query(MachineBreakdown).one()
        assert breakdown.amount_control_shift_id is None


def test_init_db_migrates_machine_breakdowns_amount_control_link(tmp_path):
    sqlite_path = tmp_path / "process_entries.sqlite3"
    connection = sqlite3.connect(sqlite_path)
    try:
        connection.execute(
            "CREATE TABLE machines ("
            "id INTEGER PRIMARY KEY, "
            "machine_code VARCHAR(16) NOT NULL, "
            "hall_number INTEGER NOT NULL, "
            "hall_name VARCHAR(32) NOT NULL, "
            "display_order INTEGER NOT NULL, "
            "created_at DATETIME NOT NULL, "
            "updated_at DATETIME NOT NULL)"
        )
        connection.execute(
            "CREATE TABLE entries ("
            "id INTEGER PRIMARY KEY, "
            "sync_status VARCHAR(32), "
            "created_at DATETIME)"
        )
        connection.execute(
            "CREATE TABLE machine_breakdowns ("
            "id INTEGER PRIMARY KEY, "
            "machine_id INTEGER NOT NULL, "
            "entry_id INTEGER, "
            "produced_product TEXT NOT NULL, "
            "stop_reason TEXT NOT NULL, "
            "duration_minutes INTEGER NOT NULL, "
            "stopped_at DATETIME, "
            "resumed_at DATETIME, "
            "created_at DATETIME NOT NULL, "
            "updated_at DATETIME NOT NULL)"
        )
        connection.execute(
            "INSERT INTO machines ("
            "id, machine_code, hall_number, hall_name, display_order, created_at, updated_at"
            ") VALUES (1, '101', 1, 'Hole 1', 1, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)"
        )
        connection.execute(
            "INSERT INTO entries (id, sync_status, created_at) "
            "VALUES (1, 'synced', CURRENT_TIMESTAMP)"
        )
        connection.execute(
            "INSERT INTO machine_breakdowns ("
            "id, machine_id, entry_id, produced_product, stop_reason, duration_minutes, "
            "created_at, updated_at"
            ") VALUES (1, 1, 1, 'Product 101', 'Hydraulic pressure fault', 45, "
            "CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)"
        )
        connection.commit()
    finally:
        connection.close()

    settings = Settings(
        excel_path="dummy.xlsx",
        app_pin="1234",
        sqlite_path=str(sqlite_path),
    )
    init_db(settings)
    init_db(settings)

    with create_session(settings) as session:
        inspector = inspect(session.get_bind())
        columns = {
            column["name"] for column in inspector.get_columns("machine_breakdowns")
        }
        foreign_keys = inspector.get_foreign_keys("machine_breakdowns")
        breakdown = session.query(MachineBreakdown).one()

    assert "amount_control_shift_id" in columns
    assert breakdown.produced_product == "Product 101"
    assert breakdown.entry_id == 1
    assert breakdown.amount_control_shift_id is None
    assert {
        (
            tuple(foreign_key["constrained_columns"]),
            foreign_key["referred_table"],
            tuple(foreign_key["referred_columns"]),
            foreign_key["options"].get("ondelete"),
        )
        for foreign_key in foreign_keys
    } >= {
        (
            ("amount_control_shift_id",),
            "amount_control_shifts",
            ("id",),
            "SET NULL",
        ),
    }


def test_session_scope_commits_and_rolls_back_safely(tmp_path):
    settings = Settings(
        excel_path="dummy.xlsx",
        app_pin="1234",
        sqlite_path=str(tmp_path / "process_entries.sqlite3"),
    )
    init_db(settings)

    with session_scope(settings) as session:
        context = TourContext(
            date="2026-06-08",
            ambient_temp="24",
            production_engineer="Engineer",
            shift_chief="Selman",
            shift="A",
        )
        session.add(context)
        session.flush()
        session.add(
            Entry(
                tour_context_id=context.id,
                col_a="sample",
                sync_status=SYNC_STATUS_PENDING_EXCEL,
            )
        )

    with create_session(settings) as session:
        assert session.query(TourContext).count() == 1
        assert session.query(Entry).count() == 1

    with pytest.raises(DatabaseCommitError), session_scope(settings) as session:
        session.add(Entry(sync_status="not_a_real_sync_status"))

    with create_session(settings) as session:
        assert session.query(Entry).count() == 1


def test_init_db_drops_legacy_machine_cache_table(tmp_path):
    settings = Settings(
        excel_path="dummy.xlsx",
        app_pin="1234",
        sqlite_path=str(tmp_path / "process_entries.sqlite3"),
    )
    init_db(settings)

    with create_session(settings) as session:
        session.execute(
            text(
                "CREATE TABLE machines_cache ("
                "machine_id TEXT PRIMARY KEY, "
                "latest_product TEXT)"
            )
        )
        session.execute(
            text(
                "INSERT INTO machines_cache (machine_id, latest_product) "
                "VALUES ('M-01', 'Product')"
            )
        )
        session.commit()

    init_db(settings)

    with create_session(settings) as session:
        inspector = inspect(session.get_bind())
        assert "machines_cache" not in set(inspector.get_table_names())
