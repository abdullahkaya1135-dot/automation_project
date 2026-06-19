from datetime import date
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from app.config import Settings
from app.database import init_db
from app.services.cycle_report_service import (
    CycleTableEntry,
    NoTodayRowsError,
    ProcessCycleRow,
    build_cycle_report_rows,
    create_cycle_report,
    normalize_machine_group,
    parse_part_description,
)
from app.services.process_excel_import import import_process_excel_to_database
from app.services.shop_order_source import ShopOrderOption

HEADERS = [
    "Date",
    "Ambient Temp",
    "Production Engineer",
    "Shift Chief",
    "Shift",
    "Machine",
    "Product",
    "Work Order",
    "Raw Material",
    "Total Cavity Count",
    "Active Cavity Count",
    "Cycle Time",
    "Cooling Time",
    "Injection Time",
    "Blow Time",
    "Conditioner Temp",
    "Dryer Temp",
    "Injection Pressure",
    "Injection Speed",
    "Holding Time",
    "Holding Speed",
    "Holding Pressure",
    "Clamp Force",
    "Oven Temperatures",
    "Mold Temperatures",
]
REPORT_HEADERS = [
    "Makine Adı",
    "Makine No.",
    "Ağız",
    "Gramaj",
    "Toplam Göz",
    "Çalışan Göz",
    "Gerçekleşen Çevrim Süresi",
    "Optimum Çevrim Süresi",
    "Kontrol Durumu",
]


def _create_process_workbook(path: Path, rows: list[list[object]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "PROSES 2026"
    worksheet.append(HEADERS)
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    workbook.close()


def _process_row(
    day: str,
    machine_no: object,
    work_order: object,
    total_cavity: object,
    active_cavity: object,
    cycle_time: object,
) -> list[object]:
    row: list[object] = [None] * 25
    row[0] = day
    row[5] = machine_no
    row[7] = work_order
    row[9] = total_cavity
    row[10] = active_cavity
    row[11] = cycle_time
    return row


def _create_cycle_table(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sayfa1"
    worksheet.append(
        [
            "YENİ MAKİNA NO - New Machine Code",
            "MAKİNA ADI Machine Name",
            "AĞIZ ÇAP  Neck Diameter",
            "GR",
            "PREFORM MOLD CODE",
            "PREFORM KALIBI ÜRETİM YILI PRODUCTION DATE",
            "ESTIMATED CYCLE TIME",
            "BPH",
        ]
    )
    worksheet.append([174, "DPH70", 28, 35, None, None, "14", None])
    worksheet.append([None, None, None, 40, None, None, "16", None])
    worksheet.merge_cells("A2:A3")
    worksheet.merge_cells("B2:B3")
    worksheet.merge_cells("C2:C3")
    worksheet.append([161, "PF 6/2", 28, 45, None, None, "24", None])
    worksheet.append([162, "PF 6/2", 28, 45, None, None, "21", None])
    workbook.save(path)
    workbook.close()


def _ifs_operations() -> list[dict[str, object]]:
    return [
        {
            "OrderNo": "2668",
            "PreferredResourceId": "175",
            "WorkCenterNo": "70DPH",
            "PartNoDesc": "PET0207-01 750ML SEFFAF SEFFAF 28MM YIVLI 35GR",
        },
        {
            "OrderNo": "PF-ORDER",
            "PreferredResourceId": "162",
            "WorkCenterNo": "PF-6",
            "PartNoDesc": "PET0001-01 400ML SEFFAF 28MM YIVLI 45GR",
        },
    ]


def _settings(tmp_path: Path) -> Settings:
    return Settings(
        excel_path=str(tmp_path / "process.xlsx"),
        app_pin="1234",
        sqlite_path=str(tmp_path / "process_entries.sqlite3"),
        backup_dir=str(tmp_path / "backups"),
        cycle_table_path=str(tmp_path / "cycle.xlsx"),
        report_output_dir=str(tmp_path / "reports"),
    )


def _seed_process_database(settings: Settings) -> None:
    init_db(settings)
    import_process_excel_to_database(settings)


def _decimal(value: object) -> Decimal:
    return Decimal(str(value))


def _cycle_entry(
    machine_no: object,
    machine_group: str,
    neck: object,
    gram: object,
    cycle: object,
) -> CycleTableEntry:
    return CycleTableEntry(
        machine_no=str(machine_no),
        machine_group_key=normalize_machine_group(machine_group),
        neck_diameter=_decimal(neck),
        gram=_decimal(gram),
        estimated_cycle_time=_decimal(cycle),
    )


def _cycle_entries(
    *entries: CycleTableEntry,
) -> dict[tuple[str, Decimal, Decimal], list[CycleTableEntry]]:
    indexed: dict[tuple[str, Decimal, Decimal], list[CycleTableEntry]] = {}
    for entry in entries:
        key = (entry.machine_group_key, entry.neck_diameter, entry.gram)
        indexed.setdefault(key, []).append(entry)
    return indexed


def _build_single_report_row(
    *,
    part_description: str,
    cycle_entries: dict[tuple[str, Decimal, Decimal], list[CycleTableEntry]],
    machine_no: str = "175",
    machine_group: str = "70DPH",
    cycle_time: object = 10,
):
    rows = build_cycle_report_rows(
        [
            ProcessCycleRow(
                row_number=1,
                machine_no=machine_no,
                work_order="WO-1",
                total_cavity_count=1,
                active_cavity_count=1,
                actual_cycle_time=_decimal(cycle_time),
            )
        ],
        [
            ShopOrderOption(
                order_no="WO-1",
                resource_id=machine_no,
                work_center_no=machine_group,
                part_description=part_description,
            )
        ],
        cycle_entries,
    )
    return rows[0]


def test_parse_part_description_extracts_neck_and_gram_values():
    assert parse_part_description(
        "ENJ0057-01 KAPAK SEFFAF SARI 44MM CAKMA 6GR"
    ) == (44, 6)
    assert parse_part_description(
        "IBM0008-01 66ML OPAK BEYAZ 18MM CAKMA 9,5GR (SEA)"
    ) == (18, pytest.approx(9.5))


def test_normalize_machine_group_aliases():
    assert normalize_machine_group("70DPH") == "DPH70"
    assert normalize_machine_group("50MB2") == "50MB"
    assert normalize_machine_group("PF-6") == "PF62"
    assert normalize_machine_group("PF-8") == "PF84"
    assert normalize_machine_group("12M1") == "12M"


def test_cycle_report_exact_cycle_match_wins_before_near_match():
    row = _build_single_report_row(
        part_description="PET 28MM 35GR",
        cycle_entries=_cycle_entries(
            _cycle_entry(175, "DPH70", 28, 35, 14),
            _cycle_entry(175, "DPH70", 29, 35, 99),
        ),
    )

    assert row.optimum_cycle_time == Decimal("14")
    assert row.control_status == "Uygun"


def test_cycle_report_uses_nearest_neck_when_gram_matches_within_range():
    row = _build_single_report_row(
        part_description="PET 30MM 35GR",
        cycle_entries=_cycle_entries(
            _cycle_entry(175, "DPH70", 28, 35, 14),
            _cycle_entry(175, "DPH70", 34, 35, 99),
        ),
    )

    assert row.optimum_cycle_time == Decimal("14")
    assert row.control_status == "Uygun"


def test_cycle_report_uses_nearest_gram_when_neck_matches_within_range():
    row = _build_single_report_row(
        part_description="PET 28MM 38GR",
        cycle_entries=_cycle_entries(
            _cycle_entry(175, "DPH70", 28, 35, 14),
            _cycle_entry(175, "DPH70", 28, 42, 99),
        ),
    )

    assert row.optimum_cycle_time == Decimal("14")
    assert row.control_status == "Uygun"


def test_cycle_report_leaves_optimum_blank_when_near_match_is_out_of_range():
    row = _build_single_report_row(
        part_description="PET 32MM 39GR",
        cycle_entries=_cycle_entries(
            _cycle_entry(175, "DPH70", 28, 39, 14),
            _cycle_entry(175, "DPH70", 32, 35, 16),
        ),
    )

    assert row.optimum_cycle_time is None
    assert row.control_status == "Optimum çevrim bulunamadı"


def test_cycle_report_near_match_tie_uses_higher_candidate_value():
    row = _build_single_report_row(
        part_description="PET 28MM 35GR",
        cycle_entries=_cycle_entries(
            _cycle_entry(175, "DPH70", 25, 35, 14),
            _cycle_entry(175, "DPH70", 31, 35, 16),
        ),
        cycle_time=17,
    )

    assert row.optimum_cycle_time == Decimal("16")
    assert row.control_status == "Uygun"


def test_cycle_report_tries_neck_fallback_before_gram_fallback():
    row = _build_single_report_row(
        part_description="PET 30MM 38GR",
        cycle_entries=_cycle_entries(
            _cycle_entry(175, "DPH70", 29, 38, 14),
            _cycle_entry(175, "DPH70", 30, 37, 16),
        ),
        cycle_time=16,
    )

    assert row.optimum_cycle_time == Decimal("14")
    assert row.control_status == "Kontrol Edilecek"


def test_create_cycle_report_matches_sources_and_uses_machine_tiebreaker(
    tmp_path,
    monkeypatch,
):
    async def fake_fetch_pet_ongoing_operations(_settings):
        return _ifs_operations()

    monkeypatch.setattr(
        "app.services.cycle_report_service.fetch_pet_ongoing_operations",
        fake_fetch_pet_ongoing_operations,
    )
    _create_process_workbook(
        tmp_path / "process.xlsx",
        [
            _process_row("09.06.2026", 175, 2668, 3, 2, 213),
            _process_row("09.06.2026", 162, "PF-ORDER", 6, 6, "23,1"),
        ],
    )
    _create_cycle_table(tmp_path / "cycle.xlsx")
    settings = _settings(tmp_path)
    _seed_process_database(settings)

    result = create_cycle_report(settings, date(2026, 6, 9))

    assert result.row_count == 2
    assert result.matched_count == 2
    assert result.warning_count == 0
    assert result.output_path.name == "09.06.2026 Cycle Report.xlsx"
    assert result.output_path.exists()

    workbook = load_workbook(result.output_path)
    worksheet = workbook["Çevrim Kontrol"]
    assert [cell.value for cell in worksheet[1]] == REPORT_HEADERS
    assert [cell.value for cell in worksheet[2]] == [
        "PF-6",
        "162",
        28,
        45,
        6,
        6,
        23.1,
        21,
        "Uygun",
    ]
    assert [cell.value for cell in worksheet[3]] == [
        "70DPH",
        "175",
        28,
        35,
        3,
        2,
        213,
        14,
        "Kontrol Edilecek",
    ]
    workbook.close()


def test_create_cycle_report_rejects_no_today_rows_without_output(tmp_path):
    _create_process_workbook(
        tmp_path / "process.xlsx",
        [_process_row("08.06.2026", 175, 2668, 3, 2, 213)],
    )
    _create_cycle_table(tmp_path / "cycle.xlsx")
    settings = _settings(tmp_path)
    _seed_process_database(settings)

    with pytest.raises(NoTodayRowsError):
        create_cycle_report(settings, date(2026, 6, 9))

    assert not (tmp_path / "reports").exists()
