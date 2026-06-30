from openpyxl import Workbook

from app.core.config import Settings


def sqlite_settings(tmp_path, **overrides) -> Settings:
    values = {
        "excel_path": "dummy.xlsx",
        "app_pin": "1234",
        "sqlite_path": str(tmp_path / "process_entries.sqlite3"),
    }
    values.update(overrides)
    return Settings(**values)


def create_cycle_seed_workbook(path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sayfa1"
    worksheet.append(
        [
            "Machine No",
            "Machine Name",
            "Neck",
            "GR",
            "Mold Code",
            "Production Date",
            "Estimated Cycle Time",
        ]
    )
    worksheet.append([174, "DPH70", 28, 35, "MOLD-A", None, 14])
    worksheet.append([None, None, None, 40, "MOLD-B", None, 16])
    worksheet.merge_cells("A2:A3")
    worksheet.merge_cells("B2:B3")
    worksheet.merge_cells("C2:C3")
    workbook.save(path)
    workbook.close()
