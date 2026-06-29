import os

import pytest
from openpyxl import Workbook

from app.features.production_planning.service import (
    MachinePlanningRow,
    ProductionPlanningError,
    latest_planning_workbook,
    read_machine_planning_rows,
    read_visible_planning_orders,
    resolve_planning_workbook,
)


def _write_candidate(path, modified_time):
    path.write_bytes(b"not a real workbook")
    os.utime(path, (modified_time, modified_time))


def test_read_visible_planning_orders_ignores_hidden_content_and_deduplicates(
    tmp_path,
):
    workbook_path = tmp_path / "planning.xlsx"
    workbook = Workbook()

    visible = workbook.active
    visible.title = "Visible"
    visible["A1"] = "0"
    visible["A2"] = "123"
    visible["A3"] = "KAPATILACAK"
    visible["A4"] = "123"
    visible["A5"] = "456-789"
    visible["A6"] = "999"
    visible.row_dimensions[6].hidden = True

    hidden_sheet = workbook.create_sheet("Hidden Sheet")
    hidden_sheet.sheet_state = "hidden"
    hidden_sheet["A1"] = "555"

    hidden_column = workbook.create_sheet("Hidden Column")
    hidden_column["A1"] = "777"
    hidden_column.column_dimensions["A"].hidden = True

    second_visible = workbook.create_sheet("Second Visible")
    second_visible["A1"] = "789"
    second_visible["A2"] = "321"

    workbook.save(workbook_path)
    workbook.close()

    orders = read_visible_planning_orders(workbook_path)

    assert [order.order_no for order in orders] == ["123", "456", "789", "321"]
    assert [
        (order.sheet_name, order.row_number, order.cell_value) for order in orders
    ] == [
        ("Visible", 2, "123"),
        ("Visible", 5, "456-789"),
        ("Visible", 5, "456-789"),
        ("Second Visible", 2, "321"),
    ]


def test_read_machine_planning_rows_groups_visible_jobs_by_machine_sequence(
    tmp_path,
):
    workbook_path = tmp_path / "planning.xlsx"
    workbook = Workbook()

    visible = workbook.active
    visible.title = "Plan"
    visible["A1"] = "999"
    visible["C2"] = "M101"
    visible["A3"] = "1001"
    visible["B3"] = "1"
    visible["C3"] = "P-100"
    visible["D3"] = "Mold A"
    visible["E3"] = 120
    visible["A4"] = "1001"
    visible["B4"] = "2"
    visible["C4"] = "P-101"
    visible["D4"] = "Mold B"
    visible["E4"] = "80"
    visible["A5"] = "0"
    visible["C5"] = "P-ZERO"
    visible["A6"] = "1002-1003"
    visible["B6"] = "3"
    visible["C6"] = "P-102"
    visible["D6"] = "Mold C"
    visible["E6"] = 40
    visible["C7"] = "M201"
    visible["A8"] = "2001"
    visible["B8"] = "1"
    visible["C8"] = "P-200"
    visible["D8"] = "Mold D"
    visible["E8"] = 60
    visible["A9"] = "2002"
    visible["C9"] = "P-HIDDEN"
    visible.row_dimensions[9].hidden = True

    second_visible = workbook.create_sheet("Second")
    second_visible["C1"] = "M101"
    second_visible["A2"] = "1004"
    second_visible["B2"] = "1"
    second_visible["C2"] = "P-103"
    second_visible["D2"] = "Mold E"
    second_visible["E2"] = 10

    hidden_sheet = workbook.create_sheet("Hidden Sheet")
    hidden_sheet.sheet_state = "hidden"
    hidden_sheet["C1"] = "M301"
    hidden_sheet["A2"] = "3001"

    hidden_column = workbook.create_sheet("Hidden Column")
    hidden_column["C1"] = "M301"
    hidden_column["A2"] = "3002"
    hidden_column.column_dimensions["A"].hidden = True

    workbook.save(workbook_path)
    workbook.close()

    rows = read_machine_planning_rows(workbook_path)

    assert rows == [
        MachinePlanningRow(
            machine_code="101",
            order_no="1001",
            sheet_name="Plan",
            row_number=3,
            sequence_index=0,
            product_no="P-100",
            mold="Mold A",
            order_quantity="120",
            cell_value="1001",
        ),
        MachinePlanningRow(
            machine_code="101",
            order_no="1001",
            sheet_name="Plan",
            row_number=4,
            sequence_index=1,
            product_no="P-101",
            mold="Mold B",
            order_quantity="80",
            cell_value="1001",
        ),
        MachinePlanningRow(
            machine_code="101",
            order_no="1002",
            sheet_name="Plan",
            row_number=6,
            sequence_index=2,
            product_no="P-102",
            mold="Mold C",
            order_quantity="40",
            cell_value="1002-1003",
        ),
        MachinePlanningRow(
            machine_code="101",
            order_no="1003",
            sheet_name="Plan",
            row_number=6,
            sequence_index=3,
            product_no="P-102",
            mold="Mold C",
            order_quantity="40",
            cell_value="1002-1003",
        ),
        MachinePlanningRow(
            machine_code="201",
            order_no="2001",
            sheet_name="Plan",
            row_number=8,
            sequence_index=0,
            product_no="P-200",
            mold="Mold D",
            order_quantity="60",
            cell_value="2001",
        ),
        MachinePlanningRow(
            machine_code="101",
            order_no="1004",
            sheet_name="Second",
            row_number=2,
            sequence_index=4,
            product_no="P-103",
            mold="Mold E",
            order_quantity="10",
            cell_value="1004",
        ),
    ]


def test_read_machine_planning_rows_ignores_temporary_excel_files(tmp_path):
    workbook_path = tmp_path / "~$planning.xlsx"
    workbook_path.write_bytes(b"not a real workbook")

    assert read_machine_planning_rows(workbook_path) == []


def test_latest_planning_workbook_prefers_newest_filename_date_over_modified_time(
    tmp_path,
):
    older_date_newer_modified = tmp_path / "12.06.2026 CIZELGE 1.xlsx"
    newer_date_older_modified = tmp_path / "13.06.2026 CIZELGE 1.xlsx"
    excel_temp_file = tmp_path / "~$14.06.2026 CIZELGE 1.xlsx"
    undated_newer_modified = tmp_path / "manual-planning.xlsx"

    _write_candidate(older_date_newer_modified, 300)
    _write_candidate(newer_date_older_modified, 100)
    _write_candidate(excel_temp_file, 500)
    _write_candidate(undated_newer_modified, 600)

    assert latest_planning_workbook(tmp_path) == newer_date_older_modified


def test_latest_planning_workbook_uses_modified_time_as_same_day_tiebreaker(
    tmp_path,
):
    first = tmp_path / "12.06.2026 CIZELGE 1.xlsx"
    second = tmp_path / "12.06.2026 CIZELGE 2.xlsx"

    _write_candidate(first, 100)
    _write_candidate(second, 200)

    assert latest_planning_workbook(tmp_path) == second


def test_latest_planning_workbook_uses_modified_time_when_names_have_no_dates(
    tmp_path,
):
    older = tmp_path / "planning-a.xlsx"
    newer = tmp_path / "planning-b.xlsx"

    _write_candidate(older, 100)
    _write_candidate(newer, 200)

    assert latest_planning_workbook(tmp_path) == newer


def test_latest_planning_workbook_raises_when_no_valid_candidates(tmp_path):
    (tmp_path / "archive").mkdir()
    _write_candidate(tmp_path / "~$12.06.2026 CIZELGE 1.xlsx", 100)
    (tmp_path / "12.06.2026 CIZELGE 1.xls").write_bytes(b"old format")

    with pytest.raises(ProductionPlanningError, match="No production planning"):
        latest_planning_workbook(tmp_path)


def test_resolve_planning_workbook_uses_static_path_when_folder_is_blank(tmp_path):
    fallback_path = tmp_path / "manual.xlsx"

    assert resolve_planning_workbook("", fallback_path) == fallback_path
