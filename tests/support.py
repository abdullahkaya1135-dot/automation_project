import sqlite3
from collections.abc import Callable, Iterator
from contextlib import contextmanager
from pathlib import Path
from typing import Any

import httpx
from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.core.config import Settings
from app.main import create_app

PROCESS_WORKBOOK_HEADERS = [
    "Date",
    "Ambient Temp",
    "Production Engineer",
    "Shift Chief",
    "Shift",
    "Machine",
    "Product",
    "Work Order",
    "Raw Material",
    "Total Cavity Count",
    "Active Cavity Count",
    "Cycle Time",
    "Cooling Time",
    "Injection Time",
    "Blow Time",
    "Conditioner Temp",
    "Dryer Temp",
    "Injection Pressure",
    "Injection Speed",
    "Holding Time",
    "Holding Speed",
    "Holding Pressure",
    "Clamp Force",
    "Oven Temperatures",
    "Mold Temperatures",
]

AUXILIARY_WORKBOOK_HEADERS = [
    "TARÄ°H",
    "MAKÄ°NE ",
    "Ã‡IKIÅ FREKANSI (HZ)",
    "SET DEÄERÄ° (BAR)",
    "GERÄ° BESLEME (BAR)",
    "ISI SET DEÄERÄ° Â©",
    "GÄ°RÄ°Å ISISI Â©",
    "Ã‡IKIÅ ISISI Â©",
    "KOMP BASINÃ‡ DEÄERÄ° (BAR)",
]


def make_settings(tmp_path: Path, **overrides: Any) -> Settings:
    values = {
        "excel_path": str(tmp_path / "process.xlsx"),
        "app_pin": "1234",
        "session_secret": "test-session-secret",
        "sqlite_path": str(tmp_path / "process_entries.sqlite3"),
        "backup_dir": str(tmp_path / "backups"),
        "cycle_table_path": str(tmp_path / "cycle.xlsx"),
        "report_output_dir": str(tmp_path / "reports"),
        "auxiliary_systems_form_path": str(tmp_path / "auxiliary_form.xlsx"),
        "auxiliary_systems_target_path": str(tmp_path / "auxiliary.xlsx"),
    } | overrides
    return Settings(**values)


def create_process_workbook(
    path: Path,
    *,
    sheet_name: str = "PROSES 2026",
    headers: list[str] | None = None,
    last_data_row: int | None = None,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(headers or PROCESS_WORKBOOK_HEADERS)
    if last_data_row is not None:
        worksheet.cell(row=last_data_row, column=1, value="existing")
    workbook.save(path)
    workbook.close()


def create_cycle_seed_workbook(
    path: Path,
    *,
    sheet_name: str = "Sayfa1",
    rows: list[list[Any]] | None = None,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(
        [
            "Machine No",
            "Machine Name",
            "Neck",
            "GR",
            "Mold Code",
            "Production Date",
            "Estimated Cycle Time",
            "BPH",
        ]
    )
    for row in rows or [[174, "DPH70", 28, 35, "MOLD-A", None, 14, None]]:
        worksheet.append(row)
    workbook.save(path)
    workbook.close()


def create_auxiliary_workbook(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "YARDIMCI TESÄ°SLER TAKÄ°P"
    worksheet.append(AUXILIARY_WORKBOOK_HEADERS)
    workbook.save(path)
    workbook.close()


def create_auxiliary_form(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "FORM"
    worksheet["A1"] = "YARDIMCI TESÄ°SLER KONTROL FORMU"
    workbook.save(path)
    workbook.close()


def apply_settings_env(monkeypatch: Any, settings: Settings) -> None:
    monkeypatch.setenv("EXCEL_PATH", settings.excel_path)
    monkeypatch.setenv("APP_PIN", settings.app_pin)
    monkeypatch.setenv("SESSION_SECRET", settings.session_secret)
    monkeypatch.setenv("SQLITE_PATH", settings.sqlite_path)
    monkeypatch.setenv("BACKUP_DIR", settings.backup_dir)
    monkeypatch.setenv("CYCLE_TABLE_PATH", settings.cycle_table_path)
    monkeypatch.setenv("REPORT_OUTPUT_DIR", settings.report_output_dir)
    monkeypatch.setenv(
        "AUXILIARY_SYSTEMS_FORM_PATH",
        settings.auxiliary_systems_form_path,
    )
    monkeypatch.setenv(
        "AUXILIARY_SYSTEMS_TARGET_PATH",
        settings.auxiliary_systems_target_path,
    )
    monkeypatch.setenv(
        "AUXILIARY_SYSTEMS_SHEET_NAME",
        settings.auxiliary_systems_sheet_name,
    )


@contextmanager
def authenticated_test_client(
    monkeypatch: Any,
    tmp_path: Path,
    **settings_overrides: Any,
) -> Iterator[TestClient]:
    settings = make_settings(tmp_path, **settings_overrides)
    create_process_workbook(Path(settings.excel_path))
    create_cycle_seed_workbook(Path(settings.cycle_table_path))
    create_auxiliary_workbook(Path(settings.auxiliary_systems_target_path))
    create_auxiliary_form(Path(settings.auxiliary_systems_form_path))
    apply_settings_env(monkeypatch, settings)

    with TestClient(create_app()) as client:
        login_response = client.post("/api/login", json={"pin": settings.app_pin})
        assert login_response.status_code == 200
        yield client


class LegacySqliteBuilder:
    def __init__(self, path: Path) -> None:
        self.path = path

    @contextmanager
    def connect(self) -> Iterator[sqlite3.Connection]:
        connection = sqlite3.connect(self.path)
        try:
            yield connection
            connection.commit()
        finally:
            connection.close()

    def executescript(self, script: str) -> None:
        with self.connect() as connection:
            connection.executescript(script)


class IFSRequestRecorder:
    def __init__(
        self,
        handler: Callable[[httpx.Request], httpx.Response],
    ) -> None:
        self.handler = handler
        self.requests: list[httpx.Request] = []

    def __call__(self, request: httpx.Request) -> httpx.Response:
        self.requests.append(request)
        return self.handler(request)

    def transport(self) -> httpx.MockTransport:
        return httpx.MockTransport(self)
