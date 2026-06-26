from __future__ import annotations

import csv
import math
import os
import statistics
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


DEFAULT_SOURCE = (
    "\\\\fileserver\\PRODUCTION\\\u00dcRET\u0130M RAPOR FK\\"
    "PROSES-YARDIMCI S\u0130STEMLER\\PROSES 2026_FK.xlsx"
)
REQUESTED_START = 1498
REQUESTED_END = 1867
COLUMNS = ("F", "G", "J", "K", "L")
OBSERVATION_COLUMNS = ("J", "K", "L")


@dataclass(frozen=True)
class RowRecord:
    scope: str
    sheet: str
    row_number: int
    machine_code: str
    product_text: str
    agiz_mm: float | None
    gramaj_gr: float | None
    values: dict[str, object]
    numeric_values: dict[str, float | None]
    parse_status: str


def normalize_machine(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_number(value: object) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        number = float(value)
    else:
        text = str(value).strip().replace(",", ".")
        if not text:
            return None
        try:
            number = float(text)
        except ValueError:
            return None
    if not math.isfinite(number) or number <= 0:
        return None
    return number


def first_regex_number(text: str, suffix: str) -> float | None:
    import re

    pattern = re.compile(r"(\d+(?:[,.]\d+)?)\s*" + suffix + r"\b", re.IGNORECASE)
    match = pattern.search(text)
    if not match:
        return None
    return parse_number(match.group(1))


def parse_product(product_text: object) -> tuple[float | None, float | None]:
    if product_text is None:
        return None, None
    text = str(product_text)
    return first_regex_number(text, "MM"), first_regex_number(text, "GR")


def parse_status(machine: str, product_text: str, agiz: float | None, gramaj: float | None) -> str:
    missing = []
    if not machine:
        missing.append("machine")
    if not product_text:
        missing.append("product")
    if agiz is None:
        missing.append("agiz_mm")
    if gramaj is None:
        missing.append("gramaj_gr")
    return "ok" if not missing else "missing_" + "_".join(missing)


def iter_scope_rows(ws, scope: str, start_row: int, end_row: int) -> Iterable[RowRecord]:
    if start_row > ws.max_row:
        return
    effective_end = min(end_row, ws.max_row)
    for row_number, row in enumerate(
        ws.iter_rows(
            min_row=start_row,
            max_row=effective_end,
            min_col=6,
            max_col=12,
            values_only=True,
        ),
        start=start_row,
    ):
        # F:L -> F, G, H, I, J, K, L
        f_value = row[0]
        g_value = row[1]
        values = {"F": f_value, "G": g_value, "J": row[4], "K": row[5], "L": row[6]}
        if not any(value not in (None, "") for value in values.values()):
            continue
        machine = normalize_machine(f_value)
        product = "" if g_value is None else str(g_value).strip()
        agiz, gramaj = parse_product(product)
        numeric_values = {col: parse_number(values[col]) for col in OBSERVATION_COLUMNS}
        yield RowRecord(
            scope=scope,
            sheet=ws.title,
            row_number=row_number,
            machine_code=machine,
            product_text=product,
            agiz_mm=agiz,
            gramaj_gr=gramaj,
            values=values,
            numeric_values=numeric_values,
            parse_status=parse_status(machine, product, agiz, gramaj),
        )


def quantile(sorted_values: list[float], p: float) -> float:
    if not sorted_values:
        raise ValueError("quantile requires at least one value")
    if len(sorted_values) == 1:
        return sorted_values[0]
    pos = (len(sorted_values) - 1) * p
    lower = math.floor(pos)
    upper = math.ceil(pos)
    if lower == upper:
        return sorted_values[int(pos)]
    weight = pos - lower
    return sorted_values[lower] * (1 - weight) + sorted_values[upper] * weight


def stdev_or_blank(values: list[float]) -> float | None:
    if len(values) < 2:
        return None
    return statistics.stdev(values)


def clean_values(values: list[float]) -> tuple[list[float], str, float | None, float | None]:
    if len(values) < 4:
        return values[:], "median; no outlier filter because raw_n < 4", None, None

    sorted_values = sorted(values)
    q1 = quantile(sorted_values, 0.25)
    q3 = quantile(sorted_values, 0.75)
    iqr = q3 - q1
    if iqr > 0:
        low = q1 - 1.5 * iqr
        high = q3 + 1.5 * iqr
        cleaned = [value for value in values if low <= value <= high]
        if cleaned:
            return cleaned, "IQR 1.5x filter; recommended_cycle = clean median", low, high

    median = statistics.median(values)
    deviations = [abs(value - median) for value in values]
    mad = statistics.median(deviations)
    if mad > 0:
        low = median - 3 * 1.4826 * mad
        high = median + 3 * 1.4826 * mad
        cleaned = [value for value in values if low <= value <= high]
        if cleaned:
            return cleaned, "MAD 3-sigma filter; recommended_cycle = clean median", low, high

    return values[:], "median; no outlier filter because spread estimate is zero", None, None


def format_number(value: float | None) -> str:
    if value is None:
        return ""
    rounded = round(float(value), 6)
    return f"{rounded:.6f}".rstrip("0").rstrip(".")


def group_stats(records: list[RowRecord], observation_columns: tuple[str, ...]) -> list[dict[str, object]]:
    grouped: dict[tuple[str, float, float], list[tuple[str, float]]] = defaultdict(list)
    for record in records:
        if not (record.machine_code and record.agiz_mm is not None and record.gramaj_gr is not None):
            continue
        for col in observation_columns:
            value = record.numeric_values.get(col)
            if value is not None:
                grouped[(record.machine_code, record.agiz_mm, record.gramaj_gr)].append((col, value))

    rows = []
    for (machine, agiz, gramaj), observations in sorted(grouped.items(), key=lambda item: (item[0][0], item[0][1], item[0][2])):
        values = [value for _, value in observations]
        cleaned, method, low, high = clean_values(values)
        by_col = Counter(col for col, _ in observations)
        raw_stdev = stdev_or_blank(values)
        clean_stdev = stdev_or_blank(cleaned)
        rows.append(
            {
                "machine_code": machine,
                "agiz_mm": agiz,
                "gramaj_gr": gramaj,
                "observation_columns": ",".join(observation_columns),
                "raw_n": len(values),
                "clean_n": len(cleaned),
                "outlier_n": len(values) - len(cleaned),
                "n_by_column": "; ".join(f"{col}:{by_col[col]}" for col in observation_columns),
                "raw_min": min(values),
                "raw_max": max(values),
                "raw_median": statistics.median(values),
                "raw_mean": statistics.mean(values),
                "raw_stdev": raw_stdev,
                "clean_min": min(cleaned),
                "clean_max": max(cleaned),
                "clean_median": statistics.median(cleaned),
                "clean_mean": statistics.mean(cleaned),
                "clean_stdev": clean_stdev,
                "recommended_cycle": statistics.median(cleaned),
                "outlier_low_bound": low,
                "outlier_high_bound": high,
                "method": method,
            }
        )
    return rows


def write_csv(path: Path, rows: list[dict[str, object]], fieldnames: list[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            formatted = {
                key: format_number(value) if isinstance(value, float) else value
                for key, value in row.items()
            }
            writer.writerow(formatted)


def row_records_to_dicts(records: list[RowRecord]) -> list[dict[str, object]]:
    rows = []
    for record in records:
        rows.append(
            {
                "scope": record.scope,
                "sheet": record.sheet,
                "row_number": record.row_number,
                "machine_code": record.machine_code,
                "product_text": record.product_text,
                "agiz_mm": record.agiz_mm,
                "gramaj_gr": record.gramaj_gr,
                "raw_F": record.values["F"],
                "raw_G": record.values["G"],
                "raw_J": record.values["J"],
                "raw_K": record.values["K"],
                "raw_L": record.values["L"],
                "numeric_J": record.numeric_values["J"],
                "numeric_K": record.numeric_values["K"],
                "numeric_L": record.numeric_values["L"],
                "parse_status": record.parse_status,
            }
        )
    return rows


def observations_to_dicts(records: list[RowRecord], observation_columns: tuple[str, ...]) -> list[dict[str, object]]:
    rows = []
    for record in records:
        groupable = bool(record.machine_code and record.agiz_mm is not None and record.gramaj_gr is not None)
        for col in observation_columns:
            value = record.numeric_values.get(col)
            if value is None:
                continue
            rows.append(
                {
                    "scope": record.scope,
                    "sheet": record.sheet,
                    "row_number": record.row_number,
                    "machine_code": record.machine_code,
                    "agiz_mm": record.agiz_mm,
                    "gramaj_gr": record.gramaj_gr,
                    "source_column": col,
                    "cycle_value": value,
                    "included_in_group_stats": "yes" if groupable else "no",
                    "parse_status": record.parse_status,
                    "product_text": record.product_text,
                }
            )
    return rows


def write_summary(
    path: Path,
    source: str,
    dimensions: list[dict[str, object]],
    requested_records: list[RowRecord],
    candidate_records: list[RowRecord],
    candidate_sheet,
    candidate_start: int,
    candidate_end: int,
    stats_all: list[dict[str, object]],
    stats_l_only: list[dict[str, object]],
    schema_sheet_name: str,
    schema_records: list[RowRecord],
    schema_stats_all: list[dict[str, object]],
    schema_stats_l_only: list[dict[str, object]],
    headers: dict[str, object],
) -> None:
    requested_count = REQUESTED_END - REQUESTED_START + 1
    parsed_rows = [r for r in candidate_records if r.parse_status == "ok"]
    numeric_all = sum(1 for r in candidate_records for v in r.numeric_values.values() if v is not None)
    numeric_l = sum(1 for r in candidate_records if r.numeric_values.get("L") is not None)
    parse_counter = Counter(r.parse_status for r in candidate_records)
    schema_parsed_rows = [r for r in schema_records if r.parse_status == "ok"]
    schema_numeric_all = sum(1 for r in schema_records for v in r.numeric_values.values() if v is not None)
    schema_numeric_l = sum(1 for r in schema_records if r.numeric_values.get("L") is not None)

    lines = [
        "# Cycle-Time Analysis Draft",
        "",
        f"Source workbook: `{source}`",
        f"Requested rows: `{REQUESTED_START}:{REQUESTED_END}` ({requested_count} rows), columns `{', '.join(COLUMNS)}`.",
        "",
        "## Workbook Availability",
        "",
        "| Sheet | Max row | Max column | Requested row start present? | Requested row end present? |",
        "|---|---:|---:|---|---|",
    ]
    for item in dimensions:
        lines.append(
            f"| {item['sheet']} | {item['max_row']} | {item['max_column']} | "
            f"{'yes' if item['max_row'] >= REQUESTED_START else 'no'} | "
            f"{'yes' if item['max_row'] >= REQUESTED_END else 'no'} |"
        )

    lines.extend(
        [
            "",
            "## Scope Result",
            "",
            f"- Requested range nonblank rows found: {len(requested_records)}.",
            f"- No worksheet in this workbook reaches row {REQUESTED_END}; the largest sheet is `{candidate_sheet.title}` at row {candidate_sheet.max_row}.",
            f"- Candidate fallback analyzed here: `{candidate_sheet.title}!{candidate_start}:{candidate_end}`. This is the latest {requested_count} rows on the largest sheet and is labeled as fallback, not as a silent replacement.",
            f"- Schema-match fallback analyzed here: `{schema_sheet_name}`. This is the only sheet where column F/G/J/K/L match the requested machine/product/cycle schema and column G contains both `MM` and `GR`.",
            "",
            "## Candidate Data Profile",
            "",
            f"- Candidate rows with any requested-column data: {len(candidate_records)}.",
            f"- Rows with machine code plus parsed agiz_mm and gramaj_gr: {len(parsed_rows)}.",
            f"- Numeric observations across J/K/L: {numeric_all}.",
            f"- Numeric observations in L only: {numeric_l}.",
            f"- Groups produced from J/K/L: {len(stats_all)}.",
            f"- Groups produced from L only: {len(stats_l_only)}.",
            "",
            "Schema-match fallback profile:",
            "",
            f"- Rows with any requested-column data: {len(schema_records)}.",
            f"- Rows with machine code plus parsed agiz_mm and gramaj_gr: {len(schema_parsed_rows)}.",
            f"- Numeric observations across J/K/L: {schema_numeric_all}.",
            f"- Numeric observations in L only: {schema_numeric_l}.",
            f"- Groups produced from J/K/L: {len(schema_stats_all)}.",
            f"- Groups produced from L only: {len(schema_stats_l_only)}.",
            "",
            "Parse status counts:",
            "",
            "| Parse status | Rows |",
            "|---|---:|",
        ]
    )
    for status, count in parse_counter.most_common():
        lines.append(f"| {status} | {count} |")

    lines.extend(
        [
            "",
            "## Method",
            "",
            "- Product text in column G is parsed with regex patterns for `<number>MM` and `<number>GR`; comma decimals such as `5,5GR` are normalized to `5.5`.",
            "- Rows are grouped by machine code from F plus parsed agiz_mm and gramaj_gr.",
            "- Numeric positive values are converted from the selected observation columns.",
            "- For each group, outliers are removed with a Tukey IQR 1.5x filter when sample size is at least 4 and IQR is nonzero. If IQR cannot estimate spread, a MAD-based 3-sigma filter is used. For small samples, no outlier filter is applied.",
            "- The recommended cycle is the median of the cleaned observations, with raw and clean sample sizes and supporting min/max/median/mean/stdev retained.",
            "",
            "## Data-Quality Notes",
            "",
            "- The exact requested row range is absent from the source workbook. Treat this draft as a range-audit plus fallback candidate, not a final production output.",
            f"- In the fallback sheet, row 1 headers for J/K/L are: J=`{headers.get('J')}`, K=`{headers.get('K')}`, L=`{headers.get('L')}`.",
            "- Those headers indicate L is the true cycle-time field in the populated fallback sheets, while J and K are cavity/eye counts. The all-J/K/L stats are included to mirror the requested columns, but the L-only stats are the more usable cycle recommendation for these fallback scopes.",
            "- Rows missing either `MM` or `GR` in product text cannot be grouped by agiz and gramaj and are excluded from group stats.",
            "- Groups with small raw sample sizes are reported, but their recommendations have limited robustness because no outlier filter can be defended with n < 4.",
        ]
    )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    output_dir = Path(__file__).resolve().parent
    source = os.environ.get("SOURCE_XLSX", DEFAULT_SOURCE)

    wb = load_workbook(source, read_only=True, data_only=True)
    try:
        dimensions = [
            {"sheet": ws.title, "max_row": ws.max_row, "max_column": ws.max_column}
            for ws in wb.worksheets
            if ws.title != "Grafik1"
        ]

        requested_records: list[RowRecord] = []
        for ws in wb.worksheets:
            if ws.title == "Grafik1":
                continue
            requested_records.extend(iter_scope_rows(ws, "requested_rows_1498_1867", REQUESTED_START, REQUESTED_END))

        requested_count = REQUESTED_END - REQUESTED_START + 1
        candidate_sheet = max(
            (ws for ws in wb.worksheets if ws.title != "Grafik1"),
            key=lambda sheet: sheet.max_row,
        )
        candidate_end = candidate_sheet.max_row
        candidate_start = max(1, candidate_end - requested_count + 1)
        candidate_records = list(
            iter_scope_rows(
                candidate_sheet,
                f"fallback_latest_{requested_count}_rows_largest_sheet",
                candidate_start,
                candidate_end,
            )
        )

        headers = {
            "F": candidate_sheet.cell(1, 6).value,
            "G": candidate_sheet.cell(1, 7).value,
            "J": candidate_sheet.cell(1, 10).value,
            "K": candidate_sheet.cell(1, 11).value,
            "L": candidate_sheet.cell(1, 12).value,
        }

        stats_all = group_stats(candidate_records, OBSERVATION_COLUMNS)
        stats_l_only = group_stats(candidate_records, ("L",))

        schema_scopes: list[tuple[int, str, list[RowRecord]]] = []
        for ws in wb.worksheets:
            if ws.title == "Grafik1":
                continue
            records = list(iter_scope_rows(ws, "schema_match_best_sheet", 1, ws.max_row))
            match_count = sum(1 for record in records if record.parse_status == "ok")
            schema_scopes.append((match_count, ws.title, records))
        schema_match_count, schema_sheet_name, schema_records = max(schema_scopes, key=lambda item: item[0])
        if schema_match_count == 0:
            schema_records = []
            schema_sheet_name = ""
        schema_stats_all = group_stats(schema_records, OBSERVATION_COLUMNS)
        schema_stats_l_only = group_stats(schema_records, ("L",))

        write_csv(
            output_dir / "workbook_row_availability.csv",
            dimensions,
            ["sheet", "max_row", "max_column"],
        )
        write_csv(
            output_dir / "requested_rows_1498_1867_extract.csv",
            row_records_to_dicts(requested_records),
            [
                "scope",
                "sheet",
                "row_number",
                "machine_code",
                "product_text",
                "agiz_mm",
                "gramaj_gr",
                "raw_F",
                "raw_G",
                "raw_J",
                "raw_K",
                "raw_L",
                "numeric_J",
                "numeric_K",
                "numeric_L",
                "parse_status",
            ],
        )
        write_csv(
            output_dir / "candidate_latest_370_rows_extract.csv",
            row_records_to_dicts(candidate_records),
            [
                "scope",
                "sheet",
                "row_number",
                "machine_code",
                "product_text",
                "agiz_mm",
                "gramaj_gr",
                "raw_F",
                "raw_G",
                "raw_J",
                "raw_K",
                "raw_L",
                "numeric_J",
                "numeric_K",
                "numeric_L",
                "parse_status",
            ],
        )
        write_csv(
            output_dir / "candidate_latest_370_observations_jkl.csv",
            observations_to_dicts(candidate_records, OBSERVATION_COLUMNS),
            [
                "scope",
                "sheet",
                "row_number",
                "machine_code",
                "agiz_mm",
                "gramaj_gr",
                "source_column",
                "cycle_value",
                "included_in_group_stats",
                "parse_status",
                "product_text",
            ],
        )
        write_csv(
            output_dir / "candidate_cycle_group_stats_jkl.csv",
            stats_all,
            [
                "machine_code",
                "agiz_mm",
                "gramaj_gr",
                "observation_columns",
                "raw_n",
                "clean_n",
                "outlier_n",
                "n_by_column",
                "raw_min",
                "raw_max",
                "raw_median",
                "raw_mean",
                "raw_stdev",
                "clean_min",
                "clean_max",
                "clean_median",
                "clean_mean",
                "clean_stdev",
                "recommended_cycle",
                "outlier_low_bound",
                "outlier_high_bound",
                "method",
            ],
        )
        write_csv(
            output_dir / "candidate_cycle_group_stats_l_only.csv",
            stats_l_only,
            [
                "machine_code",
                "agiz_mm",
                "gramaj_gr",
                "observation_columns",
                "raw_n",
                "clean_n",
                "outlier_n",
                "n_by_column",
                "raw_min",
                "raw_max",
                "raw_median",
                "raw_mean",
                "raw_stdev",
                "clean_min",
                "clean_max",
                "clean_median",
                "clean_mean",
                "clean_stdev",
                "recommended_cycle",
                "outlier_low_bound",
                "outlier_high_bound",
                "method",
            ],
        )
        write_csv(
            output_dir / "schema_match_best_sheet_extract.csv",
            row_records_to_dicts(schema_records),
            [
                "scope",
                "sheet",
                "row_number",
                "machine_code",
                "product_text",
                "agiz_mm",
                "gramaj_gr",
                "raw_F",
                "raw_G",
                "raw_J",
                "raw_K",
                "raw_L",
                "numeric_J",
                "numeric_K",
                "numeric_L",
                "parse_status",
            ],
        )
        write_csv(
            output_dir / "schema_match_best_sheet_observations_jkl.csv",
            observations_to_dicts(schema_records, OBSERVATION_COLUMNS),
            [
                "scope",
                "sheet",
                "row_number",
                "machine_code",
                "agiz_mm",
                "gramaj_gr",
                "source_column",
                "cycle_value",
                "included_in_group_stats",
                "parse_status",
                "product_text",
            ],
        )
        write_csv(
            output_dir / "schema_match_cycle_group_stats_jkl.csv",
            schema_stats_all,
            [
                "machine_code",
                "agiz_mm",
                "gramaj_gr",
                "observation_columns",
                "raw_n",
                "clean_n",
                "outlier_n",
                "n_by_column",
                "raw_min",
                "raw_max",
                "raw_median",
                "raw_mean",
                "raw_stdev",
                "clean_min",
                "clean_max",
                "clean_median",
                "clean_mean",
                "clean_stdev",
                "recommended_cycle",
                "outlier_low_bound",
                "outlier_high_bound",
                "method",
            ],
        )
        write_csv(
            output_dir / "schema_match_cycle_group_stats_l_only.csv",
            schema_stats_l_only,
            [
                "machine_code",
                "agiz_mm",
                "gramaj_gr",
                "observation_columns",
                "raw_n",
                "clean_n",
                "outlier_n",
                "n_by_column",
                "raw_min",
                "raw_max",
                "raw_median",
                "raw_mean",
                "raw_stdev",
                "clean_min",
                "clean_max",
                "clean_median",
                "clean_mean",
                "clean_stdev",
                "recommended_cycle",
                "outlier_low_bound",
                "outlier_high_bound",
                "method",
            ],
        )
        write_summary(
            output_dir / "cycle_report_summary.md",
            source,
            dimensions,
            requested_records,
            candidate_records,
            candidate_sheet,
            candidate_start,
            candidate_end,
            stats_all,
            stats_l_only,
            schema_sheet_name,
            schema_records,
            schema_stats_all,
            schema_stats_l_only,
            headers,
        )

        print(f"created={output_dir}")
        print(f"requested_nonblank_rows={len(requested_records)}")
        print(f"candidate_sheet={candidate_sheet.title}")
        print(f"candidate_rows={candidate_start}:{candidate_end}")
        print(f"candidate_records={len(candidate_records)}")
        print(f"groups_jkl={len(stats_all)}")
        print(f"groups_l_only={len(stats_l_only)}")
        print(f"schema_sheet={schema_sheet_name}")
        print(f"schema_records={len(schema_records)}")
        print(f"schema_groups_jkl={len(schema_stats_all)}")
        print(f"schema_groups_l_only={len(schema_stats_l_only)}")
    finally:
        wb.close()


if __name__ == "__main__":
    main()
