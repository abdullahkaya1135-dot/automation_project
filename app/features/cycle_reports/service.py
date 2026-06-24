import asyncio
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import date as Date
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import Integer, cast, select

from ...core.config import Settings
from ...core.database import create_session
from ...integrations.ifs.client import fetch_pet_ongoing_operations
from ...models import Machine
from ...services.shop_order_source import (
    ShopOrderOption,
    shop_order_options_from_ifs_operations,
)
from ...services.text_normalization import stripped_text as _clean_text
from ...services.workbook_utils import merged_cell_values, merged_value
from .models import Entry, MachineCycleTableRow

REPORT_HEADERS = (
    "Makine Adı",
    "Makine No.",
    "Ağız",
    "Gramaj",
    "Toplam Göz",
    "Çalışan Göz",
    "Gerçekleşen Çevrim Süresi",
    "Optimum Çevrim Süresi",
    "Kontrol Durumu",
    "Sonuç",
)
SOURCE_COLUMN_COUNT = 12
REPORT_HALL_NUMBERS = (1, 2)
PART_NECK_PATTERN = re.compile(r"(\d+(?:[,.]\d+)?)\s*MM\b", re.IGNORECASE)
PART_GRAM_PATTERN = re.compile(r"(\d+(?:[,.]\d+)?)\s*GR\b", re.IGNORECASE)
STATUS_OK = "Uygun"
STATUS_CHECK = "Kontrol Edilecek"
COMPLETE_STATUSES = frozenset({STATUS_OK, STATUS_CHECK})
CYCLE_TOLERANCE_MULTIPLIER = Decimal("1.10")
NEAR_MATCH_RANGE = Decimal("3")


class CycleReportError(RuntimeError):
    """Raised when a cycle-control report cannot be created."""


class NoTodayRowsError(CycleReportError):
    """Raised when there are no source rows for the requested report date."""


@dataclass(frozen=True)
class ProcessCycleRow:
    row_number: int
    machine_no: str
    work_order: str
    total_cavity_count: Any
    active_cavity_count: Any
    actual_cycle_time: Decimal | None


@dataclass(frozen=True)
class CycleTableEntry:
    machine_no: str
    machine_group_key: str
    neck_diameter: Decimal
    gram: Decimal
    estimated_cycle_time: Decimal


@dataclass(frozen=True)
class CycleReportRow:
    machine_group: str | None
    machine_no: str
    neck_diameter: Decimal | None
    gram: Decimal | None
    total_cavity_count: Any
    active_cavity_count: Any
    actual_cycle_time: Decimal | None
    optimum_cycle_time: Decimal | None
    control_status: str
    result: str | None


@dataclass(frozen=True)
class CycleReportResult:
    output_path: Path
    row_count: int
    matched_count: int
    warning_count: int
    report_date: Date

    def as_dict(self) -> dict[str, Any]:
        return {
            "output_path": str(self.output_path),
            "row_count": self.row_count,
            "matched_count": self.matched_count,
            "warning_count": self.warning_count,
            "date": self.report_date.isoformat(),
        }


def _identifier(value: Any) -> str:
    text = _clean_text(value)
    if not text:
        return ""
    if re.fullmatch(r"\d+\.0+", text):
        return text.split(".", 1)[0]
    return text


def _number(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int):
        return Decimal(value)
    if isinstance(value, float):
        return Decimal(str(value))

    text = _clean_text(value).replace(",", ".")
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _excel_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        if value == value.to_integral_value():
            return int(value)
        return float(value)
    return value


def _parse_date(value: Any) -> Date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, Date):
        return value

    text = _clean_text(value)
    if not text:
        return None
    for date_format in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, date_format).date()
        except ValueError:
            continue
    return None


def parse_part_description(part_description: str) -> tuple[Decimal | None, Decimal | None]:
    neck_match = PART_NECK_PATTERN.search(part_description)
    gram_match = PART_GRAM_PATTERN.search(part_description)
    neck = _number(neck_match.group(1)) if neck_match else None
    gram = _number(gram_match.group(1)) if gram_match else None
    return neck, gram


def normalize_machine_group(value: Any) -> str:
    text = _clean_text(value).upper()
    compact = re.sub(r"[\s\-/]+", "", text)
    aliases = {
        "70DPH": "DPH70",
        "DPH70": "DPH70",
        "50MB": "50MB",
        "50MB1": "50MB",
        "50MB2": "50MB",
        "50MB3": "50MB",
        "PF6": "PF62",
        "PF62": "PF62",
        "PF8": "PF84",
        "PF84": "PF84",
        "12M1": "12M",
        "12M": "12M",
        "TP250": "PREFORM",
        "TP380": "PREFORM",
        "PREFORM": "PREFORM",
    }
    return aliases.get(compact, compact)


def _read_process_rows(settings: Settings, report_date: Date) -> list[ProcessCycleRow]:
    with create_session(settings) as session:
        entries = session.scalars(
            select(Entry)
            .join(Machine, Entry.machine_code == Machine.machine_code)
            .where(Entry.process_date == report_date.isoformat())
            .where(Machine.hall_number.in_(REPORT_HALL_NUMBERS))
            .order_by(
                cast(Entry.machine_code, Integer).asc(),
                Entry.machine_code.asc(),
                Entry.production_engineer_id.asc(),
                Entry.id.asc(),
            )
        ).all()

    return [
        ProcessCycleRow(
            row_number=entry.id,
            machine_no=_identifier(entry.col_f),
            work_order=_identifier(entry.col_h),
            total_cavity_count=_number(entry.col_j),
            active_cavity_count=_number(entry.col_k),
            actual_cycle_time=_number(entry.col_l),
        )
        for entry in entries
    ]


def _indexed_shop_orders(
    options: list[ShopOrderOption],
) -> dict[str, list[ShopOrderOption]]:
    by_order: dict[str, list[ShopOrderOption]] = defaultdict(list)
    for option in options:
        order_no = _identifier(option.order_no)
        if order_no:
            by_order[order_no].append(option)
    return by_order


def _matching_shop_order(
    row: ProcessCycleRow,
    by_order: dict[str, list[ShopOrderOption]],
) -> tuple[ShopOrderOption | None, str | None]:
    candidates = by_order.get(row.work_order, [])
    if not candidates:
        return None, "HTML eşleşmesi bulunamadı"
    if len(candidates) == 1:
        return candidates[0], None

    machine_matches = [
        candidate
        for candidate in candidates
        if _identifier(candidate.resource_id) == row.machine_no
    ]
    if len(machine_matches) == 1:
        return machine_matches[0], None
    if machine_matches:
        return None, "HTML eşleşmesi belirsiz"
    return None, "HTML makine eşleşmesi bulunamadı"


def _read_cycle_table(settings: Settings) -> dict[tuple[str, Decimal, Decimal], list[CycleTableEntry]]:
    database_entries = _read_cycle_table_from_database(settings)
    if database_entries:
        return database_entries
    return _read_cycle_table_from_workbook(settings)


def _read_cycle_table_from_database(
    settings: Settings,
) -> dict[tuple[str, Decimal, Decimal], list[CycleTableEntry]]:
    with create_session(settings) as session:
        seed_rows = session.scalars(
            select(MachineCycleTableRow).order_by(
                MachineCycleTableRow.source_row_number.asc()
            )
        ).all()

    current_machine_no = ""
    current_group = ""
    current_neck: Decimal | None = None
    entries: dict[tuple[str, Decimal, Decimal], list[CycleTableEntry]] = defaultdict(list)

    for seed_row in seed_rows:
        machine_no = _identifier(seed_row.col_a)
        previous_machine_no = current_machine_no
        if machine_no:
            current_machine_no = machine_no

        group_key = normalize_machine_group(seed_row.col_b)
        previous_group = current_group
        if group_key:
            current_group = group_key

        neck = _number(seed_row.col_c)
        if (
            neck is None
            and (
                (machine_no and machine_no != previous_machine_no)
                or (group_key and group_key != previous_group)
            )
        ):
            current_neck = None
        if neck is not None:
            current_neck = neck
        else:
            neck = current_neck

        gram = _number(seed_row.col_d)
        cycle = _number(seed_row.col_g)
        if not current_group or neck is None or gram is None or cycle is None:
            continue

        key = (current_group, neck, gram)
        entries[key].append(
            CycleTableEntry(
                machine_no=current_machine_no,
                machine_group_key=current_group,
                neck_diameter=neck,
                gram=gram,
                estimated_cycle_time=cycle,
            )
        )
    return dict(entries)


def _read_cycle_table_from_workbook(settings: Settings) -> dict[tuple[str, Decimal, Decimal], list[CycleTableEntry]]:
    path = Path(settings.cycle_table_path).expanduser()
    try:
        workbook = load_workbook(path, read_only=False, data_only=True)
    except FileNotFoundError as exc:
        raise CycleReportError(f"Makine çevrim tablosu bulunamadı: {path}") from exc
    except OSError as exc:
        raise CycleReportError(f"Makine çevrim tablosu okunamadı: {path} ({exc})") from exc

    try:
        worksheet = workbook.worksheets[0]
        merged_values = merged_cell_values(worksheet)
        current_machine_no = ""
        current_group = ""
        current_neck: Decimal | None = None
        entries: dict[tuple[str, Decimal, Decimal], list[CycleTableEntry]] = defaultdict(list)

        for row_index in range(2, worksheet.max_row + 1):
            machine_no_value = merged_value(worksheet, merged_values, row_index, 1)
            group_value = merged_value(worksheet, merged_values, row_index, 2)
            neck_value = merged_value(worksheet, merged_values, row_index, 3)
            gram_value = merged_value(worksheet, merged_values, row_index, 4)
            cycle_value = merged_value(worksheet, merged_values, row_index, 7)

            machine_no = _identifier(machine_no_value)
            previous_machine_no = current_machine_no
            if machine_no:
                current_machine_no = machine_no

            group_key = normalize_machine_group(group_value)
            previous_group = current_group
            if group_key:
                current_group = group_key

            neck = _number(neck_value)
            if (
                neck is None
                and (
                    (machine_no and machine_no != previous_machine_no)
                    or (group_key and group_key != previous_group)
                )
            ):
                current_neck = None
            if neck is not None:
                current_neck = neck
            else:
                neck = current_neck

            gram = _number(gram_value)
            cycle = _number(cycle_value)
            if not current_group or neck is None or gram is None or cycle is None:
                continue

            key = (current_group, neck, gram)
            entries[key].append(
                CycleTableEntry(
                    machine_no=current_machine_no,
                    machine_group_key=current_group,
                    neck_diameter=neck,
                    gram=gram,
                    estimated_cycle_time=cycle,
                )
            )
        return dict(entries)
    finally:
        workbook.close()


def _lookup_optimum_cycle(
    cycle_entries: dict[tuple[str, Decimal, Decimal], list[CycleTableEntry]],
    *,
    machine_group: str,
    machine_no: str,
    neck_diameter: Decimal,
    gram: Decimal,
) -> tuple[Decimal | None, str | None]:
    group_key = normalize_machine_group(machine_group)
    key = (group_key, neck_diameter, gram)
    entries = cycle_entries.get(key, [])
    if entries:
        return _resolve_cycle_entries(entries, machine_no)

    entries = _nearest_cycle_entries(
        cycle_entries,
        group_key=group_key,
        fixed_gram=gram,
        target_neck=neck_diameter,
    )
    if entries:
        return _resolve_cycle_entries(entries, machine_no)

    entries = _nearest_cycle_entries(
        cycle_entries,
        group_key=group_key,
        fixed_neck=neck_diameter,
        target_gram=gram,
    )
    if entries:
        return _resolve_cycle_entries(entries, machine_no)

    return None, "Optimum çevrim bulunamadı"


def _resolve_cycle_entries(
    entries: list[CycleTableEntry],
    machine_no: str,
) -> tuple[Decimal | None, str | None]:
    cycles = {entry.estimated_cycle_time for entry in entries}
    if len(cycles) == 1:
        return next(iter(cycles)), None

    machine_matches = [
        entry
        for entry in entries
        if _identifier(entry.machine_no) == _identifier(machine_no)
    ]
    machine_cycles = {entry.estimated_cycle_time for entry in machine_matches}
    if len(machine_cycles) == 1:
        return next(iter(machine_cycles)), None
    return None, "Optimum çevrim belirsiz"


def _nearest_cycle_entries(
    cycle_entries: dict[tuple[str, Decimal, Decimal], list[CycleTableEntry]],
    *,
    group_key: str,
    fixed_neck: Decimal | None = None,
    fixed_gram: Decimal | None = None,
    target_neck: Decimal | None = None,
    target_gram: Decimal | None = None,
) -> list[CycleTableEntry]:
    candidates: list[tuple[Decimal, Decimal, list[CycleTableEntry]]] = []
    for (entry_group, entry_neck, entry_gram), entries in cycle_entries.items():
        if entry_group != group_key:
            continue
        if fixed_neck is not None and entry_neck != fixed_neck:
            continue
        if fixed_gram is not None and entry_gram != fixed_gram:
            continue

        moving_value: Decimal
        target_value: Decimal
        if target_neck is not None:
            moving_value = entry_neck
            target_value = target_neck
        elif target_gram is not None:
            moving_value = entry_gram
            target_value = target_gram
        else:
            continue

        distance = abs(moving_value - target_value)
        if distance <= NEAR_MATCH_RANGE:
            candidates.append((distance, moving_value, entries))

    if not candidates:
        return []
    candidates.sort(key=lambda candidate: (candidate[0], -candidate[1]))
    return candidates[0][2]


def build_cycle_report_rows(
    process_rows: list[ProcessCycleRow],
    shop_orders: list[ShopOrderOption],
    cycle_entries: dict[tuple[str, Decimal, Decimal], list[CycleTableEntry]],
) -> list[CycleReportRow]:
    by_order = _indexed_shop_orders(shop_orders)
    report_rows: list[CycleReportRow] = []

    for process_row in process_rows:
        shop_order, status = _matching_shop_order(process_row, by_order)
        machine_group = shop_order.work_center_no if shop_order else None
        neck: Decimal | None = None
        gram: Decimal | None = None
        optimum_cycle: Decimal | None = None

        if shop_order is not None:
            neck, gram = parse_part_description(shop_order.part_description)
            if neck is None or gram is None:
                status = "Ağız/Gramaj bulunamadı"

        if status is None and shop_order is not None and neck is not None and gram is not None:
            optimum_cycle, status = _lookup_optimum_cycle(
                cycle_entries,
                machine_group=shop_order.work_center_no,
                machine_no=process_row.machine_no,
                neck_diameter=neck,
                gram=gram,
            )

        if status is None:
            if process_row.actual_cycle_time is None:
                status = "Gerçekleşen çevrim eksik"
            elif optimum_cycle is None:
                status = "Optimum çevrim bulunamadı"
            elif process_row.actual_cycle_time <= optimum_cycle * CYCLE_TOLERANCE_MULTIPLIER:
                status = STATUS_OK
            else:
                status = STATUS_CHECK

        report_rows.append(
            CycleReportRow(
                machine_group=machine_group,
                machine_no=process_row.machine_no,
                neck_diameter=neck,
                gram=gram,
                total_cavity_count=process_row.total_cavity_count,
                active_cavity_count=process_row.active_cavity_count,
                actual_cycle_time=process_row.actual_cycle_time,
                optimum_cycle_time=optimum_cycle,
                control_status=status,
                result=None if status == STATUS_CHECK else status,
            )
        )
    return report_rows


def _machine_code_sort_key(value: Any) -> tuple[int, int | str]:
    machine_code = _identifier(value)
    if machine_code.isdigit():
        return (0, int(machine_code))
    return (1, machine_code)


def _ordered_report_rows(rows: list[CycleReportRow]) -> list[CycleReportRow]:
    return sorted(
        rows,
        key=lambda row: (
            row.control_status,
            _machine_code_sort_key(row.machine_no),
        ),
    )


def _read_shop_order_options(settings: Settings) -> list[ShopOrderOption]:
    try:
        rows = asyncio.run(fetch_pet_ongoing_operations(settings))
    except Exception as exc:
        raise CycleReportError(str(exc) or exc.__class__.__name__) from exc
    return shop_order_options_from_ifs_operations(rows)


def _write_report(output_path: Path, rows: list[CycleReportRow]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Çevrim Kontrol"
    worksheet.append(REPORT_HEADERS)

    for row in rows:
        worksheet.append(
            [
                row.machine_group,
                row.machine_no,
                _excel_value(row.neck_diameter),
                _excel_value(row.gram),
                row.total_cavity_count,
                row.active_cavity_count,
                _excel_value(row.actual_cycle_time),
                _excel_value(row.optimum_cycle_time),
                row.control_status,
                row.result,
            ]
        )

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(bottom=Side(style="thin", color="D9E2F3"))
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    for column_cells in worksheet.columns:
        column_letter = column_cells[0].column_letter
        max_length = max(len(_clean_text(cell.value)) for cell in column_cells)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 34)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    workbook.save(output_path)
    workbook.close()


def create_cycle_report(settings: Settings, report_date: Date) -> CycleReportResult:
    process_rows = _read_process_rows(settings, report_date)
    if not process_rows:
        raise NoTodayRowsError(
            f"{report_date.isoformat()} tarihi için PROSES 2026 kaydı bulunamadı."
        )

    shop_orders = _read_shop_order_options(settings)

    cycle_entries = _read_cycle_table(settings)
    report_rows = _ordered_report_rows(
        build_cycle_report_rows(process_rows, shop_orders, cycle_entries)
    )

    output_dir = Path(settings.report_output_dir).expanduser()
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{report_date.strftime('%d.%m.%Y')} Cycle Report.xlsx"
    _write_report(output_path, report_rows)

    warning_count = sum(
        1 for row in report_rows if row.control_status not in COMPLETE_STATUSES
    )
    matched_count = len(report_rows) - warning_count
    return CycleReportResult(
        output_path=output_path,
        row_count=len(report_rows),
        matched_count=matched_count,
        warning_count=warning_count,
        report_date=report_date,
    )
