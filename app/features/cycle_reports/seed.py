from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import text
from sqlalchemy.engine import Connection

from ...core.config import Settings

CYCLE_TABLE_COLUMNS = (1, 2, 3, 4, 5, 7)
CYCLE_TABLE_FIELD_NAMES = ("col_a", "col_b", "col_c", "col_d", "col_e", "col_g")


@dataclass(frozen=True)
class CycleTableSeedResult:
    source_path: Path
    scanned_rows: int = 0
    inserted_rows: int = 0
    skipped: bool = False
    error: str = ""


def seed_cycle_table_from_workbook(
    connection: Connection,
    settings: Settings,
) -> CycleTableSeedResult:
    path = Path(settings.cycle_table_path).expanduser()
    try:
        rows = _read_seed_rows(path)
    except FileNotFoundError:
        return CycleTableSeedResult(
            source_path=path,
            skipped=True,
            error=f"Cycle table workbook not found: {path}",
        )
    except OSError as exc:
        return CycleTableSeedResult(
            source_path=path,
            skipped=True,
            error=f"Cycle table workbook could not be read: {path} ({exc})",
        )

    connection.execute(text("DELETE FROM machine_cycle_table_rows"))
    if rows:
        connection.execute(
            text(
                "INSERT INTO machine_cycle_table_rows ("
                "source_row_number, col_a, col_b, col_c, col_d, col_e, col_g, "
                "created_at, updated_at"
                ") VALUES ("
                ":source_row_number, :col_a, :col_b, :col_c, :col_d, :col_e, :col_g, "
                "CURRENT_TIMESTAMP, CURRENT_TIMESTAMP"
                ")"
            ),
            rows,
        )

    return CycleTableSeedResult(
        source_path=path,
        scanned_rows=len(rows),
        inserted_rows=len(rows),
    )


def _read_seed_rows(path: Path) -> list[dict[str, str | int | None]]:
    workbook = load_workbook(path, read_only=False, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        merged_values = _merged_cell_values(worksheet)
        seed_rows: list[dict[str, str | int | None]] = []

        for row_index in range(2, worksheet.max_row + 1):
            values = [
                _stored_text(
                    _merged_value(worksheet, merged_values, row_index, column_index)
                )
                for column_index in CYCLE_TABLE_COLUMNS
            ]
            if not any(values):
                continue

            row = {
                "source_row_number": row_index,
                **dict(zip(CYCLE_TABLE_FIELD_NAMES, values, strict=True)),
            }
            seed_rows.append(row)

        return seed_rows
    finally:
        workbook.close()


def _merged_cell_values(worksheet: Worksheet) -> dict[tuple[int, int], Any]:
    merged_values: dict[tuple[int, int], Any] = {}
    for merged_range in worksheet.merged_cells.ranges:
        min_column, min_row, max_column, max_row = merged_range.bounds
        value = worksheet.cell(min_row, min_column).value
        for row_index in range(min_row, max_row + 1):
            for column_index in range(min_column, max_column + 1):
                merged_values[(row_index, column_index)] = value
    return merged_values


def _merged_value(
    worksheet: Worksheet,
    merged_values: dict[tuple[int, int], Any],
    row_index: int,
    column_index: int,
) -> Any:
    value = worksheet.cell(row_index, column_index).value
    if value is not None:
        return value
    return merged_values.get((row_index, column_index))


def _stored_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    text_value = " ".join(str(value).strip().split())
    return text_value or None
