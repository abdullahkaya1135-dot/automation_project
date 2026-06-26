from __future__ import annotations

import asyncio
import json
import re
from collections import defaultdict
from collections.abc import Iterable, Mapping, Sequence
from dataclasses import dataclass, field, replace
from datetime import UTC, datetime, time, timedelta
from datetime import date as Date
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import select
from sqlalchemy.orm import selectinload

from ...core.config import Settings
from ...core.database import create_session
from ...domain import shifts
from ...features.cycle_reports.service import parse_part_description
from ...features.process_entries.normalization import (
    normalize_machine_code,
    normalize_process_date,
)
from ...integrations.ifs.client import (
    fetch_inventory_part_descriptions,
    fetch_operation_history_rows,
    fetch_production_label_stock_rows,
    fetch_shop_order_operation_actual_rows,
)
from ...services.text_normalization import (
    collapsed_whitespace_text as _clean_text,
)
from ...services.text_normalization import (
    optional_collapsed_whitespace_text as _optional_text,
)
from ..serialization import timestamp as _timestamp
from .models import (
    AmountControlShift,
    Entry,
    Machine,
    MachineBreakdown,
    ProductionLossIfsActual,
    ProductionLossLabelEvent,
    ProductionLossReport,
    ProductionLossReportRow,
)

SHIFT_FIELD_BY_LABEL = {
    "24.00-08.00": "shift_2400_0800_quantity",
    "08.00-16.00": "shift_0800_1600_quantity",
    "16.00-24.00": "shift_1600_2400_quantity",
}
SHIFT_ACTUAL_FIELD_BY_LABEL = {
    "24.00-08.00": "shift_2400_0800_actual_quantity",
    "08.00-16.00": "shift_0800_1600_actual_quantity",
    "16.00-24.00": "shift_1600_2400_actual_quantity",
}
SHIFT_MACHINE_MINUTES_FIELD_BY_LABEL = {
    "24.00-08.00": "shift_2400_0800_machine_minutes",
    "08.00-16.00": "shift_0800_1600_machine_minutes",
    "16.00-24.00": "shift_1600_2400_machine_minutes",
}
SHIFT_OPTIMUM_FIELD_BY_LABEL = {
    "24.00-08.00": "shift_2400_0800_optimum_quantity",
    "08.00-16.00": "shift_0800_1600_optimum_quantity",
    "16.00-24.00": "shift_1600_2400_optimum_quantity",
}
SHIFT_LOSS_FIELD_BY_LABEL = {
    "24.00-08.00": "shift_2400_0800_loss_quantity",
    "08.00-16.00": "shift_0800_1600_loss_quantity",
    "16.00-24.00": "shift_1600_2400_loss_quantity",
}
REPORT_HEADERS = (
    "Date",
    "Machine Name",
    "Machine Code",
    "Job Order",
    "Lot/Parti No",
    "Product",
    "Gr",
    "24.00-08.00 Actual",
    "24.00-08.00 Machine Minutes",
    "24.00-08.00 Expected",
    "24.00-08.00 Loss",
    "08.00-16.00 Actual",
    "08.00-16.00 Machine Minutes",
    "08.00-16.00 Expected",
    "08.00-16.00 Loss",
    "16.00-24.00 Actual",
    "16.00-24.00 Machine Minutes",
    "16.00-24.00 Expected",
    "16.00-24.00 Loss",
    "Daily Total",
    "Active Cavities",
    "Cycle Time (s)",
    "Net Machine Time (min)",
    "Gross Elapsed (min)",
    "Expected Net Qty",
    "Production Loss Net",
    "Expected Gross Qty",
    "Production Loss Gross",
    "Break Loss Qty",
    "Local Breakdown (min)",
    "Warnings",
)
DATETIME_FIELDS_START = (
    "RealStart",
    "RealStartDate",
    "ActualStartDate",
    "ActualStartTime",
    "ActualStart",
    "StartedAt",
    "ProductionStartTime",
)
DATETIME_FIELDS_FINISH = (
    "RealFinished",
    "RealFinish",
    "RealFinishDate",
    "RealFinishedDate",
    "ActualFinishDate",
    "ActualFinishTime",
    "ActualFinish",
    "StoppedAt",
    "ProductionFinishTime",
)
NET_MACHINE_DURATION_FIELDS = (
    "RealMachRunTime",
    "NetMachineMinutes",
    "MachineMinutes",
    "RunTimeMinutes",
    "ProductionMinutes",
    "NetMachineHours",
    "MachineHours",
    "RunTimeHours",
    "ProductionHours",
    "NetMachineSeconds",
    "MachineSeconds",
    "RunTimeSeconds",
    "ProductionSeconds",
)
COMPLETED_QTY_FIELDS = (
    "CompletedQty",
    "QtyComplete",
    "QtyCompleted",
    "ReportedQty",
    "QtyReported",
)
SCRAP_QTY_FIELDS = ("ScrapQty", "QtyScrapped", "ScrappedQty")
OPERATION_HISTORY_TIME_FIELDS = (
    "TimeOfProduction",
    "time_of_production",
    "timeOfProduction",
)
OPERATION_HISTORY_MACHINE_FIELDS = (
    "ResourceId",
    "resource_id",
    "PreferredResourceId",
    "preferred_resource_id",
)
OPERATION_HISTORY_ORDER_FIELDS = (
    "OrderNo",
    "order_no",
    "ShopOrderNo",
    "shop_order_no",
)
OPERATION_HISTORY_PART_FIELDS = ("PartNo", "part_no")
OPERATION_HISTORY_QUANTITY_FIELDS = ("QtyComplete", "qty_complete")
OPERATION_HISTORY_TRANSACTION_FIELDS = (
    "TransactionId",
    "transaction_id",
    "TransactionID",
    "TransactionNo",
    "transaction_no",
    "HistorySeq",
    "OperationHistoryId",
    "ObjId",
    "Objid",
)
GRAM_PATTERN = re.compile(r"(\d+(?:[,.]\d+)?)\s*GR\b", re.IGNORECASE)
REALIZED_CYCLE_SOURCE = "process-entry-col-l"
REALIZED_CYCLE_MISSING_MESSAGE = "Gerceklesen cevrim eksik"
REALIZED_CYCLE_INVALID_MESSAGE = "Gerceklesen cevrim gecersiz"
LABEL_STOCK_REPORT_ID = "InventoryPartInStockSet"
LABEL_STOCK_SOURCE = "ifs-inventory-label-stock"
LABEL_STOCK_PRODUCTION_LOCATION = "U01"
UNKNOWN_MACHINE_CODE = "UNKNOWN"
IFS_REALIZED_OPERATION_SOURCE = "ifs-operation-statistics-realized"
LABEL_STOCK_HANDLING_UNIT_FIELDS = ("HandlingUnitId", "Taşıma Birim No")
LABEL_STOCK_LOCATION_FIELDS = ("LocationNo", "Lokasyon No")
LABEL_STOCK_LOT_FIELDS = ("LotBatchNo", "Lot/Parti No")
LABEL_STOCK_PART_FIELDS = ("PartNo", "Malzeme")
LABEL_STOCK_PRODUCT_FIELDS = ("PartNoDesc", "Malzeme Açıklama")
LABEL_STOCK_ONHAND_FIELDS = ("QtyOnhand", "Eldeki Miktar")
LABEL_STOCK_PACK_QUANTITY_FIELDS = ("Cf_Palet_Ici_Miktar", "Palet Ici Miktar")
LABEL_STOCK_RECEIPT_FIELDS = ("ReceiptDate", "Teslim Alma Tarihi")
LABEL_STOCK_UOM_FIELDS = ("UoM", "ÖB")
LABEL_STOCK_UNIT_TYPE_FIELDS = (
    "HandlingUnitTypeDesc",
    "Taşıma Birim Türü Açıklama",
)


class ProductionLossReportError(RuntimeError):
    """Raised when a production-loss report cannot be created."""


async def fetch_shop_order_operation_history_rows(
    settings: Settings,
    *,
    date_from_utc: datetime,
    date_to_utc: datetime,
) -> list[dict[str, Any]]:
    return await fetch_operation_history_rows(
        settings,
        date_from_utc,
        date_to_utc,
    )


@dataclass
class ShiftAggregate:
    record_date: str
    machine_id: int | None
    machine_code: str
    job_order: str
    shift_quantities: dict[str, int] = field(
        default_factory=lambda: {shift: 0 for shift in shifts.SHIFT_OPTIONS}
    )
    shift_breakdown_minutes: dict[str, int] = field(
        default_factory=lambda: {shift: 0 for shift in shifts.SHIFT_OPTIONS}
    )
    local_breakdown_minutes: int = 0
    product_hint: str | None = None
    part_no_hint: str | None = None
    machine_group_hint: str | None = None
    lot_batch_no: str | None = None
    label_event_count: int = 0

    @property
    def daily_total_quantity(self) -> int:
        return sum(self.shift_quantities.values())

    def breakdown_minutes_for_shift(self, shift_label: str) -> int:
        return self.shift_breakdown_minutes.get(shift_label, 0)


@dataclass(frozen=True)
class StandaloneBreakdownContext:
    record_date: str
    machine_code: str
    fallback_shift: str
    job_order: str | None


@dataclass(frozen=True)
class ProcessMeta:
    product: str | None
    active_cavities: Decimal | None
    realized_cycle_time: Decimal | None
    raw_cycle_time: str | None
    submitted_at: datetime | None
    entry_id: int


@dataclass
class IfsActual:
    job_order: str
    machine_code: str | None = None
    release_no: str | None = None
    sequence_no: str | None = None
    operation_no: int | None = None
    work_center_no: str | None = None
    part_no: str | None = None
    part_description: str | None = None
    actual_start_at: datetime | None = None
    actual_finish_at: datetime | None = None
    net_machine_minutes: Decimal | None = None
    completed_quantity: Decimal | None = None
    scrap_quantity: Decimal | None = None
    raw_rows: list[dict[str, Any]] = field(default_factory=list)


@dataclass(frozen=True)
class LabelProductionEvent:
    result_key: str
    report_id: str
    record_date: str
    shift: str
    machine_code: str
    job_order: str
    quantity: int
    label_time: datetime | None = None
    archive_exec_time: datetime | None = None
    print_job_id: str | None = None
    part_no: str | None = None
    product_description: str | None = None
    package_id: str | None = None
    lot_batch_no: str | None = None
    pallet_no: str | None = None
    sequence_no: str | None = None
    raw_xml: str | None = None


@dataclass(frozen=True)
class OperationHistoryProductionEvent:
    record_date: str
    shift: str
    machine_code: str
    job_order: str
    part_no: str | None
    quantity: int
    transaction_id: str
    time_of_production: datetime
    product_description: str | None = None
    work_center_no: str | None = None
    resource_description: str | None = None


@dataclass(frozen=True)
class ProductionLossReportResult:
    report_id: int
    date_from: str
    date_to: str
    output_path: Path
    row_count: int
    warning_count: int
    generated_at: datetime
    rows: list[dict[str, Any]]
    source_summary: dict[str, Any]

    def as_dict(self) -> dict[str, Any]:
        return {
            "id": self.report_id,
            "date_from": self.date_from,
            "date_to": self.date_to,
            "output_path": str(self.output_path),
            "row_count": self.row_count,
            "warning_count": self.warning_count,
            "generated_at": _timestamp(self.generated_at),
            "source_summary": self.source_summary,
            "rows": self.rows,
        }


def create_production_loss_report(
    settings: Settings,
    *,
    date_from: str,
    date_to: str,
    refresh_ifs: bool = True,
    refresh_labels: bool = True,
) -> ProductionLossReportResult:
    start_date, end_date = _date_range(date_from, date_to)
    aggregates, quantity_summary = _load_label_shift_aggregates(
        settings,
        start_date,
        end_date,
        refresh_labels=refresh_labels,
    )
    quantity_source = LABEL_STOCK_SOURCE
    if not aggregates:
        operation_aggregates, operation_history_summary = (
            _load_operation_history_shift_aggregates(
                settings,
                start_date,
                end_date,
                refresh_operation_history=refresh_labels,
            )
        )
        if operation_aggregates:
            aggregates = operation_aggregates
            quantity_source = "ifs-operation-history"
            quantity_summary = {
                **quantity_summary,
                **operation_history_summary,
            }
        else:
            quantity_summary = {
                **quantity_summary,
                **operation_history_summary,
            }
            raise ProductionLossReportError(
                _label_stock_empty_report_message(quantity_summary)
            )

    process_meta = _read_process_metadata(settings, start_date, end_date)
    if quantity_source != LABEL_STOCK_SOURCE:
        realized_cycle_precheck = _realized_cycle_summary_for_aggregates(
            aggregates,
            process_meta,
        )
        if not realized_cycle_precheck["realized_cycle_valid_count"]:
            raise ProductionLossReportError(
                _realized_cycle_empty_report_message(realized_cycle_precheck)
            )

    order_numbers = sorted({aggregate.job_order for aggregate in aggregates})
    ifs_actuals, ifs_summary = _load_ifs_actuals(
        settings,
        order_numbers,
        refresh_ifs=refresh_ifs,
    )
    if quantity_source == LABEL_STOCK_SOURCE:
        aggregates = _apply_effective_machines_to_aggregates(
            settings,
            aggregates,
            ifs_actuals,
        )
        _apply_amount_control_breakdowns_to_unique_aggregates(
            settings,
            start_date,
            end_date,
            aggregates,
        )
        _apply_standalone_breakdowns_to_unique_aggregates(
            settings,
            start_date,
            end_date,
            aggregates,
        )
        aggregates = _sorted_aggregate_list(aggregates)
    row_models, realized_cycle_summary = _build_report_rows(
        settings,
        aggregates,
        process_meta,
        ifs_actuals,
    )
    if not row_models:
        raise ProductionLossReportError(
            _realized_cycle_empty_report_message(realized_cycle_summary)
        )

    generated_at = datetime.now(UTC).replace(tzinfo=None)
    warning_count = sum(1 for row in row_models if _clean_text(row.warnings)) + int(
        realized_cycle_summary["realized_cycle_skipped_count"]
    )
    source_summary = {
        "quantity_source": quantity_source,
        **quantity_summary,
        **ifs_summary,
        **realized_cycle_summary,
        "quantity_group_count": len(aggregates),
        "process_match_count": _process_match_count(
            aggregates,
            process_meta,
            ifs_actuals,
        ),
    }

    session = create_session(settings)
    try:
        report = ProductionLossReport(
            date_from=start_date.isoformat(),
            date_to=end_date.isoformat(),
            generated_at=generated_at,
            row_count=len(row_models),
            warning_count=warning_count,
            source_summary_json=json.dumps(source_summary, ensure_ascii=False),
        )
        session.add(report)
        session.flush()
        for row_model in row_models:
            row_model.report_id = report.id
            report.rows.append(row_model)

        output_path = _report_output_path(settings, report.id, start_date, end_date)
        _write_workbook(output_path, row_models)
        report.output_path = str(output_path)
        session.commit()
        payload_rows = [serialize_production_loss_row(row) for row in row_models]
        return ProductionLossReportResult(
            report_id=report.id,
            date_from=report.date_from,
            date_to=report.date_to,
            output_path=output_path,
            row_count=report.row_count,
            warning_count=report.warning_count,
            generated_at=report.generated_at,
            rows=payload_rows,
            source_summary=source_summary,
        )
    except Exception:
        session.rollback()
        raise
    finally:
        session.close()


def list_production_loss_reports(
    settings: Settings,
    *,
    limit: int = 20,
) -> list[dict[str, Any]]:
    with create_session(settings) as session:
        reports = session.scalars(
            select(ProductionLossReport)
            .order_by(
                ProductionLossReport.generated_at.desc(),
                ProductionLossReport.id.desc(),
            )
            .limit(limit)
        ).all()
        return [serialize_production_loss_report(report) for report in reports]


def get_production_loss_report(
    settings: Settings,
    report_id: int,
) -> dict[str, Any] | None:
    with create_session(settings) as session:
        report = session.scalars(
            select(ProductionLossReport)
            .options(selectinload(ProductionLossReport.rows))
            .where(ProductionLossReport.id == report_id)
        ).first()
        if report is None:
            return None
        return serialize_production_loss_report(report, include_rows=True)


def serialize_production_loss_report(
    report: ProductionLossReport,
    *,
    include_rows: bool = False,
) -> dict[str, Any]:
    payload = {
        "id": report.id,
        "date_from": report.date_from,
        "date_to": report.date_to,
        "generated_at": _timestamp(report.generated_at),
        "output_path": report.output_path,
        "row_count": report.row_count,
        "warning_count": report.warning_count,
        "source_summary": _json_object(report.source_summary_json),
    }
    if include_rows:
        payload["rows"] = [
            serialize_production_loss_row(row)
            for row in sorted(
                report.rows,
                key=lambda item: (
                    item.record_date,
                    _natural_machine_key(item.machine_code),
                    item.job_order,
                    item.lot_batch_no or "",
                    item.id,
                ),
            )
        ]
    return payload


def serialize_production_loss_row(row: ProductionLossReportRow) -> dict[str, Any]:
    return {
        "id": row.id,
        "record_date": row.record_date,
        "machine_id": row.machine_id,
        "machine_code": row.machine_code,
        "machine_name": row.machine_name,
        "job_order": row.job_order,
        "lot_batch_no": row.lot_batch_no,
        "product_no": row.product_no,
        "product_description": row.product_description,
        "gram": row.gram,
        "active_cavities": row.active_cavities,
        "cycle_time_seconds": row.cycle_time_seconds,
        "shift_2400_0800_quantity": row.shift_2400_0800_quantity,
        "shift_0800_1600_quantity": row.shift_0800_1600_quantity,
        "shift_1600_2400_quantity": row.shift_1600_2400_quantity,
        "shift_2400_0800_actual_quantity": row.shift_2400_0800_actual_quantity,
        "shift_2400_0800_machine_minutes": row.shift_2400_0800_machine_minutes,
        "shift_2400_0800_optimum_quantity": row.shift_2400_0800_optimum_quantity,
        "shift_2400_0800_loss_quantity": row.shift_2400_0800_loss_quantity,
        "shift_0800_1600_actual_quantity": row.shift_0800_1600_actual_quantity,
        "shift_0800_1600_machine_minutes": row.shift_0800_1600_machine_minutes,
        "shift_0800_1600_optimum_quantity": row.shift_0800_1600_optimum_quantity,
        "shift_0800_1600_loss_quantity": row.shift_0800_1600_loss_quantity,
        "shift_1600_2400_actual_quantity": row.shift_1600_2400_actual_quantity,
        "shift_1600_2400_machine_minutes": row.shift_1600_2400_machine_minutes,
        "shift_1600_2400_optimum_quantity": row.shift_1600_2400_optimum_quantity,
        "shift_1600_2400_loss_quantity": row.shift_1600_2400_loss_quantity,
        "daily_total_quantity": row.daily_total_quantity,
        "net_machine_minutes": row.net_machine_minutes,
        "gross_elapsed_minutes": row.gross_elapsed_minutes,
        "gross_start_at": _timestamp(row.gross_start_at),
        "gross_finish_at": _timestamp(row.gross_finish_at),
        "optimum_net_quantity": row.optimum_net_quantity,
        "production_loss_net": row.production_loss_net,
        "optimum_gross_quantity": row.optimum_gross_quantity,
        "production_loss_gross": row.production_loss_gross,
        "break_loss_quantity": row.break_loss_quantity,
        "local_breakdown_minutes": row.local_breakdown_minutes,
        "warnings": row.warnings,
    }


def normalize_ifs_actual_rows(rows: Sequence[Mapping[str, Any]]) -> list[IfsActual]:
    by_key: dict[tuple[str, str | None], IfsActual] = {}
    for row in rows:
        order_no = _first_text(row, ("OrderNo", "order_no", "ShopOrderNo"))
        if not order_no:
            continue
        machine_code = _ifs_actual_machine_code(row)
        key = (order_no, machine_code)
        actual = by_key.setdefault(
            key,
            IfsActual(job_order=order_no, machine_code=machine_code),
        )
        _merge_ifs_actual(actual, row)
    return list(by_key.values())


def _ifs_actual_machine_code(row: Mapping[str, Any]) -> str | None:
    direct = normalize_machine_code(
        _first_text(
            row,
            (
                "PreferredResourceId",
                "preferred_resource_id",
                "ResourceId",
                "resource_id",
                "Machine",
                "machine",
            ),
        )
    )
    if direct is not None:
        return direct

    return _machine_code_from_work_center_description(
        _nested_text(row, "WorkCenterNoRef", "Description")
    )


def _machine_code_from_work_center_description(value: Any) -> str | None:
    text = _optional_text(value)
    if text is None:
        return None
    match = re.search(r"(?:^|[_\s-])(\d{3,4})\s*$", text)
    if match is None:
        return None
    return match.group(1)


def _ifs_actual_part_description(row: Mapping[str, Any]) -> str | None:
    return _first_text(
        row,
        ("PartDescription", "PartNoDesc", "part_description", "Description"),
    ) or _nested_text(row, "PartNoRef", "Description")


def _build_report_rows(
    settings: Settings,
    aggregates: list[ShiftAggregate],
    process_meta: dict[tuple[str, str, str], ProcessMeta],
    ifs_actuals: dict[tuple[str, str | None], IfsActual],
) -> tuple[list[ProductionLossReportRow], dict[str, Any]]:
    by_order = _ifs_actuals_by_order(ifs_actuals.values())
    process_by_date_order = _process_meta_by_date_order(process_meta)
    machine_ids = _machine_ids_by_code(settings)
    report_rows: list[ProductionLossReportRow] = []
    realized_cycle_summary = _empty_realized_cycle_summary()

    for aggregate in aggregates:
        warnings: list[str] = []
        actual = _matching_ifs_actual(aggregate, ifs_actuals, by_order)
        row_aggregate = _aggregate_with_effective_machine(
            aggregate,
            actual,
            machine_ids,
        )
        process = _matching_process_meta(
            row_aggregate,
            process_meta,
            process_by_date_order,
        )
        cycle_time, skip_detail = _realized_cycle_time_for_row(process)
        if skip_detail is not None:
            _record_realized_cycle_skip(
                realized_cycle_summary,
                row_aggregate,
                skip_detail,
            )
            continue
        realized_cycle_summary["realized_cycle_valid_count"] += 1
        realized_cycle_summary["realized_cycle_included_quantity_total"] += (
            row_aggregate.daily_total_quantity
        )

        if actual is None:
            warnings.append("IFS actual timing not found")

        product_description = (
            actual.part_description
            if actual and actual.part_description
            else row_aggregate.product_hint
            if row_aggregate.product_hint
            else process.product
            if process and process.product
            else None
        )
        gram = _gram_from_product_description(product_description or "")
        _, parsed_gram = parse_part_description(product_description or "")
        if gram is None and parsed_gram is not None:
            gram = parsed_gram

        machine_name = (
            actual.work_center_no
            if actual and actual.work_center_no
            else row_aggregate.machine_group_hint
        )

        active_cavities = process.active_cavities if process else None
        if active_cavities is None:
            warnings.append("Active cavity count not found")

        if actual is not None and actual.actual_start_at is None:
            warnings.append("IFS actual start not found")
        if actual is not None and actual.actual_finish_at is None:
            warnings.append(
                "IFS actual finish not found; continuing production shifts only"
            )

        shift_calculations = _shift_calculations_for_row(
            settings,
            row_aggregate,
            actual,
            cycle_time,
            active_cavities,
        )
        total_shift_minutes = _sum_decimal_values(
            item["machine_minutes"] for item in shift_calculations.values()
        )
        gross_minutes = _sum_decimal_values(
            item["gross_minutes"] for item in shift_calculations.values()
        )
        optimum_net = _sum_decimal_values(
            item["optimum"] for item in shift_calculations.values()
        )
        loss_net = _sum_decimal_values(
            item["loss"] for item in shift_calculations.values()
        )
        optimum_gross = _sum_decimal_values(
            item["gross_optimum"] for item in shift_calculations.values()
        )
        loss_gross = _sum_decimal_values(
            item["gross_loss"] for item in shift_calculations.values()
        )
        break_loss = (
            loss_gross - loss_net
            if loss_gross is not None and loss_net is not None
            else None
        )

        report_rows.append(
            ProductionLossReportRow(
                report_id=0,
                record_date=row_aggregate.record_date,
                machine_id=row_aggregate.machine_id,
                machine_code=row_aggregate.machine_code,
                machine_name=machine_name,
                job_order=row_aggregate.job_order,
                lot_batch_no=row_aggregate.lot_batch_no,
                product_no=actual.part_no
                if actual and actual.part_no
                else row_aggregate.part_no_hint,
                product_description=product_description,
                gram=_decimal_text(gram),
                active_cavities=_decimal_text(active_cavities),
                cycle_time_seconds=_decimal_text(cycle_time),
                shift_2400_0800_quantity=row_aggregate.shift_quantities["24.00-08.00"],
                shift_0800_1600_quantity=row_aggregate.shift_quantities["08.00-16.00"],
                shift_1600_2400_quantity=row_aggregate.shift_quantities["16.00-24.00"],
                shift_2400_0800_actual_quantity=row_aggregate.shift_quantities[
                    "24.00-08.00"
                ],
                shift_2400_0800_machine_minutes=_decimal_text(
                    shift_calculations["24.00-08.00"]["machine_minutes"]
                ),
                shift_2400_0800_optimum_quantity=_decimal_text(
                    shift_calculations["24.00-08.00"]["optimum"]
                ),
                shift_2400_0800_loss_quantity=_decimal_text(
                    shift_calculations["24.00-08.00"]["loss"]
                ),
                shift_0800_1600_actual_quantity=row_aggregate.shift_quantities[
                    "08.00-16.00"
                ],
                shift_0800_1600_machine_minutes=_decimal_text(
                    shift_calculations["08.00-16.00"]["machine_minutes"]
                ),
                shift_0800_1600_optimum_quantity=_decimal_text(
                    shift_calculations["08.00-16.00"]["optimum"]
                ),
                shift_0800_1600_loss_quantity=_decimal_text(
                    shift_calculations["08.00-16.00"]["loss"]
                ),
                shift_1600_2400_actual_quantity=row_aggregate.shift_quantities[
                    "16.00-24.00"
                ],
                shift_1600_2400_machine_minutes=_decimal_text(
                    shift_calculations["16.00-24.00"]["machine_minutes"]
                ),
                shift_1600_2400_optimum_quantity=_decimal_text(
                    shift_calculations["16.00-24.00"]["optimum"]
                ),
                shift_1600_2400_loss_quantity=_decimal_text(
                    shift_calculations["16.00-24.00"]["loss"]
                ),
                daily_total_quantity=row_aggregate.daily_total_quantity,
                net_machine_minutes=_decimal_text(total_shift_minutes),
                gross_elapsed_minutes=_decimal_text(gross_minutes),
                gross_start_at=actual.actual_start_at if actual else None,
                gross_finish_at=actual.actual_finish_at if actual else None,
                optimum_net_quantity=_decimal_text(optimum_net),
                production_loss_net=_decimal_text(loss_net),
                optimum_gross_quantity=_decimal_text(optimum_gross),
                production_loss_gross=_decimal_text(loss_gross),
                break_loss_quantity=_decimal_text(break_loss),
                local_breakdown_minutes=row_aggregate.local_breakdown_minutes,
                warnings="; ".join(dict.fromkeys(warnings)) or None,
            )
        )

    _set_realized_cycle_skip_message(realized_cycle_summary)

    return sorted(
        report_rows,
        key=lambda row: (
            row.record_date,
            _natural_machine_key(row.machine_code),
            row.job_order,
            row.lot_batch_no or "",
        ),
    ), realized_cycle_summary


def _machine_ids_by_code(settings: Settings) -> dict[str, int]:
    with create_session(settings) as session:
        machines = session.scalars(select(Machine)).all()
    return {
        machine_code: machine.id
        for machine in machines
        if (machine_code := normalize_machine_code(machine.machine_code))
    }


def _apply_effective_machines_to_aggregates(
    settings: Settings,
    aggregates: Sequence[ShiftAggregate],
    ifs_actuals: Mapping[tuple[str, str | None], IfsActual],
) -> list[ShiftAggregate]:
    by_order = _ifs_actuals_by_order(ifs_actuals.values())
    machine_ids = _machine_ids_by_code(settings)
    actuals = dict(ifs_actuals)
    return [
        _aggregate_with_effective_machine(
            aggregate,
            _matching_ifs_actual(aggregate, actuals, by_order),
            machine_ids,
        )
        for aggregate in aggregates
    ]


def _sorted_aggregate_list(
    aggregates: Iterable[ShiftAggregate],
) -> list[ShiftAggregate]:
    return sorted(
        aggregates,
        key=lambda item: (
            item.record_date,
            _natural_machine_key(item.machine_code),
            item.job_order,
            item.lot_batch_no or "",
        ),
    )


def _apply_amount_control_breakdowns_to_unique_aggregates(
    settings: Settings,
    start_date: Date,
    end_date: Date,
    aggregates: Sequence[ShiftAggregate],
) -> None:
    _apply_amount_control_breakdowns(
        settings,
        start_date,
        end_date,
        _unique_aggregates_by_date_machine_job(aggregates),
    )


def _apply_standalone_breakdowns_to_unique_aggregates(
    settings: Settings,
    start_date: Date,
    end_date: Date,
    aggregates: Sequence[ShiftAggregate],
) -> None:
    if not aggregates:
        return
    _apply_standalone_breakdowns(
        settings,
        start_date,
        end_date,
        _unique_aggregates_by_date_machine_job(aggregates),
        _unique_aggregates_by_date_machine_shift(aggregates),
    )


def _unique_aggregates_by_date_machine_job(
    aggregates: Sequence[ShiftAggregate],
) -> dict[tuple[str, str, str], ShiftAggregate]:
    grouped: dict[tuple[str, str, str], list[ShiftAggregate]] = defaultdict(list)
    for aggregate in aggregates:
        grouped[
            (aggregate.record_date, aggregate.machine_code, aggregate.job_order)
        ].append(aggregate)
    return {key: values[0] for key, values in grouped.items() if len(values) == 1}


def _unique_aggregates_by_date_machine_shift(
    aggregates: Sequence[ShiftAggregate],
) -> dict[tuple[str, str, str], ShiftAggregate]:
    grouped: dict[tuple[str, str, str], list[ShiftAggregate]] = defaultdict(list)
    for aggregate in aggregates:
        for shift_label, quantity in aggregate.shift_quantities.items():
            if quantity > 0:
                key = (aggregate.record_date, aggregate.machine_code, shift_label)
                grouped[key].append(aggregate)
    return {key: values[0] for key, values in grouped.items() if len(values) == 1}


def _aggregate_with_effective_machine(
    aggregate: ShiftAggregate,
    actual: IfsActual | None,
    machine_ids: Mapping[str, int],
) -> ShiftAggregate:
    if not _unknown_machine_code(aggregate.machine_code):
        return aggregate
    machine_code = normalize_machine_code(actual.machine_code if actual else None)
    if machine_code is None:
        return aggregate
    return replace(
        aggregate,
        machine_code=machine_code,
        machine_id=machine_ids.get(machine_code),
    )


def _unknown_machine_code(value: str | None) -> bool:
    return not value or value == UNKNOWN_MACHINE_CODE


def _process_meta_sort_key(process: ProcessMeta) -> tuple[datetime, int]:
    return process.submitted_at or datetime.min, process.entry_id


def _process_meta_by_date_order(
    process_meta: Mapping[tuple[str, str, str], ProcessMeta],
) -> dict[tuple[str, str], ProcessMeta]:
    by_date_order: dict[tuple[str, str], ProcessMeta] = {}
    for (record_date, _machine_code, job_order), process in process_meta.items():
        key = (record_date, job_order)
        existing = by_date_order.get(key)
        if existing is None or _process_meta_sort_key(process) > _process_meta_sort_key(
            existing
        ):
            by_date_order[key] = process
    return by_date_order


def _matching_process_meta(
    aggregate: ShiftAggregate,
    process_meta: Mapping[tuple[str, str, str], ProcessMeta],
    process_by_date_order: Mapping[tuple[str, str], ProcessMeta],
) -> ProcessMeta | None:
    exact = process_meta.get(
        (aggregate.record_date, aggregate.machine_code, aggregate.job_order)
    )
    if exact is not None:
        return exact
    return process_by_date_order.get((aggregate.record_date, aggregate.job_order))


def _process_match_count(
    aggregates: Sequence[ShiftAggregate],
    process_meta: Mapping[tuple[str, str, str], ProcessMeta],
    ifs_actuals: Mapping[tuple[str, str | None], IfsActual],
) -> int:
    by_order = _ifs_actuals_by_order(ifs_actuals.values())
    process_by_date_order = _process_meta_by_date_order(process_meta)
    actuals = dict(ifs_actuals)
    count = 0
    for aggregate in aggregates:
        actual = _matching_ifs_actual(aggregate, actuals, by_order)
        machine_code = (
            normalize_machine_code(actual.machine_code if actual else None)
            if _unknown_machine_code(aggregate.machine_code)
            else aggregate.machine_code
        )
        row_aggregate = (
            replace(aggregate, machine_code=machine_code)
            if machine_code is not None
            else aggregate
        )
        if (
            _matching_process_meta(
                row_aggregate,
                process_meta,
                process_by_date_order,
            )
            is not None
        ):
            count += 1
    return count


def _realized_cycle_summary_for_aggregates(
    aggregates: Sequence[ShiftAggregate],
    process_meta: Mapping[tuple[str, str, str], ProcessMeta],
) -> dict[str, Any]:
    summary = _empty_realized_cycle_summary()
    process_by_date_order = _process_meta_by_date_order(process_meta)
    for aggregate in aggregates:
        process = _matching_process_meta(
            aggregate,
            process_meta,
            process_by_date_order,
        )
        _, skip_detail = _realized_cycle_time_for_row(process)
        if skip_detail is not None:
            _record_realized_cycle_skip(summary, aggregate, skip_detail)
            continue
        summary["realized_cycle_valid_count"] += 1
        summary["realized_cycle_included_quantity_total"] += (
            aggregate.daily_total_quantity
        )
    _set_realized_cycle_skip_message(summary)
    return summary


def _empty_realized_cycle_summary() -> dict[str, Any]:
    return {
        "realized_cycle_source": REALIZED_CYCLE_SOURCE,
        "realized_cycle_valid_count": 0,
        "realized_cycle_missing_count": 0,
        "realized_cycle_invalid_count": 0,
        "realized_cycle_skipped_count": 0,
        "realized_cycle_included_quantity_total": 0,
        "realized_cycle_skipped_quantity_total": 0,
        "realized_cycle_skip_message": None,
        "realized_cycle_skip_details": [],
    }


def _realized_cycle_time_for_row(
    process: ProcessMeta | None,
) -> tuple[Decimal | None, dict[str, Any] | None]:
    if process is None:
        return None, {
            "entry_id": None,
            "raw_cycle": None,
            "reason": "missing-process-entry",
            "message": REALIZED_CYCLE_MISSING_MESSAGE,
        }

    if process.raw_cycle_time is None:
        return None, {
            "entry_id": process.entry_id,
            "raw_cycle": None,
            "reason": "missing-cycle-time",
            "message": REALIZED_CYCLE_MISSING_MESSAGE,
        }

    if process.realized_cycle_time is None:
        return None, {
            "entry_id": process.entry_id,
            "raw_cycle": process.raw_cycle_time,
            "reason": "invalid-cycle-time",
            "message": REALIZED_CYCLE_INVALID_MESSAGE,
        }

    if process.realized_cycle_time <= 0:
        return None, {
            "entry_id": process.entry_id,
            "raw_cycle": process.raw_cycle_time,
            "reason": "non-positive-cycle-time",
            "message": REALIZED_CYCLE_INVALID_MESSAGE,
        }

    return process.realized_cycle_time, None


def _record_realized_cycle_skip(
    summary: dict[str, Any],
    aggregate: ShiftAggregate,
    detail: dict[str, Any],
) -> None:
    if detail["message"] == REALIZED_CYCLE_MISSING_MESSAGE:
        summary["realized_cycle_missing_count"] += 1
    else:
        summary["realized_cycle_invalid_count"] += 1
    summary["realized_cycle_skipped_count"] += 1
    summary["realized_cycle_skipped_quantity_total"] += aggregate.daily_total_quantity
    summary["realized_cycle_skip_details"].append(
        {
            "record_date": aggregate.record_date,
            "machine_code": aggregate.machine_code,
            "job_order": aggregate.job_order,
            "entry_id": detail["entry_id"],
            "raw_cycle": detail["raw_cycle"],
            "reason": detail["reason"],
            "message": detail["message"],
        }
    )


def _set_realized_cycle_skip_message(summary: dict[str, Any]) -> None:
    if summary["realized_cycle_skipped_count"]:
        summary["realized_cycle_skip_message"] = (
            f"{summary['realized_cycle_skipped_count']} "
            "Production Loss satiri gerceklesen cevrim eksik/gecersiz oldugu "
            "icin atlandi."
        )


def _realized_cycle_empty_report_message(summary: Mapping[str, Any]) -> str:
    skipped = summary["realized_cycle_skipped_count"]
    missing = summary["realized_cycle_missing_count"]
    invalid = summary["realized_cycle_invalid_count"]
    return (
        "Hesaplanabilir Production Loss satiri yok. "
        f"{skipped} satir gerceklesen cevrim eksik/gecersiz oldugu icin atlandi "
        f"({missing} eksik, {invalid} gecersiz)."
    )


def _load_operation_history_shift_aggregates(
    settings: Settings,
    start_date: Date,
    end_date: Date,
    *,
    refresh_operation_history: bool,
) -> tuple[list[ShiftAggregate], dict[str, Any]]:
    summary: dict[str, Any] = {
        "operation_history_refreshed": False,
        "operation_history_ifs_error": None,
        "operation_history_row_count": 0,
        "operation_history_event_count": 0,
        "operation_history_quantity_total": 0,
        "operation_history_parse_error_count": 0,
        "operation_history_out_of_range_count": 0,
        "operation_history_shift_group_count": 0,
        "operation_history_product_description_count": 0,
        "operation_history_product_description_error": None,
    }
    events: list[OperationHistoryProductionEvent] = []
    if refresh_operation_history:
        try:
            date_from_utc, date_to_utc = _operation_history_window(
                settings,
                start_date,
                end_date,
            )
            raw_rows = asyncio.run(
                fetch_shop_order_operation_history_rows(
                    settings,
                    date_from_utc=date_from_utc,
                    date_to_utc=date_to_utc,
                )
            )
            product_descriptions: dict[str, str] = {}
            part_numbers = sorted(
                {
                    part_no
                    for row in raw_rows
                    if (part_no := _first_text(row, OPERATION_HISTORY_PART_FIELDS))
                }
            )
            if part_numbers:
                try:
                    product_descriptions = asyncio.run(
                        fetch_inventory_part_descriptions(settings, part_numbers)
                    )
                    summary["operation_history_product_description_count"] = len(
                        product_descriptions
                    )
                except Exception as exc:
                    summary["operation_history_product_description_error"] = (
                        str(exc) or exc.__class__.__name__
                    )
            events, parse_error_count, out_of_range_count = (
                normalize_operation_history_rows(
                    settings,
                    raw_rows,
                    start_date,
                    end_date,
                    product_descriptions=product_descriptions,
                )
            )
            summary["operation_history_refreshed"] = True
            summary["operation_history_row_count"] = len(raw_rows)
            summary["operation_history_event_count"] = len(events)
            summary["operation_history_quantity_total"] = sum(
                event.quantity for event in events
            )
            summary["operation_history_parse_error_count"] = parse_error_count
            summary["operation_history_out_of_range_count"] = out_of_range_count
            summary["operation_history_shift_group_count"] = len(
                _operation_history_shift_group_keys(events)
            )
        except Exception as exc:
            summary["operation_history_ifs_error"] = str(exc) or exc.__class__.__name__

    aggregates = _aggregates_from_operation_history_events(settings, events)
    _apply_amount_control_breakdowns(settings, start_date, end_date, aggregates)
    _apply_standalone_breakdowns_to_unique_aggregates(
        settings,
        start_date,
        end_date,
        list(aggregates.values()),
    )
    return (
        sorted(
            aggregates.values(),
            key=lambda item: (
                item.record_date,
                _natural_machine_key(item.machine_code),
                item.job_order,
            ),
        ),
        summary,
    )


def _operation_history_empty_report_message(summary: Mapping[str, Any]) -> str:
    message = "Secilen tarih araliginda IFS operasyon gecmisi uretim kaydi bulunamadi."
    ifs_error = _optional_text(summary.get("operation_history_ifs_error"))
    if ifs_error:
        return f"{message} IFS operasyon gecmisi alinamadi: {ifs_error}"
    return message


def _operation_history_window(
    settings: Settings,
    start_date: Date,
    end_date: Date,
) -> tuple[datetime, datetime]:
    tzinfo = shifts.request_timezone(settings)
    start_local = datetime.combine(start_date, time.min, tzinfo=tzinfo)
    end_local = datetime.combine(
        end_date + timedelta(days=1),
        time.min,
        tzinfo=tzinfo,
    )
    return start_local.astimezone(UTC), end_local.astimezone(UTC)


def normalize_operation_history_rows(
    settings: Settings,
    rows: Sequence[Mapping[str, Any]],
    start_date: Date,
    end_date: Date,
    *,
    product_descriptions: Mapping[str, str] | None = None,
) -> tuple[list[OperationHistoryProductionEvent], int, int]:
    events: list[OperationHistoryProductionEvent] = []
    seen_transactions: set[str] = set()
    parse_error_count = 0
    out_of_range_count = 0
    for row in rows:
        event = _operation_history_event_from_row(
            settings,
            row,
            product_descriptions=product_descriptions or {},
        )
        if event is None:
            parse_error_count += 1
            continue
        if event.transaction_id in seen_transactions:
            continue
        seen_transactions.add(event.transaction_id)
        event_date = Date.fromisoformat(event.record_date)
        if event_date < start_date or event_date > end_date:
            out_of_range_count += 1
            continue
        events.append(event)
    return events, parse_error_count, out_of_range_count


def _operation_history_event_from_row(
    settings: Settings,
    row: Mapping[str, Any],
    *,
    product_descriptions: Mapping[str, str],
) -> OperationHistoryProductionEvent | None:
    time_of_production_utc = _first_datetime(row, OPERATION_HISTORY_TIME_FIELDS)
    time_of_production = _stored_utc_to_local(settings, time_of_production_utc)
    machine_code = normalize_machine_code(
        _first_text(row, OPERATION_HISTORY_MACHINE_FIELDS)
    )
    job_order = _identifier(_first_value(row, OPERATION_HISTORY_ORDER_FIELDS))
    part_no = _first_text(row, OPERATION_HISTORY_PART_FIELDS)
    quantity = _int_decimal(_first_value(row, OPERATION_HISTORY_QUANTITY_FIELDS))
    work_center_no = _first_text(row, ("WorkCenterNo", "work_center_no"))
    resource_description = _first_text(
        row,
        ("ResourceDescription", "resource_description"),
    )
    if (
        time_of_production is None
        or machine_code is None
        or not job_order
        or not part_no
        or quantity is None
        or quantity <= 0
    ):
        return None

    transaction_id = _identifier(
        _first_value(row, OPERATION_HISTORY_TRANSACTION_FIELDS)
    )
    if not transaction_id:
        transaction_id = "|".join(
            (
                job_order,
                machine_code,
                time_of_production.isoformat(timespec="seconds"),
                str(quantity),
                part_no or "",
            )
        )

    return OperationHistoryProductionEvent(
        record_date=time_of_production.date().isoformat(),
        shift=shifts.shift_for_request_time(time_of_production),
        machine_code=machine_code,
        job_order=job_order,
        part_no=part_no,
        quantity=quantity,
        transaction_id=transaction_id,
        time_of_production=time_of_production,
        product_description=_optional_text(product_descriptions.get(part_no)),
        work_center_no=work_center_no,
        resource_description=resource_description,
    )


def _operation_history_shift_group_keys(
    events: Sequence[OperationHistoryProductionEvent],
) -> set[tuple[str, str, str, str]]:
    return {
        (event.record_date, event.shift, event.machine_code, event.job_order)
        for event in events
    }


def _aggregates_from_operation_history_events(
    settings: Settings,
    events: Sequence[OperationHistoryProductionEvent],
) -> dict[tuple[str, str, str], ShiftAggregate]:
    with create_session(settings) as session:
        machines = session.scalars(select(Machine)).all()
        machine_ids = {
            normalize_machine_code(machine.machine_code): machine.id
            for machine in machines
        }

    shift_quantities: dict[tuple[str, str, str, str], int] = defaultdict(int)
    part_hints: dict[tuple[str, str, str], str] = {}
    product_hints: dict[tuple[str, str, str], str] = {}
    machine_group_hints: dict[tuple[str, str, str], str] = {}
    event_counts: dict[tuple[str, str, str], int] = defaultdict(int)
    for event in events:
        aggregate_key = (event.record_date, event.machine_code, event.job_order)
        shift_key = (
            event.record_date,
            event.shift,
            event.machine_code,
            event.job_order,
        )
        shift_quantities[shift_key] += event.quantity
        event_counts[aggregate_key] += 1
        if event.part_no and aggregate_key not in part_hints:
            part_hints[aggregate_key] = event.part_no
        if event.product_description and aggregate_key not in product_hints:
            product_hints[aggregate_key] = event.product_description
        machine_group_hint = event.work_center_no or event.resource_description
        if machine_group_hint and aggregate_key not in machine_group_hints:
            machine_group_hints[aggregate_key] = machine_group_hint

    aggregates: dict[tuple[str, str, str], ShiftAggregate] = {}
    for (
        record_date,
        shift,
        machine_code,
        job_order,
    ), quantity in shift_quantities.items():
        aggregate_key = (record_date, machine_code, job_order)
        aggregate = aggregates.setdefault(
            aggregate_key,
            ShiftAggregate(
                record_date=record_date,
                machine_id=machine_ids.get(machine_code),
                machine_code=machine_code,
                job_order=job_order,
                product_hint=product_hints.get(aggregate_key),
                part_no_hint=part_hints.get(aggregate_key),
                machine_group_hint=machine_group_hints.get(aggregate_key),
            ),
        )
        if shift in aggregate.shift_quantities:
            aggregate.shift_quantities[shift] += quantity
        aggregate.label_event_count = event_counts[aggregate_key]

    return aggregates


def _load_label_shift_aggregates(
    settings: Settings,
    start_date: Date,
    end_date: Date,
    *,
    refresh_labels: bool,
) -> tuple[list[ShiftAggregate], dict[str, Any]]:
    summary: dict[str, Any] = {
        "labels_refreshed": False,
        "label_ifs_error": None,
        "label_archive_count": 0,
        "label_cached_count": 0,
        "label_event_count": 0,
        "label_quantity_total": 0,
        "label_parse_error_count": 0,
        "label_out_of_range_count": 0,
        "label_missing_u01_count": 0,
        "label_zero_quantity_count": 0,
        "label_stock_row_count": 0,
        "label_physical_unit_count": 0,
        "label_source": LABEL_STOCK_SOURCE,
    }
    if refresh_labels:
        try:
            raw_rows = asyncio.run(
                fetch_production_label_stock_rows(
                    settings,
                    date_from=start_date,
                    date_to=end_date,
                )
            )
            events, parse_error_count, out_of_range_count, detail_counts = (
                normalize_inventory_label_rows(
                    settings,
                    raw_rows,
                    start_date,
                    end_date,
                )
            )
            _cache_label_events(settings, events)
            summary["labels_refreshed"] = True
            summary["label_stock_row_count"] = len(raw_rows)
            summary["label_physical_unit_count"] = detail_counts["physical_unit_count"]
            summary["label_event_count"] = len(events)
            summary["label_quantity_total"] = sum(event.quantity for event in events)
            summary["label_parse_error_count"] = parse_error_count
            summary["label_out_of_range_count"] = out_of_range_count
            summary["label_missing_u01_count"] = detail_counts["missing_u01_count"]
            summary["label_zero_quantity_count"] = detail_counts["zero_quantity_count"]
        except Exception as exc:
            summary["label_ifs_error"] = str(exc) or exc.__class__.__name__

    cached_events = _read_cached_label_events(settings, start_date, end_date)
    summary["label_cached_count"] = len(cached_events)
    if not summary["label_event_count"]:
        summary["label_event_count"] = len(cached_events)
        summary["label_quantity_total"] = sum(event.quantity for event in cached_events)
    aggregates = _aggregates_from_label_events(settings, cached_events)
    return (
        sorted(
            aggregates.values(),
            key=lambda item: (
                item.record_date,
                _natural_machine_key(item.machine_code),
                item.job_order,
            ),
        ),
        summary,
    )


def _label_stock_empty_report_message(summary: Mapping[str, Any]) -> str:
    message = "Secilen tarih araliginda IFS etiket stok hareketi bulunamadi."
    label_error = _optional_text(summary.get("label_ifs_error"))
    if label_error:
        return f"{message} IFS etiket stok hareketleri alinamadi: {label_error}"
    operation_error = _optional_text(summary.get("operation_history_ifs_error"))
    if operation_error:
        return f"{message} Yedek IFS operasyon gecmisi de alinamadi: {operation_error}"
    return message


def normalize_inventory_label_rows(
    settings: Settings,
    rows: Sequence[Mapping[str, Any]],
    start_date: Date,
    end_date: Date,
) -> tuple[list[LabelProductionEvent], int, int, dict[str, int]]:
    grouped_rows: dict[tuple[str, str, str], list[Mapping[str, Any]]] = defaultdict(
        list
    )
    parse_error_count = 0
    for row in rows:
        group_key = _label_stock_group_key(row)
        if group_key is None:
            parse_error_count += 1
            continue
        grouped_rows[group_key].append(row)

    events: list[LabelProductionEvent] = []
    out_of_range_count = 0
    missing_u01_count = 0
    zero_quantity_count = 0
    for group_key, group_rows in grouped_rows.items():
        event, skip_reason = _inventory_label_event_from_rows(
            settings,
            group_key,
            group_rows,
        )
        if event is None:
            if skip_reason == "missing-u01":
                missing_u01_count += 1
            elif skip_reason == "zero-quantity":
                zero_quantity_count += 1
            else:
                parse_error_count += 1
            continue
        event_date = _iso_date(event.record_date, "record_date")
        if event_date < start_date or event_date > end_date:
            out_of_range_count += 1
            continue
        events.append(event)

    return (
        events,
        parse_error_count,
        out_of_range_count,
        {
            "physical_unit_count": len(grouped_rows),
            "missing_u01_count": missing_u01_count,
            "zero_quantity_count": zero_quantity_count,
        },
    )


def _label_stock_group_key(row: Mapping[str, Any]) -> tuple[str, str, str] | None:
    handling_unit_id = _identifier(_first_value(row, LABEL_STOCK_HANDLING_UNIT_FIELDS))
    part_no = _identifier(_first_value(row, LABEL_STOCK_PART_FIELDS))
    lot_batch_no = _identifier(_first_value(row, LABEL_STOCK_LOT_FIELDS))
    if not handling_unit_id or not part_no or not lot_batch_no:
        return None
    return handling_unit_id, part_no, lot_batch_no


def _inventory_label_event_from_rows(
    settings: Settings,
    group_key: tuple[str, str, str],
    rows: Sequence[Mapping[str, Any]],
) -> tuple[LabelProductionEvent | None, str | None]:
    handling_unit_id, part_no, lot_batch_no = group_key
    u01_rows = [row for row in rows if _is_label_production_location(row)]
    if not u01_rows:
        return None, "missing-u01"
    production_row = min(
        u01_rows,
        key=lambda row: _label_stock_time(settings, row) or datetime.max,
    )
    label_time = _label_stock_time(settings, production_row)
    if label_time is None:
        return None, "missing-time"

    quantity = _label_stock_quantity(rows)
    if quantity is None or quantity <= 0:
        return None, "zero-quantity"

    job_order = _job_order_from_lot_batch_no(lot_batch_no)
    if not job_order:
        return None, "missing-job-order"

    return LabelProductionEvent(
        result_key=_label_stock_result_key(group_key),
        report_id=LABEL_STOCK_REPORT_ID,
        record_date=label_time.date().isoformat(),
        shift=shifts.shift_for_request_time(label_time),
        machine_code=UNKNOWN_MACHINE_CODE,
        job_order=job_order,
        quantity=quantity,
        label_time=label_time,
        archive_exec_time=None,
        print_job_id=None,
        part_no=part_no,
        product_description=_first_text(rows[0], LABEL_STOCK_PRODUCT_FIELDS),
        package_id=handling_unit_id,
        lot_batch_no=lot_batch_no,
        pallet_no=None,
        sequence_no=None,
        raw_xml=json.dumps(list(rows), ensure_ascii=False, default=str),
    ), None


def _is_label_production_location(row: Mapping[str, Any]) -> bool:
    location = _first_text(row, LABEL_STOCK_LOCATION_FIELDS)
    return _canonical_location(location) == _canonical_location(
        LABEL_STOCK_PRODUCTION_LOCATION
    )


def _canonical_location(value: Any) -> str:
    text = _clean_text(value).upper()
    match = re.fullmatch(r"U0*(\d+)", text)
    if match is not None:
        return f"U{int(match.group(1))}"
    return text


def _label_stock_time(settings: Settings, row: Mapping[str, Any]) -> datetime | None:
    return _local_datetime_value(
        settings,
        _first_value(row, LABEL_STOCK_RECEIPT_FIELDS),
    )


def _label_stock_quantity(rows: Sequence[Mapping[str, Any]]) -> int | None:
    onhand_values = [
        value
        for row in rows
        if (value := _decimal(_first_value(row, LABEL_STOCK_ONHAND_FIELDS))) is not None
        and value > 0
    ]
    if onhand_values:
        return _int_decimal(sum(onhand_values, Decimal("0")))

    pack_values = [
        value
        for row in rows
        if (value := _decimal(_first_value(row, LABEL_STOCK_PACK_QUANTITY_FIELDS)))
        is not None
        and value > 0
    ]
    if pack_values:
        return _int_decimal(max(pack_values))
    return None


def _job_order_from_lot_batch_no(lot_batch_no: str) -> str:
    return (
        lot_batch_no.split("-", 1)[0].strip() if "-" in lot_batch_no else lot_batch_no
    )


def _label_stock_result_key(group_key: tuple[str, str, str]) -> str:
    handling_unit_id, part_no, lot_batch_no = group_key
    return f"stock:{handling_unit_id}:{part_no}:{lot_batch_no}"


def _label_archive_window(
    settings: Settings,
    start_date: Date,
    end_date: Date,
) -> tuple[datetime, datetime]:
    tzinfo = shifts.request_timezone(settings)
    start_local = datetime.combine(start_date, time.min, tzinfo=tzinfo)
    end_local = datetime.combine(
        end_date + timedelta(days=1),
        time.min,
        tzinfo=tzinfo,
    )
    return start_local.astimezone(UTC), end_local.astimezone(UTC)


def _normalize_label_rows(
    settings: Settings,
    rows: Sequence[Mapping[str, Any]],
    start_date: Date,
    end_date: Date,
) -> tuple[list[LabelProductionEvent], int, int]:
    events: list[LabelProductionEvent] = []
    parse_error_count = 0
    out_of_range_count = 0
    for row in rows:
        event = _label_event_from_row(settings, row)
        if event is None:
            parse_error_count += 1
            continue
        event_date = _iso_date(event.record_date, "record_date")
        if event_date < start_date or event_date > end_date:
            out_of_range_count += 1
            continue
        events.append(event)
    return events, parse_error_count, out_of_range_count


def _label_event_from_row(
    settings: Settings,
    row: Mapping[str, Any],
) -> LabelProductionEvent | None:
    result_key = _identifier(row.get("ResultKey"))
    report_id = _identifier(row.get("ReportId"))
    label_time = _parse_label_time(settings, row.get("label_time"))
    quantity = _int_decimal(row.get("quantity"))
    machine_code = normalize_machine_code(row.get("machine_code"))
    job_order = _identifier(row.get("job_order"))
    if (
        not result_key
        or not report_id
        or label_time is None
        or quantity is None
        or machine_code is None
        or not job_order
    ):
        return None
    return LabelProductionEvent(
        result_key=result_key,
        report_id=report_id,
        print_job_id=_optional_text(row.get("PrintJobId")),
        archive_exec_time=_datetime_value(row.get("ExecTime")),
        label_time=label_time,
        record_date=label_time.date().isoformat(),
        shift=shifts.shift_for_request_time(label_time),
        machine_code=machine_code,
        job_order=job_order,
        part_no=_optional_text(row.get("part_no")),
        product_description=_optional_text(row.get("product_description")),
        quantity=quantity,
        package_id=_optional_text(row.get("package_id")),
        lot_batch_no=_optional_text(row.get("lot_batch_no")),
        pallet_no=_optional_text(row.get("pallet_no")),
        sequence_no=_optional_text(row.get("sequence_no")),
        raw_xml=_optional_text(row.get("raw_xml")),
    )


def _parse_label_time(settings: Settings, value: Any) -> datetime | None:
    text = _clean_text(value)
    if not text:
        return None
    for date_format in ("%d.%m.%Y %H:%M:%S", "%d.%m.%Y %H:%M"):
        try:
            return datetime.strptime(text, date_format)
        except ValueError:
            continue
    return _local_datetime_value(settings, text)


def _local_datetime_value(settings: Settings, value: Any) -> datetime | None:
    if isinstance(value, datetime):
        parsed = value
    elif value is None:
        return None
    else:
        text = str(value).strip()
        if not text:
            return None
        try:
            parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
        except ValueError:
            return None

    if parsed.tzinfo is None:
        return parsed
    return parsed.astimezone(shifts.request_timezone(settings)).replace(tzinfo=None)


def _int_decimal(value: Any) -> int | None:
    number = _decimal(value)
    if number is None:
        return None
    return int(number)


def _cache_label_events(
    settings: Settings,
    events: Sequence[LabelProductionEvent],
) -> None:
    if not events:
        return
    now = datetime.now(UTC).replace(tzinfo=None)
    session = create_session(settings)
    try:
        for event in events:
            existing = session.scalars(
                select(ProductionLossLabelEvent).where(
                    ProductionLossLabelEvent.result_key == event.result_key
                )
            ).first()
            if existing is None:
                existing = ProductionLossLabelEvent(result_key=event.result_key)
                session.add(existing)
            existing.report_id = event.report_id
            existing.print_job_id = event.print_job_id
            existing.archive_exec_time = event.archive_exec_time
            existing.label_time = event.label_time
            existing.record_date = event.record_date
            existing.shift = event.shift
            existing.machine_code = event.machine_code
            existing.job_order = event.job_order
            existing.part_no = event.part_no
            existing.product_description = event.product_description
            existing.quantity = event.quantity
            existing.package_id = event.package_id
            existing.lot_batch_no = event.lot_batch_no
            existing.pallet_no = event.pallet_no
            existing.sequence_no = event.sequence_no
            existing.raw_xml = event.raw_xml
            existing.source_updated_at = now
        session.commit()
    except Exception:
        session.rollback()
        raise
    finally:
        session.close()


def _read_cached_label_events(
    settings: Settings,
    start_date: Date,
    end_date: Date,
) -> list[LabelProductionEvent]:
    with create_session(settings) as session:
        rows = session.scalars(
            select(ProductionLossLabelEvent)
            .where(ProductionLossLabelEvent.record_date >= start_date.isoformat())
            .where(ProductionLossLabelEvent.record_date <= end_date.isoformat())
            .order_by(
                ProductionLossLabelEvent.record_date.asc(),
                ProductionLossLabelEvent.machine_code.asc(),
                ProductionLossLabelEvent.job_order.asc(),
                ProductionLossLabelEvent.label_time.asc(),
                ProductionLossLabelEvent.result_key.asc(),
            )
        ).all()
        return [
            LabelProductionEvent(
                result_key=row.result_key,
                report_id=row.report_id,
                print_job_id=row.print_job_id,
                archive_exec_time=row.archive_exec_time,
                label_time=row.label_time,
                record_date=row.record_date,
                shift=row.shift,
                machine_code=row.machine_code,
                job_order=row.job_order,
                part_no=row.part_no,
                product_description=row.product_description,
                quantity=row.quantity,
                package_id=row.package_id,
                lot_batch_no=row.lot_batch_no,
                pallet_no=row.pallet_no,
                sequence_no=row.sequence_no,
                raw_xml=row.raw_xml,
            )
            for row in rows
        ]


def _aggregates_from_label_events(
    settings: Settings,
    events: Sequence[LabelProductionEvent],
) -> dict[tuple[str, str, str, str], ShiftAggregate]:
    with create_session(settings) as session:
        machines = session.scalars(select(Machine)).all()
        machine_ids = {
            normalize_machine_code(machine.machine_code): machine.id
            for machine in machines
        }

    aggregates: dict[tuple[str, str, str, str], ShiftAggregate] = {}
    for event in events:
        lot_batch_no = event.lot_batch_no or ""
        key = (
            event.record_date,
            event.machine_code,
            event.job_order,
            lot_batch_no,
        )
        aggregate = aggregates.setdefault(
            key,
            ShiftAggregate(
                record_date=event.record_date,
                machine_id=machine_ids.get(event.machine_code),
                machine_code=event.machine_code,
                job_order=event.job_order,
                product_hint=event.product_description,
                part_no_hint=event.part_no,
                lot_batch_no=event.lot_batch_no,
            ),
        )
        if event.shift in aggregate.shift_quantities:
            aggregate.shift_quantities[event.shift] += event.quantity
        if aggregate.product_hint is None:
            aggregate.product_hint = event.product_description
        if aggregate.part_no_hint is None:
            aggregate.part_no_hint = event.part_no
        aggregate.label_event_count += 1

    return aggregates


def _apply_amount_control_breakdowns(
    settings: Settings,
    start_date: Date,
    end_date: Date,
    aggregates: dict[tuple[str, str, str], ShiftAggregate],
) -> None:
    if not aggregates:
        return
    tzinfo = shifts.request_timezone(settings)
    with create_session(settings) as session:
        rows = session.scalars(
            select(AmountControlShift)
            .options(
                selectinload(AmountControlShift.machine),
                selectinload(AmountControlShift.machine_breakdowns),
            )
            .join(AmountControlShift.machine)
            .where(AmountControlShift.record_date >= start_date.isoformat())
            .where(AmountControlShift.record_date <= end_date.isoformat())
        ).all()

    for amount_shift in rows:
        machine_code = normalize_machine_code(amount_shift.machine.machine_code)
        job_order = _identifier(amount_shift.job_order)
        if machine_code is None or not job_order:
            continue
        aggregate = aggregates.get((amount_shift.record_date, machine_code, job_order))
        if aggregate is None:
            continue
        for breakdown in amount_shift.machine_breakdowns:
            for shift_label, minutes in _breakdown_minutes_by_shift(
                amount_shift.record_date,
                amount_shift.shift,
                breakdown,
                tzinfo,
            ).items():
                if shift_label in aggregate.shift_breakdown_minutes:
                    aggregate.shift_breakdown_minutes[shift_label] += minutes
                    aggregate.local_breakdown_minutes += minutes


def _apply_standalone_breakdowns(
    settings: Settings,
    start_date: Date,
    end_date: Date,
    aggregates_by_job: Mapping[tuple[str, str, str], ShiftAggregate],
    aggregates_by_shift: Mapping[tuple[str, str, str], ShiftAggregate],
) -> None:
    if not aggregates_by_job and not aggregates_by_shift:
        return
    tzinfo = shifts.request_timezone(settings)
    with create_session(settings) as session:
        rows = session.scalars(
            select(MachineBreakdown)
            .options(
                selectinload(MachineBreakdown.machine),
                selectinload(MachineBreakdown.entry).selectinload(
                    Entry.tour_context
                ),
            )
            .where(MachineBreakdown.amount_control_shift_id.is_(None))
        ).all()

    for breakdown in rows:
        context = _standalone_breakdown_context(breakdown, tzinfo)
        if context is None:
            continue
        breakdown_date = Date.fromisoformat(context.record_date)
        if breakdown_date < start_date or breakdown_date > end_date:
            continue
        minutes_by_shift = _breakdown_minutes_by_shift(
            context.record_date,
            context.fallback_shift,
            breakdown,
            tzinfo,
        )
        if context.job_order:
            aggregate = aggregates_by_job.get(
                (context.record_date, context.machine_code, context.job_order)
            )
            if aggregate is None:
                continue
            _add_breakdown_minutes(aggregate, minutes_by_shift)
            continue

        for shift_label, minutes in minutes_by_shift.items():
            aggregate = aggregates_by_shift.get(
                (context.record_date, context.machine_code, shift_label)
            )
            if aggregate is not None:
                _add_breakdown_minutes(aggregate, {shift_label: minutes})


def _standalone_breakdown_context(
    breakdown: MachineBreakdown,
    tzinfo: Any,
) -> StandaloneBreakdownContext | None:
    machine_code = normalize_machine_code(breakdown.machine.machine_code)
    if machine_code is None:
        return None

    record_date = _standalone_breakdown_record_date(breakdown, tzinfo)
    fallback_shift = _standalone_breakdown_shift(breakdown)
    has_timing = _has_usable_breakdown_timing(breakdown, tzinfo)
    if record_date is None or (
        fallback_shift not in shifts.SHIFT_OPTIONS and not has_timing
    ):
        return None

    return StandaloneBreakdownContext(
        record_date=record_date,
        machine_code=machine_code,
        fallback_shift=fallback_shift,
        job_order=_standalone_breakdown_job_order(breakdown),
    )


def _standalone_breakdown_record_date(
    breakdown: MachineBreakdown,
    tzinfo: Any,
) -> str | None:
    entry = breakdown.entry
    candidates = [
        getattr(breakdown, "record_date", None),
        entry.process_date if entry is not None else None,
        entry.tour_context.date
        if entry is not None and entry.tour_context is not None
        else None,
        entry.col_a if entry is not None else None,
    ]
    for candidate in candidates:
        record_date = _normalized_breakdown_record_date(candidate)
        if record_date:
            return record_date

    stopped = _stored_utc_to_local_timezone(breakdown.stopped_at, tzinfo)
    if stopped is not None:
        return stopped.date().isoformat()
    return None


def _normalized_breakdown_record_date(value: Any) -> str | None:
    record_date = normalize_process_date(value)
    if not record_date:
        return None
    try:
        Date.fromisoformat(record_date)
    except ValueError:
        return None
    return record_date


def _standalone_breakdown_shift(breakdown: MachineBreakdown) -> str:
    entry = breakdown.entry
    shift_label = _clean_text(getattr(breakdown, "shift", None))
    if shift_label in shifts.SHIFT_OPTIONS:
        return shift_label
    if entry is not None and entry.tour_context is not None:
        shift_label = _clean_text(entry.tour_context.shift)
        if shift_label in shifts.SHIFT_OPTIONS:
            return shift_label
    return ""


def _standalone_breakdown_job_order(breakdown: MachineBreakdown) -> str | None:
    job_order = _identifier(getattr(breakdown, "job_order", None))
    if job_order:
        return job_order
    if breakdown.entry is not None:
        return _identifier(breakdown.entry.col_h)
    return None


def _has_usable_breakdown_timing(
    breakdown: MachineBreakdown,
    tzinfo: Any,
) -> bool:
    stopped = _stored_utc_to_local_timezone(breakdown.stopped_at, tzinfo)
    resumed = _stored_utc_to_local_timezone(breakdown.resumed_at, tzinfo)
    return stopped is not None and resumed is not None and resumed > stopped


def _add_breakdown_minutes(
    aggregate: ShiftAggregate,
    minutes_by_shift: Mapping[str, int],
) -> None:
    for shift_label, minutes in minutes_by_shift.items():
        if shift_label in aggregate.shift_breakdown_minutes:
            aggregate.shift_breakdown_minutes[shift_label] += minutes
            aggregate.local_breakdown_minutes += minutes


def _breakdown_minutes_by_shift(
    record_date: str,
    fallback_shift: str,
    breakdown: MachineBreakdown,
    tzinfo: Any,
) -> dict[str, int]:
    duration = int(breakdown.duration_minutes or 0)
    if duration <= 0:
        return {}

    stopped = _stored_utc_to_local_timezone(breakdown.stopped_at, tzinfo)
    resumed = _stored_utc_to_local_timezone(breakdown.resumed_at, tzinfo)
    if stopped is None or resumed is None or resumed <= stopped:
        return (
            {fallback_shift: duration} if fallback_shift in shifts.SHIFT_OPTIONS else {}
        )

    minutes_by_shift: dict[str, int] = {}
    for shift_label in shifts.SHIFT_OPTIONS:
        start, finish = _shift_window_for_date(record_date, shift_label)
        overlap = _overlap_minutes(stopped, resumed, start, finish)
        minutes = int(overlap) if overlap is not None else 0
        if minutes > 0:
            minutes_by_shift[shift_label] = minutes
    return minutes_by_shift


def _read_shift_aggregates(
    settings: Settings,
    start_date: Date,
    end_date: Date,
) -> list[ShiftAggregate]:
    aggregates: dict[tuple[str, str, str], ShiftAggregate] = {}
    with create_session(settings) as session:
        rows = session.scalars(
            select(AmountControlShift)
            .options(
                selectinload(AmountControlShift.machine),
                selectinload(AmountControlShift.machine_breakdowns).selectinload(
                    MachineBreakdown.machine
                ),
            )
            .join(AmountControlShift.machine)
            .where(AmountControlShift.record_date >= start_date.isoformat())
            .where(AmountControlShift.record_date <= end_date.isoformat())
            .order_by(
                AmountControlShift.record_date.asc(),
                Machine.machine_code.asc(),
                AmountControlShift.job_order.asc(),
                AmountControlShift.shift.asc(),
            )
        ).all()

    for amount_shift in rows:
        machine_code = normalize_machine_code(amount_shift.machine.machine_code)
        if machine_code is None:
            continue
        job_order = _identifier(amount_shift.job_order)
        if not job_order:
            continue
        key = (amount_shift.record_date, machine_code, job_order)
        aggregate = aggregates.setdefault(
            key,
            ShiftAggregate(
                record_date=amount_shift.record_date,
                machine_id=amount_shift.machine_id,
                machine_code=machine_code,
                job_order=job_order,
            ),
        )
        if amount_shift.shift in aggregate.shift_quantities:
            aggregate.shift_quantities[amount_shift.shift] += (
                amount_shift.produced_quantity
            )
        aggregate.local_breakdown_minutes += sum(
            breakdown.duration_minutes
            for breakdown in amount_shift.machine_breakdowns
            if breakdown.duration_minutes
        )
        if aggregate.product_hint is None:
            aggregate.product_hint = _first_breakdown_product(
                amount_shift.machine_breakdowns
            )

    return sorted(
        aggregates.values(),
        key=lambda item: (
            item.record_date,
            _natural_machine_key(item.machine_code),
            item.job_order,
        ),
    )


def _read_process_metadata(
    settings: Settings,
    start_date: Date,
    end_date: Date,
) -> dict[tuple[str, str, str], ProcessMeta]:
    metadata: dict[tuple[str, str, str], ProcessMeta] = {}
    sort_keys: dict[tuple[str, str, str], tuple[datetime, int]] = {}
    with create_session(settings) as session:
        entries = session.scalars(
            select(Entry)
            .where(Entry.process_date >= start_date.isoformat())
            .where(Entry.process_date <= end_date.isoformat())
            .order_by(
                Entry.process_date.asc(), Entry.machine_code.asc(), Entry.id.asc()
            )
        ).all()

    for entry in entries:
        process_date = normalize_process_date(entry.process_date or entry.col_a)
        machine_code = normalize_machine_code(entry.machine_code or entry.col_f)
        job_order = _identifier(entry.col_h)
        if not process_date or not machine_code or not job_order:
            continue
        key = (process_date, machine_code, job_order)
        submitted_at = entry.submitted_at or entry.created_at
        sort_key = (submitted_at or datetime.min, int(entry.id or 0))
        if sort_key <= sort_keys.get(key, (datetime.min, 0)):
            continue
        sort_keys[key] = sort_key
        metadata[key] = ProcessMeta(
            product=_optional_text(entry.col_g),
            active_cavities=_decimal(entry.col_k),
            realized_cycle_time=_decimal(entry.col_l),
            raw_cycle_time=_optional_text(entry.col_l),
            submitted_at=submitted_at,
            entry_id=entry.id,
        )
    return metadata


def _load_ifs_actuals(
    settings: Settings,
    order_numbers: Sequence[str],
    *,
    refresh_ifs: bool,
) -> tuple[dict[tuple[str, str | None], IfsActual], dict[str, Any]]:
    summary: dict[str, Any] = {
        "ifs_refreshed": False,
        "ifs_error": None,
        "ifs_actual_count": 0,
        "ifs_cached_count": 0,
    }
    if refresh_ifs:
        try:
            rows = asyncio.run(
                fetch_shop_order_operation_actual_rows(settings, order_numbers)
            )
            actuals = normalize_ifs_actual_rows(rows)
            _cache_ifs_actuals(settings, actuals)
            summary["ifs_refreshed"] = True
            summary["ifs_actual_count"] = len(actuals)
        except Exception as exc:
            summary["ifs_error"] = str(exc) or exc.__class__.__name__

    cached_actuals = _read_cached_ifs_actuals(settings, order_numbers)
    summary["ifs_cached_count"] = len(cached_actuals)
    return {
        (actual.job_order, actual.machine_code): actual for actual in cached_actuals
    }, summary


def _cache_ifs_actuals(settings: Settings, actuals: Sequence[IfsActual]) -> None:
    if not actuals:
        return
    now = datetime.now(UTC).replace(tzinfo=None)
    session = create_session(settings)
    try:
        for actual in actuals:
            existing = session.scalars(
                select(ProductionLossIfsActual)
                .where(ProductionLossIfsActual.job_order == actual.job_order)
                .where(ProductionLossIfsActual.machine_code == actual.machine_code)
            ).first()
            if existing is None:
                existing = ProductionLossIfsActual(
                    job_order=actual.job_order,
                    machine_code=actual.machine_code,
                )
                session.add(existing)
            existing.release_no = actual.release_no
            existing.sequence_no = actual.sequence_no
            existing.operation_no = actual.operation_no
            existing.work_center_no = actual.work_center_no
            existing.part_no = actual.part_no
            existing.part_description = actual.part_description
            existing.actual_start_at = actual.actual_start_at
            existing.actual_finish_at = actual.actual_finish_at
            existing.net_machine_minutes = _decimal_text(actual.net_machine_minutes)
            existing.completed_quantity = _decimal_text(actual.completed_quantity)
            existing.scrap_quantity = _decimal_text(actual.scrap_quantity)
            existing.source = IFS_REALIZED_OPERATION_SOURCE
            existing.raw_json = json.dumps(
                actual.raw_rows, ensure_ascii=False, default=str
            )
            existing.source_updated_at = now
        session.commit()
    except Exception:
        session.rollback()
        raise
    finally:
        session.close()


def _read_cached_ifs_actuals(
    settings: Settings,
    order_numbers: Sequence[str],
) -> list[IfsActual]:
    order_set = {
        str(order_no).strip() for order_no in order_numbers if str(order_no).strip()
    }
    if not order_set:
        return []
    with create_session(settings) as session:
        cached = session.scalars(
            select(ProductionLossIfsActual)
            .where(ProductionLossIfsActual.job_order.in_(order_set))
            .where(ProductionLossIfsActual.source == IFS_REALIZED_OPERATION_SOURCE)
            .order_by(
                ProductionLossIfsActual.job_order.asc(),
                ProductionLossIfsActual.machine_code.asc(),
            )
        ).all()
        return [
            IfsActual(
                job_order=row.job_order,
                machine_code=row.machine_code,
                release_no=row.release_no,
                sequence_no=row.sequence_no,
                operation_no=row.operation_no,
                work_center_no=row.work_center_no,
                part_no=row.part_no,
                part_description=row.part_description,
                actual_start_at=row.actual_start_at,
                actual_finish_at=row.actual_finish_at,
                net_machine_minutes=_decimal(row.net_machine_minutes),
                completed_quantity=_decimal(row.completed_quantity),
                scrap_quantity=_decimal(row.scrap_quantity),
                raw_rows=_json_array(row.raw_json),
            )
            for row in cached
        ]


def _merge_ifs_actual(actual: IfsActual, row: Mapping[str, Any]) -> None:
    actual.raw_rows.append(dict(row))
    actual.release_no = actual.release_no or _first_text(
        row, ("ReleaseNo", "release_no")
    )
    actual.sequence_no = actual.sequence_no or _first_text(
        row,
        ("SequenceNo", "sequence_no"),
    )
    actual.operation_no = actual.operation_no or _int_value(
        _first_value(row, ("OperationNo", "operation_no"))
    )
    actual.work_center_no = actual.work_center_no or _first_text(
        row,
        ("WorkCenterNo", "work_center_no"),
    )
    actual.part_no = actual.part_no or _first_text(row, ("PartNo", "part_no"))
    actual.part_description = actual.part_description or _ifs_actual_part_description(row)
    actual.actual_start_at = _earliest(
        actual.actual_start_at,
        _first_datetime(row, DATETIME_FIELDS_START),
    )
    actual.actual_finish_at = _latest(
        actual.actual_finish_at,
        _first_datetime(row, DATETIME_FIELDS_FINISH),
    )
    actual.net_machine_minutes = _add_decimal(
        actual.net_machine_minutes,
        _duration_minutes(row, NET_MACHINE_DURATION_FIELDS),
    )
    actual.completed_quantity = _add_decimal(
        actual.completed_quantity,
        _first_decimal(row, COMPLETED_QTY_FIELDS),
    )
    actual.scrap_quantity = _add_decimal(
        actual.scrap_quantity,
        _first_decimal(row, SCRAP_QTY_FIELDS),
    )


def _optimum_quantity(
    minutes: Decimal | None,
    cycle_seconds: Decimal | None,
    active_cavities: Decimal | None,
) -> Decimal | None:
    if minutes is None or cycle_seconds is None or active_cavities is None:
        return None
    if cycle_seconds <= 0 or active_cavities <= 0:
        return None
    return (minutes * Decimal(60) / cycle_seconds) * active_cavities


def _sum_decimal_values(values: Iterable[Decimal | None]) -> Decimal | None:
    total: Decimal | None = None
    for value in values:
        if value is None:
            continue
        total = value if total is None else total + value
    return total


def _shift_calculations_for_row(
    settings: Settings,
    aggregate: ShiftAggregate,
    actual: IfsActual | None,
    cycle_time: Decimal | None,
    active_cavities: Decimal | None,
) -> dict[str, dict[str, Decimal | None]]:
    calculations: dict[str, dict[str, Decimal | None]] = {}
    for shift_label in shifts.SHIFT_OPTIONS:
        gross_minutes = _shift_gross_minutes(settings, aggregate, actual, shift_label)
        breakdown = Decimal(aggregate.breakdown_minutes_for_shift(shift_label))
        minutes = (
            max(gross_minutes - min(breakdown, gross_minutes), Decimal("0"))
            if gross_minutes is not None
            else None
        )
        optimum = _optimum_quantity(minutes, cycle_time, active_cavities)
        gross_optimum = _optimum_quantity(gross_minutes, cycle_time, active_cavities)
        actual_quantity = Decimal(aggregate.shift_quantities.get(shift_label, 0))
        if actual_quantity == 0:
            optimum = Decimal("0") if optimum is not None else None
            gross_optimum = Decimal("0") if gross_optimum is not None else None
        loss = optimum - actual_quantity if optimum is not None else None
        gross_loss = (
            gross_optimum - actual_quantity if gross_optimum is not None else None
        )
        calculations[shift_label] = {
            "gross_minutes": gross_minutes,
            "machine_minutes": minutes,
            "optimum": optimum,
            "loss": loss,
            "gross_optimum": gross_optimum,
            "gross_loss": gross_loss,
        }
    return calculations


def _shift_gross_minutes(
    settings: Settings,
    aggregate: ShiftAggregate,
    actual: IfsActual | None,
    shift_label: str,
) -> Decimal | None:
    if actual is None or actual.actual_start_at is None:
        return None
    actual_start = _stored_utc_to_local(settings, actual.actual_start_at)
    if actual_start is None:
        return None
    start, finish = _shift_window(settings, aggregate.record_date, shift_label)
    actual_finish = _stored_utc_to_local(settings, actual.actual_finish_at)
    if actual_finish is None:
        current_local = _current_local_datetime(settings)
        if current_local < finish:
            return None
        actual_finish = finish
    if actual_finish < actual_start:
        return None
    return _overlap_minutes(actual_start, actual_finish, start, finish)


def _current_local_datetime(settings: Settings) -> datetime:
    return datetime.now(shifts.request_timezone(settings)).replace(tzinfo=None)


def _stored_utc_to_local(settings: Settings, value: datetime | None) -> datetime | None:
    return _stored_utc_to_local_timezone(value, shifts.request_timezone(settings))


def _stored_utc_to_local_timezone(
    value: datetime | None,
    tzinfo: Any,
) -> datetime | None:
    if value is None:
        return None
    if value.tzinfo is not None:
        return value.astimezone(tzinfo).replace(tzinfo=None)
    return value.replace(tzinfo=UTC).astimezone(tzinfo).replace(tzinfo=None)


def _overlap_minutes(
    start: datetime,
    finish: datetime,
    window_start: datetime,
    window_finish: datetime,
) -> Decimal:
    overlap_start = max(start, window_start)
    overlap_finish = min(finish, window_finish)
    if overlap_finish <= overlap_start:
        return Decimal("0")
    return Decimal(str((overlap_finish - overlap_start).total_seconds())) / Decimal(60)


def _shift_window(
    settings: Settings,
    record_date: str,
    shift_label: str,
) -> tuple[datetime, datetime]:
    return _shift_window_for_date(record_date, shift_label)


def _shift_window_for_date(
    record_date: str,
    shift_label: str,
) -> tuple[datetime, datetime]:
    date_value = Date.fromisoformat(record_date)
    if shift_label == "24.00-08.00":
        start_time = time(0, 0)
        end_time = time(8, 0)
        end_date = date_value
    elif shift_label == "08.00-16.00":
        start_time = time(8, 0)
        end_time = time(16, 0)
        end_date = date_value
    else:
        start_time = time(16, 0)
        end_time = time(0, 0)
        end_date = date_value + timedelta(days=1)

    start = datetime.combine(date_value, start_time)
    finish = datetime.combine(end_date, end_time)
    return start, finish


def _write_workbook(output_path: Path, rows: Sequence[ProductionLossReportRow]) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Production Loss"
    worksheet.append(REPORT_HEADERS)

    for row in rows:
        worksheet.append(
            [
                row.record_date,
                row.machine_name,
                row.machine_code,
                row.job_order,
                row.lot_batch_no,
                row.product_description,
                _excel_decimal(row.gram),
                row.shift_2400_0800_quantity,
                _excel_decimal(row.shift_2400_0800_machine_minutes),
                _excel_decimal(row.shift_2400_0800_optimum_quantity),
                _excel_decimal(row.shift_2400_0800_loss_quantity),
                row.shift_0800_1600_quantity,
                _excel_decimal(row.shift_0800_1600_machine_minutes),
                _excel_decimal(row.shift_0800_1600_optimum_quantity),
                _excel_decimal(row.shift_0800_1600_loss_quantity),
                row.shift_1600_2400_quantity,
                _excel_decimal(row.shift_1600_2400_machine_minutes),
                _excel_decimal(row.shift_1600_2400_optimum_quantity),
                _excel_decimal(row.shift_1600_2400_loss_quantity),
                row.daily_total_quantity,
                _excel_decimal(row.active_cavities),
                _excel_decimal(row.cycle_time_seconds),
                _excel_decimal(row.net_machine_minutes),
                _excel_decimal(row.gross_elapsed_minutes),
                _excel_decimal(row.optimum_net_quantity),
                _excel_decimal(row.production_loss_net),
                _excel_decimal(row.optimum_gross_quantity),
                _excel_decimal(row.production_loss_gross),
                _excel_decimal(row.break_loss_quantity),
                row.local_breakdown_minutes,
                row.warnings,
            ]
        )

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(bottom=Side(style="thin", color="D9E2F3"))
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_cells in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row_cells:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    for column_cells in worksheet.columns:
        column_letter = column_cells[0].column_letter
        max_length = max(len(_clean_text(cell.value)) for cell in column_cells)
        worksheet.column_dimensions[column_letter].width = min(
            max(max_length + 2, 12),
            36,
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    workbook.save(output_path)
    workbook.close()


def _report_output_path(
    settings: Settings,
    report_id: int,
    start_date: Date,
    end_date: Date,
) -> Path:
    output_dir = Path(settings.report_output_dir).expanduser()
    return (
        output_dir
        / f"{start_date.isoformat()}_{end_date.isoformat()} Production Loss {report_id}.xlsx"
    )


def _date_range(date_from: str, date_to: str) -> tuple[Date, Date]:
    start = _iso_date(date_from, "date_from")
    end = _iso_date(date_to, "date_to")
    if end < start:
        raise ProductionLossReportError("date_to date_from degerinden once olamaz.")
    return start, end


def _iso_date(value: Any, field_name: str) -> Date:
    normalized = normalize_process_date(value)
    if not normalized:
        raise ProductionLossReportError(f"{field_name} zorunludur.")
    try:
        return Date.fromisoformat(normalized)
    except ValueError as exc:
        raise ProductionLossReportError(
            f"{field_name} YYYY-MM-DD formatinda olmalidir."
        ) from exc


def _matching_ifs_actual(
    aggregate: ShiftAggregate,
    actuals: dict[tuple[str, str | None], IfsActual],
    by_order: dict[str, list[IfsActual]],
) -> IfsActual | None:
    exact = actuals.get((aggregate.job_order, aggregate.machine_code))
    if exact is not None:
        return exact
    order_actuals = by_order.get(aggregate.job_order, [])
    if len(order_actuals) == 1:
        return order_actuals[0]
    machine_actuals = [
        actual
        for actual in order_actuals
        if normalize_machine_code(actual.machine_code) is not None
    ]
    if len(machine_actuals) == 1:
        return machine_actuals[0]
    return None


def _ifs_actuals_by_order(actuals: Iterable[IfsActual]) -> dict[str, list[IfsActual]]:
    by_order: dict[str, list[IfsActual]] = defaultdict(list)
    for actual in actuals:
        by_order[actual.job_order].append(actual)
    return by_order


def _gram_from_product_description(value: str) -> Decimal | None:
    match = GRAM_PATTERN.search(value)
    if match is None:
        return None
    return _decimal(match.group(1))


def _minutes_between(
    start: datetime | None,
    finish: datetime | None,
) -> Decimal | None:
    if start is None or finish is None or finish < start:
        return None
    return Decimal(str((finish - start).total_seconds())) / Decimal(60)


def _duration_minutes(
    row: Mapping[str, Any],
    field_names: Sequence[str],
) -> Decimal | None:
    for field_name in field_names:
        value = _first_value(row, (field_name,))
        number = _decimal(value)
        if number is None:
            continue
        lowered = field_name.casefold()
        if "second" in lowered:
            return number / Decimal(60)
        if "hour" in lowered:
            return number * Decimal(60)
        if field_name == "RealMachRunTime":
            return number
        return number
    return None


def _first_datetime(
    row: Mapping[str, Any],
    field_names: Sequence[str],
) -> datetime | None:
    for field_name in field_names:
        parsed = _datetime_value(_first_value(row, (field_name,)))
        if parsed is not None:
            return parsed
    return None


def _datetime_value(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        parsed = value
    elif value is None:
        return None
    else:
        text = str(value).strip()
        if not text:
            return None
        try:
            parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
        except ValueError:
            return None

    if parsed.tzinfo is not None:
        parsed = parsed.astimezone(UTC).replace(tzinfo=None)
    return parsed


def _first_decimal(
    row: Mapping[str, Any],
    field_names: Sequence[str],
) -> Decimal | None:
    for field_name in field_names:
        number = _decimal(_first_value(row, (field_name,)))
        if number is not None:
            return number
    return None


def _first_text(row: Mapping[str, Any], field_names: Sequence[str]) -> str | None:
    for field_name in field_names:
        text = _optional_text(_first_value(row, (field_name,)))
        if text is not None:
            return text
    return None


def _nested_text(row: Mapping[str, Any], field_name: str, nested_field: str) -> str | None:
    nested = row.get(field_name)
    if not isinstance(nested, Mapping):
        return None
    return _optional_text(nested.get(nested_field))


def _first_value(row: Mapping[str, Any], field_names: Sequence[str]) -> Any:
    for field_name in field_names:
        if field_name in row:
            return row[field_name]
    return None


def _first_breakdown_product(
    breakdowns: Sequence[MachineBreakdown],
) -> str | None:
    for breakdown in breakdowns:
        product = _optional_text(breakdown.produced_product)
        if product is not None:
            return product
    return None


def _identifier(value: Any) -> str:
    text = _clean_text(value)
    if re.fullmatch(r"\d+\.0+", text):
        return text.split(".", 1)[0]
    return text


def _decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, bool):
        return None
    if isinstance(value, int | float):
        return Decimal(str(value))
    text = _clean_text(value).replace(",", ".")
    if not text:
        return None
    try:
        number = Decimal(text)
    except InvalidOperation:
        return None
    if not number.is_finite():
        return None
    return number


def _positive_decimal(value: Any) -> Decimal | None:
    number = _decimal(value)
    if number is None or not number.is_finite() or number <= 0:
        return None
    return number


def _int_value(value: Any) -> int | None:
    if value is None:
        return None
    try:
        return int(value)
    except TypeError, ValueError:
        return None


def _add_decimal(left: Decimal | None, right: Decimal | None) -> Decimal | None:
    if right is None:
        return left
    if left is None:
        return right
    return left + right


def _earliest(left: datetime | None, right: datetime | None) -> datetime | None:
    if right is None:
        return left
    if left is None or right < left:
        return right
    return left


def _latest(left: datetime | None, right: datetime | None) -> datetime | None:
    if right is None:
        return left
    if left is None or right > left:
        return right
    return left


def _decimal_text(value: Decimal | None) -> str | None:
    if value is None:
        return None
    rounded = value.quantize(Decimal("0.01"))
    return format(rounded, "f").rstrip("0").rstrip(".") or "0"


def _excel_decimal(value: str | None) -> float | int | None:
    number = _decimal(value)
    if number is None:
        return None
    if number == number.to_integral_value():
        return int(number)
    return float(number)


def _json_object(value: str | None) -> dict[str, Any]:
    if not value:
        return {}
    try:
        payload = json.loads(value)
    except json.JSONDecodeError:
        return {}
    return payload if isinstance(payload, dict) else {}


def _json_array(value: str | None) -> list[dict[str, Any]]:
    if not value:
        return []
    try:
        payload = json.loads(value)
    except json.JSONDecodeError:
        return []
    if not isinstance(payload, list):
        return []
    return [row for row in payload if isinstance(row, dict)]


def _natural_machine_key(value: str) -> tuple[int, str]:
    try:
        return (0, f"{int(value):08d}")
    except TypeError, ValueError:
        return (1, str(value))
