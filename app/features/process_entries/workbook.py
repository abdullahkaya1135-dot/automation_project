import re
from collections.abc import Mapping, Sequence
from typing import Any

from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from ...core.config import Settings
from ...services.excel_service import (
    EXCEL_COLUMN_COUNT,
    EXCEL_COLUMN_LETTERS,
    HEADER_ROW,
    create_workbook_backup,
    detect_last_value_row,
    open_workbook_sheet,
    save_workbook,
    validate_headers,
)
from ...services.excel_write_lock import serialized_excel_write
from .fields import (
    ENTRY_FIELD_NAMES,
    EXCEL_FIELD_NAME_BY_COLUMN,
    blank_excluded_section_fields,
)

NUMERIC_COLUMN_LETTERS = frozenset(
    {
        "B",
        "J",
        "K",
        "L",
        "M",
        "N",
        "O",
        "P",
        "Q",
        "R",
        "S",
        "T",
        "U",
        "V",
        "W",
        "X",
        "Y",
    }
)
DECIMAL_NUMBER_PATTERN = re.compile(r"^[+-]?\d+(?:[,.]\d+)?$")
HYPHEN_RANGE_PATTERN = re.compile(r"\d\s*-\s*\d")


def normalize_excel_value(value: Any, column_letter: str) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        if HYPHEN_RANGE_PATTERN.search(stripped):
            return stripped
        if (
            column_letter in NUMERIC_COLUMN_LETTERS
            and DECIMAL_NUMBER_PATTERN.fullmatch(stripped)
        ):
            normalized = stripped.replace(",", ".")
            number = float(normalized)
            if number.is_integer() and "." not in normalized:
                return int(number)
            return number
        return stripped
    return value


def _normalized_existing_excel_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None
    return value


def _excel_values_match(existing: Any, expected: Any) -> bool:
    return _normalized_existing_excel_value(existing) == expected


def _row_matches_expected(
    existing_values: tuple[Any, ...],
    expected_values: list[Any],
) -> bool:
    return all(
        _excel_values_match(existing, expected)
        for existing, expected in zip(existing_values, expected_values, strict=True)
    )


def _existing_entry_row_lookup(
    worksheet: Worksheet,
    *,
    last_value_row: int,
) -> dict[tuple[Any, ...], int]:
    if last_value_row <= HEADER_ROW:
        return {}

    lookup: dict[tuple[Any, ...], int] = {}
    for row_number, values in enumerate(
        worksheet.iter_rows(
            min_row=HEADER_ROW + 1,
            max_row=last_value_row,
            max_col=EXCEL_COLUMN_COUNT,
            values_only=True,
        ),
        start=HEADER_ROW + 1,
    ):
        key = tuple(_normalized_existing_excel_value(value) for value in values)
        lookup.setdefault(key, row_number)
    return lookup


def _payload_value(payload: Mapping[str, Any] | object, field_name: str) -> Any:
    if isinstance(payload, Mapping):
        return payload.get(field_name)
    return getattr(payload, field_name, None)


def build_excel_row(payload: Mapping[str, Any] | object) -> list[Any]:
    if isinstance(payload, Mapping):
        section_payload = blank_excluded_section_fields(payload)
    else:
        section_payload = blank_excluded_section_fields(
            {
                field_name: getattr(payload, field_name, None)
                for field_name in ENTRY_FIELD_NAMES
            }
        )

    row_values: list[Any] = []
    for column_letter in EXCEL_COLUMN_LETTERS:
        value = _payload_value(
            section_payload,
            EXCEL_FIELD_NAME_BY_COLUMN[column_letter],
        )
        row_values.append(normalize_excel_value(value, column_letter))
    return row_values


def append_entry_to_workbook(
    settings: Settings,
    payload: Mapping[str, Any] | object,
    *,
    reuse_existing_match: bool = False,
) -> int:
    with serialized_excel_write("process_entry_append"):
        if reuse_existing_match:
            existing_row = _find_matching_entry_row_unlocked(settings, payload)
            if existing_row is not None:
                return existing_row
        return _append_entry_to_workbook_unlocked(settings, payload)


def append_entries_to_workbook(
    settings: Settings,
    payloads: Sequence[Mapping[str, Any] | object],
    *,
    reuse_existing_matches: bool = False,
) -> list[int]:
    if not payloads:
        return []

    with serialized_excel_write("process_entry_bulk_append"):
        return _append_entries_to_workbook_unlocked(
            settings,
            payloads,
            reuse_existing_matches=reuse_existing_matches,
        )


def _find_matching_entry_row_unlocked(
    settings: Settings,
    payload: Mapping[str, Any] | object,
) -> int | None:
    workbook: Workbook | None = None
    try:
        workbook, worksheet = open_workbook_sheet(settings)
        validate_headers(worksheet)
        last_value_row = detect_last_value_row(worksheet)
        if last_value_row <= HEADER_ROW:
            return None

        expected_values = build_excel_row(payload)
        for row_number, values in enumerate(
            worksheet.iter_rows(
                min_row=HEADER_ROW + 1,
                max_row=last_value_row,
                max_col=EXCEL_COLUMN_COUNT,
                values_only=True,
            ),
            start=HEADER_ROW + 1,
        ):
            if _row_matches_expected(values, expected_values):
                return row_number
        return None
    finally:
        if workbook is not None:
            workbook.close()


def _write_row_values(
    worksheet: Worksheet,
    row_number: int,
    row_values: list[Any],
) -> None:
    for column_index, value in enumerate(row_values, start=1):
        worksheet.cell(row=row_number, column=column_index, value=value)


def _append_entry_to_workbook_unlocked(
    settings: Settings,
    payload: Mapping[str, Any] | object,
) -> int:
    workbook: Workbook | None = None
    try:
        workbook, worksheet = open_workbook_sheet(
            settings,
            read_only=False,
            data_only=False,
        )
        validate_headers(worksheet)

        last_value_row = detect_last_value_row(worksheet)
        target_row = max(HEADER_ROW, last_value_row) + 1
        row_values = build_excel_row(payload)

        create_workbook_backup(settings)
        _write_row_values(worksheet, target_row, row_values)
        save_workbook(settings, workbook)

        return target_row
    finally:
        if workbook is not None:
            workbook.close()


def _append_entries_to_workbook_unlocked(
    settings: Settings,
    payloads: Sequence[Mapping[str, Any] | object],
    *,
    reuse_existing_matches: bool,
) -> list[int]:
    workbook: Workbook | None = None
    try:
        workbook, worksheet = open_workbook_sheet(
            settings,
            read_only=False,
            data_only=False,
        )
        validate_headers(worksheet)

        last_value_row = detect_last_value_row(worksheet)
        next_row = max(HEADER_ROW, last_value_row) + 1
        existing_rows = (
            _existing_entry_row_lookup(worksheet, last_value_row=last_value_row)
            if reuse_existing_matches
            else {}
        )

        row_numbers: list[int] = []
        rows_to_write: list[tuple[int, list[Any]]] = []
        for payload in payloads:
            row_values = build_excel_row(payload)
            existing_row = existing_rows.get(tuple(row_values))
            if existing_row is not None:
                row_numbers.append(existing_row)
                continue

            row_numbers.append(next_row)
            rows_to_write.append((next_row, row_values))
            next_row += 1

        if rows_to_write:
            create_workbook_backup(settings)
            for row_number, row_values in rows_to_write:
                _write_row_values(worksheet, row_number, row_values)
            save_workbook(settings, workbook)

        return row_numbers
    finally:
        if workbook is not None:
            workbook.close()
