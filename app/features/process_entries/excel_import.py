from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from ...core.config import Settings
from ...core.database import commit_session, create_session
from ...services.excel_service import (
    EXCEL_COLUMN_COUNT,
    EXCEL_COLUMN_LETTERS,
    HEADER_ROW,
    detect_last_value_row,
    normalize_excel_value,
    validate_headers,
)
from .fields import ENTRY_FIELD_NAMES
from .models import Entry
from .normalization import (
    normalize_machine_code,
    normalize_process_date,
    normalize_production_engineer_name,
    production_engineer_display_order,
)
from .service import apply_entry_process_metadata


@dataclass(frozen=True)
class ProcessExcelImportResult:
    source_path: Path
    sheet_name: str
    scanned_rows: int
    inserted_rows: int
    skipped_duplicates: int


def import_process_excel_to_database(
    settings: Settings,
    *,
    source_path: str | Path | None = None,
    sheet_name: str | None = None,
) -> ProcessExcelImportResult:
    path = Path(source_path or settings.excel_path)
    target_sheet_name = sheet_name or settings.sheet_name
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if target_sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sayfa bulunamadı: {target_sheet_name}")
        worksheet = workbook[target_sheet_name]
        validate_headers(worksheet)
        last_row = detect_last_value_row(worksheet)
        imported_payloads = []
        for values in worksheet.iter_rows(
            min_row=HEADER_ROW + 1,
            max_row=last_row,
            max_col=EXCEL_COLUMN_COUNT,
            values_only=True,
        ):
            payload = _payload_from_excel_values(values)
            if not any(payload.values()):
                continue
            imported_payloads.append(payload)
    finally:
        workbook.close()

    with create_session(settings) as session:
        existing_signatures = _existing_entry_signatures(session)
        sorted_payloads = sorted(imported_payloads, key=_payload_sort_key)
        inserted_rows = 0
        skipped_duplicates = 0
        for payload in sorted_payloads:
            signature = _entry_signature(payload)
            if signature in existing_signatures:
                skipped_duplicates += 1
                continue
            entry = Entry(sync_status="synced", **payload)
            apply_entry_process_metadata(session, entry)
            session.add(entry)
            session.flush()
            existing_signatures.add(_entry_signature(_entry_payload(entry)))
            inserted_rows += 1
        commit_session(session)

    return ProcessExcelImportResult(
        source_path=path,
        sheet_name=target_sheet_name,
        scanned_rows=len(imported_payloads),
        inserted_rows=inserted_rows,
        skipped_duplicates=skipped_duplicates,
    )


def _payload_from_excel_values(values: tuple[Any, ...]) -> dict[str, str | None]:
    payload: dict[str, str | None] = {}
    for field_name, column_letter, value in zip(
        ENTRY_FIELD_NAMES,
        EXCEL_COLUMN_LETTERS,
        values,
        strict=True,
    ):
        payload[field_name] = _stored_text(normalize_excel_value(value, column_letter))
    payload["col_c"] = normalize_production_engineer_name(payload.get("col_c"))
    return payload


def _stored_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = " ".join(str(value).strip().split())
    return text or None


def _entry_payload(entry: Entry) -> dict[str, str | None]:
    return {field_name: getattr(entry, field_name) for field_name in ENTRY_FIELD_NAMES}


def _entry_signature(payload: dict[str, str | None]) -> tuple[str, ...]:
    return tuple(payload.get(field_name) or "" for field_name in ENTRY_FIELD_NAMES)


def _existing_entry_signatures(session: Session) -> set[tuple[str, ...]]:
    return {
        _entry_signature(_entry_payload(entry))
        for entry in session.scalars(select(Entry)).all()
    }


def _payload_sort_key(payload: dict[str, str | None]) -> tuple[str, int, int, str]:
    process_date = normalize_process_date(payload.get("col_a")) or ""
    machine_code = normalize_machine_code(payload.get("col_f")) or "0"
    engineer_name = normalize_production_engineer_name(payload.get("col_c")) or ""
    return (
        process_date,
        int(machine_code) if machine_code.isdigit() else 0,
        production_engineer_display_order(engineer_name),
        engineer_name,
    )
