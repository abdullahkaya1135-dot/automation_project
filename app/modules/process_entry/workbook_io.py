import shutil
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from ...config import Settings
from ...shared.workbook_utils import (
    backup_filename,
    clean_text,
    normalize_header,
    prune_backups,
)
from ..workbook_rows import detect_last_value_row_in_columns
from .excel_schema import (
    EXCEL_COLUMN_COUNT,
    EXCEL_COLUMN_LETTERS,
    HEADER_KEYWORDS_BY_COLUMN,
    HEADER_ROW,
)


class ExcelServiceError(RuntimeError):
    """Base error for workbook access and validation failures."""


class WorkbookUnavailableError(ExcelServiceError):
    """Raised when the configured workbook cannot be opened."""


class WorkbookMissingError(WorkbookUnavailableError):
    """Raised when the configured workbook path does not exist."""


class WorkbookPermissionError(WorkbookUnavailableError):
    """Raised when the configured workbook cannot be read or written."""


class WorkbookLockedError(WorkbookUnavailableError):
    """Raised when the workbook appears locked by Excel or another process."""


class SheetNotFoundError(ExcelServiceError):
    """Raised when the configured sheet is absent from the workbook."""


class WorkbookStructureError(ExcelServiceError):
    """Raised when workbook columns no longer match the expected shape."""


class WorkbookSaveError(ExcelServiceError):
    """Raised when workbook save fails for an unexpected reason."""


@dataclass(frozen=True)
class ExcelStatus:
    available: bool
    error: str = ""


def open_workbook_sheet(
    settings: Settings,
    *,
    read_only: bool = True,
    data_only: bool = True,
) -> tuple[Workbook, Worksheet]:
    try:
        workbook = load_workbook(
            settings.excel_path,
            read_only=read_only,
            data_only=data_only,
        )
    except FileNotFoundError as exc:
        raise WorkbookMissingError(
            f"Çalışma kitabı bulunamadı: {settings.excel_path}"
        ) from exc
    except PermissionError as exc:
        raise WorkbookPermissionError(
            f"Çalışma kitabı izinler veya kilit nedeniyle okunamıyor: {settings.excel_path}"
        ) from exc
    except OSError as exc:
        raise WorkbookUnavailableError(
            f"Çalışma kitabı açılamadı: {settings.excel_path} ({exc})"
        ) from exc

    if settings.sheet_name not in workbook.sheetnames:
        workbook.close()
        raise SheetNotFoundError(f"Sayfa bulunamadı: {settings.sheet_name}")

    return workbook, workbook[settings.sheet_name]


def read_headers(worksheet: Worksheet) -> list[str]:
    return [
        clean_text(worksheet.cell(row=HEADER_ROW, column=column_index).value)
        for column_index in range(1, EXCEL_COLUMN_COUNT + 1)
    ]


def validate_headers(worksheet: Worksheet) -> list[str]:
    headers = read_headers(worksheet)
    missing_columns = [
        EXCEL_COLUMN_LETTERS[index]
        for index, header in enumerate(headers)
        if not header
    ]
    if missing_columns:
        raise WorkbookStructureError(
            "Çalışma kitabında A:Y başlıkları bulunmalıdır. Eksik başlık(lar): "
            + ", ".join(missing_columns)
        )

    normalized_headers = {
        column_letter: normalize_header(header)
        for column_letter, header in zip(EXCEL_COLUMN_LETTERS, headers, strict=True)
    }
    mismatches: list[str] = []
    for column_letter, keywords in HEADER_KEYWORDS_BY_COLUMN.items():
        header = normalized_headers[column_letter]
        if not any(keyword in header for keyword in keywords):
            mismatches.append(
                f"{column_letter} için beklenenlerden biri {keywords}, bulunan {headers[EXCEL_COLUMN_LETTERS.index(column_letter)]!r}"
            )
    if mismatches:
        raise WorkbookStructureError(
            "Çalışma kitabı başlıkları beklenen A:Y düzeniyle eşleşmiyor: "
            + "; ".join(mismatches)
        )

    return headers


def detect_last_value_row(worksheet: Worksheet) -> int:
    return detect_last_value_row_in_columns(worksheet, EXCEL_COLUMN_COUNT)


def create_workbook_backup(settings: Settings) -> Path:
    source_path = Path(settings.excel_path)
    backup_dir = Path(settings.backup_dir)
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup_path = backup_dir / backup_filename(source_path, "workbook")

    try:
        shutil.copyfile(source_path, backup_path)
    except FileNotFoundError as exc:
        raise WorkbookMissingError(
            f"Yedek oluşturulurken çalışma kitabı bulunamadı: {settings.excel_path}"
        ) from exc
    except PermissionError as exc:
        raise WorkbookPermissionError(
            f"İzinler veya kilit nedeniyle çalışma kitabı yedeği oluşturulamadı: {settings.excel_path}"
        ) from exc
    except OSError as exc:
        raise WorkbookUnavailableError(
            f"Çalışma kitabı yedeği oluşturulamadı: {settings.excel_path} ({exc})"
        ) from exc

    prune_backups(backup_dir, source_path, settings.backup_keep_count, "workbook")
    return backup_path


def check_excel_reachable(settings: Settings) -> ExcelStatus:
    workbook: Workbook | None = None
    try:
        workbook, worksheet = open_workbook_sheet(settings)
        validate_headers(worksheet)
        return ExcelStatus(available=True)
    except ExcelServiceError as exc:
        return ExcelStatus(available=False, error=str(exc))
    finally:
        if workbook is not None:
            workbook.close()
