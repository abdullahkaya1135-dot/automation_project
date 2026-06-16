from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .cycle_types import REPORT_HEADERS, CycleReportRow
from .cycle_values import clean_text, excel_value


def write_cycle_report(output_path: Path, rows: list[CycleReportRow]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Çevrim Kontrol"
    worksheet.append(REPORT_HEADERS)

    for row in rows:
        worksheet.append(
            [
                row.machine_group,
                row.machine_no,
                excel_value(row.neck_diameter),
                excel_value(row.gram),
                row.total_cavity_count,
                row.active_cavity_count,
                excel_value(row.actual_cycle_time),
                excel_value(row.optimum_cycle_time),
                row.control_status,
            ]
        )

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(bottom=Side(style="thin", color="D9E2F3"))
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    for column_cells in worksheet.columns:
        column_letter = column_cells[0].column_letter
        max_length = max(len(clean_text(cell.value)) for cell in column_cells)
        worksheet.column_dimensions[column_letter].width = min(
            max(max_length + 2, 12),
            34,
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    workbook.save(output_path)
    workbook.close()
