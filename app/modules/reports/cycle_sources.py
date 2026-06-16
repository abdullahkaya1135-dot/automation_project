import asyncio
from collections import defaultdict
from collections.abc import Callable, Coroutine
from datetime import date as Date
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from ...config import Settings
from ..process_entry.workbook_io import (
    detect_last_value_row,
    open_workbook_sheet,
    validate_headers,
)
from ..shop_orders.source import (
    ShopOrderOption,
    shop_order_options_from_ifs_operations,
)
from .cycle_types import CycleReportError, CycleTableEntry, ProcessCycleRow
from .cycle_values import identifier, normalize_machine_group, number, parse_date

SOURCE_COLUMN_COUNT = 12
OperationFetcher = Callable[[Settings], Coroutine[Any, Any, list[dict[str, Any]]]]


def read_process_rows(settings: Settings, report_date: Date) -> list[ProcessCycleRow]:
    workbook: Workbook | None = None
    try:
        workbook, worksheet = open_workbook_sheet(
            settings,
            read_only=False,
            data_only=True,
        )
        validate_headers(worksheet)
        last_row = detect_last_value_row(worksheet)
        rows: list[ProcessCycleRow] = []
        for row_number, values in enumerate(
            worksheet.iter_rows(
                min_row=2,
                max_row=last_row,
                max_col=SOURCE_COLUMN_COUNT,
                values_only=True,
            ),
            start=2,
        ):
            if parse_date(values[0]) != report_date:
                continue

            rows.append(
                ProcessCycleRow(
                    row_number=row_number,
                    machine_no=identifier(values[5]),
                    work_order=identifier(values[7]),
                    total_cavity_count=values[9],
                    active_cavity_count=values[10],
                    actual_cycle_time=number(values[11]),
                )
            )
        return rows
    finally:
        if workbook is not None:
            workbook.close()


def read_shop_order_options(
    settings: Settings,
    *,
    fetch_operations: OperationFetcher,
) -> list[ShopOrderOption]:
    try:
        rows = asyncio.run(fetch_operations(settings))
    except Exception as exc:
        raise CycleReportError(str(exc) or exc.__class__.__name__) from exc
    return shop_order_options_from_ifs_operations(rows)


def merged_cell_values(worksheet: Worksheet) -> dict[tuple[int, int], Any]:
    merged_values: dict[tuple[int, int], Any] = {}
    for merged_range in worksheet.merged_cells.ranges:
        min_column, min_row, max_column, max_row = merged_range.bounds
        value = worksheet.cell(min_row, min_column).value
        for row_index in range(min_row, max_row + 1):
            for column_index in range(min_column, max_column + 1):
                merged_values[(row_index, column_index)] = value
    return merged_values


def merged_value(
    worksheet: Worksheet,
    merged_values: dict[tuple[int, int], Any],
    row_index: int,
    column_index: int,
) -> Any:
    value = worksheet.cell(row_index, column_index).value
    if value is not None:
        return value
    return merged_values.get((row_index, column_index))


def read_cycle_table(
    settings: Settings,
) -> dict[tuple[str, Decimal, Decimal], list[CycleTableEntry]]:
    path = Path(settings.cycle_table_path).expanduser()
    try:
        workbook = load_workbook(path, read_only=False, data_only=True)
    except FileNotFoundError as exc:
        raise CycleReportError(f"Makine çevrim tablosu bulunamadı: {path}") from exc
    except OSError as exc:
        raise CycleReportError(f"Makine çevrim tablosu okunamadı: {path} ({exc})") from exc

    try:
        worksheet = workbook.worksheets[0]
        merged_values = merged_cell_values(worksheet)
        current_machine_no = ""
        current_group = ""
        current_neck: Decimal | None = None
        entries: dict[tuple[str, Decimal, Decimal], list[CycleTableEntry]] = (
            defaultdict(list)
        )

        for row_index in range(2, worksheet.max_row + 1):
            machine_no_value = merged_value(worksheet, merged_values, row_index, 1)
            group_value = merged_value(worksheet, merged_values, row_index, 2)
            neck_value = merged_value(worksheet, merged_values, row_index, 3)
            gram_value = merged_value(worksheet, merged_values, row_index, 4)
            cycle_value = merged_value(worksheet, merged_values, row_index, 7)

            machine_no = identifier(machine_no_value)
            previous_machine_no = current_machine_no
            if machine_no:
                current_machine_no = machine_no

            group_key = normalize_machine_group(group_value)
            previous_group = current_group
            if group_key:
                current_group = group_key

            neck = number(neck_value)
            if (
                neck is None
                and (
                    (machine_no and machine_no != previous_machine_no)
                    or (group_key and group_key != previous_group)
                )
            ):
                current_neck = None
            if neck is not None:
                current_neck = neck
            else:
                neck = current_neck

            gram = number(gram_value)
            cycle = number(cycle_value)
            if not current_group or neck is None or gram is None or cycle is None:
                continue

            key = (current_group, neck, gram)
            entries[key].append(
                CycleTableEntry(
                    machine_no=current_machine_no,
                    machine_group_key=current_group,
                    neck_diameter=neck,
                    gram=gram,
                    estimated_cycle_time=cycle,
                )
            )
        return dict(entries)
    finally:
        workbook.close()
