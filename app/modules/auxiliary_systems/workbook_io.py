import shutil
from copy import copy
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from ...config import Settings
from ...shared.workbook_utils import (
    backup_filename,
    clean_text,
    normalize_header,
    prune_backups,
)
from ..process_entry.workbook_io import (
    ExcelServiceError,
    SheetNotFoundError,
    WorkbookMissingError,
    WorkbookPermissionError,
    WorkbookStructureError,
    WorkbookUnavailableError,
)
from ..workbook_rows import detect_last_value_row_in_columns
from .fields import (
    AUXILIARY_COLUMN_COUNT,
    AUXILIARY_COLUMN_LETTERS,
    AUXILIARY_HEADER_KEYWORDS_BY_COLUMN,
    AUXILIARY_HEADER_ROW,
)


@dataclass(frozen=True)
class AuxiliarySystemsStatus:
    form_available: bool
    target_available: bool
    form_error: str = ""
    target_error: str = ""


def open_auxiliary_target_sheet(
    settings: Settings,
    *,
    read_only: bool = True,
    data_only: bool = True,
) -> tuple[Workbook, Worksheet]:
    try:
        workbook = load_workbook(
            settings.auxiliary_systems_target_path,
            read_only=read_only,
            data_only=data_only,
        )
    except FileNotFoundError as exc:
        raise WorkbookMissingError(
            "Yardımcı sistemler çalışma kitabı bulunamadı: "
            f"{settings.auxiliary_systems_target_path}"
        ) from exc
    except PermissionError as exc:
        raise WorkbookPermissionError(
            "Yardımcı sistemler çalışma kitabı izinler veya kilit nedeniyle "
            f"okunamıyor: {settings.auxiliary_systems_target_path}"
        ) from exc
    except OSError as exc:
        raise WorkbookUnavailableError(
            "Yardımcı sistemler çalışma kitabı açılamadı: "
            f"{settings.auxiliary_systems_target_path} ({exc})"
        ) from exc

    if settings.auxiliary_systems_sheet_name not in workbook.sheetnames:
        workbook.close()
        raise SheetNotFoundError(
            f"Yardımcı sistemler sayfası bulunamadı: "
            f"{settings.auxiliary_systems_sheet_name}"
        )

    return workbook, workbook[settings.auxiliary_systems_sheet_name]


def read_auxiliary_headers(worksheet: Worksheet) -> list[str]:
    return [
        clean_text(worksheet.cell(row=AUXILIARY_HEADER_ROW, column=column_index).value)
        for column_index in range(1, AUXILIARY_COLUMN_COUNT + 1)
    ]


def validate_auxiliary_headers(worksheet: Worksheet) -> list[str]:
    headers = read_auxiliary_headers(worksheet)
    missing_columns = [
        AUXILIARY_COLUMN_LETTERS[index]
        for index, header in enumerate(headers)
        if not header
    ]
    if missing_columns:
        raise WorkbookStructureError(
            "Yardımcı sistemler çalışma kitabında A:I başlıkları bulunmalıdır. "
            "Eksik başlık(lar): " + ", ".join(missing_columns)
        )

    normalized_headers = {
        column_letter: normalize_header(header)
        for column_letter, header in zip(
            AUXILIARY_COLUMN_LETTERS,
            headers,
            strict=True,
        )
    }
    mismatches: list[str] = []
    for column_letter, keywords in AUXILIARY_HEADER_KEYWORDS_BY_COLUMN.items():
        header = normalized_headers[column_letter]
        if not any(keyword in header for keyword in keywords):
            mismatches.append(
                f"{column_letter} için beklenenlerden biri {keywords}, "
                f"bulunan {headers[AUXILIARY_COLUMN_LETTERS.index(column_letter)]!r}"
            )
    if mismatches:
        raise WorkbookStructureError(
            "Yardımcı sistemler çalışma kitabı başlıkları beklenen A:I "
            "düzeniyle eşleşmiyor: " + "; ".join(mismatches)
        )

    return headers


def detect_auxiliary_last_value_row(worksheet: Worksheet) -> int:
    return detect_last_value_row_in_columns(worksheet, AUXILIARY_COLUMN_COUNT)


def create_auxiliary_workbook_backup(settings: Settings) -> Path:
    source_path = Path(settings.auxiliary_systems_target_path)
    backup_dir = Path(settings.backup_dir)
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup_path = backup_dir / backup_filename(source_path, "yardimci_sistemler")

    try:
        shutil.copyfile(source_path, backup_path)
    except FileNotFoundError as exc:
        raise WorkbookMissingError(
            "Yardımcı sistemler yedeği oluşturulurken çalışma kitabı bulunamadı: "
            f"{settings.auxiliary_systems_target_path}"
        ) from exc
    except PermissionError as exc:
        raise WorkbookPermissionError(
            "İzinler veya kilit nedeniyle yardımcı sistemler yedeği "
            f"oluşturulamadı: {settings.auxiliary_systems_target_path}"
        ) from exc
    except OSError as exc:
        raise WorkbookUnavailableError(
            "Yardımcı sistemler yedeği oluşturulamadı: "
            f"{settings.auxiliary_systems_target_path} ({exc})"
        ) from exc

    prune_backups(
        backup_dir,
        source_path,
        settings.backup_keep_count,
        "yardimci_sistemler",
    )
    return backup_path


def copy_auxiliary_row_format(
    worksheet: Worksheet,
    source_row: int,
    target_row: int,
) -> None:
    for column_index in range(1, AUXILIARY_COLUMN_COUNT + 1):
        source_cell = worksheet.cell(row=source_row, column=column_index)
        target_cell = worksheet.cell(row=target_row, column=column_index)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)
        if source_cell.border:
            target_cell.border = copy(source_cell.border)
        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)
        if source_cell.font:
            target_cell.font = copy(source_cell.font)
        if source_cell.protection:
            target_cell.protection = copy(source_cell.protection)

    source_dimension = worksheet.row_dimensions[source_row]
    target_dimension = worksheet.row_dimensions[target_row]
    target_dimension.height = source_dimension.height


def check_auxiliary_systems_reachable(settings: Settings) -> AuxiliarySystemsStatus:
    form_available = True
    form_error = ""
    if not Path(settings.auxiliary_systems_form_path).exists():
        form_available = False
        form_error = (
            "Yardımcı sistemler form dosyası bulunamadı: "
            f"{settings.auxiliary_systems_form_path}"
        )

    target_available = True
    target_error = ""
    workbook: Workbook | None = None
    try:
        workbook, worksheet = open_auxiliary_target_sheet(settings)
        validate_auxiliary_headers(worksheet)
    except ExcelServiceError as exc:
        target_available = False
        target_error = str(exc)
    finally:
        if workbook is not None:
            workbook.close()

    return AuxiliarySystemsStatus(
        form_available=form_available,
        target_available=target_available,
        form_error=form_error,
        target_error=target_error,
    )
