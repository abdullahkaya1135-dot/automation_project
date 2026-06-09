import re
import shutil
import unicodedata
from collections.abc import Mapping
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .config import Settings


EXCEL_COLUMN_COUNT = 25
HEADER_ROW = 1
EXCEL_COLUMN_LETTERS = tuple(
    get_column_letter(column_index)
    for column_index in range(1, EXCEL_COLUMN_COUNT + 1)
)
ENTRY_FIELD_NAMES = tuple(
    f"col_{letter}" for letter in "abcdefghijklmnopq"
)
EXCEL_FIELD_NAME_BY_COLUMN = {
    "A": "col_a",
    "B": "col_b",
    "C": "col_c",
    "D": "col_d",
    "E": "col_e",
    "F": "col_f",
    "G": None,
    "H": "col_g",
    "I": None,
    "J": "col_h",
    "K": "col_i",
    "L": "col_j",
    "M": "col_k",
    "N": "col_l",
    "O": "col_m",
    "P": "col_n",
    "Q": "col_o",
    "R": None,
    "S": None,
    "T": None,
    "U": None,
    "V": None,
    "W": None,
    "X": "col_p",
    "Y": "col_q",
}
NUMERIC_COLUMN_LETTERS = frozenset(
    {
        "B",
        "J",
        "K",
        "L",
        "M",
        "N",
        "O",
        "P",
        "Q",
        "R",
        "S",
        "T",
        "U",
        "V",
        "W",
        "X",
        "Y",
    }
)
HEADER_KEYWORDS_BY_COLUMN = {
    "A": ("date", "tarih"),
    "B": ("ambient", "ortam"),
    "C": ("production", "engineer", "uretim", "muh"),
    "D": ("shift chief", "vardiya amiri"),
    "E": ("shift", "vardiya"),
    "F": ("machine", "makine"),
    "G": ("product", "urun"),
    "H": ("work order", "is emri"),
    "I": ("raw material", "hammadde"),
    "J": ("total", "toplam"),
    "K": ("working", "active", "calisan"),
    "L": ("cycle", "cevrim"),
    "M": ("cooling", "sogutma"),
    "N": ("injection", "enjeksiyon"),
    "O": ("blow", "ufleme", "sisirme"),
    "P": ("conditioner", "sartlandirici"),
    "Q": ("dryer", "kurutucu"),
    "R": ("pressure", "basinc"),
    "S": ("speed", "hiz"),
    "T": ("holding", "utuleme", "zaman"),
    "U": ("holding", "utuleme", "hiz"),
    "V": ("holding", "utuleme", "basinc"),
    "W": ("clamp", "mengene"),
    "X": ("oven", "barrel", "ocak"),
    "Y": ("mold", "kalip"),
}
DECIMAL_NUMBER_PATTERN = re.compile(r"^[+-]?\d+(?:[,.]\d+)?$")
HYPHEN_RANGE_PATTERN = re.compile(r"\d\s*-\s*\d")
INVALID_BACKUP_FILENAME_CHARS = re.compile(r'[<>:"/\\|?*]+')


class ExcelServiceError(RuntimeError):
    """Base error for workbook access and validation failures."""


class WorkbookUnavailableError(ExcelServiceError):
    """Raised when the configured workbook cannot be opened."""


class WorkbookMissingError(WorkbookUnavailableError):
    """Raised when the configured workbook path does not exist."""


class WorkbookPermissionError(WorkbookUnavailableError):
    """Raised when the configured workbook cannot be read or written."""


class WorkbookLockedError(WorkbookUnavailableError):
    """Raised when the workbook appears locked by Excel or another process."""


class SheetNotFoundError(ExcelServiceError):
    """Raised when the configured sheet is absent from the workbook."""


class WorkbookStructureError(ExcelServiceError):
    """Raised when workbook columns no longer match the expected shape."""


class WorkbookSaveError(ExcelServiceError):
    """Raised when workbook save fails for an unexpected reason."""


@dataclass(frozen=True)
class ExcelStatus:
    available: bool
    error: str = ""


def open_workbook_sheet(
    settings: Settings,
    *,
    read_only: bool = True,
    data_only: bool = True,
) -> tuple[Workbook, Worksheet]:
    try:
        workbook = load_workbook(
            settings.excel_path,
            read_only=read_only,
            data_only=data_only,
        )
    except FileNotFoundError as exc:
        raise WorkbookMissingError(
            f"Çalışma kitabı bulunamadı: {settings.excel_path}"
        ) from exc
    except PermissionError as exc:
        raise WorkbookPermissionError(
            f"Çalışma kitabı izinler veya kilit nedeniyle okunamıyor: {settings.excel_path}"
        ) from exc
    except OSError as exc:
        raise WorkbookUnavailableError(
            f"Çalışma kitabı açılamadı: {settings.excel_path} ({exc})"
        ) from exc

    if settings.sheet_name not in workbook.sheetnames:
        workbook.close()
        raise SheetNotFoundError(f"Sayfa bulunamadı: {settings.sheet_name}")

    return workbook, workbook[settings.sheet_name]


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _has_meaningful_value(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    return True


def _normalize_header(value: Any) -> str:
    text = _clean_text(value).casefold()
    text = "".join(
        character
        for character in unicodedata.normalize("NFKD", text)
        if not unicodedata.combining(character)
    )
    translation = str.maketrans(
        {
            "\u00e7": "c",
            "\u011f": "g",
            "\u0131": "i",
            "\u00f6": "o",
            "\u015f": "s",
            "\u00fc": "u",
        }
    )
    normalized = text.translate(translation)
    normalized = re.sub(r"[^a-z0-9]+", " ", normalized)
    return re.sub(r"\s+", " ", normalized).strip()


def read_headers(worksheet: Worksheet) -> list[str]:
    return [
        _clean_text(worksheet.cell(row=HEADER_ROW, column=column_index).value)
        for column_index in range(1, EXCEL_COLUMN_COUNT + 1)
    ]


def validate_headers(worksheet: Worksheet) -> list[str]:
    headers = read_headers(worksheet)
    missing_columns = [
        EXCEL_COLUMN_LETTERS[index]
        for index, header in enumerate(headers)
        if not header
    ]
    if missing_columns:
        raise WorkbookStructureError(
            "Çalışma kitabında A:Y başlıkları bulunmalıdır. Eksik başlık(lar): "
            + ", ".join(missing_columns)
        )

    normalized_headers = {
        column_letter: _normalize_header(header)
        for column_letter, header in zip(EXCEL_COLUMN_LETTERS, headers, strict=True)
    }
    mismatches: list[str] = []
    for column_letter, keywords in HEADER_KEYWORDS_BY_COLUMN.items():
        header = normalized_headers[column_letter]
        if not any(keyword in header for keyword in keywords):
            mismatches.append(
                f"{column_letter} için beklenenlerden biri {keywords}, bulunan {headers[EXCEL_COLUMN_LETTERS.index(column_letter)]!r}"
            )
    if mismatches:
        raise WorkbookStructureError(
            "Çalışma kitabı başlıkları beklenen A:Y düzeniyle eşleşmiyor: "
            + "; ".join(mismatches)
        )

    return headers


def detect_last_value_row(worksheet: Worksheet) -> int:
    cells = getattr(worksheet, "_cells", None)
    if cells is None:
        last_row = 0
        for row_index, row in enumerate(
            worksheet.iter_rows(
                min_row=1,
                max_col=EXCEL_COLUMN_COUNT,
                values_only=True,
            ),
            start=1,
        ):
            if any(_has_meaningful_value(value) for value in row):
                last_row = row_index
        return last_row

    if not cells:
        return 0

    last_row = 0
    for (row_index, column_index), cell in cells.items():
        if column_index > EXCEL_COLUMN_COUNT:
            continue
        if _has_meaningful_value(cell.value):
            last_row = max(last_row, row_index)
    return last_row


def normalize_excel_value(value: Any, column_letter: str) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        if HYPHEN_RANGE_PATTERN.search(stripped):
            return stripped
        if (
            column_letter in NUMERIC_COLUMN_LETTERS
            and DECIMAL_NUMBER_PATTERN.fullmatch(stripped)
        ):
            normalized = stripped.replace(",", ".")
            number = float(normalized)
            if number.is_integer() and "." not in normalized:
                return int(number)
            return number
        return stripped
    return value


def _payload_value(payload: Mapping[str, Any] | object, field_name: str | None) -> Any:
    if field_name is None:
        return None
    if isinstance(payload, Mapping):
        return payload.get(field_name)
    return getattr(payload, field_name, None)


def build_excel_row(payload: Mapping[str, Any] | object) -> list[Any]:
    row_values: list[Any] = []
    for column_letter in EXCEL_COLUMN_LETTERS:
        value = _payload_value(payload, EXCEL_FIELD_NAME_BY_COLUMN[column_letter])
        row_values.append(normalize_excel_value(value, column_letter))
    return row_values


def _backup_filename(source_path: Path) -> str:
    safe_stem = INVALID_BACKUP_FILENAME_CHARS.sub("_", source_path.stem).strip()
    if not safe_stem:
        safe_stem = "workbook"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    return f"{safe_stem}_{timestamp}{source_path.suffix or '.xlsx'}"


def _prune_backups(backup_dir: Path, source_path: Path, keep_count: int) -> None:
    safe_stem = INVALID_BACKUP_FILENAME_CHARS.sub("_", source_path.stem).strip()
    pattern = f"{safe_stem}_*{source_path.suffix or '.xlsx'}"
    backups = sorted(backup_dir.glob(pattern), reverse=True)
    for old_backup in backups[keep_count:]:
        old_backup.unlink(missing_ok=True)


def create_workbook_backup(settings: Settings) -> Path:
    source_path = Path(settings.excel_path)
    backup_dir = Path(settings.backup_dir)
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup_path = backup_dir / _backup_filename(source_path)

    try:
        shutil.copyfile(source_path, backup_path)
    except FileNotFoundError as exc:
        raise WorkbookMissingError(
            f"Yedek oluşturulurken çalışma kitabı bulunamadı: {settings.excel_path}"
        ) from exc
    except PermissionError as exc:
        raise WorkbookPermissionError(
            f"İzinler veya kilit nedeniyle çalışma kitabı yedeği oluşturulamadı: {settings.excel_path}"
        ) from exc
    except OSError as exc:
        raise WorkbookUnavailableError(
            f"Çalışma kitabı yedeği oluşturulamadı: {settings.excel_path} ({exc})"
        ) from exc

    _prune_backups(backup_dir, source_path, settings.backup_keep_count)
    return backup_path


def append_entry_to_workbook(
    settings: Settings,
    payload: Mapping[str, Any] | object,
) -> int:
    workbook: Workbook | None = None
    try:
        workbook, worksheet = open_workbook_sheet(
            settings,
            read_only=False,
            data_only=False,
        )
        validate_headers(worksheet)

        last_value_row = detect_last_value_row(worksheet)
        target_row = max(HEADER_ROW, last_value_row) + 1
        row_values = build_excel_row(payload)

        create_workbook_backup(settings)
        for column_index, value in enumerate(row_values, start=1):
            worksheet.cell(row=target_row, column=column_index, value=value)

        try:
            workbook.save(settings.excel_path)
        except PermissionError as exc:
            raise WorkbookLockedError(
                f"Çalışma kitabı kilitli olabileceği için kaydedilemedi: {settings.excel_path}"
            ) from exc
        except OSError as exc:
            raise WorkbookSaveError(
                f"Çalışma kitabı kaydedilemedi: {settings.excel_path} ({exc})"
            ) from exc

        return target_row
    finally:
        if workbook is not None:
            workbook.close()


def check_excel_reachable(settings: Settings) -> ExcelStatus:
    workbook: Workbook | None = None
    try:
        workbook, worksheet = open_workbook_sheet(settings)
        validate_headers(worksheet)
        return ExcelStatus(available=True)
    except ExcelServiceError as exc:
        return ExcelStatus(available=False, error=str(exc))
    finally:
        if workbook is not None:
            workbook.close()
