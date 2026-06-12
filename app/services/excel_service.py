import re
import shutil
from collections.abc import Mapping
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..config import Settings
from ..domain.entry_fields import (
    ENTRY_FIELD_NAMES,
    EXCEL_FIELD_NAME_BY_COLUMN,
    blank_excluded_section_fields,
)
from .excel_write_lock import serialized_excel_write
from .workbook_utils import (
    backup_filename,
    clean_text,
    has_meaningful_value,
    normalize_header,
    prune_backups,
)

EXCEL_COLUMN_COUNT = 25
HEADER_ROW = 1
EXCEL_COLUMN_LETTERS = tuple(
    get_column_letter(column_index)
    for column_index in range(1, EXCEL_COLUMN_COUNT + 1)
)
NUMERIC_COLUMN_LETTERS = frozenset(
    {
        "B",
        "J",
        "K",
        "L",
        "M",
        "N",
        "O",
        "P",
        "Q",
        "R",
        "S",
        "T",
        "U",
        "V",
        "W",
        "X",
        "Y",
    }
)
HEADER_KEYWORDS_BY_COLUMN = {
    "A": ("date", "tarih"),
    "B": ("ambient", "ortam"),
    "C": ("production", "engineer", "uretim", "muh"),
    "D": ("shift chief", "vardiya amiri"),
    "E": ("shift", "vardiya"),
    "F": ("machine", "makine"),
    "G": ("product", "urun"),
    "H": ("work order", "is emri"),
    "I": ("raw material", "hammadde"),
    "J": ("total", "toplam"),
    "K": ("working", "active", "calisan"),
    "L": ("cycle", "cevrim"),
    "M": ("cooling", "sogutma"),
    "N": ("injection", "enjeksiyon"),
    "O": ("blow", "ufleme", "sisirme"),
    "P": ("conditioner", "sartlandirici"),
    "Q": ("dryer", "kurutucu"),
    "R": ("pressure", "basinc"),
    "S": ("speed", "hiz"),
    "T": ("holding", "utuleme", "zaman"),
    "U": ("holding", "utuleme", "hiz"),
    "V": ("holding", "utuleme", "basinc"),
    "W": ("clamp", "mengene"),
    "X": ("oven", "barrel", "ocak"),
    "Y": ("mold", "kalip"),
}
DECIMAL_NUMBER_PATTERN = re.compile(r"^[+-]?\d+(?:[,.]\d+)?$")
HYPHEN_RANGE_PATTERN = re.compile(r"\d\s*-\s*\d")


class ExcelServiceError(RuntimeError):
    """Base error for workbook access and validation failures."""


class WorkbookUnavailableError(ExcelServiceError):
    """Raised when the configured workbook cannot be opened."""


class WorkbookMissingError(WorkbookUnavailableError):
    """Raised when the configured workbook path does not exist."""


class WorkbookPermissionError(WorkbookUnavailableError):
    """Raised when the configured workbook cannot be read or written."""


class WorkbookLockedError(WorkbookUnavailableError):
    """Raised when the workbook appears locked by Excel or another process."""


class SheetNotFoundError(ExcelServiceError):
    """Raised when the configured sheet is absent from the workbook."""


class WorkbookStructureError(ExcelServiceError):
    """Raised when workbook columns no longer match the expected shape."""


class WorkbookSaveError(ExcelServiceError):
    """Raised when workbook save fails for an unexpected reason."""


@dataclass(frozen=True)
class ExcelStatus:
    available: bool
    error: str = ""


def open_workbook_sheet(
    settings: Settings,
    *,
    read_only: bool = True,
    data_only: bool = True,
) -> tuple[Workbook, Worksheet]:
    try:
        workbook = load_workbook(
            settings.excel_path,
            read_only=read_only,
            data_only=data_only,
        )
    except FileNotFoundError as exc:
        raise WorkbookMissingError(
            f"Çalışma kitabı bulunamadı: {settings.excel_path}"
        ) from exc
    except PermissionError as exc:
        raise WorkbookPermissionError(
            f"Çalışma kitabı izinler veya kilit nedeniyle okunamıyor: {settings.excel_path}"
        ) from exc
    except OSError as exc:
        raise WorkbookUnavailableError(
            f"Çalışma kitabı açılamadı: {settings.excel_path} ({exc})"
        ) from exc

    if settings.sheet_name not in workbook.sheetnames:
        workbook.close()
        raise SheetNotFoundError(f"Sayfa bulunamadı: {settings.sheet_name}")

    return workbook, workbook[settings.sheet_name]


def read_headers(worksheet: Worksheet) -> list[str]:
    return [
        clean_text(worksheet.cell(row=HEADER_ROW, column=column_index).value)
        for column_index in range(1, EXCEL_COLUMN_COUNT + 1)
    ]


def validate_headers(worksheet: Worksheet) -> list[str]:
    headers = read_headers(worksheet)
    missing_columns = [
        EXCEL_COLUMN_LETTERS[index]
        for index, header in enumerate(headers)
        if not header
    ]
    if missing_columns:
        raise WorkbookStructureError(
            "Çalışma kitabında A:Y başlıkları bulunmalıdır. Eksik başlık(lar): "
            + ", ".join(missing_columns)
        )

    normalized_headers = {
        column_letter: normalize_header(header)
        for column_letter, header in zip(EXCEL_COLUMN_LETTERS, headers, strict=True)
    }
    mismatches: list[str] = []
    for column_letter, keywords in HEADER_KEYWORDS_BY_COLUMN.items():
        header = normalized_headers[column_letter]
        if not any(keyword in header for keyword in keywords):
            mismatches.append(
                f"{column_letter} için beklenenlerden biri {keywords}, bulunan {headers[EXCEL_COLUMN_LETTERS.index(column_letter)]!r}"
            )
    if mismatches:
        raise WorkbookStructureError(
            "Çalışma kitabı başlıkları beklenen A:Y düzeniyle eşleşmiyor: "
            + "; ".join(mismatches)
        )

    return headers


def detect_last_value_row(worksheet: Worksheet) -> int:
    cells = getattr(worksheet, "_cells", None)
    if cells is None:
        last_row = 0
        for row_index, row in enumerate(
            worksheet.iter_rows(
                min_row=1,
                max_col=EXCEL_COLUMN_COUNT,
                values_only=True,
            ),
            start=1,
        ):
            if any(has_meaningful_value(value) for value in row):
                last_row = row_index
        return last_row

    if not cells:
        return 0

    last_row = 0
    for (row_index, column_index), cell in cells.items():
        if column_index > EXCEL_COLUMN_COUNT:
            continue
        if has_meaningful_value(cell.value):
            last_row = max(last_row, row_index)
    return last_row


def normalize_excel_value(value: Any, column_letter: str) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        if HYPHEN_RANGE_PATTERN.search(stripped):
            return stripped
        if (
            column_letter in NUMERIC_COLUMN_LETTERS
            and DECIMAL_NUMBER_PATTERN.fullmatch(stripped)
        ):
            normalized = stripped.replace(",", ".")
            number = float(normalized)
            if number.is_integer() and "." not in normalized:
                return int(number)
            return number
        return stripped
    return value


def _normalized_existing_excel_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None
    return value


def _excel_values_match(existing: Any, expected: Any) -> bool:
    return _normalized_existing_excel_value(existing) == expected


def _row_matches_expected(
    existing_values: tuple[Any, ...],
    expected_values: list[Any],
) -> bool:
    return all(
        _excel_values_match(existing, expected)
        for existing, expected in zip(existing_values, expected_values, strict=True)
    )


def _payload_value(payload: Mapping[str, Any] | object, field_name: str) -> Any:
    if isinstance(payload, Mapping):
        return payload.get(field_name)
    return getattr(payload, field_name, None)


def build_excel_row(payload: Mapping[str, Any] | object) -> list[Any]:
    if isinstance(payload, Mapping):
        section_payload = blank_excluded_section_fields(payload)
    else:
        section_payload = blank_excluded_section_fields(
            {
                field_name: getattr(payload, field_name, None)
                for field_name in ENTRY_FIELD_NAMES
            }
        )

    row_values: list[Any] = []
    for column_letter in EXCEL_COLUMN_LETTERS:
        value = _payload_value(
            section_payload,
            EXCEL_FIELD_NAME_BY_COLUMN[column_letter],
        )
        row_values.append(normalize_excel_value(value, column_letter))
    return row_values


def create_workbook_backup(settings: Settings) -> Path:
    source_path = Path(settings.excel_path)
    backup_dir = Path(settings.backup_dir)
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup_path = backup_dir / backup_filename(source_path, "workbook")

    try:
        shutil.copyfile(source_path, backup_path)
    except FileNotFoundError as exc:
        raise WorkbookMissingError(
            f"Yedek oluşturulurken çalışma kitabı bulunamadı: {settings.excel_path}"
        ) from exc
    except PermissionError as exc:
        raise WorkbookPermissionError(
            f"İzinler veya kilit nedeniyle çalışma kitabı yedeği oluşturulamadı: {settings.excel_path}"
        ) from exc
    except OSError as exc:
        raise WorkbookUnavailableError(
            f"Çalışma kitabı yedeği oluşturulamadı: {settings.excel_path} ({exc})"
        ) from exc

    prune_backups(backup_dir, source_path, settings.backup_keep_count, "workbook")
    return backup_path


def append_entry_to_workbook(
    settings: Settings,
    payload: Mapping[str, Any] | object,
    *,
    reuse_existing_match: bool = False,
) -> int:
    with serialized_excel_write("process_entry_append"):
        if reuse_existing_match:
            existing_row = _find_matching_entry_row_unlocked(settings, payload)
            if existing_row is not None:
                return existing_row
        return _append_entry_to_workbook_unlocked(settings, payload)


def _find_matching_entry_row_unlocked(
    settings: Settings,
    payload: Mapping[str, Any] | object,
) -> int | None:
    workbook: Workbook | None = None
    try:
        workbook, worksheet = open_workbook_sheet(settings)
        validate_headers(worksheet)
        last_value_row = detect_last_value_row(worksheet)
        if last_value_row <= HEADER_ROW:
            return None

        expected_values = build_excel_row(payload)
        for row_number, values in enumerate(
            worksheet.iter_rows(
                min_row=HEADER_ROW + 1,
                max_row=last_value_row,
                max_col=EXCEL_COLUMN_COUNT,
                values_only=True,
            ),
            start=HEADER_ROW + 1,
        ):
            if _row_matches_expected(values, expected_values):
                return row_number
        return None
    finally:
        if workbook is not None:
            workbook.close()


def _append_entry_to_workbook_unlocked(
    settings: Settings,
    payload: Mapping[str, Any] | object,
) -> int:
    workbook: Workbook | None = None
    try:
        workbook, worksheet = open_workbook_sheet(
            settings,
            read_only=False,
            data_only=False,
        )
        validate_headers(worksheet)

        last_value_row = detect_last_value_row(worksheet)
        target_row = max(HEADER_ROW, last_value_row) + 1
        row_values = build_excel_row(payload)

        create_workbook_backup(settings)
        for column_index, value in enumerate(row_values, start=1):
            worksheet.cell(row=target_row, column=column_index, value=value)

        try:
            workbook.save(settings.excel_path)
        except PermissionError as exc:
            raise WorkbookLockedError(
                f"Çalışma kitabı kilitli olabileceği için kaydedilemedi: {settings.excel_path}"
            ) from exc
        except OSError as exc:
            raise WorkbookSaveError(
                f"Çalışma kitabı kaydedilemedi: {settings.excel_path} ({exc})"
            ) from exc

        return target_row
    finally:
        if workbook is not None:
            workbook.close()


def check_excel_reachable(settings: Settings) -> ExcelStatus:
    workbook: Workbook | None = None
    try:
        workbook, worksheet = open_workbook_sheet(settings)
        validate_headers(worksheet)
        return ExcelStatus(available=True)
    except ExcelServiceError as exc:
        return ExcelStatus(available=False, error=str(exc))
    finally:
        if workbook is not None:
            workbook.close()
