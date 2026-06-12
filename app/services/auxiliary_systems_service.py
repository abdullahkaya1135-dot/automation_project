import re
import shutil
from collections.abc import Mapping
from copy import copy
from dataclasses import dataclass
from datetime import date as Date
from datetime import datetime
from pathlib import Path
from typing import Any, TypedDict

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..config import Settings
from .excel_service import (
    ExcelServiceError,
    SheetNotFoundError,
    WorkbookLockedError,
    WorkbookMissingError,
    WorkbookPermissionError,
    WorkbookSaveError,
    WorkbookStructureError,
    WorkbookUnavailableError,
)
from .excel_write_lock import serialized_excel_write
from .workbook_utils import (
    backup_filename,
    clean_text,
    has_meaningful_value,
    normalize_header,
    prune_backups,
)

AUXILIARY_COLUMN_COUNT = 9
AUXILIARY_HEADER_ROW = 1
AUXILIARY_COLUMN_LETTERS = tuple(
    get_column_letter(column_index)
    for column_index in range(1, AUXILIARY_COLUMN_COUNT + 1)
)
AUXILIARY_MEASUREMENT_FIELD_NAMES = (
    "tower_frequency",
    "tower_set_pressure",
    "tower_feedback_pressure",
    "chiller_motor_frequency",
    "chiller_motor_set_pressure",
    "chiller_motor_feedback_pressure",
    "termokar_chiller_1_temp_set",
    "termokar_chiller_1_inlet_temp",
    "termokar_chiller_1_outlet_temp",
    "termokar_chiller_2_temp_set",
    "termokar_chiller_2_inlet_temp",
    "termokar_chiller_2_outlet_temp",
    "planer_temp_set",
    "planer_inlet_temp",
    "planer_outlet_temp",
    "itech_temp_set",
    "itech_current_temp",
    "compressor_high_708_pressure",
    "compressor_high_709_pressure",
    "compressor_high_710_pressure",
    "compressor_high_711_pressure",
    "compressor_low_712_pressure",
    "compressor_low_713_pressure",
    "compressor_low_714_pressure",
    "compressor_low_715_pressure",
    "compressor_low_716_pressure",
)
AUXILIARY_CHECK_FIELD_NAMES = (
    "oil_cooling_water_tank_checked",
    "chiller_water_tank_checked",
    "air_tank_1_drained",
    "air_tank_2_drained",
    "cleanliness_checked",
)
AUXILIARY_FIELD_NAMES = (
    *AUXILIARY_MEASUREMENT_FIELD_NAMES,
    *AUXILIARY_CHECK_FIELD_NAMES,
)


class AuxiliaryRowSpec(TypedDict):
    machine: str
    fields: dict[str, str]


AUXILIARY_ROW_SPECS: tuple[AuxiliaryRowSpec, ...] = (
    {
        "machine": "1- ELEKTROMOTOR (KULE)",
        "fields": {
            "C": "tower_frequency",
            "D": "tower_set_pressure",
            "E": "tower_feedback_pressure",
        },
    },
    {
        "machine": "2- ELEKTROMOTOR (CHİLLER)",
        "fields": {
            "C": "chiller_motor_frequency",
            "D": "chiller_motor_set_pressure",
            "E": "chiller_motor_feedback_pressure",
        },
    },
    {
        "machine": "TERMOKAR CHİLLER - 1",
        "fields": {
            "F": "termokar_chiller_1_temp_set",
            "G": "termokar_chiller_1_inlet_temp",
            "H": "termokar_chiller_1_outlet_temp",
        },
    },
    {
        "machine": "TERMOKAR CHİLLER - 2",
        "fields": {
            "F": "termokar_chiller_2_temp_set",
            "G": "termokar_chiller_2_inlet_temp",
            "H": "termokar_chiller_2_outlet_temp",
        },
    },
    {
        "machine": "PLANER SOĞUTUCU",
        "fields": {
            "F": "planer_temp_set",
            "G": "planer_inlet_temp",
            "H": "planer_outlet_temp",
        },
    },
    {
        "machine": "ITECH SOĞUTUCU",
        "fields": {
            "F": "itech_temp_set",
            "G": "itech_current_temp",
        },
    },
    {
        "machine": "YÜKSEK BASINÇ 708",
        "fields": {"I": "compressor_high_708_pressure"},
    },
    {
        "machine": "YÜKSEK BASINÇ 709",
        "fields": {"I": "compressor_high_709_pressure"},
    },
    {
        "machine": "YÜKSEK BASINÇ 710",
        "fields": {"I": "compressor_high_710_pressure"},
    },
    {
        "machine": "YÜKSEK BASINÇ 711",
        "fields": {"I": "compressor_high_711_pressure"},
    },
    {
        "machine": "ALÇAK BASINÇ 712",
        "fields": {"I": "compressor_low_712_pressure"},
    },
    {
        "machine": "ALÇAK BASINÇ 713",
        "fields": {"I": "compressor_low_713_pressure"},
    },
    {
        "machine": "ALÇAK BASINÇ 714",
        "fields": {"I": "compressor_low_714_pressure"},
    },
    {
        "machine": "ALÇAK BASINÇ 715 (150 ÖZEN)",
        "fields": {"I": "compressor_low_715_pressure"},
    },
    {
        "machine": "ALÇAK BASINÇ 716 (ATLAS)",
        "fields": {"I": "compressor_low_716_pressure"},
    },
)
AUXILIARY_HEADER_KEYWORDS_BY_COLUMN = {
    "A": ("tarih",),
    "B": ("makine",),
    "C": ("cikis frekansi", "frekans"),
    "D": ("set degeri",),
    "E": ("geri besleme",),
    "F": ("isi set",),
    "G": ("giris isisi", "giris isi"),
    "H": ("cikis isisi", "cikis isi"),
    "I": ("komp basinc", "basinc degeri"),
}
DECIMAL_NUMBER_PATTERN = re.compile(r"^[+-]?\d+(?:[,.]\d+)?$")


@dataclass(frozen=True)
class AuxiliarySystemsStatus:
    form_available: bool
    target_available: bool
    form_error: str = ""
    target_error: str = ""


def open_auxiliary_target_sheet(
    settings: Settings,
    *,
    read_only: bool = True,
    data_only: bool = True,
) -> tuple[Workbook, Worksheet]:
    try:
        workbook = load_workbook(
            settings.auxiliary_systems_target_path,
            read_only=read_only,
            data_only=data_only,
        )
    except FileNotFoundError as exc:
        raise WorkbookMissingError(
            "Yardımcı sistemler çalışma kitabı bulunamadı: "
            f"{settings.auxiliary_systems_target_path}"
        ) from exc
    except PermissionError as exc:
        raise WorkbookPermissionError(
            "Yardımcı sistemler çalışma kitabı izinler veya kilit nedeniyle "
            f"okunamıyor: {settings.auxiliary_systems_target_path}"
        ) from exc
    except OSError as exc:
        raise WorkbookUnavailableError(
            "Yardımcı sistemler çalışma kitabı açılamadı: "
            f"{settings.auxiliary_systems_target_path} ({exc})"
        ) from exc

    if settings.auxiliary_systems_sheet_name not in workbook.sheetnames:
        workbook.close()
        raise SheetNotFoundError(
            f"Yardımcı sistemler sayfası bulunamadı: "
            f"{settings.auxiliary_systems_sheet_name}"
        )

    return workbook, workbook[settings.auxiliary_systems_sheet_name]


def read_auxiliary_headers(worksheet: Worksheet) -> list[str]:
    return [
        clean_text(worksheet.cell(row=AUXILIARY_HEADER_ROW, column=column_index).value)
        for column_index in range(1, AUXILIARY_COLUMN_COUNT + 1)
    ]


def validate_auxiliary_headers(worksheet: Worksheet) -> list[str]:
    headers = read_auxiliary_headers(worksheet)
    missing_columns = [
        AUXILIARY_COLUMN_LETTERS[index]
        for index, header in enumerate(headers)
        if not header
    ]
    if missing_columns:
        raise WorkbookStructureError(
            "Yardımcı sistemler çalışma kitabında A:I başlıkları bulunmalıdır. "
            "Eksik başlık(lar): " + ", ".join(missing_columns)
        )

    normalized_headers = {
        column_letter: normalize_header(header)
        for column_letter, header in zip(
            AUXILIARY_COLUMN_LETTERS,
            headers,
            strict=True,
        )
    }
    mismatches: list[str] = []
    for column_letter, keywords in AUXILIARY_HEADER_KEYWORDS_BY_COLUMN.items():
        header = normalized_headers[column_letter]
        if not any(keyword in header for keyword in keywords):
            mismatches.append(
                f"{column_letter} için beklenenlerden biri {keywords}, "
                f"bulunan {headers[AUXILIARY_COLUMN_LETTERS.index(column_letter)]!r}"
            )
    if mismatches:
        raise WorkbookStructureError(
            "Yardımcı sistemler çalışma kitabı başlıkları beklenen A:I "
            "düzeniyle eşleşmiyor: " + "; ".join(mismatches)
        )

    return headers


def detect_auxiliary_last_value_row(worksheet: Worksheet) -> int:
    cells = getattr(worksheet, "_cells", None)
    if cells is None:
        last_row = 0
        for row_index, row in enumerate(
            worksheet.iter_rows(
                min_row=1,
                max_col=AUXILIARY_COLUMN_COUNT,
                values_only=True,
            ),
            start=1,
        ):
            if any(has_meaningful_value(value) for value in row):
                last_row = row_index
        return last_row

    if not cells:
        return 0

    last_row = 0
    for (row_index, column_index), cell in cells.items():
        if column_index > AUXILIARY_COLUMN_COUNT:
            continue
        if has_meaningful_value(cell.value):
            last_row = max(last_row, row_index)
    return last_row


def normalize_auxiliary_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, int | float):
        return value
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        if DECIMAL_NUMBER_PATTERN.fullmatch(stripped):
            normalized = stripped.replace(",", ".")
            number = float(normalized)
            if number.is_integer() and "." not in normalized:
                return int(number)
            return number
        return stripped
    return value


def _normalized_existing_auxiliary_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None
    return value


def _auxiliary_values_match(existing: Any, expected: Any) -> bool:
    return _normalized_existing_auxiliary_value(existing) == expected


def _auxiliary_row_matches_expected(
    existing_values: tuple[Any, ...],
    expected_values: list[Any],
) -> bool:
    return all(
        _auxiliary_values_match(existing, expected)
        for existing, expected in zip(existing_values, expected_values, strict=True)
    )


def date_text(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, Date):
        return value.strftime("%d.%m.%Y")
    text = clean_text(value)
    if not text:
        raise ValueError("tarih zorunludur.")
    try:
        return datetime.fromisoformat(text).strftime("%d.%m.%Y")
    except ValueError:
        return text


def _payload_value(payload: Mapping[str, Any] | object, field_name: str) -> Any:
    if isinstance(payload, Mapping):
        return payload.get(field_name)
    return getattr(payload, field_name, None)


def build_auxiliary_rows(
    payload: Mapping[str, Any] | object,
    recorded_date: Any,
) -> list[list[Any]]:
    rows: list[list[Any]] = []
    recorded_date_text = date_text(recorded_date)
    for spec in AUXILIARY_ROW_SPECS:
        row_values = [recorded_date_text, spec["machine"], None, None, None, None, None, None, None]
        fields = spec["fields"]
        for column_letter, field_name in fields.items():
            column_index = AUXILIARY_COLUMN_LETTERS.index(column_letter)
            row_values[column_index] = normalize_auxiliary_value(
                _payload_value(payload, field_name)
            )
        rows.append(row_values)
    return rows


def create_auxiliary_workbook_backup(settings: Settings) -> Path:
    source_path = Path(settings.auxiliary_systems_target_path)
    backup_dir = Path(settings.backup_dir)
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup_path = backup_dir / backup_filename(source_path, "yardimci_sistemler")

    try:
        shutil.copyfile(source_path, backup_path)
    except FileNotFoundError as exc:
        raise WorkbookMissingError(
            "Yardımcı sistemler yedeği oluşturulurken çalışma kitabı bulunamadı: "
            f"{settings.auxiliary_systems_target_path}"
        ) from exc
    except PermissionError as exc:
        raise WorkbookPermissionError(
            "İzinler veya kilit nedeniyle yardımcı sistemler yedeği "
            f"oluşturulamadı: {settings.auxiliary_systems_target_path}"
        ) from exc
    except OSError as exc:
        raise WorkbookUnavailableError(
            "Yardımcı sistemler yedeği oluşturulamadı: "
            f"{settings.auxiliary_systems_target_path} ({exc})"
        ) from exc

    prune_backups(
        backup_dir,
        source_path,
        settings.backup_keep_count,
        "yardimci_sistemler",
    )
    return backup_path


def _copy_row_format(
    worksheet: Worksheet,
    source_row: int,
    target_row: int,
) -> None:
    for column_index in range(1, AUXILIARY_COLUMN_COUNT + 1):
        source_cell = worksheet.cell(row=source_row, column=column_index)
        target_cell = worksheet.cell(row=target_row, column=column_index)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)
        if source_cell.border:
            target_cell.border = copy(source_cell.border)
        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)
        if source_cell.font:
            target_cell.font = copy(source_cell.font)
        if source_cell.protection:
            target_cell.protection = copy(source_cell.protection)

    source_dimension = worksheet.row_dimensions[source_row]
    target_dimension = worksheet.row_dimensions[target_row]
    target_dimension.height = source_dimension.height


def append_auxiliary_systems_to_workbook(
    settings: Settings,
    payload: Mapping[str, Any] | object,
    recorded_date: Any,
    *,
    reuse_existing_match: bool = False,
) -> tuple[int, int]:
    with serialized_excel_write("auxiliary_systems_append"):
        if reuse_existing_match:
            existing_block = _find_matching_auxiliary_block_unlocked(
                settings,
                payload,
                recorded_date,
            )
            if existing_block is not None:
                return existing_block
        return _append_auxiliary_systems_to_workbook_unlocked(
            settings,
            payload,
            recorded_date,
        )


def _find_matching_auxiliary_block_unlocked(
    settings: Settings,
    payload: Mapping[str, Any] | object,
    recorded_date: Any,
) -> tuple[int, int] | None:
    workbook: Workbook | None = None
    try:
        workbook, worksheet = open_auxiliary_target_sheet(settings)
        validate_auxiliary_headers(worksheet)
        last_value_row = detect_auxiliary_last_value_row(worksheet)
        expected_rows = build_auxiliary_rows(payload, recorded_date)
        expected_count = len(expected_rows)
        if last_value_row < AUXILIARY_HEADER_ROW + expected_count:
            return None

        max_start_row = last_value_row - expected_count + 1
        for start_row in range(AUXILIARY_HEADER_ROW + 1, max_start_row + 1):
            matched = True
            for row_offset, expected_values in enumerate(expected_rows):
                existing_values = next(
                    worksheet.iter_rows(
                        min_row=start_row + row_offset,
                        max_row=start_row + row_offset,
                        max_col=AUXILIARY_COLUMN_COUNT,
                        values_only=True,
                    )
                )
                if not _auxiliary_row_matches_expected(
                    existing_values,
                    expected_values,
                ):
                    matched = False
                    break
            if matched:
                return start_row, start_row + expected_count - 1
        return None
    finally:
        if workbook is not None:
            workbook.close()


def _append_auxiliary_systems_to_workbook_unlocked(
    settings: Settings,
    payload: Mapping[str, Any] | object,
    recorded_date: Any,
) -> tuple[int, int]:
    workbook: Workbook | None = None
    try:
        workbook, worksheet = open_auxiliary_target_sheet(
            settings,
            read_only=False,
            data_only=False,
        )
        validate_auxiliary_headers(worksheet)

        rows = build_auxiliary_rows(payload, recorded_date)
        last_value_row = detect_auxiliary_last_value_row(worksheet)
        target_start_row = max(AUXILIARY_HEADER_ROW, last_value_row) + 1
        template_start_row = target_start_row - len(rows)

        create_auxiliary_workbook_backup(settings)
        for row_offset, row_values in enumerate(rows):
            target_row = target_start_row + row_offset
            template_row = template_start_row + row_offset
            if template_row > AUXILIARY_HEADER_ROW:
                _copy_row_format(worksheet, template_row, target_row)
            for column_index, value in enumerate(row_values, start=1):
                worksheet.cell(row=target_row, column=column_index, value=value)

        try:
            workbook.save(settings.auxiliary_systems_target_path)
        except PermissionError as exc:
            raise WorkbookLockedError(
                "Yardımcı sistemler çalışma kitabı kilitli olabileceği için "
                f"kaydedilemedi: {settings.auxiliary_systems_target_path}"
            ) from exc
        except OSError as exc:
            raise WorkbookSaveError(
                "Yardımcı sistemler çalışma kitabı kaydedilemedi: "
                f"{settings.auxiliary_systems_target_path} ({exc})"
            ) from exc

        return target_start_row, target_start_row + len(rows) - 1
    finally:
        if workbook is not None:
            workbook.close()


def check_auxiliary_systems_reachable(settings: Settings) -> AuxiliarySystemsStatus:
    form_available = True
    form_error = ""
    if not Path(settings.auxiliary_systems_form_path).exists():
        form_available = False
        form_error = (
            "Yardımcı sistemler form dosyası bulunamadı: "
            f"{settings.auxiliary_systems_form_path}"
        )

    target_available = True
    target_error = ""
    workbook: Workbook | None = None
    try:
        workbook, worksheet = open_auxiliary_target_sheet(settings)
        validate_auxiliary_headers(worksheet)
    except ExcelServiceError as exc:
        target_available = False
        target_error = str(exc)
    finally:
        if workbook is not None:
            workbook.close()

    return AuxiliarySystemsStatus(
        form_available=form_available,
        target_available=target_available,
        form_error=form_error,
        target_error=target_error,
    )
